"""Certificación, avance físico y facturación desde el circuito Cert_* (v8_6).

El modelo viejo (2_Certificaciones, ledger plano) se reemplazó por 5 hojas
relacionales. Este módulo las lee y arma:

  - items: ledger de certificados (Cert_Calculo agrupado por ID Certif, unido a
    Cert_Facturacion por ID Cert+Fact para estado/cobro). Mismo SHAPE que el
    contrato viejo (id/base/cac/iva/total/estado/cobrado/saldo) para que el
    frontend no cambie.
  - agregados oficiales (total certificado / facturado / cobrado / pendientes y
    % avance físico) desde la fila TOTAL de Cert_Control_OC (rollup cacheado y
    autoritativo; reconcilia el circuito).
  - avance por OC (≈ presupuesto/etapa): una fila por OC de Cert_Control_OC.
  - apertura fiscal: Blanco/Negro desde Cert_Calculo.

Nota: el avance físico (% de obra ejecutada) es Cert_Control_OC col E, NO
certificado$/venta — el certificado financiero incluye CAC y anticipos.
"""
from . import nav
from openpyxl.utils.datetime import from_excel


def _to_iso(v):
    if v is None or v == "":
        return None
    if hasattr(v, "date"):            # datetime
        return v.date().isoformat()
    if hasattr(v, "isoformat"):       # date
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # serial de Excel (GDR/SVD guardan la fecha de cert como número, sin formato fecha)
        try:
            dt = from_excel(v)
            return dt.date().isoformat() if hasattr(dt, "date") else str(dt)
        except Exception:
            return str(v)
    return str(v)


# ----------------------------------------------------- Cert_Control_OC (rollup)
def _control_oc(wb):
    """(total, ocs) desde Cert_Control_OC. total = fila TOTAL; ocs = filas por OC.
    Cada uno: {avance, certificado, facturado, cobrado, contrato, ...}."""
    try:
        ws = wb["Cert_Control_OC"]
    except KeyError:
        return {}, []
    hr = nav.find_row(ws, "id oc") or 1
    col = nav.map_columns(ws, hr, {
        "id_oc":     {"all": ["id", "oc"]},
        "presup":    {"all": ["presupuesto"]},
        "desc":      {"all": ["descripcion"]},
        "contrato":  {"all": ["contrato"]},
        "avance":    {"all": ["avance", "fisico"]},
        "avance_sc": {"all": ["avance", "certificar"]},   # col J '$ Avance s/certificar' (Gap A)
        "certificado": {"all": ["certificado"], "not": ["%", "anticipo"]},
        "anticipo":  {"all": ["anticipo"]},
        "facturado": {"all": ["facturado"]},
        "cobrado":   {"all": ["cobrado"]},
        "saldo_fact": {"all": ["saldo", "facturar"]},
        "saldo_cobr": {"all": ["saldo", "cobrar"]},
        "estado":    {"all": ["estado"]},
    })

    def row(r):
        g = lambda k: nav.num(ws.cell(row=r, column=col[k]).value) if k in col else 0.0
        return {
            "avance": g("avance"), "avance_sin_cert": g("avance_sc"),
            "certificado": g("certificado"),
            "anticipo": g("anticipo"),
            "facturado": g("facturado"), "cobrado": g("cobrado"),
            "contrato": g("contrato"),
            "saldo_facturar": g("saldo_fact"), "saldo_cobrar": g("saldo_cobr"),
        }

    total, ocs = {}, []
    for r in range(hr + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=col["id_oc"]).value if "id_oc" in col else None
        an = nav.norm(a)
        if an.startswith("total"):
            total = row(r)
            break  # las filas siguientes son la sección "PENDIENTES DE CERTIFICAR"
        if not an or an.startswith("▌") or an.startswith("#"):
            continue
        rec = row(r)
        rec["id_oc"] = str(a).strip()
        rec["presupuesto"] = (str(ws.cell(row=r, column=col["presup"]).value).strip()
                              if "presup" in col and ws.cell(row=r, column=col["presup"]).value else str(a).strip())
        rec["descripcion"] = (str(ws.cell(row=r, column=col["desc"]).value).strip()
                              if "desc" in col and ws.cell(row=r, column=col["desc"]).value else "")
        ocs.append(rec)
    return total, ocs


# ----------------------------------------------------- Cert_Calculo + Facturacion (ledger)
def _ledger(wb):
    """items del ledger: un certificado por ID Certif. Cada item trae el desglose por parte
    fiscal (`partes`: Blanco/Negro con base/CAC/IVA/USD/total/facturado/cobrado), unido a
    Cert_Facturacion. Devuelve (items, fiscal, oc_bn) — oc_bn = B/N del certificado por OC."""
    try:
        wc = wb["Cert_Calculo"]
    except KeyError:
        return [], [], {}
    col = nav.map_columns(wc, 1, {
        "id":      {"all": ["id", "certif"], "not": ["fact"]},
        "idfact":  {"all": ["cert", "fact"]},
        "tipo":    {"all": ["tipo"]},
        "oc":      {"all": ["id", "oc"]},
        "fecha":   {"all": ["fecha"]},
        "base":    {"all": ["base", "neta"]},
        "cac":     {"all": ["cac"], "not": ["indice", "ratio"]},
        "iva":     {"all": ["iva"], "not": ["%"]},
        "total":   {"all": ["total"], "not": ["base", "u$", "usd"]},
        "usd":     {"all": ["u$"]},                 # 'U$ Total' (sólo filas Negro)
        "pfact":   {"all": ["facturacion"]},         # % facturación (split B/N)
        "iva_pct": {"all": ["%", "iva"]},            # % IVA
        "cac_ratio": {"all": ["ratio", "cac"]},      # índice de actualización CAC
        "tc":      {"any": ["mep"]},                 # TC USD (MEP a la fecha)
    })
    fact = _facturacion(wb)  # idfact -> {tipo, monto, cobrado, retencion, conciliado, comprob, fecha_cobro}

    def _parte(tipo):
        return {"tipo": tipo, "base": 0.0, "cac": 0.0, "iva": 0.0, "usd": 0.0, "total": 0.0,
                "facturado": 0.0, "cobrado": 0.0, "retencion": 0.0, "comprob": "",
                "pct_fact": 0.0, "pct_iva": 0.0}

    grp = {}    # id certif -> acumulador
    oc_res = {}  # id_oc -> resumen del certificado por OC (anticipo/cert B/N, CAC, IVA)
    for r in range(2, wc.max_row + 1):
        idv = wc.cell(row=r, column=col["id"]).value if "id" in col else None
        if idv is None or str(idv).strip() == "":
            continue
        idc = str(idv).strip()
        idfact = str(wc.cell(row=r, column=col["idfact"]).value or "").strip() if "idfact" in col else ""
        tipo = str(wc.cell(row=r, column=col["tipo"]).value or "").strip() if "tipo" in col else ""
        oc = str(wc.cell(row=r, column=col["oc"]).value or "").strip() if "oc" in col else ""
        base = nav.num(wc.cell(row=r, column=col["base"]).value) if "base" in col else 0.0
        cac = nav.num(wc.cell(row=r, column=col["cac"]).value) if "cac" in col else 0.0
        iva = nav.num(wc.cell(row=r, column=col["iva"]).value) if "iva" in col else 0.0
        total = nav.num(wc.cell(row=r, column=col["total"]).value) if "total" in col else 0.0
        usd = nav.num(wc.cell(row=r, column=col["usd"]).value) if "usd" in col else 0.0
        pfact = nav.num(wc.cell(row=r, column=col["pfact"]).value) if "pfact" in col else 0.0
        ivap = nav.num(wc.cell(row=r, column=col["iva_pct"]).value) if "iva_pct" in col else 0.0
        ratio = nav.num(wc.cell(row=r, column=col["cac_ratio"]).value) if "cac_ratio" in col else 0.0
        tc = nav.num(wc.cell(row=r, column=col["tc"]).value) if "tc" in col else 0.0
        g = grp.setdefault(idc, {
            "id": idc, "oc": oc,
            "fecha": _to_iso(wc.cell(row=r, column=col["fecha"]).value) if "fecha" in col else None,
            "base": 0.0, "cac": 0.0, "iva": 0.0, "usd": 0.0, "total": 0.0,
            "facturado": 0.0, "monto_cobrado": 0.0, "retencion": 0.0,
            "fecha_cobro": None, "nro_factura": "", "partes": {},
            "cac_ratio": 0.0, "tc_usd": 0.0,
        })
        g["base"] += base; g["cac"] += cac; g["iva"] += iva; g["usd"] += usd; g["total"] += total
        if ratio:
            g["cac_ratio"] = ratio
        if tc:
            g["tc_usd"] = tc
        p = g["partes"].setdefault(tipo, _parte(tipo))
        p["base"] += base; p["cac"] += cac; p["iva"] += iva; p["usd"] += usd; p["total"] += total
        if pfact:
            p["pct_fact"] = pfact
        if ivap:
            p["pct_iva"] = ivap
        f = fact.get(idfact)
        if f:
            g["facturado"] += f["monto"]; g["monto_cobrado"] += f["cobrado"]; g["retencion"] += f["retencion"]
            p["facturado"] += f["monto"]; p["cobrado"] += f["cobrado"]; p["retencion"] += f["retencion"]
            if not p["comprob"] and f["comprob"]:
                p["comprob"] = f["comprob"]
            if tipo.lower().startswith("b") and f["comprob"] and not g["nro_factura"]:
                g["nro_factura"] = f["comprob"]
            if f["fecha_cobro"] and (g["fecha_cobro"] is None or f["fecha_cobro"] > g["fecha_cobro"]):
                g["fecha_cobro"] = f["fecha_cobro"]
        # resumen por OC (para el drill de OC): descomposición ADITIVA del total certificado por
        # tipo fiscal — base anticipo + base avance + CAC + IVA = total certificado.
        if oc:
            o = oc_res.setdefault(oc, {"ant_b": 0.0, "ant_n": 0.0, "av_b": 0.0, "av_n": 0.0,
                                       "cac_b": 0.0, "cac_n": 0.0, "iva_b": 0.0, "iva_n": 0.0,
                                       "tot_b": 0.0, "tot_n": 0.0, "certs": set()})
            es_ant = "ant" in idc.lower()
            k = "b" if tipo.lower().startswith("b") else ("n" if tipo.lower().startswith("n") else None)
            if k:
                o[("ant_" if es_ant else "av_") + k] += base   # base neta (col L), a precio de ppto
                o["cac_" + k] += cac
                o["iva_" + k] += iva
                o["tot_" + k] += total
            if not es_ant:
                o["certs"].add(idc)

    items = []
    for g in grp.values():
        facturado_full = g["facturado"] >= g["total"] - 1
        cobrado_full = g["monto_cobrado"] >= g["total"] - 1
        estado = "Cobrado" if cobrado_full else ("Facturado" if g["facturado"] > 1 else "Certificado")
        b = g["partes"].get("Blanco", {}); n = g["partes"].get("Negro", {})
        bl = round((b.get("total", 0) / g["total"] * 100)) if g["total"] else 0
        ng = round((n.get("total", 0) / g["total"] * 100)) if g["total"] else 0
        items.append({
            "id": g["id"], "fecha": g["fecha"],
            "tipo_fiscal": f"B{bl}/N{ng}" if (bl or ng) else "",
            "concepto": g["oc"],
            "base": g["base"], "cac": g["cac"], "iva": g["iva"], "usd": g["usd"], "total": g["total"],
            "estado": estado,
            "nro_factura": g["nro_factura"],
            "facturado": g["facturado"],
            "sin_facturar": max(g["total"] - g["facturado"], 0.0),
            "fecha_factura": g["fecha_cobro"],
            "fecha_cobro": g["fecha_cobro"] if g["monto_cobrado"] else None,
            "retenciones": g["retencion"],
            "monto_cobrado": g["monto_cobrado"],
            "saldo": g["total"] - g["monto_cobrado"],
            "cac_ratio": g["cac_ratio"], "tc_usd": g["tc_usd"],
            "partes": sorted(g["partes"].values(), key=lambda x: x["tipo"]),
        })
    items.sort(key=lambda x: (str(x["concepto"] or ""), str(x["fecha"] or ""), x["id"]))

    # apertura fiscal Blanco/Negro (compat; el frontend ya no la muestra como card)
    fiscal = {}
    for g in grp.values():
        for p in g["partes"].values():
            f = fiscal.setdefault(p["tipo"], {"tipo": p["tipo"], "n": 0, "base": 0.0, "total": 0.0})
            f["n"] += 1; f["base"] += p["base"]; f["total"] += p["total"]
    return items, sorted(fiscal.values(), key=lambda f: f["base"], reverse=True), oc_res


def _facturacion(wb):
    """idfact -> {tipo, monto, cobrado, retencion, conciliado(bool), comprob, fecha_cobro}.
    AGREGA por idfact (relación 1:N: una parte B/N puede tener varios comprobantes/parciales).
    monto = Σ facturado; cobrado = Σ de los conciliados.

    Layout v8_9: monto en moneda del cobro (D) + moneda (E) + monto_ars_equiv (G,
    el ARS para sumar); 'Conciliado' manual reemplazado por 'Estado conciliado'
    computado (✅/🔴); 'Retención' como dato. Se lee monto_ars_equiv (no la col
    multimoneda) y se deriva conciliado del estado."""
    try:
        ws = wb["Cert_Facturacion"]
    except KeyError:
        return {}
    hr = nav.find_row(ws, "id cert") or nav.find_row_scan(ws, "comprobante") or 1
    col = nav.map_columns(ws, hr, {
        "idfact":     {"all": ["id", "cert", "fact"]},
        "comprob":    {"all": ["comprobante"]},
        "tipo":       {"all": ["tipo"]},                # Blanco / Negro / Anticipo
        "monto":      {"all": ["monto", "ars"]},        # monto_ars_equiv (ARS)
        "retencion":  {"all": ["retencion"]},
        "fecha_cobro": {"all": ["fecha", "cobro"]},
        "estado":     {"all": ["estado"]},              # "Estado conciliado" computado
    })
    out = {}
    for r in range(hr + 1, ws.max_row + 1):
        idf = ws.cell(row=r, column=col["idfact"]).value if "idfact" in col else None
        if idf is None or str(idf).strip() == "":
            continue
        est = nav.norm(ws.cell(row=r, column=col["estado"]).value) if "estado" in col else ""
        monto = nav.num(ws.cell(row=r, column=col["monto"]).value) if "monto" in col else 0.0
        concil = "conciliado" in est
        fc = _to_iso(ws.cell(row=r, column=col["fecha_cobro"]).value) if "fecha_cobro" in col else None
        e = out.setdefault(str(idf).strip(), {
            "tipo": str(ws.cell(row=r, column=col["tipo"]).value or "").strip() if "tipo" in col else "",
            "monto": 0.0, "cobrado": 0.0, "retencion": 0.0, "conciliado": True,
            "comprob": "", "fecha_cobro": None,
        })
        e["monto"] += monto
        if concil:
            e["cobrado"] += monto
        else:
            e["conciliado"] = False
        e["retencion"] += nav.num(ws.cell(row=r, column=col["retencion"]).value) if "retencion" in col else 0.0
        cmp = str(ws.cell(row=r, column=col["comprob"]).value or "").strip() if "comprob" in col else ""
        if cmp and not e["comprob"]:
            e["comprob"] = cmp
        if fc and (e["fecha_cobro"] is None or fc > e["fecha_cobro"]):
            e["fecha_cobro"] = fc
    return out


# ----------------------------------------------------- ensamble público
def read_cert(wb):
    """Devuelve el bloque de certificaciones + avance (contrato del dashboard)."""
    total, ocs = _control_oc(wb)
    items, fiscal, oc_res = _ledger(wb)

    cert_tot = total.get("certificado", 0.0)   # incluye el anticipo (Cert_Control_OC!F)
    anticipo = total.get("anticipo", 0.0)       # parte de anticipo (Cert_Control_OC!I)
    cert_avance = cert_tot - anticipo           # certificación de AVANCE (sin anticipo)
    facturado = total.get("facturado", 0.0)
    cobrado = total.get("cobrado", 0.0)
    avance = total.get("avance", 0.0)
    avance_sin_cert = max(total.get("avance_sin_cert", 0.0), 0.0)   # Gap A (col J); clamp ruido de float

    # apertura B/N de facturado y cobrado, desde Cert_Facturacion (por Tipo). Anticipo aparte.
    fact = _facturacion(wb)
    fb = {"blanco": 0.0, "negro": 0.0}   # facturado por tipo fiscal
    cb = {"blanco": 0.0, "negro": 0.0}   # cobrado por tipo fiscal
    for f in fact.values():
        t = (f.get("tipo") or "").lower()
        key = "blanco" if t.startswith("b") else ("negro" if t.startswith("n") else None)
        if key is None:
            continue   # anticipo: fuera del split B/N
        fb[key] += f["monto"]
        cb[key] += f["cobrado"]

    # avance por OC (≈ presupuesto/etapa) — anticipo separado del certificado de avance
    def _etapa(o):
        res = oc_res.get(o["id_oc"], {})
        return {
            "etapa": o.get("presupuesto") or o["id_oc"],
            "id_oc": o["id_oc"],
            "descripcion": o.get("descripcion", ""),
            "ppto_venta": o.get("contrato", 0.0),
            "certificado": o.get("certificado", 0.0),                            # total (incl. anticipo)
            "anticipo": o.get("anticipo", 0.0),
            "certificado_avance": o.get("certificado", 0.0) - o.get("anticipo", 0.0),
            "avance_pct": o.get("avance", 0.0),   # avance FÍSICO (col E)
            "n_certs": len(res.get("certs", ())),
            # descomposición aditiva del total certificado por tipo fiscal (a precio de ppto + ajustes)
            "fiscal": {
                "anticipo": {"blanco": res.get("ant_b", 0.0), "negro": res.get("ant_n", 0.0)},
                "avance":   {"blanco": res.get("av_b", 0.0),  "negro": res.get("av_n", 0.0)},
                "cac":      {"blanco": res.get("cac_b", 0.0), "negro": res.get("cac_n", 0.0)},
                "iva":      {"blanco": res.get("iva_b", 0.0), "negro": res.get("iva_n", 0.0)},
                "total":    {"blanco": res.get("tot_b", 0.0), "negro": res.get("tot_n", 0.0)},
            },
        }
    etapas = [_etapa(o) for o in ocs]
    etapa_total = {
        "ppto_venta": sum(e["ppto_venta"] for e in etapas),
        "certificado": cert_tot,
        "anticipo": anticipo,
        "certificado_avance": cert_avance,
        "avance_pct": avance,
    }

    certificaciones = {
        "items": items,
        "total_certificado": cert_tot,           # certificado total (incl. anticipo) — para reconciliación
        "anticipo": anticipo,                    # anticipo facturado/cobrado (se recupera vía desacopio)
        "certificado_avance": cert_avance,       # certificación de avance de obra (sin anticipo) — KPI headline
        "facturado": facturado,
        "facturado_bn": fb,                      # {blanco, negro} de lo facturado (Cert_Facturacion)
        "cobrado": cobrado,
        "cobrado_bn": cb,                        # {blanco, negro} de lo cobrado (conciliado)
        "cobrado_efectivo": sum(it["monto_cobrado"] for it in items),
        "avance_sin_certificar": avance_sin_cert,  # Gap A: avance ejecutado pendiente de certificar
        "certificado_pendiente": max(cert_tot - facturado, 0.0),  # Gap B: certificado sin facturar
        "facturado_pendiente": max(facturado - cobrado, 0.0),     # facturado sin cobrar
        "apertura_fiscal": fiscal,
        "resumen_estado": _por_estado(items),
    }
    return {
        "certificaciones": certificaciones,
        "avance_etapa": {"items": etapas, "total": etapa_total},
        "avance_fisico_pct": avance,
    }


def _por_estado(items):
    est = {}
    for it in items:
        e = nav.norm(it["estado"]) or "sin estado"
        est.setdefault(e, {"estado": it["estado"] or "Sin estado", "n": 0, "total": 0.0})
        est[e]["n"] += 1
        est[e]["total"] += it["total"]
    return list(est.values())
