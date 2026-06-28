"""
Bloque 1 — Parseo del APU original ARQING 09-24 para el APU Unificado SVD 4140.

Lee 'ARQING - APU 09-24.xlsx', extrae:
  - Tablas maestras: Materiales (desc->precio), Mano de Obra (categoria->jornal),
    Equipos (efectivo via G/D), Sub CONTRATOS.
  - Composicion de cada hoja P-XX: cant_obra, rend, grado, materiales (cant,desp),
    equipos (cant,precio efectivo), MO (cant,categoria).
Resuelve precios a VALOR usando data_only (evita reimplementar las cadenas VLOOKUP).
Guarda el resultado en scratchpad/apu_parsed.json + imprime un resumen a archivo.

Sigue el Manual Tecnico APU Unificado v3 (Parte II §5, Regla 5 sumar repetidos,
A.2 mapear por descripcion).
"""
import openpyxl, warnings, json, os, re
warnings.filterwarnings('ignore')

APU = 'archivos/fuente/ARQING - APU 09-24.xlsx'
OUT_JSON = os.path.join(os.environ.get('SCRATCH', '.'), 'apu_parsed.json')

wb_f = openpyxl.load_workbook(APU, data_only=False)
wb_v = openpyxl.load_workbook(APU, data_only=True)

def norm(s):
    return re.sub(r'\s+', ' ', str(s)).strip() if s is not None else ''

# ---------- Tablas maestras ----------
def parse_materiales():
    """Materiales!B8:D118 -> {desc_norm: {codigo, ud, precio}} (key por descripcion)."""
    wsf, wsv = wb_f['Materiales'], wb_v['Materiales']
    mats = {}
    dup = []
    for r in range(8, wsf.max_row + 1):
        desc = norm(wsf.cell(r, 2).value)
        if not desc:
            continue
        precio = wsv.cell(r, 4).value
        if precio is None or not isinstance(precio, (int, float)):
            continue
        cod = norm(wsf.cell(r, 1).value)
        ud = norm(wsf.cell(r, 3).value)
        key = desc.lower()
        if key in mats and abs(mats[key]['precio'] - precio) > 0.01:
            dup.append((desc, mats[key]['precio'], precio))
        mats[key] = {'codigo': cod, 'desc': desc, 'ud': ud, 'precio': round(precio, 4)}
    return mats, dup

def parse_mano_obra():
    """Mano de Obra!B8:C101 -> {desc_norm: {codigo, jornal}}; jornal resuelto en col C (data_only)."""
    wsf, wsv = wb_f['Mano de Obra'], wb_v['Mano de Obra']
    mo = {}
    for r in range(8, 27):  # filas de personal (8..26), evita la tabla lateral basura abajo
        desc = norm(wsf.cell(r, 2).value)
        if not desc or desc == '0':
            continue
        jornal = wsv.cell(r, 3).value  # col C = Salario (Dia) resuelto
        if not isinstance(jornal, (int, float)):
            continue
        cod = norm(wsf.cell(r, 1).value)
        mo[desc.lower()] = {'codigo': cod, 'desc': desc, 'jornal': round(jornal, 4)}
    return mo

def parse_subcontratos():
    """Sub CONTRATOS!B8.. -> lista {rubro, tarea, contratista, ud, pu}."""
    wsf, wsv = wb_f['Sub CONTRATOS'], wb_v['Sub CONTRATOS']
    subs = []
    for r in range(8, wsf.max_row + 1):
        tarea = norm(wsf.cell(r, 2).value)
        if not tarea:
            continue
        pu = wsv.cell(r, 5).value
        subs.append({
            'row': r,
            'rubro': norm(wsf.cell(r, 1).value),
            'tarea': tarea,
            'contratista': norm(wsf.cell(r, 3).value),
            'ud': norm(wsf.cell(r, 4).value),
            'pu': round(pu, 4) if isinstance(pu, (int, float)) else None,
        })
    return subs

# ---------- Hojas P-XX ----------
def parse_partida(name):
    wsf, wsv = wb_f[name], wb_v[name]
    pnum = wsf['G4'].value
    desc = norm(wsf['B7'].value)
    ud = norm(wsf['C9'].value)
    cant_obra = wsv['E9'].value
    rend = wsv['F9'].value

    # localizar filas de marcadores
    markers = {}
    for r in range(1, wsf.max_row + 1):
        for c in (1, 6):  # col A y F
            v = norm(wsf.cell(r, c).value).lower()
            if v.startswith('1.- materiales'): markers['mat_h'] = r
            elif v.startswith('2.- equipos'): markers['eq_h'] = r
            elif v.startswith('3.- mano de obra'): markers['mo_h'] = r
            elif v.startswith('total materiales'): markers['mat_end'] = r
            elif v.startswith('total equipos'): markers['eq_end'] = r
            elif v.startswith('total mano de obra menor'): markers['mo_end'] = r
            elif v.startswith('grado de dificultad'): markers['grado'] = r

    grado = wsv.cell(markers['grado'], 7).value if 'grado' in markers else 1

    # Materiales: filas (mat_h+2 .. mat_end-1). header esta en mat_h+1.
    # desp se DERIVA del total real G (col7): el APU mezcla convenciones
    #   - entero-porcentaje (F=10, G=D*E*((F+100)/100))   [albanileria]
    #   - fraccion        (F=0.05, G=D*E*(1+F))           [instalaciones]
    # desp_frac = G/(cant*precio) - 1  -> robusto ante ambas.
    mats = {}        # key desc.lower() -> agregado (Regla 5: sumar repetidos)
    consumibles = 0.0
    for r in range(markers['mat_h'] + 2, markers.get('mat_end', markers['mat_h'] + 2)):
        d_desc = norm(wsf.cell(r, 2).value)
        cant = wsv.cell(r, 4).value
        if not d_desc or not isinstance(cant, (int, float)):
            continue  # subtitulo interno (sin cantidad)
        precio = wsv.cell(r, 5).value  # E resuelto
        g = wsv.cell(r, 7).value       # Total $ resuelto
        raw = wsv.cell(r, 6).value or 0
        if isinstance(g, (int, float)) and isinstance(precio, (int, float)) and cant * precio:
            desp = round(g / (cant * precio) - 1, 6)
            if abs(desp) < 1e-6:
                desp = 0.0
        else:  # G no calculable: normalizar raw (>1 => %, else fraccion)
            desp = (raw / 100) if (isinstance(raw, (int, float)) and raw > 1) else (raw or 0)
        cod = norm(wsf.cell(r, 1).value)
        key = d_desc.lower()
        if 'consumible' in key:
            consumibles += (precio or 0) * cant * (1 + desp)
            continue
        if key in mats:  # Regla 5: misma desc en dos subsecciones -> sumar cant
            mats[key]['cant'] += cant
            mats[key]['desp'] = max(mats[key]['desp'], desp)
        else:
            mats[key] = {'codigo': cod, 'desc': d_desc, 'ud': norm(wsf.cell(r, 3).value),
                         'cant': cant, 'desp': desp,
                         'precio': round(precio, 4) if isinstance(precio, (int, float)) else None}

    # Equipos: precio efectivo = G/D (incluye deprec); cant=D
    equipos = []
    eq_na = []  # equipos con G=#N/A en el APU (VLOOKUP roto) -> flag
    if 'eq_h' in markers:
        for r in range(markers['eq_h'] + 2, markers.get('eq_end', markers['eq_h'] + 2)):
            d_desc = norm(wsf.cell(r, 2).value)
            cant = wsv.cell(r, 4).value
            if not d_desc or not isinstance(cant, (int, float)) or cant == 0:
                continue
            total = wsv.cell(r, 7).value  # G
            precio_ef = (total / cant) if (isinstance(total, (int, float)) and cant) else None
            if precio_ef is None:  # VLOOKUP roto en la partida -> buscar en Rend. Equipos
                fb = REND_EQ_MAP.get(d_desc.lower())
                if fb:
                    precio_ef = fb
                else:
                    eq_na.append(d_desc)
            equipos.append({'codigo': norm(wsf.cell(r, 1).value), 'desc': d_desc,
                            'ud': norm(wsf.cell(r, 3).value), 'cant': cant,
                            'precio': round(precio_ef, 4) if precio_ef is not None else None})

    # MO: cant=D, categoria=desc; precio=E (salario) resuelto
    mo = []
    if 'mo_h' in markers:
        for r in range(markers['mo_h'] + 2, markers.get('mo_end', markers['mo_h'] + 2)):
            d_desc = norm(wsf.cell(r, 2).value)
            cant = wsv.cell(r, 4).value
            if not d_desc or not isinstance(cant, (int, float)) or cant == 0:
                continue
            precio = wsv.cell(r, 5).value
            mo.append({'codigo': norm(wsf.cell(r, 1).value), 'desc': d_desc, 'cant': cant,
                       'precio': round(precio, 4) if isinstance(precio, (int, float)) else None})

    # totales resueltos (para auditoria Δ)
    def g_of(label_row):
        return wsv.cell(label_row, 7).value if label_row else None
    cd_row = None
    for r in range(1, wsf.max_row + 1):
        if norm(wsf.cell(r, 6).value).lower().startswith('costo directo por unidad'):
            cd_row = r; break
    cd_ud = wsv.cell(cd_row, 7).value if cd_row else None

    return {
        'partida': pnum, 'sheet': name, 'desc': desc, 'ud': ud,
        'cant_obra': cant_obra, 'rend': rend, 'grado': grado,
        'materiales': list(mats.values()), 'equipos': equipos, 'mo': mo,
        'consumibles_total': round(consumibles, 4) if consumibles else 0,
        'cd_ud_apu': round(cd_ud, 4) if isinstance(cd_ud, (int, float)) else None,
        'eq_na': eq_na,
    }

# Mapa Rend. Equipos: desc.lower -> costo referencial (col E). Fallback para equipos cuyo
# VLOOKUP esta roto en la partida (Fase D): el precio igual existe en la tabla del APU 09-24.
def parse_rend_equipos():
    wsf, wsv = wb_f['Rend. Equipos'], wb_v['Rend. Equipos']
    m = {}
    for r in range(7, wsf.max_row + 1):
        desc = norm(wsf.cell(r, 2).value)
        costo = wsv.cell(r, 5).value  # col E = Costo (referencial)
        if desc and isinstance(costo, (int, float)) and costo > 0:
            m.setdefault(desc.lower(), round(costo, 4))
    return m

REND_EQ_MAP = parse_rend_equipos()

# ---------- Run ----------
materiales, dup = parse_materiales()
mano_obra = parse_mano_obra()
subs = parse_subcontratos()

p_sheets = [s for s in wb_f.sheetnames if re.fullmatch(r'P-\d+( bis)?', s)]
partidas = []
errs = []
for s in p_sheets:
    try:
        partidas.append(parse_partida(s))
    except Exception as e:
        errs.append((s, str(e)))

data = {'materiales': materiales, 'mano_obra': mano_obra, 'subcontratos': subs,
        'partidas': partidas}
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=1)

# ---------- Resumen a archivo (evita cp1252 en consola) ----------
lines = []
lines.append(f'APU 09-24 parseado -> {OUT_JSON}')
lines.append(f'Materiales maestros: {len(materiales)}  (dup precio: {len(dup)})')
for d in dup[:10]:
    lines.append(f'   DUP {d[0]}: {d[1]} vs {d[2]}')
lines.append(f'Mano de Obra categorias: {len(mano_obra)}')
lines.append(f'Subcontratos: {len(subs)}')
lines.append(f'Partidas P-XX: {len(partidas)} (errores parseo: {len(errs)})')
for e in errs:
    lines.append(f'   ERR {e[0]}: {e[1]}')
tot_mat = sum(len(p['materiales']) for p in partidas)
tot_eq = sum(len(p['equipos']) for p in partidas)
tot_mo = sum(len(p['mo']) for p in partidas)
con = sum(1 for p in partidas if p['consumibles_total'])
lines.append(f'Insumos totales: MAT={tot_mat}  EQ={tot_eq}  MO={tot_mo}  partidas c/consumibles={con}')
# muestra P-44 (regla 5) y P-70 (memb dup)
for nm in ('P-44', 'P-70'):
    p = next((x for x in partidas if x['sheet'] == nm), None)
    if p:
        lines.append(f'--- {nm}: {p["desc"][:50]} | ud={p["ud"]} cant_obra={p["cant_obra"]} rend={p["rend"]}')
        for m in p['materiales']:
            lines.append(f'     MAT {m["desc"][:34]:34} cant={m["cant"]} desp={m["desp"]} $={m["precio"]}')
        for m in p['mo']:
            lines.append(f'     MO  {m["desc"][:34]:34} cant={m["cant"]} $={m["precio"]}')

with open('scripts/_apu_parse_report.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('OK - reporte en scripts/_apu_parse_report.txt')
