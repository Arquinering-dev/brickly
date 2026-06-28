"""Bloque 7c — Loader del avance real de GDR (historial completo) al circuito Cert_*.
Parsea (anclado por texto, layouts variables) las certs de avance de los 3 documentos:
  OC01: GDR 3760_Pto 01_Cert. 09 02.xlsx  (Cert 01..08 + Cert 09 (2))
  OC02: GDR 3760_Pto 02_Cert. 15.xlsx     (Cert 01..15)
  OC03: GDR_Adicionales _03.xlsx           (Adic N°1..3)
Genera:
  - Cert_App_Output: 1 fila por (cert, tarea) con %actual<>0 (incremento). PV STORED del doc (col I)
    -> soporta OC03 adicionales (fuera del ppto base). %total = F+G (formula).
  - Cert_Cabecera: 1 fila por madre (avance + ANT). desacopio madre = % anticipo de la OC (0 en ANT).
  - Cert_Calculo: Blanco+Negro por madre (OC03: solo Negro). ANT con base HARDCODEADA (= monto anticipo).
    %fact = split sugerido de la OC (provisional; refinable por cert).

Modo: `python scripts/cargar_cert_avance_gdr.py` (dry-run, reconcilia, NO guarda)
      `python scripts/cargar_cert_avance_gdr.py --write` (escribe v8_12)
"""
import sys
import re
import datetime
import openpyxl
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
FU = "archivos/fuente/"
WRITE = "--write" in sys.argv

# OC -> (archivo, datos OC). Las hojas se descubren por patrón.
OCS = {
    "GDR-OC01": dict(file="GDR 3760_Pto 01_Cert. 09 02.xlsx", ant=0.4, blanco=0.7, negro=0.3,
                     monto=500907720, solo_negro=False),
    "GDR-OC02": dict(file="GDR 3760_Pto 02_Cert. 15.xlsx", ant=0.1, blanco=0.7, negro=0.3,
                     monto=844845248, solo_negro=False),
    "GDR-OC03": dict(file="GDR_Adicionales _03.xlsx", ant=0.0, blanco=0.0, negro=1.0,
                     monto=96300019, solo_negro=True),
}


def descubrir_hojas(wb, oc):
    """Devuelve [(sheet, 'C01'), ...] en orden de cert. Maneja 'Cert. 09 (2)' y 'Adic. N°k'."""
    out = []
    if oc == "GDR-OC03":
        for s in wb.sheetnames:
            m = re.search(r"Adic\.?\s*N°?\s*(\d+)", s)
            if m:
                out.append((s, int(m.group(1))))
    else:
        best = {}  # num -> sheet (prefiere la variante '(2)')
        for s in wb.sheetnames:
            m = re.search(r"Cert\.?\s*0*(\d+)", s)
            if not m:
                continue
            num = int(m.group(1))
            if num not in best or "(2)" in s:  # '(2)' es la versión final del cert
                best[num] = s
        out = [(best[n], n) for n in sorted(best)]
    return [(s, "C%02d" % n) for s, n in out]


def localizar(ws):
    """Anclado por texto: devuelve (hr, col_cod, col_ant, col_act, col_acum, col_pv)."""
    hr = None
    for r in range(1, 13):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        if any("descrip" in v for v in vals) and any(v == "acumulado" for v in vals):
            hr = r
            break
    if hr is None:
        return None
    hdr = {c: str(ws.cell(hr, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)}
    # dos tríos Anterior/Actual/Acumulado: el 1º (col menor) = Avance (%); el 2º = Importe ($)
    acum = sorted(c for c, t in hdr.items() if t == "acumulado")
    sub = [c for c, t in hdr.items() if t == "subtotal"]
    if len(acum) < 2 or not sub:
        return None
    best_c, best_n = 1, -1
    for c in range(1, 5):
        n = sum(1 for r in range(hr + 1, ws.max_row + 1)
                if re.match(r"^\d+\.\d+$", str(ws.cell(r, c).value or "").strip()))
        if n > best_n:
            best_c, best_n = c, n
    # cod, acum% (1er Acumulado), acum$ (2º Acumulado = Importe), pv (Subtotal)
    return hr, best_c, acum[0], acum[-1], min(sub)


def parse_cert(ws):
    """Devuelve {cod: (acum_pct, acum_imp, pv)} de un sheet de cert."""
    loc = localizar(ws)
    if not loc:
        return None
    hr, cc, cacum_pct, cacum_imp, cpv = loc
    out = {}
    for r in range(hr + 1, ws.max_row + 1):
        raw = str(ws.cell(r, cc).value or "").strip()
        if not re.match(r"^\d+\.\d+$", raw):
            continue
        cod = float(raw)
        apct = ws.cell(r, cacum_pct).value or 0
        aimp = ws.cell(r, cacum_imp).value or 0
        pv = ws.cell(r, cpv).value or 0
        out[cod] = (float(apct), float(aimp), float(pv))
    return out


def fnum(x):
    return f"{x:,.2f}"


def main():
    # ---- parseo + reconciliación: incremento %=Δacum%, incremento $=ΔImporteAcum ----
    app_rows = []        # (oc, idc, fecha, cod, prev_pct, incr_pct, acum_pct, pv, incr_imp)
    madres = []          # (idc, oc) en orden, solo avance
    for oc, info in OCS.items():
        wb = openpyxl.load_workbook(FU + info["file"], data_only=True)
        hojas = descubrir_hojas(wb, oc)
        print(f"\n[{oc}] {info['file']} — {len(hojas)} certs: {[h[1] for h in hojas]}")
        prev_pct, prev_imp, pv_last, final_imp = {}, {}, {}, {}
        base_oc = 0.0
        for sheet, cnum in hojas:
            ws = wb[sheet]
            data = parse_cert(ws)
            if data is None:
                print(f"   ⚠ {sheet}: no se pudo anclar — SALTEADO")
                continue
            idc = f"{oc}-{cnum}"
            madres.append((idc, oc))
            fecha = None
            for rr in range(3, 7):
                for cc2 in range(1, ws.max_column + 1):
                    v = ws.cell(rr, cc2).value
                    if isinstance(v, datetime.datetime) and v.year >= 2025:
                        fecha = v
            for cod, (apct, aimp, pv) in sorted(data.items()):
                pv_last[cod] = pv
                final_imp[cod] = apct  # acum %
                ipct = apct - prev_pct.get(cod, 0.0)   # incremento % = Δ acum%
                if abs(ipct) > 1e-9:
                    app_rows.append((oc, idc, fecha, cod, prev_pct.get(cod, 0.0), ipct, apct, pv))
                prev_pct[cod] = apct
        # avance proxy ponderado por PV del doc (verificación fina = 7d con budget Z)
        wsum = sum(final_imp[c] * pv_last[c] for c in final_imp)
        psum = sum(pv_last.values())
        print(f"   tareas={len(final_imp)}  avance ponderado(doc PV)={wsum/psum*100:.2f}%  base(sin CAC, %acum×PVdoc)=${fnum(wsum)}")

    print(f"\n>>> Total filas Cert_App_Output a cargar: {len(app_rows)}")
    print(f">>> Madres de avance: {len(madres)}  (OC01 {sum(1 for _,o in madres if o=='GDR-OC01')},"
          f" OC02 {sum(1 for _,o in madres if o=='GDR-OC02')}, OC03 {sum(1 for _,o in madres if o=='GDR-OC03')})")

    if not WRITE:
        print("\n[DRY-RUN] no se escribió nada. Re-correr con --write para aplicar.")
        return

    # ---- escritura ----
    wb = openpyxl.load_workbook(PATH)
    _write(wb, app_rows, madres)
    wb.save(PATH)
    print(f"\n✓ Escrito → {PATH}")


def _write(wb, app_rows, madres):
    # === Cert_App_Output ===
    ao = wb["Cert_App_Output"]
    for r in range(2, ao.max_row + 1):
        for c in range(1, ao.max_column + 1):
            ao.cell(r, c).value = None
    for i, (oc, idc, fecha, cod, prev_pct, incr_pct, acum_pct, pv) in enumerate(app_rows):
        r = 2 + i
        ao.cell(r, 1, "GDR 3760"); ao.cell(r, 2, oc); ao.cell(r, 3, idc)
        if fecha:
            ao.cell(r, 4, fecha).number_format = "dd/mm/yyyy"
        ao.cell(r, 5, cod)
        ao.cell(r, 6, prev_pct); ao.cell(r, 7, incr_pct)   # % anterior, % actual (Δacum%)
        ao.cell(r, 8, f"=F{r}+G{r}")                        # % total (=acum%)
        # PV (col I): OC01/02 -> SUMIFS al budget Z (base sin CAC, como CH); OC03 adic -> PV del doc (fuera del budget)
        if oc == "GDR-OC03":
            ao.cell(r, 9, pv).number_format = "\\$#,##0"
        else:
            ao.cell(r, 9, f"=SUMIFS('1_Presupuesto'!$Z:$Z,'1_Presupuesto'!$AS:$AS,K{r},'1_Presupuesto'!$F:$F,E{r})").number_format = "\\$#,##0"
        ao.cell(r, 10, f"=G{r}*I{r}").number_format = "\\$#,##0"  # $ base = %actual × PV (sin CAC)
        ao.cell(r, 11, f"=_xlfn.XLOOKUP(B{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$L:$L)")
        ao.cell(r, 12, f'=IF($E{r}="","",IF(AND(ISNUMBER($I{r}),$H{r}<=1.0001),"✓","⚠"))')

    # === madres: avance (orden) + ANT por OC ===
    av_madres = [idc for idc, _ in madres]
    cab = wb["Cert_Cabecera"]
    cal = wb["Cert_Calculo"]
    for ws in (cab, cal):
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).value = None

    # Cert_Cabecera: ANT (por OC) + avance
    cab_rows = []
    for oc in OCS:
        cab_rows.append((f"{oc}-ANT", 0.0))           # desac madre = 0 en ANT
    for idc, oc in madres:
        cab_rows.append((idc, OCS[oc]["ant"]))        # desac = % anticipo OC
    for i, (idc, desac) in enumerate(cab_rows):
        r = 2 + i
        cab.cell(r, 1, idc)
        cab.cell(r, 2, f'=LEFT($A{r},FIND("|",SUBSTITUTE($A{r},"-","|",2))-1)')
        cab.cell(r, 3, f'=IFERROR(_xlfn.XLOOKUP($A{r},Cert_App_Output!$C:$C,Cert_App_Output!$D:$D),"")')
        cab.cell(r, 4, f"=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$C:$C,$A{r})")
        cab.cell(r, 5, desac).number_format = "0%"
        cab.cell(r, 6, f"=$D{r}*$E{r}")
        cab.cell(r, 7, f"=$D{r}-$F{r}")
        cab.cell(r, 8, f'=IF(COUNTIF(Cert_Calculo!$A:$A,$A{r})>0,"OK","FALTA en Calculo")')

    # Cert_Calculo: B/N por madre. ANT con base HARDCODEADA = monto anticipo.
    cal_specs = []  # (idc, oc, es_ant)
    for oc in OCS:
        cal_specs.append((f"{oc}-ANT", oc, True))
    for idc, oc in madres:
        cal_specs.append((idc, oc, False))
    r = 2
    for idc, oc, es_ant in cal_specs:
        info = OCS[oc]
        tipos = [("Negro", info["negro"])] if info["solo_negro"] else \
                [("Blanco", info["blanco"]), ("Negro", info["negro"])]
        ant_monto = info["monto"] * info["ant"]
        for tipo, pfac in tipos:
            suf = "N" if tipo == "Negro" else "B"
            cal.cell(r, 1, idc)
            cal.cell(r, 2, f"{idc}-{suf}")
            cal.cell(r, 3, tipo)
            cal.cell(r, 4, f"=_xlfn.XLOOKUP(A{r},Cert_App_Output!$C:$C,Cert_App_Output!$A:$A)" if not es_ant else "GDR 3760")
            cal.cell(r, 5, oc)
            cal.cell(r, 6, f"=_xlfn.XLOOKUP(A{r},Cert_App_Output!$C:$C,Cert_App_Output!$D:$D)" if not es_ant else None)
            # G: avance = SUMIFS; ANT = hardcode anticipo
            if es_ant:
                cal.cell(r, 7, ant_monto).number_format = "\\$#,##0"
            else:
                cal.cell(r, 7, f"=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$C:$C,$A{r})")
            cal.cell(r, 8, pfac).number_format = "0%"
            cal.cell(r, 9, f"=$G{r}*H{r}")
            cal.cell(r, 10, f"=_xlfn.XLOOKUP($A{r},Cert_Cabecera!$A:$A,Cert_Cabecera!$E:$E)")
            cal.cell(r, 11, f"=-I{r}*J{r}")
            cal.cell(r, 12, f"=I{r}+K{r}")
            cal.cell(r, 13, f"=_xlfn.XLOOKUP($E{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$G:$G)")
            cal.cell(r, 14, f"=IF($Y{r}=\"\",_xlfn.XLOOKUP(DATE(YEAR(F{r}),MONTH(F{r}),1),'0_Indice_CAC'!$A:$A,'0_Indice_CAC'!$B:$B),$Y{r})" if not es_ant else f"=IF($Y{r}=\"\",'0_Indice_CAC'!$B$4,$Y{r})")
            cal.cell(r, 15, f"=N{r}/M{r}")
            cal.cell(r, 16, f"=L{r}*(O{r}-1)")
            cal.cell(r, 17, f"=L{r}+P{r}")
            cal.cell(r, 18, 0.21 if tipo == "Blanco" else 0).number_format = "0%"  # IVA solo Blanco
            cal.cell(r, 19, f"=Q{r}*R{r}")
            cal.cell(r, 20, f"=Q{r}+S{r}")
            cal.cell(r, 21, 1170)  # USD MEP (provisional; DÓLAR MEP del doc)
            cal.cell(r, 22, f'=IF(C{r}="Negro",T{r}/U{r},"")')
            cal.cell(r, 24, f'=IF(ABS(SUMIFS($H:$H,$A:$A,$A{r})-1)<0.0001,"✓","⚠ ≠100%")')
            r += 1


if __name__ == "__main__":
    main()
