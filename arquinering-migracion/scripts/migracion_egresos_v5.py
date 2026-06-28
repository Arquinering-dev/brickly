"""
migracion_egresos_v5.py — Quincenas como movimientos; fuente única de egresos
Arquinering S.R.L. — CH 2171

Sobre v8_4 -> v8_5:
  1) Escribe las quincenas como filas nuevas en 2_Movimientos (append al final),
     agregadas por (Mes, Período, Rubro). Monto = valor pegado (simula dato Tezamat).
     D (Cuenta) = nº de cuenta del rubro; S/T = fórmulas estándar de la hoja.
     Observaciones = "QUINCENA <YYYY-MM> <período>".
  2) 3_Control_Ppto y 3_Cash_Flow dejan de leer 2_Quincenas -> fuente única
     (2_Movimientos). Sin doble conteo.
  2_Quincenas queda como staging (3_Control_Jornales sigue leyendo sus horas).

Un solo save. Luego excel_recalc.py + recalc.py.
"""
import os
import shutil
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.comments import Comment

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_4.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_5.xlsx"
NEW = "2_Movimientos"


def main():
    shutil.copyfile(SRC, DST)
    # valores calculados (para N de quincenas y mapa cuenta)
    wbv = openpyxl.load_workbook(DST, data_only=True)
    qv = wbv["2_Quincenas"]
    mvv = wbv[NEW]

    # mapa Desc Cuenta -> Nº Cuenta (de movimientos existentes)
    emap = {}
    for r in range(2, mvv.max_row + 1):
        d = mvv.cell(r, 4).value
        e = mvv.cell(r, 5).value
        if e and d is not None:
            emap.setdefault(str(e), d)

    # agregación quincenas por (mes_date, periodo, rubro) -> N real
    agg = defaultdict(float)
    for r in range(4, qv.max_row + 1):
        a = qv.cell(r, 1).value
        if a in (None, ""):
            continue
        per = qv.cell(r, 2).value
        rub = qv.cell(r, 5).value
        N = qv.cell(r, 14).value or 0
        if not isinstance(N, (int, float)):
            N = 0
        mes = datetime(a.year, a.month, 1)
        agg[(mes, per, rub)] += N

    # ordenar por (mes, periodo, rubro)
    combos = sorted(agg.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2]))

    # ----- escribir sobre el wb con fórmulas -----
    wb = openpyxl.load_workbook(DST)
    mv = wb[NEW]
    log = []

    # última fila de datos
    LAST = 0
    for r in range(2, mv.max_row + 1):
        if mv.cell(r, 4).value is not None or mv.cell(r, 12).value is not None:
            LAST = r

    sin_cuenta = set()
    nr = LAST
    for (mes, per, rub), N in combos:
        nr += 1
        cuenta = emap.get(str(rub))
        if cuenta is None:
            sin_cuenta.add(rub)
        mes_str = mes.strftime("%Y-%m")
        mv.cell(nr, 3).value = f"=EOMONTH(F{nr},-1)+1"          # C Mes
        mv.cell(nr, 4).value = cuenta                           # D Cuenta (puede ser None)
        mv.cell(nr, 5).value = rub                              # E Desc Cuenta = rubro token
        mv.cell(nr, 6).value = mes                              # F Fecha (1° de mes)
        mv.cell(nr, 8).value = f"QUINCENA {mes_str} {per}"      # H Observaciones
        mv.cell(nr, 12).value = round(N, 2)                     # L Debe = monto (valor pegado)
        mv.cell(nr, 13).value = 0                               # M Haber
        mv.cell(nr, 18).value = 0                               # R Saldo Inicial
        mv.cell(nr, 19).value = (                               # S Monto Descontado
            f"=(L{nr}-M{nr})*SUMIF('0_Indice_CAC'!$A:$A,C{nr},'0_Indice_CAC'!$D:$D)"
        )
        mv.cell(nr, 20).value = f"=L{nr}-M{nr}"                 # T Monto Real
        mv.cell(nr, 24).value = (                              # X En 3_Control_Ppto
            f"=IF(ISNUMBER(MATCH(E{nr},'3_Control_Ppto'!$A$4:$A$35,0)),\"SÍ\",\"NO\")"
        )
        mv.cell(nr, 25).value = f'=IF(LEFT(D{nr}&"",1)="4","INGRESO","EGRESO")'  # Y Tipo
        mv.cell(nr, 26).value = f"(generado de 2_Quincenas {mes_str} {per})"     # Z
    log.append(f"2_Movimientos: {len(combos)} filas de quincena (append filas {LAST+1}-{nr})")
    if sin_cuenta:
        mv.cell(LAST + 1, 4).comment = Comment(
            "Falta nº de cuenta Tezamat para: " + ", ".join(sorted(sin_cuenta)) +
            ". Completar cuando Tezamat lo provea.", "migracion")
        log.append("PENDIENTE nº de cuenta para: " + ", ".join(sorted(sin_cuenta)))

    # ----- 3_Control_Ppto: fuente única 2_Movimientos (quitar término quincenas) -----
    cp = wb["3_Control_Ppto"]
    for r in range(4, 31):
        if cp.cell(r, 1).value in (None, ""):
            continue
        cp.cell(r, 5).value = f"=SUMIFS('{NEW}'!$S:$S,'{NEW}'!$E:$E,$A{r})"
        cp.cell(r, 6).value = f"=SUMIFS('{NEW}'!$T:$T,'{NEW}'!$E:$E,$A{r})"
    log.append("3_Control_Ppto E/F: solo 2_Movimientos (sin término 2_Quincenas)")

    # ----- 3_Cash_Flow: egresos sin término quincenas (ya entran por movimientos) -----
    cf = wb["3_Cash_Flow"]
    for c in range(2, 20):
        n = c - 2
        y, m = 2026 + n // 12, n % 12 + 1
        cf.cell(6, c).value = (
            f"=SUMIFS('{NEW}'!$T:$T,'{NEW}'!$Y:$Y,\"EGRESO\",'{NEW}'!$C:$C,DATE({y},{m},1))"
        )
    log.append("3_Cash_Flow fila 6: egresos solo de 2_Movimientos (sin doble conteo)")

    # ----- CHANGELOG -----
    ch = wb["CHANGELOG"]
    base = ch.max_row + 2
    extra = [
        ["v8_5 (2026-06-18) — quincenas como movimientos; fuente única de egresos"],
        ["CAMBIO", "DETALLE", "MOTIVO"],
        ["Quincenas -> 2_Movimientos", f"{len(combos)} filas (Mes,Período,Rubro); monto pegado, D=cuenta del rubro, Obs 'QUINCENA aaaa-mm NQ'",
         "MO entra a Tezamat con cuenta y rubro; una sola fuente de egresos"],
        ["Control_Ppto fuente única", "E/F = SUMIFS(2_Movimientos S/T por rubro), sin término 2_Quincenas",
         "Elimina la doble fuente de costo de MO"],
        ["Cash_Flow sin doble conteo", "Fila 6 egresos = solo 2_Movimientos EGRESO por mes",
         "La quincena ya entra por movimientos"],
        ["2_Quincenas = staging", "Se mantiene: 3_Control_Jornales lee sus horas (G/H/I) para horas vs ppto",
         "Sigue siendo origen del dato; no se elimina"],
        ["Pendiente nº cuenta", "Eléctrico MO y Herrería MO sin cuenta Tezamat (D vacía)",
         "Completar cuando Tezamat provea el plan de cuentas"],
    ]
    for i, row in enumerate(extra):
        for j, val in enumerate(row, 1):
            ch.cell(base + i, j).value = val

    wb.save(DST)
    print("✅ Guardado:", DST)
    for x in log:
        print("  -", x)


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    main()
