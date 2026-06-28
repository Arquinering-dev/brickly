"""svd_agregar_ids_tezamat.py — Agrega los subcontratos de SVD 4140 al archivo de IDs
para Arquinering (docs/Carga_Tezamat_IDs_GDR_CH.xlsx, hoja 'Subcontratos (egresos)').
Lee los datos exactos del 2_Subcontratos del archivo v8 de SVD. Preserva el formato
copiando el estilo de la última fila existente. Tambien deja una nota en 'Pendientes'.
"""
import sys
import copy
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DOC = "docs/Carga_Tezamat_IDs_GDR_CH.xlsx"
SVD = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"


def copiar_estilo(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def main():
    # leer subcontratos SVD (2_Subcontratos filas 4..13)
    sc = openpyxl.load_workbook(SVD, data_only=True)["2_Subcontratos"]
    filas = []
    for r in range(4, 14):
        cid = sc.cell(r, 1).value
        if not cid:
            continue
        prov = sc.cell(r, 2).value
        rubro = sc.cell(r, 3).value
        presup = sc.cell(r, 5).value or 0
        ajusta = sc.cell(r, 6).value          # SI / NO
        ant = sc.cell(r, 7).value or 0
        tipos = "BASE / CAC / CS" if str(ajusta).upper() == "SI" else "BASE / CS"
        if ant and float(ant) > 0:
            tipos = tipos.replace("BASE /", "BASE / ANT /")
        ejemplo = f"{cid} | BASE | (descripción del pago)"
        estado = "Sin taguear (pagado $0) → al taguear concilia solo"
        if (presup or 0) == 0:
            estado = "⚠ Contrato con monto $0 — confirmar presupuesto antes de taguear"
        filas.append((cid, prov, rubro, presup, ant, tipos, ejemplo, estado))

    wb = openpyxl.load_workbook(DOC)
    ws = wb["Subcontratos (egresos)"]
    # ultima fila con datos
    last = ws.max_row
    while last > 3 and ws.cell(last, 2).value is None:
        last -= 1
    style_row = last   # copiar estilo de la última fila existente (CH-SC-004)

    r = last + 1
    for (cid, prov, rubro, presup, ant, tipos, ejemplo, estado) in filas:
        vals = ["SVD", cid, prov, rubro, presup, ant, tipos, ejemplo, estado]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            copiar_estilo(ws.cell(style_row, c), cell)
        r += 1

    # nota en Pendientes
    try:
        pend = wb["Pendientes"]
        pr = pend.max_row
        while pr > 3 and pend.cell(pr, 1).value is None:
            pr -= 1
        notas = [
            ("SVD", "Subcontratos sin taguear",
             "Ningún pago a subcontratistas tiene ID en Tezamat. Taguear con SVD-SC-NNN | TIPO | desc."),
            ("SVD", "Subcontrato con monto $0",
             "SVD-SC-002 (Sereno de Obra) tiene presupuesto $0 pero pagos > 0 — confirmar monto de contrato."),
            ("SVD", "Pagos > presupuesto",
             "Varios SC con pagos > presup (ej. CELSI SVD-SC-001 presup 20,6M vs pagos 41M) — revisar montos en Contratistas."),
        ]
        for i, (o, t, d) in enumerate(notas):
            pend.cell(pr + 1 + i, 1, o)
            pend.cell(pr + 1 + i, 2, t)
            pend.cell(pr + 1 + i, 3, d)
            for c in (1, 2, 3):
                copiar_estilo(pend.cell(pr, c), pend.cell(pr + 1 + i, c))
    except KeyError:
        pass

    wb.save(DOC)
    print(f"✓ Agregados {len(filas)} subcontratos SVD a {DOC}")
    for (cid, prov, rubro, presup, ant, tipos, ej, est) in filas:
        print(f"   {cid}  {str(prov)[:24]:25} {str(rubro)[:16]:17} presup=${presup:>13,.0f}  TIPOS: {tipos}")


if __name__ == "__main__":
    main()
