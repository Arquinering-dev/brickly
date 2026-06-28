"""aplicar_plan_cuentas_ch.py — Alinea los rubros de CH 2171 al Plan de Cuentas Tezamat.

Decisiones (Pedro, 2026-06-19):
  - Nombres alineados al plan exacto (incluido el typo 'Homigón MO').
  - _Listas = rama 53 OBRA del plan (con código) + Mov. Variables (52302) +
    Supervisión de Obra MO (52209, marcada indirecto, queda como rubro por ahora).
  - Merges/renames en 1_Presupuesto (cols A/B/C/D) y 2_Quincenas (col E):
      Agrimensura MT/MO              -> Preliminares            (53001)
      Gastos Generales Obra MT/MO   -> Gastos Generales         (53022)
      Mantenimiento MO              -> Gastos Generales         (53022)
      Seguridad e Higiene MO        -> Seguridad e Higiene      (53026)
  - Data validation (dropdown) en celdas de rubro + sanity check (rubros ∈ plan).

Uso: python scripts/aplicar_plan_cuentas_ch.py <ruta_v8_7.xlsx>
"""
import sys
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

# --- rama 53 OBRA del plan (orden del plan) + especiales ---------------------
# (rubro, código, tipo)  tipo = MT/MO/'' (los sin split quedan '')
CANONICO = [
    ("Preliminares", "53001", ""),
    ("Demolición", "53002", ""),
    ("Movimiento de Suelos", "53003", ""),
    ("Hormigón MT", "53004", "MT"),
    ("Homigón MO", "53005", "MO"),
    ("Metálica MT", "53006", "MT"),
    ("Metálica MO", "53007", "MO"),
    ("Albañilería MT", "53008", "MT"),
    ("Albañilería MO", "53009", "MO"),
    ("Durlock MT", "53010", "MT"),
    ("Durlock MO", "53011", "MO"),
    ("Aberturas", "53012", ""),
    ("Revestimiento MT", "53013", "MT"),
    ("Revestimiento MO", "53014", "MO"),
    ("Pintura MT", "53015", "MT"),
    ("Pintura MO", "53016", "MO"),
    ("Sanitaria MT", "53017", "MT"),
    ("Sanitaria MO", "53018", "MO"),
    ("Eléctrico MT", "53019", "MT"),
    ("Eléctrico MO", "53020", "MO"),
    ("Provisiones", "53021", ""),
    ("Gastos Generales", "53022", ""),
    ("Varios Ferreteria", "53023", ""),
    ("Termomecánica MT", "53024", "MT"),
    ("Termomecánica MO", "53025", "MO"),
    ("Seguridad e Higiene", "53026", ""),
    ("Herrería MT", "53027", "MT"),
    ("Herrería MO", "53028", "MO"),
    ("Alquiler de Equipos", "53980", ""),
    ("Gastos a Reintegrar", "53990", ""),
    ("Gastos GCBA Construccion", "53999", ""),
    # especiales: cuentas no-53 usadas como rubro de obra
    ("Mov. Variables", "52302", ""),
    ("Supervisión de Obra MO", "52209", "MO"),  # indirecto, queda como rubro por ahora
]

RENAME = {
    "Agrimensura MT": "Preliminares",
    "Agrimensura MO": "Preliminares",
    "Gastos Generales Obra MT": "Gastos Generales",
    "Gastos Generales Obra MO": "Gastos Generales",
    "Mantenimiento MO": "Gastos Generales",
    "Seguridad e Higiene MO": "Seguridad e Higiene",
}


def rename_col(ws, cols, r0, r1):
    n = 0
    for r in range(r0, r1 + 1):
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() in RENAME:
                ws.cell(row=r, column=c).value = RENAME[v.strip()]
                n += 1
    return n


def apply(path):
    wb = openpyxl.load_workbook(path)

    # 1) renombrar/fusionar en 1_Presupuesto (A=1,B=2,C=3,D=4) y 2_Quincenas (E=5)
    p = wb["1_Presupuesto"]
    n_p = rename_col(p, (1, 2, 3, 4), 4, p.max_row)
    q = wb["2_Quincenas"]
    n_q = rename_col(q, (5,), 4, q.max_row)

    # 2) reescribir _Listas como espejo del plan
    L = wb["_Listas"]
    # limpiar cols A,B,C hasta fila 60
    for r in range(1, 61):
        for c in (1, 2, 3):
            L.cell(row=r, column=c).value = None
    L["A1"] = "RUBROS CANÓNICOS (Plan de Cuentas Tezamat — rama 53 OBRA + especiales)"
    L["B1"] = "Código"
    L["C1"] = "TIPOS"
    L["C2"] = "MT"
    L["C3"] = "MO"
    for i, (rub, cod, _t) in enumerate(CANONICO):
        L.cell(row=2 + i, column=1).value = rub
        L.cell(row=2 + i, column=2).value = cod
    last = 1 + len(CANONICO)  # última fila de la lista (A2..A{last})
    rng = f"$A$2:$A${last}"
    # nombre definido para la dropdown (DV cross-sheet estándar, no extensión x14)
    nm = "RUBROS_PLAN"
    try:
        del wb.defined_names[nm]
    except (KeyError, TypeError):
        pass
    wb.defined_names[nm] = DefinedName(nm, attr_text=f"_Listas!{rng}")

    # 3) sanity check en _Listas (rubros de tarea ∈ plan)
    L["E1"] = "VALIDACIÓN — rubros de tarea ∈ plan de cuentas"
    pmax = p.max_row
    qmax = q.max_row
    def sp(sheet, colrange):
        # cuenta celdas no vacías y != '-' cuyo valor NO está en la lista canónica
        return ("+".join(
            f"SUMPRODUCT(('{sheet}'!{cl}4:{cl}{mx}<>\"\")*('{sheet}'!{cl}4:{cl}{mx}<>\"-\")"
            f"*ISNA(MATCH('{sheet}'!{cl}4:{cl}{mx},{rng},0)))"
            for cl, mx in colrange))
    L["E2"] = "Rubros fuera de plan — 1_Presupuesto (A/B/C/D)"
    L["F2"] = "=" + sp("1_Presupuesto", [("A", pmax), ("B", pmax), ("C", pmax), ("D", pmax)])
    L["G2"] = '=IF(F2=0,"✓","⚠")'
    L["E3"] = "Rubros fuera de plan — 2_Quincenas (E)"
    L["F3"] = "=" + sp("2_Quincenas", [("E", qmax)])
    L["G3"] = '=IF(F3=0,"✓","⚠")'

    # 4) data validation (dropdown) en celdas de rubro
    def add_dv(ws, sqref):
        dv = DataValidation(type="list", formula1=nm,
                            allow_blank=True, showErrorMessage=False)
        dv.error = "Rubro fuera del plan de cuentas"
        dv.prompt = "Elegí un rubro del plan de cuentas"
        ws.add_data_validation(dv)
        dv.add(sqref)
    add_dv(p, f"A4:D{pmax}")
    add_dv(q, f"E4:E{qmax}")

    wb.save(path)
    print(f"✓ Plan de cuentas aplicado a {path}")
    print(f"  · renombres/fusiones: 1_Presupuesto={n_p} celdas · 2_Quincenas={n_q} celdas")
    print(f"  · _Listas reescrito: {len(CANONICO)} rubros canónicos (A2:A{last})")
    print(f"  · dropdowns en 1_Presupuesto!A4:D{pmax} y 2_Quincenas!E4:E{qmax}")
    print(f"  · sanity checks en _Listas!F2/F3 (✓/⚠ en G2/G3)")


if __name__ == "__main__":
    apply(sys.argv[1])
