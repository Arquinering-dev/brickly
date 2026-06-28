"""
Alinea el split MO/OTR–ALB de SVD_4140_Resumen_de_Obra_v8_2.xlsx!1_Presupuesto al criterio
auditado de la composición (provisorio, pendiente Arquinering).

Por tarea single-row (Cod.Ítem float): K (MO/OTR) y L (MO/ALB) se re-reparten según la composición,
preservando el total K+L (N no cambia). El rubro de etapa MO (único por tarea) se mueve entre las
columnas C (Rubro MO/OTR) y D (Rubro MO/ALB) según qué portción quede > 0. La venta P/Q se
recalcula sola (fórmulas). NO toca filas partidas (2.01.01/02, 24.10.01/02) ni MT/EQ.
"""
import openpyxl, warnings, sys, json, os
warnings.filterwarnings('ignore'); sys.stdout.reconfigure(encoding='utf-8')

RES = 'archivos/output/SVD_4140_Resumen_de_Obra_v8_2.xlsx'
SCRATCH = os.environ.get('SCRATCH', '.')
d = json.load(open(os.path.join(SCRATCH, 'budget_parsed.json'), encoding='utf-8'))

# split auditado por ítem (per-unit)
aud = {}
for it in d['items']:
    moo = sum((r['cant'] or 0) * (r['precio'] or 0) for r in it['rows'] if r['tipo'] == 'MO' and r['subtipo'] == 'OTR')
    moa = sum((r['cant'] or 0) * (r['precio'] or 0) for r in it['rows'] if r['tipo'] == 'MO' and r['subtipo'] == 'ALB')
    aud[round(it['item'], 4)] = (moo, moa)

wb = openpyxl.load_workbook(RES)
wbv = openpyxl.load_workbook(RES, data_only=True)
ws = wb['1_Presupuesto']; wsv = wbv['1_Presupuesto']

cambiadas = 0; saltadas = []; sin_rubro = []
for r in range(5, ws.max_row + 1):
    e = wsv.cell(r, 5).value
    if not isinstance(e, (int, float)) or float(e) == int(e):
        continue  # headers de rubro y filas partidas (string) se saltan
    k = round(e, 4)
    if k not in aud:
        continue
    K_old = wsv.cell(r, 11).value or 0
    L_old = wsv.cell(r, 12).value or 0
    total = K_old + L_old
    moo, moa = aud[k]
    if total < 1:  # tarea sin MO
        continue
    if abs((moo + moa) - total) > max(50, 0.02 * total):
        saltadas.append((e, round(total), round(moo + moa)))  # totales no reconcilian -> no tocar
        continue
    if abs(K_old - moo) < 50 and abs(L_old - moa) < 50:
        continue  # ya alineada
    # rubro de etapa MO (único): el que esté no-"-" en C o D
    C = ws.cell(r, 3).value; D = ws.cell(r, 4).value
    mo_rubro = C if (C and C != '-') else (D if (D and D != '-') else None)
    if mo_rubro is None:
        sin_rubro.append(e)
        continue
    # preservar total exacto: K=moo, L=total-moo
    K_new = round(moo, 6)
    L_new = round(total - moo, 6)
    ws.cell(r, 11, K_new)
    ws.cell(r, 12, L_new)
    ws.cell(r, 3, mo_rubro if K_new > 1 else '-')   # C (Rubro MO/OTR)
    ws.cell(r, 4, mo_rubro if L_new > 1 else '-')   # D (Rubro MO/ALB)
    cambiadas += 1

wb.save(RES)
print(f'Tareas alineadas (K/L + C/D): {cambiadas}')
print(f'Saltadas (totales no reconcilian, sin tocar): {saltadas}')
print(f'Sin rubro MO para asignar (revisar): {sin_rubro}')
