"""
Consolida en 1_Presupuesto las tareas que el Resumen original abrió en 2 (para asignar 2 rubros,
porque no tenía la apertura de 4 columnas de la v8): 2.01.01+2.01.02 -> 2.01 y 24.10.01+24.10.02
-> 24.10. Un único código por tarea; los rubros se asignan en las 4 columnas A/B/C/D de la v8
según el Resumen original (Pto. Costos). Preserva costo, venta y margen. NO toca otras filas.
"""
import openpyxl, warnings, sys
warnings.filterwarnings('ignore'); sys.stdout.reconfigure(encoding='utf-8')

RES = 'archivos/output/SVD_4140_Resumen_de_Obra_v8_2.xlsx'
wb = openpyxl.load_workbook(RES)
wbv = openpyxl.load_workbook(RES, data_only=True)
ws = wb['1_Presupuesto']; wsv = wbv['1_Presupuesto']

# localizar filas por código string
rows = {}
for r in range(5, ws.max_row + 1):
    e = str(wsv.cell(r, 5).value or '')
    if e in ('2.01.01', '2.01.02', '24.10.01', '24.10.02'):
        rows[e] = r
print('filas:', rows)

def setc(r, col, val):
    ws.cell(r, col).value = val   # OJO: ws.cell(r,col,None) NO borra (None=="sin valor"); usar .value

def blank(r):
    # vacía la fila (deja las fórmulas N/O-V que computan 0); rubros a "-", código/desc/costos vacíos
    for col in (1, 2, 3, 4):   # A-D rubros
        ws.cell(r, col).value = '-'
    for col in (5, 7, 9, 10, 11, 12, 13, 19):  # E cod, G desc, I cant, J-M costos, S venta
        ws.cell(r, col).value = None

# ---- 2.01 (merge en la fila de 2.01.01) ----
r01, r02 = rows['2.01.01'], rows['2.01.02']
s01 = wsv.cell(r01, 19).value or 0
s02 = wsv.cell(r02, 19).value or 0
l02 = wsv.cell(r02, 12).value or 0          # MO/ALB de 2.01.02
setc(r01, 5, 2.01)                          # E = 2.01 (float, matchea otros códigos y 1_Composicion)
setc(r01, 12, l02)                          # L = MO/ALB
setc(r01, 4, 'Homigón MO')                  # D = Rubro MO/ALB
setc(r01, 1, '-'); setc(r01, 2, '-')        # A/B = -
# C (Rubro MO/OTR)='Movimiento de Suelos' y K(MO/OTR)/M(EQ) ya están en r01
setc(r01, 19, round(s01 + s02, 6))          # S venta = suma de ambas porciones
blank(r02)

# ---- 24.10 (merge en la fila de 24.10.01, que tiene la MO/ALB) ----
r1, r2 = rows['24.10.01'], rows['24.10.02']
s1 = wsv.cell(r1, 19).value or 0
s2 = wsv.cell(r2, 19).value or 0
j2 = wsv.cell(r2, 10).value or 0            # MT de 24.10.02
a2 = wsv.cell(r2, 1).value                  # Rubro MT de 24.10.02 ('Gastos Generales')
setc(r1, 5, 24.10)                          # E = 24.10 (float -> 24.1)
setc(r1, 10, j2)                            # J = MT
setc(r1, 1, a2)                             # A = Rubro MT ('Gastos Generales')
setc(r1, 2, '-'); setc(r1, 3, '-')          # B/C = -
# L(MO/ALB) y D='Albañilería MO' ya están en r1
setc(r1, 19, round(s1 + s2, 6))             # S venta = suma
blank(r2)

wb.save(RES)
print('OK - merge guardado. 2.01 en r%d, 24.10 en r%d.' % (r01, r1))
