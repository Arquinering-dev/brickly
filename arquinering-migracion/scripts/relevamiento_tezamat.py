# -*- coding: utf-8 -*-
"""Genera el relevamiento de movimientos a modificar en Tezamat (CH 2171).

Lee 2_Movimientos (extracto crudo, con col P = Observaciones original) y
2_Subcontratos (maestro confirmado), clasifica cada movimiento según las reglas
de carga nuevas y emite docs/Relevamiento_Tezamat_CH2171.md con, para cada
movimiento, su formato ORIGINAL (como está hoy en Tezamat) y la versión AJUSTADA.

Reglas de carga (definidas con Arquinering):
  1. Pagos a subcontratos -> Observaciones con "CH-SC-NNN AVANCE" + rubro del SC.
  2. SC que ajustan CAC -> separar monto CAC del base, con "CH-SC-NNN CAC".
  3. Quincenas (nómina propia) -> cargar el costo como movimiento en Tezamat.
  4. (pendiente plan de cuentas) Desc Cuenta debe ser un rubro canónico.
"""
import io
import re
import sys
import unicodedata

import openpyxl

SRC = "archivos/referencia/CH_2171_Resumen_de_Obra_v8_6.xlsx"
OUT = "docs/Relevamiento_Tezamat_CH2171.md"

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SC_RE = re.compile(r"(CH-SC-\d+)", re.I)
SC_TIPOS = ("CARGAS SOCIALES", "AVANCE", "QUINCENA", "CAC")
# cuentas Tezamat sin rubro canónico (se cierran con el plan de cuentas)
SIN_RUBRO = {"aguas", "varios ferreteria", "gastos generales", "h. gestoria",
             "h. seguridad e higiene", "gastos en personal", "seguridad e higiene"}


def norm(v):
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def money(x):
    return f"${x:,.0f}".replace(",", ".")


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # maestro de subcontratos: id -> (proveedor, rubro, ajusta_cac)
    sc = {}
    ws = wb["2_Subcontratos"]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[0] and str(r[0]).strip().upper().startswith("CH-SC"):
            sc[str(r[0]).strip().upper()] = {
                "prov": str(r[1]).strip() if r[1] else "",
                "rubro": str(r[2]).strip() if r[2] else "",
                "ajusta_cac": str(r[5]).strip().upper() == "SI" if r[5] else False,
            }

    ws = wb["2_Movimientos"]
    movs = []
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if r[0] is None and r[1] is None and not (r[8] or r[9]):
            continue
        obs_e = str(r[4]).strip() if r[4] else ""
        obs_p = str(r[15]).strip() if len(r) > 15 and r[15] else ""
        movs.append({
            "row": i,
            "cuenta": str(r[0]).strip() if r[0] is not None else "",
            "desc": str(r[1]).strip() if r[1] else "",
            "fecha": r[2].date().isoformat() if hasattr(r[2], "date") else (str(r[2]) if r[2] else ""),
            "prov": str(r[5]).strip() if r[5] else "",
            "real": (r[8] or 0) - (r[9] or 0),
            "obs_e": obs_e,
            "obs_p": obs_p,
        })

    # clasificar
    for m in movs:
        m.update(_clasificar(m, sc))

    wb.close()
    _emit(movs, sc)


def _clasificar(m, sc):
    e_up = m["obs_e"].upper()
    p = m["obs_p"]
    scm = SC_RE.search(m["obs_e"])
    es_sintetico_cac = p.lower().startswith("(cac desagregado")
    es_sintetico_q = p.lower().startswith("(generado de 2_quincenas")

    if es_sintetico_cac:
        sid = SC_RE.search(m["obs_e"]).group(1).upper()
        base_row = re.search(r"fila (\d+)", p)
        return {
            "tipo": "2-CAC",
            "accion": f"CREAR en Tezamat un movimiento de CAC separado del pago base ({sid}). "
                      f"Hoy el pago base y su CAC van juntos en un solo asiento.",
            "desc_adj": m["desc"], "obs_adj": f"{sid} CAC",
            "orig_label": f"(no existe como asiento propio — CAC incluido en el pago base, "
                          f"fila {base_row.group(1) if base_row else '?'})",
        }
    if es_sintetico_q:
        per = re.search(r"(20\d\d-\d\d)\s*(\dQ)", m["obs_e"])
        return {
            "tipo": "3-QUINCENA",
            "accion": "CARGAR en Tezamat el costo de la quincena (nómina propia UOCRA). "
                      "Hoy no existe como asiento; se toma de la planilla 2_Quincenas.",
            "desc_adj": m["desc"], "obs_adj": m["obs_e"],
            "orig_label": "(no existe en Tezamat — costo de quincena no asentado)",
        }
    if scm:
        sid = scm.group(1).upper()
        tipo_pago = next((t for t in SC_TIPOS if t in e_up), "AVANCE")
        master = sc.get(sid, {})
        rubro_master = master.get("rubro", "")
        # ¿el Desc Cuenta actual ya corresponde al rubro del SC?
        rfix = bool(rubro_master) and norm(rubro_master).split()[0] not in norm(m["desc"])
        acc = [f'Agregar "{sid} {tipo_pago}" en Observaciones (hoy: proveedor en texto libre)']
        desc_adj = m["desc"]
        if rfix:
            acc.append(f'Reclasificar Desc Cuenta a rubro del SC: "{rubro_master}" '
                       f'(hoy "{m["desc"]}", genérico)')
            desc_adj = rubro_master
        if master.get("ajusta_cac") and tipo_pago == "AVANCE":
            acc.append("SC ajusta CAC: verificar si este pago lleva CAC a separar")
        return {"tipo": "1-SC", "accion": " · ".join(acc),
                "desc_adj": desc_adj, "obs_adj": f"{sid} {tipo_pago}",
                "orig_label": m["obs_p"] or "(sin observación)"}
    if norm(m["desc"]) in SIN_RUBRO:
        return {
            "tipo": "4-RUBRO",
            "accion": "Pendiente plan de cuentas Tezamat: esta cuenta no mapea a un rubro "
                      "canónico de obra (cae como gasto sin presupuesto en el control).",
            "desc_adj": "(definir con plan de cuentas)", "obs_adj": m["obs_e"],
            "orig_label": m["obs_p"] or m["obs_e"],
        }
    return {"tipo": "0-OK", "accion": "", "desc_adj": m["desc"], "obs_adj": m["obs_e"],
            "orig_label": m["obs_p"] or m["obs_e"]}


def _emit(movs, sc):
    from collections import Counter
    cnt = Counter(m["tipo"] for m in movs)
    L = []
    P = L.append
    P("# Relevamiento de movimientos a modificar en Tezamat — CH 2171")
    P("")
    P("> Generado por `scripts/relevamiento_tezamat.py` desde "
      "`CH_2171_Resumen_de_Obra_v8_6.xlsx` (hoja `2_Movimientos`, con `Observaciones "
      "original` = carga cruda de Tezamat) y el maestro `2_Subcontratos`.")
    P("")
    P("Objetivo: que Arquinering corrija la carga en Tezamat para que el extracto salga "
      "correcto **sin curado manual** y el Excel/dashboard funcionen sobre datos históricos.")
    P("")
    P("## Reglas de carga nuevas (definidas con Arquinering)")
    P("1. **Pago a subcontrato** → Observaciones con `CH-SC-NNN AVANCE` (o `CAC`/`QUINCENA`/"
      "`CARGAS SOCIALES`) y Desc Cuenta = rubro del subcontrato.")
    P("2. **SC que ajusta CAC** → el monto de CAC va en un asiento separado del base, con "
      "`CH-SC-NNN CAC` en Observaciones.")
    P("3. **Quincenas (nómina propia UOCRA)** → el costo se asienta como movimiento en Tezamat.")
    P("4. **Rubro** → Desc Cuenta debe ser un rubro canónico (pendiente del plan de cuentas).")
    P("")
    P("## Subcontratos confirmados (maestro `2_Subcontratos`)")
    P("")
    P("| ID | Proveedor | Rubro | Ajusta CAC |")
    P("|----|-----------|-------|:----------:|")
    for sid, s in sc.items():
        P(f"| {sid} | {s['prov']} | {s['rubro']} | {'SÍ' if s['ajusta_cac'] else 'no'} |")
    P("")
    P("## Resumen del relevamiento")
    P("")
    P(f"- **Total movimientos analizados:** {len(movs)}")
    P(f"- **Tipo 1 — Pagos a subcontrato (agregar ID + rubro):** {cnt.get('1-SC',0)}")
    P(f"- **Tipo 2 — Separar CAC del base (crear asiento):** {cnt.get('2-CAC',0)}")
    P(f"- **Tipo 3 — Cargar quincenas en Tezamat (crear asiento):** {cnt.get('3-QUINCENA',0)}")
    P(f"- **Tipo 4 — Rubro sin mapear (pendiente plan de cuentas):** {cnt.get('4-RUBRO',0)}")
    P(f"- **Sin cambios:** {cnt.get('0-OK',0)}")
    P("")
    secciones = [
        ("1-SC", "Tipo 1 — Pagos a subcontratos: agregar ID de SC (+ corregir rubro)",
         "Existen en Tezamat con el proveedor en texto libre. Hay que etiquetarlos con el ID "
         "del subcontrato y, donde el Desc Cuenta es genérico (Preliminares, H. Gestoria), "
         "reclasificarlos al rubro del SC. El nombre exacto del rubro sigue el maestro "
         "`2_Subcontratos`; su forma canónica final (p. ej. \"Electricidad\"→\"Eléctrico\") "
         "se fija con el plan de cuentas."),
        ("2-CAC", "Tipo 2 — Separar el CAC del monto base",
         "Para los SC que ajustan CAC, el ajuste debe ir en un asiento aparte. Hoy el CAC "
         "está embebido en el pago base; hay que crear el movimiento de CAC. **Confirmar los "
         "montos exactos contra Tezamat / la certificación del SC.**"),
        ("3-QUINCENA", "Tipo 3 — Cargar quincenas (nómina propia) en Tezamat",
         "El costo de las quincenas UOCRA hoy NO está asentado en Tezamat (se toma de la "
         "planilla `2_Quincenas`). Hay que cargarlo como movimiento, con su rubro MO y la "
         "referencia de período."),
        ("4-RUBRO", "Tipo 4 — Cuentas sin rubro canónico (pendiente plan de cuentas)",
         "Estas cuentas Tezamat no mapean a un rubro de obra y caen como gasto sin "
         "presupuesto. La mayoría son **gastos indirectos / generales** (honorarios de "
         "seguridad e higiene, gestoría, volquetes, baños químicos, agua, EPP, gastos en "
         "personal) que probablemente vayan a Gastos Generales / indirectos, no a un rubro "
         "de obra. Se resuelven al recibir el plan de cuentas de Tezamat — NO modificar aún. "
         "Incluye un ajuste en negativo (fila 85) a revisar."),
    ]
    for tipo, titulo, intro in secciones:
        items = [m for m in movs if m["tipo"] == tipo]
        if not items:
            continue
        P(f"## {titulo}")
        P("")
        P(intro)
        P("")
        P("| # fila | Fecha | Monto | Desc Cuenta (hoy) | Observación original (hoy) | → Desc Cuenta ajustado | → Observación ajustada | Acción |")
        P("|:------:|-------|------:|-------------------|----------------------------|------------------------|------------------------|--------|")
        for m in items:
            P(f"| {m['row']} | {m['fecha']} | {money(m['real'])} | {m['desc']} | "
              f"{m['orig_label'][:70]} | {m['desc_adj']} | {m['obs_adj']} | {m['accion']} |")
        P("")
        P(f"**Subtotal Tipo {tipo[0]}: {len(items)} movimientos · {money(sum(x['real'] for x in items))}.**")
        P("")

    # apéndice: todos los movimientos sin cambio (resumen por cuenta)
    ok = [m for m in movs if m["tipo"] == "0-OK"]
    P("## Movimientos sin cambios")
    P("")
    P(f"{len(ok)} movimientos ya están cargados correctamente (cuenta con rubro válido, sin SC "
      "ni quincena ni CAC pendiente). Resumen por Desc Cuenta:")
    P("")
    from collections import defaultdict
    agg = defaultdict(lambda: [0, 0.0])
    for m in ok:
        agg[m["desc"]][0] += 1
        agg[m["desc"]][1] += m["real"]
    P("| Desc Cuenta | # mov | Monto |")
    P("|-------------|:-----:|------:|")
    for d, (n, t) in sorted(agg.items(), key=lambda x: -x[1][1]):
        P(f"| {d} | {n} | {money(t)} |")
    P("")

    with open(OUT, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L))
    print(f"OK -> {OUT}  ({len(movs)} movimientos, "
          f"{sum(1 for m in movs if m['tipo']!='0-OK')} con cambios)")


if __name__ == "__main__":
    main()
