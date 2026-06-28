# -*- coding: utf-8 -*-
"""
formato_v8.py — Motor de formato v8 FIEL A GDR.

Filosofía: para cada celda de datos, copia de la hoja espejo de GDR el formato real
(bordes, formato de número, nombre y tamaño de fuente, relleno, alineación) tomando una
"plantilla por columna" de GDR. Encima aplica:
  - Color de FUENTE semántico (definición v8): fórmula→verde, input→azul, pendiente→rojo, resto→negro.
  - Chrome definido: título (barra celeste), header (banda 1F4E78 + blanco), sección (▌).
  - Formato condicional copiado de GDR (resalta mes anterior, escalas de color, etc.).
  - Gridlines según GDR; anchos/altos; dropdowns propios de CH.

Uso:  python scripts/formato_v8.py <destino.xlsx> [referencia_GDR.xlsx]
"""
import sys, shutil
from copy import copy, deepcopy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

REF_DEFAULT = r"archivos\referencia\GDR_3760_Resumen_de_Obra_v8.xlsx"

TAB = {
    "0_CONFIG": "FFFFFF00", "0_Indice_CAC": "FFFFFF00", "0_Jornales_MO": "FFFFFF00",
    "1_Presupuesto": "FF595959", "1_GGBB": "FF595959", "1_Composicion": "FF595959",
    "2_Gastos": "FFFFC000", "2_Gastos_DirInd": "FFFFC000", "2_Quincenas": "FFFFC000",
    "2_Subcontratos": "FF2E75B6", "2_Pagos_Subc": "FF2E75B6", "2_Pagos_Quincena_SC": "FF2E75B6",
    "2_Certificaciones": "FF548235",
    "3_Dashboard": "FF1F4E78", "3_Control_Ppto": "FF1F4E78",
    "3_Control_Jornales": "FF1F4E78", "3_Cash_Flow": "FF1F4E78",
    "_Listas": "FFA6A6A6",
}
# CH: (title_row, header_row, data_row)
LAYOUT = {
    "0_CONFIG": (1, None, 3), "0_Indice_CAC": (1, 3, 4), "0_Jornales_MO": (1, 3, 4),
    "1_Presupuesto": (1, 4, 5), "1_GGBB": (None, None, 4), "1_Composicion": (None, 1, 2),
    "2_Gastos": (None, 1, 2), "2_Gastos_DirInd": (1, 4, 5), "2_Quincenas": (1, 3, 4),
    "2_Subcontratos": (1, 3, 4), "2_Pagos_Subc": (1, 3, 4), "2_Pagos_Quincena_SC": (1, 3, 4),
    "2_Certificaciones": (1, 4, 5), "3_Dashboard": (1, None, None),
    "3_Control_Ppto": (1, 3, 4), "3_Control_Jornales": (1, 3, 4), "3_Cash_Flow": (1, 2, 4),
    "_Listas": (None, 1, 2),
}
# GDR: fila de header (para ubicar la plantilla de datos). None = sin plantilla por columna.
GDR_HEADER = {
    "0_Indice_CAC": 3, "0_Jornales_MO": 3, "1_Presupuesto": 3, "1_Composicion": 1,
    "2_Gastos": 1, "2_Gastos_DirInd": 3, "2_Quincenas": 3, "2_Subcontratos": 3,
    "2_Pagos_Subc": 3, "2_Pagos_Quincena_SC": 3, "2_Certificaciones": 4,
    "3_Control_Ppto": 3, "3_Control_Jornales": 3, "3_Cash_Flow": 4, "_Listas": 1,
}
FREEZE_DEFAULT = {"2_Gastos": "A2", "1_Composicion": "A2"}
# Columnas de CH que están desalineadas respecto a GDR: usar la plantilla de otra columna GDR.
# 0_Jornales_MO: CH tiene F vacía, así que /hora (G-J) toma el formato de las tarifas $/día (B).
TEMPLATE_FROM = {"0_Jornales_MO": {7: 2, 8: 2, 9: 2, 10: 2}}
INPUT_FILL = {"DDEEFF", "FFDDEEFF"}
PEND_FILL = {"FFFF99", "FFFFFF99", "FFFF00", "FFFFFF00"}
TITLE_FILL = PatternFill("solid", fgColor="D6E4F0")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DVS = {
    "2_Subcontratos": [("F", '"SI,NO"', 4, 200)],
    "2_Certificaciones": [("D", '"B65,N35,GDN"', 5, 200), ("C", '"1,2,3"', 5, 200)],
}


def fill_rgb(c):
    if c.fill and c.fill.patternType == "solid":
        fg = c.fill.fgColor
        if fg.type == "rgb" and isinstance(fg.rgb, str):
            return fg.rgb.upper()
    return None


def gdr_input_cols(g):
    cols = set()
    for row in g.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if (c.font and c.font.color and c.font.color.type == "rgb"
                    and str(c.font.color.rgb) == "FF0000FF"):
                cols.add(c.column)
    return cols


def font_color_for(c, input_cols, data_row):
    if isinstance(c.value, str) and c.value.startswith("="):
        return "FF008000"
    if fill_rgb(c) in PEND_FILL:
        return "FFFF0000"
    if fill_rgb(c) in INPUT_FILL:
        return "FF0000FF"
    if c.column in input_cols and (not data_row or c.row >= data_row):
        return "FF0000FF"
    return "FF000000"


def has_border(c):
    b = c.border
    return any(s and s.style for s in (b.top, b.bottom, b.left, b.right))


def col_template(g, gdr_data_row, col, scan=80):
    """Plantilla de formato de GDR para una columna. Prefiere una celda con borde (para tomar el
    rango de la grilla); si la hoja no usa bordes (usa gridlines), cae en la primera celda con
    valor para copiar igual el formato de número y la fuente."""
    tmpl_b = first_b = last_b = tmpl_any = None
    for r in range(gdr_data_row, min(gdr_data_row + scan, g.max_row) + 1):
        c = g.cell(r, col)
        if c.value is not None and tmpl_any is None:
            tmpl_any = c
        if has_border(c):
            if tmpl_b is None:
                tmpl_b, first_b = c, r
            last_b = r
    return (tmpl_b or tmpl_any), first_b, last_b


def last_data_row(ws, data_row, keycols=(1, 2, 3, 4, 5)):
    last = data_row
    for r in range(data_row, ws.max_row + 1):
        if any(ws.cell(r, c).value is not None for c in keycols):
            last = r
    return last


def last_col_with_value(ws, row):
    last = 1
    for col in range(1, ws.max_column + 1):
        if ws.cell(row, col).value is not None:
            last = col
    return last


def aplicar(destino, ref):
    shutil.copyfile(destino, destino.replace(".xlsx", "_bak_pre_formato.xlsx"))
    wb = openpyxl.load_workbook(destino)
    gdr = openpyxl.load_workbook(ref)
    resumen = []

    for sh in wb.sheetnames:
        ws = wb[sh]
        g = gdr[sh] if sh in gdr.sheetnames else None
        title_row, header_row, data_row = LAYOUT.get(sh, (None, None, None))
        input_cols = gdr_input_cols(g) if g else set()
        special = {r for r in (title_row, header_row) if r}

        # ---- chrome de hoja ----
        if sh in TAB:
            ws.sheet_properties.tabColor = TAB[sh]
        if g:
            ws.sheet_view.showGridLines = g.sheet_view.showGridLines  # fiel a GDR
        fz = f"A{data_row}" if data_row else (g.freeze_panes if g else None) or FREEZE_DEFAULT.get(sh)
        if fz:
            ws.freeze_panes = fz
        if g:
            for letra, dim in g.column_dimensions.items():
                if dim.width:
                    ws.column_dimensions[letra].width = dim.width

        # ---- copia de formato por columna (plantilla GDR) ----
        gdr_h = GDR_HEADER.get(sh)
        if g and gdr_h is not None and data_row:
            gdr_data_row = gdr_h + 1
            ch_last = last_data_row(ws, data_row)
            maxcol = max(ws.max_column, g.max_column)
            # altura de fila de datos de GDR
            gh = g.row_dimensions.get(gdr_data_row)
            row_h = gh.height if gh else None
            # Pase 1: regla de fuente sobre TODAS las celdas con valor (incl. subtítulos) — mata el gris
            for r in range(1, ch_last + 1):
                if r in special:
                    continue
                for col in range(1, maxcol + 1):
                    cc = ws.cell(r, col)
                    if cc.value is None:
                        continue
                    f = cc.font
                    cc.font = Font(name="Aptos Narrow", size=f.sz or 10, bold=f.b,
                                   italic=f.i, color=font_color_for(cc, input_cols, data_row))
            # Pase 2: plantilla de formato por columna desde GDR (bordes/numfmt/fill/fuente)
            tfrom = TEMPLATE_FROM.get(sh, {})
            for col in range(1, maxcol + 1):
                tmpl, gf, gl = col_template(g, gdr_data_row, tfrom.get(col, col))
                if tmpl is None:
                    continue
                end = max(gl, ch_last) if gl else ch_last
                for r in range(data_row, end + 1):
                    if r in special:
                        continue
                    cc = ws.cell(r, col)
                    cc.border = copy(tmpl.border)
                    cc.number_format = tmpl.number_format
                    cc.alignment = copy(tmpl.alignment)
                    if fill_rgb(cc) not in PEND_FILL:
                        cc.fill = copy(tmpl.fill)
                    cc.font = Font(name=tmpl.font.name, size=tmpl.font.sz,
                                   bold=tmpl.font.b, italic=tmpl.font.i,
                                   color=font_color_for(cc, input_cols, data_row))
            if row_h:
                for r in range(data_row, ch_last + 1):
                    ws.row_dimensions[r].height = row_h
        else:
            # hojas irregulares (CONFIG/GGBB/Dashboard): solo regla de fuente
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None or c.row in special:
                        continue
                    f = c.font
                    c.font = Font(name=f.name or "Aptos Narrow", size=f.sz or 11,
                                  bold=f.b, italic=f.i,
                                  color=font_color_for(c, input_cols, data_row))

        width = max(last_col_with_value(ws, header_row) if header_row else 1,
                    last_col_with_value(ws, title_row) if title_row else 1)

        # ---- chrome: título / header / sección ----
        if title_row:
            for col in range(1, width + 1):
                c = ws.cell(title_row, col)
                c.font = Font(name="Aptos Narrow", size=12, bold=True, color="FF1F3864")
                c.fill = TITLE_FILL
        if header_row:
            for col in range(1, width + 1):
                c = ws.cell(header_row, col)
                c.font = Font(name="Aptos Narrow", size=(c.font.sz or 10), bold=True, color="FFFFFFFF")
                c.fill = HEADER_FILL
                c.alignment = HEADER_ALIGN
        for row in ws.iter_rows():
            for c in row:
                if c.row in special:
                    continue
                if isinstance(c.value, str) and c.value.strip().startswith("▌"):
                    for col in range(1, width + 1):
                        cc = ws.cell(c.row, col)
                        cc.fill = SECTION_FILL
                        cc.font = Font(name="Aptos Narrow", size=(cc.font.sz or 11),
                                       bold=True, color="FFFFFFFF")
                    break

        # ---- formato condicional copiado de GDR ----
        if g:
            for cf in g.conditional_formatting:
                rng = str(cf.sqref)
                for rule in cf.rules:
                    try:
                        ws.conditional_formatting.add(rng, deepcopy(rule))
                    except Exception:
                        pass

        # ---- dropdowns propios de CH ----
        for col, f1, r0, r1 in DVS.get(sh, []):
            dv = DataValidation(type="list", formula1=f1, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col}{r0}:{col}{r1}")

        resumen.append((sh, str(g.sheet_view.showGridLines) if g else "-",
                        title_row or "-", header_row or "-"))

    wb.save(destino)
    wb.close(); gdr.close()
    print(f"{'Hoja':<22}{'Grid':<8}{'Titulo':<8}{'Header'}")
    for r in resumen:
        print(f"{r[0]:<22}{str(r[1]):<8}{str(r[2]):<8}{r[3]}")
    print(f"\nOK — formato fiel a GDR en {len(resumen)} hojas.")


if __name__ == "__main__":
    destino = sys.argv[1]
    ref = sys.argv[2] if len(sys.argv) > 2 else REF_DEFAULT
    aplicar(destino, ref)
