"""
recalc.py — Verificador de fórmulas post-save
Arquinering S.R.L. — Proyecto Migración Resumen de Obra

Uso:
    python scripts/recalc.py <ruta_archivo.xlsx>

Reporta todas las celdas con fórmulas que contienen errores de Excel
(#REF!, #NAME?, #VALUE!, #DIV/0!, #N/A, #NULL!, #NUM!).

IMPORTANTE: openpyxl no ejecuta cálculos de Excel. Este script detecta
errores que quedaron "grabados" en el archivo al último guardado desde Excel,
o errores que openpyxl puede identificar en la estructura de la fórmula.

Para una validación completa, abrir el archivo en Excel y usar
Fórmulas → Comprobación de errores después de ejecutar este script.
"""

import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter

# La consola de Windows (PowerShell/cmd) usa cp1252 por defecto y no puede
# codificar los emojis de los mensajes. Forzar UTF-8 en stdout/stderr evita
# el UnicodeEncodeError sin necesidad de setear PYTHONIOENCODING a mano.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

# Strings que representan errores en celdas Excel
ERRORES_EXCEL = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}

def verificar_archivo(ruta):
    if not os.path.exists(ruta):
        print(f"\n❌ Archivo no encontrado: {ruta}")
        sys.exit(1)

    nombre = os.path.basename(ruta)
    print(f"\n🔍 Verificando: {nombre}")
    print("=" * 60)

    try:
        # Cargar con data_only=True para leer valores calculados
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        print(f"❌ Error al abrir el archivo: {e}")
        sys.exit(1)

    errores_encontrados = []
    total_celdas = 0
    total_formulas = 0

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        for fila in ws.iter_rows():
            for celda in fila:
                total_celdas += 1
                valor = celda.value

                # Detectar errores en valores de celda
                if isinstance(valor, str) and valor.strip().upper() in ERRORES_EXCEL:
                    errores_encontrados.append({
                        "hoja": nombre_hoja,
                        "celda": f"{get_column_letter(celda.column)}{celda.row}",
                        "error": valor.strip()
                    })

    wb.close()

    # También verificar fórmulas en el archivo sin data_only para detectar
    # referencias estructuralmente inválidas
    try:
        wb2 = openpyxl.load_workbook(ruta, data_only=False)
        for nombre_hoja in wb2.sheetnames:
            ws = wb2[nombre_hoja]
            for fila in ws.iter_rows():
                for celda in fila:
                    valor = celda.value
                    if isinstance(valor, str) and valor.startswith("="):
                        total_formulas += 1
                        # Detectar referencias claramente rotas en la fórmula
                        if "#REF!" in valor:
                            coord = f"{get_column_letter(celda.column)}{celda.row}"
                            # Evitar duplicados
                            ya_reportado = any(
                                e["hoja"] == nombre_hoja and e["celda"] == coord
                                for e in errores_encontrados
                            )
                            if not ya_reportado:
                                errores_encontrados.append({
                                    "hoja": nombre_hoja,
                                    "celda": coord,
                                    "error": "#REF! (en fórmula)",
                                    "formula": valor[:80] + ("..." if len(valor) > 80 else "")
                                })
        wb2.close()
    except Exception:
        pass  # Si falla la segunda lectura, reportar lo que tenemos

    # Reporte final
    print(f"📊 Total celdas escaneadas: {total_celdas:,}")
    print(f"📐 Fórmulas detectadas:     {total_formulas:,}")
    print()

    if not errores_encontrados:
        print("✅ 0 errores de fórmula — archivo listo para continuar")
        print()
        return 0
    else:
        print(f"❌ {len(errores_encontrados)} error(es) encontrado(s):")
        print()
        for e in errores_encontrados:
            linea = f"   Hoja: {e['hoja']:<25} | Celda: {e['celda']:<8} | Error: {e['error']}"
            if "formula" in e:
                linea += f"\n   {'':>5}Fórmula: {e['formula']}"
            print(linea)
        print()
        print("⚠️  Resolver todos los errores antes de continuar con el trabajo.")
        print()
        return 1


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/recalc.py <ruta_archivo.xlsx>")
        print("Ejemplo: python scripts/recalc.py archivos/output/CH_2171_v8_WIP.xlsx")
        sys.exit(1)

    ruta = sys.argv[1]
    codigo_salida = verificar_archivo(ruta)
    sys.exit(codigo_salida)
