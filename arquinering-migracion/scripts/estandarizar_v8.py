"""
estandarizar_v8.py — Emprolijado estructural de los Resumen v8 (sesión 2026-06-14).

Aplica DOS cambios seguros y aditivos a cada obra (no toca cadenas de fórmula
existentes; solo renombra un header y AGREGA columnas-fórmula al final):

  1) 1_Composicion: renombra el header "en PPTO?" -> "Cod_Item_Ppto"
     (la columna contiene el código de ítem del presupuesto para las partidas
      usadas; el nombre viejo era engañoso). Detección por texto del header.

  2) 1_Presupuesto: agrega al final dos columnas-fórmula con HEADER IDÉNTICO en
     ambas obras, que normalizan la divergencia totales(GDR) vs por-unidad(CH):
       - "Costo_ud"    = costo por unidad
       - "Costo_total" = costo total (= costo/ud × cant)
     GDR guarda O como TOTAL  -> Costo_ud = O/Cant ; Costo_total = O
     CH  guarda O como /UD    -> Costo_ud = O      ; Costo_total = O*Cant

Las columnas-fórmula quedan VACÍAS de valor hasta recalcular en Excel
(scripts/excel_recalc.py). NO usar delete/insert: solo append al final.
"""
import sys
import unicodedata
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def norm(v):
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().replace("\n", " ").split())


# estilos estándar v8
F_HEADER = Font(name="Aptos Narrow", bold=True, color="FFFFFFFF")
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
ALIGN_HEADER = Alignment(horizontal="center", vertical="center")
F_FORMULA = Font(name="Aptos Narrow", color="FF008000")  # verde = fórmula


def find_col_in_row(ws, row, pred):
    for c in range(1, ws.max_column + 1):
        if pred(norm(ws.cell(row=row, column=c).value)):
            return c
    return None


def find_header_row(ws, must_have):
    """Primera fila (1..15) que tenga una celda cuyo header matchea must_have."""
    for r in range(1, 16):
        for c in range(1, ws.max_column + 1):
            if must_have(norm(ws.cell(row=r, column=c).value)):
                return r
    return None


def rename_en_ppto(wb):
    ws = wb["1_Composicion"]
    # header está en la fila 1
    col = find_col_in_row(ws, 1, lambda h: h == "en ppto?" or h == "en ppto")
    if col is None:
        # ya renombrado?
        col = find_col_in_row(ws, 1, lambda h: "cod_item" in h or "cod item" in h)
        if col:
            return f"1_Composicion: ya estaba renombrado (col {get_column_letter(col)})"
        raise RuntimeError("1_Composicion: no encontré el header 'en PPTO?'")
    ws.cell(row=1, column=col).value = "Cod_Item_Ppto"
    return f"1_Composicion: header col {get_column_letter(col)} -> 'Cod_Item_Ppto'"


def add_costo_cols(wb, obra):
    ws = wb["1_Presupuesto"]
    hr = find_header_row(ws, lambda h: h == "cant")
    if hr is None:
        raise RuntimeError("1_Presupuesto: no encontré la fila de header (Cant)")
    cJ = find_col_in_row(ws, hr, lambda h: h == "cant")
    cO = find_col_in_row(ws, hr, lambda h: h.startswith("costo") and ("unit" in h or "total" in h))
    if not cJ or not cO:
        raise RuntimeError(f"1_Presupuesto: falta Cant({cJ}) u O/Costo({cO})")
    Jl, Ol = get_column_letter(cJ), get_column_letter(cO)

    # idempotencia: si ya existen, no duplicar
    if find_col_in_row(ws, hr, lambda h: h == "costo_ud"):
        return f"1_Presupuesto: columnas estándar ya existían (skip)"

    c1 = ws.max_column + 2      # gap de 1 col, append al final
    c2 = c1 + 1
    for c, name in ((c1, "Costo_ud"), (c2, "Costo_total")):
        hc = ws.cell(row=hr, column=c)
        hc.value = name
        hc.font = F_HEADER
        hc.fill = FILL_HEADER
        hc.alignment = ALIGN_HEADER

    n = 0
    for r in range(hr + 1, ws.max_row + 1):
        oval = ws.cell(row=r, column=cO).value
        if oval in (None, ""):           # filas de grupo/blank: sin O -> skip
            continue
        if obra == "GDR":   # O = total
            f_ud = f"=IFERROR({Ol}{r}/{Jl}{r},0)"
            f_tot = f"=+{Ol}{r}"
        else:               # CH: O = /ud
            f_ud = f"=+{Ol}{r}"
            f_tot = f"=+{Ol}{r}*{Jl}{r}"
        nf = ws.cell(row=r, column=cO).number_format
        for c, fx in ((c1, f_ud), (c2, f_tot)):
            cell = ws.cell(row=r, column=c)
            cell.value = fx
            cell.font = F_FORMULA
            cell.number_format = nf
        n += 1
    return (f"1_Presupuesto[{obra}]: +Costo_ud({get_column_letter(c1)}) "
            f"+Costo_total({get_column_letter(c2)}) en {n} filas "
            f"(Cant={Jl}, Costo={Ol})")


def procesar(path, obra):
    print(f"\n=== {obra}: {path} ===")
    wb = openpyxl.load_workbook(path)  # con fórmulas
    print("  " + rename_en_ppto(wb))
    print("  " + add_costo_cols(wb, obra))
    wb.save(path)
    print("  ✓ guardado (openpyxl) — pendiente recalc en Excel")


if __name__ == "__main__":
    TARGETS = [
        ("dashboard/data/GDR_3760_Resumen_de_Obra_v8.xlsx", "GDR"),
        ("dashboard/data/CH_2171_Resumen_de_Obra_v8_1.xlsx", "CH"),
    ]
    for path, obra in TARGETS:
        procesar(path, obra)
    print("\nListo. Ahora correr: python scripts/excel_recalc.py <archivos>")
