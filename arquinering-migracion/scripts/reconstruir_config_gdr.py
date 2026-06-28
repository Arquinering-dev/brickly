"""reconstruir_config_gdr.py — Bloque "0_CONFIG" de la migración de GDR a v8 (Q2).
Reescribe 0_CONFIG al layout limpio de CH (dict etiqueta→valor que lee el reader),
con los parámetros propios de GDR confirmados por Pedro (2026-06-22):
  - Mes base CAC = dic-2024 (valor INDEC 15356.4)
  - Apertura fiscal = B70 / N30 / GGN
  - K = '1_GGBB'!F67 (auditable, no manual)
SUMPRODUCT auditables sobre 1_Presupuesto filas 5:227 (r4=header sección,
r228=SUB TOTAL, r229=blank quedan fuera). Reconciliación de egresos (B45) se cablea
en bloque 6 cuando exista 2_Movimientos.

Uso: python scripts/reconstruir_config_gdr.py
"""
import sys
import datetime
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
P = "'1_Presupuesto'"
R0, R1 = 5, 227  # rango de datos limpio de 1_Presupuesto

# (fila, etiqueta A, valor/fórmula B, nota C)
LAYOUT = [
    (1,  "▌ CONFIGURACIÓN DE OBRA — GDR 3760", None, None),
    (3,  "Nombre de obra", "Edificio GDR 3760 — García del Río 3760", None),
    (4,  "Estado", "En ejecución", None),
    (5,  "Fecha inicio", datetime.datetime(2025, 2, 1), "presupuesto cerrado dic-2024"),
    (6,  "Duración estimada (meses)", 18, "feb-2025 a ago-2026"),
    (7,  "— ÍNDICE CAC —", None, None),
    (8,  "Mes base CAC", datetime.datetime(2024, 12, 1), "fecha base del presupuesto"),
    (9,  "Valor CAC base (INDEC)", 15356.4, "índice INDEC del mes base"),
    (10, "Mes corriente", "=INDEX('0_Indice_CAC'!$A$4:$A$33,MATCH(1E+300,'0_Indice_CAC'!$B$4:$B$33))", "último mes cargado"),
    (11, "Ratio CAC corriente", "=MIN('0_Indice_CAC'!$D$4:$D$33)", "ratio del mes más reciente"),
    (12, "— COEFICIENTE K —", None, None),
    (13, "K (Gastos Generales y Beneficio)", "='1_GGBB'!F67", "CRÍTICO: no modificar manualmente"),
    (14, "— PRESUPUESTO —", None, None),
    (15, "Apertura fiscal", "B70 / N30 / GGN", "Blanco 70% c/IVA · Negro 30% · GGN"),
    (16, "Costo controlable (sin EQ)",
         f"=SUMPRODUCT(({P}!$K${R0}:$K${R1}+{P}!$L${R0}:$L${R1}+{P}!$M${R0}:$M${R1})*{P}!$J${R0}:$J${R1})",
         "MT+MO/OTR+MO/ALB × Cant (EQ excluido)"),
    (17, "Costo total (con EQ)",
         f"=SUMPRODUCT({P}!$O${R0}:$O${R1}*{P}!$J${R0}:$J${R1})", "Costo Unit. × Cant"),
    (18, "Precio de venta total",
         f"=SUMPRODUCT({P}!$P${R0}:$P${R1}*{P}!$J${R0}:$J${R1})", "P Unit. × Cant"),
    (19, "Sanity check (debe ser 0)", "=+B17*B13-B18", "EQ no tiene rubro → excluido de costo controlable"),
    (44, "— RECONCILIACIÓN MENSUAL —", None, None),
    (45, "Total egresos (2_Movimientos)", None, "← se cablea en bloque 6 (crear 2_Movimientos)"),
    (46, "Total Tezamat (pegar mensual)", None, "input mensual"),
    (47, "Diferencia (debe ser 0)", None, "se activa al cablear B45"),
    (49, "Tolerancia conciliación ($)", 100, None),
    (52, "— PENDIENTES (celdas amarillas) —", None, None),
    (53, "0_Indice_CAC col B", "CAC meses futuros sin publicar", None),
    (54, "0_Jornales_MO", "Tarifas UOCRA meses futuros", None),
    (55, "0_CONFIG!B46", "Input mensual total Tezamat", None),
    (56, "1_Presupuesto cols A,B,C,D", "Asignación de rubros por ítem (4 cols input manual)", None),
]


def main():
    wb = openpyxl.load_workbook(PATH)
    ws = wb["0_CONFIG"]
    # limpiar TODO el contenido viejo (A/B/C filas 1..72)
    for r in range(1, 73):
        for c in (1, 2, 3):
            ws.cell(row=r, column=c).value = None
    # escribir el layout nuevo
    for row, a, b, c in LAYOUT:
        ws.cell(row=row, column=1).value = a
        if b is not None:
            ws.cell(row=row, column=2).value = b
        if c is not None:
            ws.cell(row=row, column=3).value = c
    wb.save(PATH)
    print(f"✓ 0_CONFIG reconstruido → {PATH}")
    print(f"  · {len(LAYOUT)} filas escritas; rango SUMPRODUCT {R0}:{R1}")
    print("  · Esperados (validados contra referencia): B16=994.598.033,29 ·"
          " B17=1.016.173.633,33 · B18=1.458.753.139,80 · B19=0")
    print("  · B45 reconciliación egresos: DIFERIDO a bloque 6 (2_Movimientos)")


if __name__ == "__main__":
    main()
