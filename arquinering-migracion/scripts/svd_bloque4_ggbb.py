"""svd_bloque4_ggbb.py — Migracion SVD 4140 a v8. Bloque 4.
Copia la hoja GGBB legacy de SVD (mismo template que GDR 1_GGBB, autocontenida) a la
hoja 1_GGBB del archivo de trabajo, y repunta el markup K (F66) en 0_CONFIG!B13 y
1_Presupuesto!P2. El reader del dashboard ancla por texto -> absorbe el offset.
"""
import sys
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
LEGACY = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"
MARKUP_CELL = "F66"   # K = 1+F64 en la GGBB legacy de SVD


def build():
    wb = openpyxl.load_workbook(DST)
    lg = openpyxl.load_workbook(LEGACY)["GGBB"]
    g = wb["1_GGBB"]

    # limpiar 1_GGBB (datos heredados de GDR) en el rango que vamos a poblar
    for r in range(1, max(g.max_row, lg.max_row) + 1):
        for c in range(1, 47):
            g.cell(row=r, column=c).value = None

    # copiar valores/formulas de la GGBB legacy (autocontenida: solo refs internas)
    ncopied = 0
    for r in range(1, lg.max_row + 1):
        for c in range(1, min(lg.max_column, 46) + 1):
            v = lg.cell(row=r, column=c).value
            if v is not None:
                g.cell(row=r, column=c).value = v
                ncopied += 1

    # repuntar el markup K
    wb["0_CONFIG"]["B13"] = f"='1_GGBB'!{MARKUP_CELL}"
    wb["1_Presupuesto"]["P2"] = f"='1_GGBB'!{MARKUP_CELL}"

    wb.save(DST)
    print(f"✓ Bloque 4 → {DST}")
    print(f"  · 1_GGBB: {ncopied} celdas copiadas de la GGBB legacy (template GDR, sin refs externas)")
    print(f"  · markup K = '1_GGBB'!{MARKUP_CELL} (CONFIG!B13, 1_Presupuesto!P2)")


if __name__ == "__main__":
    build()
