"""Bloque 8 — Elimina las hojas legacy de GDR que no van en la estructura v8 nueva.
Q1 (Pedro 2026-06-22): 0_Indice_CAC y 0_Jornales_MO SE MANTIENEN locales (el circuito Cert_*
lee 0_Indice_CAC!B; nada de links vivos entre archivos). Solo se borran las de análisis/legacy.
Pre-verificado: 0 refs desde hojas sobrevivientes hacia las borradas.
"""
import sys
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
DEL = ["3_Dashboard", "3_Control_Ppto", "3_Control_Jornales", "3_Cash_Flow",
       "2_Certificaciones", "2_Gastos", "2_Pagos_Subc", "2_Pagos_Quincena_SC"]

wb = openpyxl.load_workbook(PATH)
borradas = []
for s in DEL:
    if s in wb.sheetnames:
        del wb[s]
        borradas.append(s)
wb.save(PATH)
print(f"✓ Eliminadas {len(borradas)} hojas legacy: {borradas}")
print(f"  Hojas restantes ({len(wb.sheetnames)}): {wb.sheetnames}")
