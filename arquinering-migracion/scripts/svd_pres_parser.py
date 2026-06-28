"""
Bloque 2 — Parseo del presupuesto SVD 4140 (El Salvador_Pres 04) y armado de la
composicion por tarea (PTOs + partidas APU) siguiendo el Manual APU Unificado v3.

Para cada item de '01' (184 decimales) lee las formulas de F/G/I/J/L, despacha por
ruta (la formula da el puntero exacto) y construye filas de COMPOSICION por-unidad:
  - Caso 3 (APU 09.24!Exx) -> usa la composicion P-XX parseada en Bloque 1.
  - Caso 2 (Tareas!Hxx)    -> expande sub-componentes de Tareas (MAT F/G, MO/ALB
                              jornales M/N/O/P, MO/OTR col L).
  - Caso 5 (ELECT/COVE...) -> expande la seccion de la hoja auxiliar.
  - Caso 4a (SUBCONTRATOS) -> MO/OTR (desc terminal).
  - Caso 4b (P.MO/MO HA AING) -> MO/ALB; MO HA SILVA -> MO/OTR.
  - Caso 1 (valor pegado)  -> lump por columna.
  - L (EQ) -> Caso 6 lump porcentual (o EQ del APU en ruta 3).

Reconcilia cada item contra sus valores resueltos en '01' (H=MAT/ud, K=MO/ud, L=EQ/ud).
Resuelve ALB vs OTR por trazabilidad; marca dudas. Salida: budget_parsed.json + reporte.
"""
import openpyxl, warnings, json, os, re
warnings.filterwarnings('ignore')

PRES = 'archivos/fuente/El Salvador_Pres 04 (sin pintura).xlsx'
SCRATCH = os.environ.get('SCRATCH', '.')
APU_JSON = os.path.join(SCRATCH, 'apu_parsed.json')
OUT_JSON = os.path.join(SCRATCH, 'budget_parsed.json')

wf = openpyxl.load_workbook(PRES, data_only=False)
wv = openpyxl.load_workbook(PRES, data_only=True)
apu = json.load(open(APU_JSON, encoding='utf-8'))
# partidas APU por numero (ignora sufijos tipo '63 bis': el indice referencia enteros)
apu_by_num = {}
for p in apu['partidas']:
    pn = p['partida']
    if pn is None:
        continue
    m = re.match(r'\s*(\d+)', str(pn))
    if m:
        apu_by_num.setdefault(int(m.group(1)), p)

def norm(s):
    return re.sub(r'\s+', ' ', str(s)).strip() if s is not None else ''

REF = re.compile(r"=?\+?'?([^'!=+]+?)'?!\$?([A-Z]+)\$?(\d+)")
def parse_refs(formula):
    """Devuelve lista de (hoja, col, fila) referenciadas en una formula."""
    if not isinstance(formula, str) or not formula.startswith('='):
        return []
    out = []
    for m in REF.finditer(formula):
        out.append((m.group(1).strip(), m.group(2), int(m.group(3))))
    return out

def col_idx(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n

# ---------- Hojas de soporte ----------
# indice APU 09.24: fila -> P-num (col B), E/H/K resueltos
def parse_indice():
    wsf, wsv = wf['APU 09.24'], wv['APU 09.24']
    idx = {}
    for r in range(11, wsf.max_row + 1):
        pnum = wsf.cell(r, 2).value  # col B
        if isinstance(pnum, (int, float)):
            idx[r] = {'pnum': int(pnum),
                      'mat': wsv.cell(r, 5).value, 'mo': wsv.cell(r, 8).value,
                      'eq': wsv.cell(r, 11).value}
    return idx

# Cotizaciones: fila -> {desc, ud, precio}
def parse_cotiz():
    wsf, wsv = wf['Cotizaciones'], wv['Cotizaciones']
    cz = {}
    for r in range(3, wsf.max_row + 1):
        desc = norm(wsf.cell(r, 2).value)  # B
        precio = wsv.cell(r, 4).value      # D
        if desc and isinstance(precio, (int, float)):
            cz[r] = {'desc': desc, 'ud': norm(wsf.cell(r, 3).value), 'precio': round(precio, 4)}
    return cz

# SUBCONTRATOS (presupuesto): fila -> {desc, ud, precio}
def parse_subc():
    wsf, wsv = wf['SUBCONTRATOS'], wv['SUBCONTRATOS']
    # detectar header
    hdr = None
    for r in range(1, 12):
        row = [norm(wsf.cell(r, c).value).upper() for c in range(1, 8)]
        if 'TAREA' in row or 'DESCRIPCION' in row or 'CONTRATISTA' in row:
            hdr = r; head = row; break
    sc = {}
    # columnas: buscar desc(TAREA/DESCRIPCION), U, PU/precio
    desc_c = next((i + 1 for i, h in enumerate(head) if h in ('TAREA', 'DESCRIPCION')), 2)
    for r in range((hdr or 7) + 1, wsf.max_row + 1):
        desc = norm(wsf.cell(r, desc_c).value)
        if not desc:
            continue
        # precio: primer numerico resuelto en cols D..H
        precio = None; ud = ''
        for c in range(3, 9):
            v = wsv.cell(r, c).value
            if isinstance(v, (int, float)) and v > 0 and precio is None:
                precio = round(v, 4)
        sc[r] = {'desc': desc, 'ud': ud, 'precio': precio,
                 'contratista': norm(wsf.cell(r, 3).value)}
    return sc

# P.MO 09.24: fila -> categoria + jornal; y mapa col->categoria base
PMO_CAT = {4: 'ESPECIALIZADO', 5: 'OFICIAL', 6: 'MEDIO OFICIAL', 7: 'AYUDANTE',
           9: 'SOBRESTANTE', 10: 'JEFE DE OBRA', 11: 'COORD OBRAS', 12: 'CAPATAZ'}
def parse_pmo():
    wsv = wv['P.MO 09.24']
    pmo = {}
    for r, cat in PMO_CAT.items():
        j = wsv.cell(r, 2).value  # B
        pmo[r] = {'cat': cat, 'jornal': round(j, 4) if isinstance(j, (int, float)) else None}
    return pmo

# MO HA AING / SILVA: fila -> {desc, precio}
def parse_moha(sheet):
    wsf, wsv = wf[sheet], wv[sheet]
    moha = {}
    for r in range(4, wsf.max_row + 1):
        desc = norm(wsf.cell(r, 1).value)  # A
        # precio actualizado en C (P UNIT ACTUALIZADO); subtotal en D
        precio = wsv.cell(r, 3).value
        if desc and isinstance(precio, (int, float)):
            moha[r] = {'desc': desc, 'precio': round(precio, 4)}
    return moha

# Auxiliares: dada (hoja, col, fila) de un subtotal de seccion, expandir materiales.
# Estructura ELECT/etc: subtotal en col (I o H) = SUM(Ha:Hb); filas mat: B=desc,
# C=spec, D=ud, E=cant, G=P.ACT, H=subtotal=E*G.
def expand_aux(sheet, cell_col, cell_row):
    wsf, wsv = wf[sheet], wv[sheet]
    f = wsf.cell(cell_row, col_idx(cell_col)).value
    rng = None
    if isinstance(f, str):
        m = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', f.replace('$', ''))
        if m:
            rng = (int(m.group(2)), int(m.group(4)))
    if not rng:
        # el valor referenciado no es un SUM: tratar como lump unico
        v = wsv.cell(cell_row, col_idx(cell_col)).value
        return [{'tipo': 'MAT', 'subtipo': 'OTR', 'desc': f'{sheet} {cell_col}{cell_row}',
                 'cant': 1, 'desp': 0, 'precio': round(v, 4) if isinstance(v, (int, float)) else 0,
                 'fuente': f'aux:{sheet}'}], (v if isinstance(v, (int, float)) else 0)
    rows = []
    total = 0.0
    for r in range(rng[0], rng[1] + 1):
        desc = norm(wsf.cell(r, 2).value)  # B
        sub = wsv.cell(r, 8).value          # H subtotal
        cant = wsv.cell(r, 5).value          # E
        pact = wsv.cell(r, 7).value          # G
        if not desc or not isinstance(sub, (int, float)) or sub == 0:
            continue
        rows.append({'tipo': 'MAT', 'subtipo': 'OTR', 'desc': desc,
                     'ud': norm(wsf.cell(r, 4).value),
                     'cant': cant if isinstance(cant, (int, float)) else 1,
                     'desp': 0,
                     'precio': round(pact, 4) if isinstance(pact, (int, float)) else round(sub, 4),
                     'fuente': f'aux:{sheet}'})
        total += sub
    return rows, total

# ---------- Resolver por evaluacion-atribucion (Caso 1/4/5/6 multi-termino) ----------
INSUMO_SHEETS = {'Cotizaciones', 'SUBCONTRATOS', 'P.MO 09.24', 'ELECT', 'COVE', 'CALEFA',
                 'INCEND', 'ENCOFRADO', 'Mamposteria', 'Mampostería', 'Locales',
                 'MO HA AING', 'MO HA SILVA', 'GGBB', 'H°A° '}
REF2 = re.compile(r"(?:(?:'(?P<s1>[^']+)'|(?P<s2>[A-Za-z][A-Za-z0-9\. ]*?))!)?\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)")

def cval(sheet, col, row):
    ws = wv[sheet] if (sheet and sheet in wv.sheetnames) else wv['01']
    try:
        v = ws.cell(row, col_idx(col)).value
    except Exception:
        return 0
    return v if isinstance(v, (int, float)) else 0

def _refs_in(formula):
    out = []
    for m in REF2.finditer(formula):
        sheet = m.group('s1') or m.group('s2') or None
        out.append((m.start(), m.end(), sheet, m.group('col'), int(m.group('row'))))
    return out

def _eval_with(formula, refs, override):
    """Evalua la formula sustituyendo cada ref por override.get(idx, valor_real).
    Sustituye sobre la formula ORIGINAL (las posiciones de refs son de ahi) y recien
    despues recorta '=+' — si no, el desfase deja letras y devuelve None (=> lump)."""
    s = formula
    for i in range(len(refs) - 1, -1, -1):
        st, en, sheet, col, row = refs[i]
        val = override.get(i, cval(sheet, col, row))
        s = s[:st] + f'({val})' + s[en:]
    s = s.lstrip('=+ ').replace('$', '').replace('^', '**')
    if re.search(r'[A-Za-z]', s):  # quedo alguna ref sin resolver
        return None
    try:
        return eval(s, {'__builtins__': {}}, {})
    except Exception:
        return None

def resolve_cost(formula, resolved_val, tipo, sub_default, item_desc):
    """Devuelve filas de composicion atribuyendo el costo a cada insumo referenciado.
    tipo='MAT'|'MO'. Maneja multi-termino, coeficientes y refs a celda especifica."""
    if not isinstance(formula, str) or not formula.startswith('='):
        # valor pegado -> lump
        return [{'tipo': tipo, 'subtipo': sub_default, 'desc': item_desc, 'cant': 1, 'desp': 0,
                 'precio': round(resolved_val, 4), 'fuente': 'lump'}]
    refs = _refs_in(formula)
    insumo_idx = [i for i, (st, en, sh, c, r) in enumerate(refs) if sh in INSUMO_SHEETS]
    if not insumo_idx:
        return [{'tipo': tipo, 'subtipo': sub_default, 'desc': item_desc, 'cant': 1, 'desp': 0,
                 'precio': round(resolved_val, 4), 'fuente': 'lump'}]
    rows = []
    for i in insumo_idx:
        st, en, sheet, col, row = refs[i]
        # contribucion de este insumo: poner los OTROS insumos en 0
        ov = {j: 0 for j in insumo_idx if j != i}
        contrib = _eval_with(formula, refs, ov)
        if contrib is None or abs(contrib) < 1e-9:
            continue
        # si el ref es subtotal SUM de hoja aux -> expandir a materiales (escala por contrib/subtotal)
        aux_rows = None
        if sheet in INSUMO_SHEETS and tipo == 'MAT' and sheet not in ('Cotizaciones', 'H°A° '):
            fcell = wf[sheet].cell(row, col_idx(col)).value if sheet in wf.sheetnames else None
            if isinstance(fcell, str) and 'SUM(' in fcell.upper():
                er, tot = expand_aux(sheet, col, row)
                if tot:
                    sc = contrib / tot
                    for x in er:
                        x = dict(x); x['cant'] = round((x['cant'] or 1) * sc, 8)
                        rows.append(x)
                    aux_rows = True
        if aux_rows:
            continue
        # fila simple: desc/ subtipo segun hoja terminal
        sub_t, desc_t = sub_default, item_desc
        cant_t, precio_t = 1, round(contrib, 4)
        if sheet == 'Cotizaciones':
            cz = COTIZ.get(row)
            desc_t = cz['desc'] if cz else f'Cotizaciones!{col}{row}'
        elif sheet == 'SUBCONTRATOS':
            sub_t = 'OTR'
            sc = SUBC.get(row)
            desc_t = sc['desc'] if sc else f'SUBCONTRATOS!{col}{row}'
        elif sheet == 'P.MO 09.24':
            sub_t = sub_default  # la columna manda: col J (albañil)->ALB, col I (otro)->OTR
            desc_t = PMO.get(row, {}).get('cat', f'P.MO!{col}{row}')
            jor = PMO.get(row, {}).get('jornal')
            if jor and abs(jor) > 1e-6:  # itemizar como N jornales x salario (detalle de horas)
                cant_t, precio_t = round(contrib / jor, 6), round(jor, 4)
        elif sheet == 'MO HA AING':
            if row in MOHA_BLOCKS:  # bloque detallado -> decomponer jornales por categoria (ALB)
                blk = MOHA_BLOCKS[row]
                scale = (contrib / blk['precio']) if blk['precio'] else 1.0
                for cat, info in blk['cats'].items():
                    cant_cat = (info['jor'] / blk['rend']) * blk['gg'] * scale
                    rows.append({'tipo': 'MO', 'subtipo': 'ALB', 'desc': cat,
                                 'cant': round(cant_cat, 6), 'desp': 0,
                                 'precio': round(info['sal'], 4), 'fuente': 'MO HA AING (jornales)'})
                continue
            if row < 20:  # seccion resumen = precios SILVA (subcontrato) -> OTR
                sub_t = 'OTR'; desc_t = MOHA_AING.get(row, {}).get('desc', f'AING resumen!{row}')
            else:
                sub_t = 'ALB'; desc_t = MOHA_AING.get(row, {}).get('desc', f'AING!{row}')
        elif sheet == 'MO HA SILVA':
            sub_t = 'OTR'; desc_t = MOHA_SILVA.get(row, {}).get('desc', f'SILVA!{row}')
        elif sheet == 'H°A° ':  # computo de hormigon: nombrar segun el elemento (col A)
            a = norm(wf['H°A° '].cell(row, 1).value)
            desc_t = ('Encofrado ' + a) if a else item_desc  # E40/M29 (hierro/H°) sin A -> desc tarea
        else:  # aux sin SUM -> seccion como una fila
            desc_t = f'{sheet}: {item_desc[:30]}'
            sub_t = 'OTR'
        rows.append({'tipo': tipo, 'subtipo': sub_t, 'desc': desc_t, 'cant': cant_t, 'desp': 0,
                     'precio': precio_t, 'fuente': sheet})
    if not rows:  # nada atribuible -> lump
        rows = [{'tipo': tipo, 'subtipo': sub_default, 'desc': item_desc, 'cant': 1, 'desp': 0,
                 'precio': round(resolved_val, 4), 'fuente': 'lump'}]
    return rows

INDICE = parse_indice()
COTIZ = parse_cotiz()
SUBC = parse_subc()
PMO = parse_pmo()
MOHA_AING = parse_moha('MO HA AING')
MOHA_SILVA = parse_moha('MO HA SILVA') if 'MO HA SILVA' in wf.sheetnames else {}

# --- Bloques detallados de MO HA AING (analisis tipo APU de la cuadrilla propia) ---
# Cada bloque termina en "PRECIO UNITARIO:" (col C) con el valor en col D — eso es lo que
# referencia '01'. Arriba: "Total Mano de Obra Menor" = SUM(D a:D b) y "Unitario Mano de Obra".
# Decompone la MO en jornales por categoria UOCRA. GG = precio_unit/unitario (varia por bloque).
_CAT_OK = {'ESPECIALIZADO', 'OFICIAL', 'MEDIO OFICIAL', 'AYUDANTE', 'CAPATAZ'}
def parse_moha_blocks(sheet):
    wsf, wsv = wf[sheet], wv[sheet]
    blocks = {}
    for r in range(20, wsf.max_row + 1):
        if not str(wsf.cell(r, 3).value or '').strip().startswith('PRECIO UNITARIO'):
            continue
        precio = wsv.cell(r, 4).value
        tot_row = uni_row = None
        for rr in range(r - 1, max(r - 9, 1), -1):
            c3 = str(wsf.cell(rr, 3).value or '')
            if c3.startswith('Unitario Mano') and uni_row is None:
                uni_row = rr
            if c3.startswith('Total Mano') and tot_row is None:
                tot_row = rr
                break
        if not tot_row:
            continue
        total = wsv.cell(tot_row, 4).value
        unitario = wsv.cell(uni_row, 4).value if uni_row else None
        m = re.search(r'SUM\(\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)\)', str(wsf.cell(tot_row, 4).value or ''))
        if not m or not (total and unitario):
            continue
        a, b = int(m.group(1)), int(m.group(2))
        rend = total / unitario
        gg = precio / unitario if precio else 1.0
        cats = {}
        for rr in range(a, b + 1):
            cat = str(wsf.cell(rr, 1).value or '').strip().upper()
            cat = ' '.join(cat.split())
            jor = wsv.cell(rr, 2).value
            sal = wsv.cell(rr, 3).value
            if cat in _CAT_OK and isinstance(jor, (int, float)) and jor and isinstance(sal, (int, float)):
                if cat not in cats:
                    cats[cat] = {'jor': 0.0, 'sal': sal}
                cats[cat]['jor'] += jor
        if cats:
            blocks[r] = {'rend': rend, 'gg': gg, 'precio': round(precio, 4), 'cats': cats}
    return blocks

MOHA_BLOCKS = parse_moha_blocks('MO HA AING')
PMO_BASE = {13: PMO[4]['jornal'], 14: PMO[5]['jornal'], 15: PMO[6]['jornal'], 16: PMO[7]['jornal']}  # M/N/O/P salary cols->row5

# Tareas: cabeceras + subrangos
wsT_f, wsT_v = wf['Tareas'], wv['Tareas']
THEADERS = [r for r in range(7, wsT_f.max_row + 1)
            if isinstance(wsT_f.cell(r, 1).value, str) and "'01'!" in wsT_f.cell(r, 1).value]
# salarios base fila 5 (M/N/O/P = cols 13/14/15/16)
SAL = {13: wsT_v.cell(5, 13).value, 14: wsT_v.cell(5, 14).value,
       15: wsT_v.cell(5, 15).value, 16: wsT_v.cell(5, 16).value}
CATMAP = {13: 'ESPECIALIZADO', 14: 'OFICIAL', 15: 'MEDIO OFICIAL', 16: 'AYUDANTE'}

def tareas_subrange(hr):
    i = THEADERS.index(hr)
    nxt = THEADERS[i + 1] if i + 1 < len(THEADERS) else wsT_f.max_row + 1
    return range(hr + 1, nxt)

def terminal_mo_desc(ref):
    """Dado un ref (hoja,col,fila) de un costo de la columna 'otro' (col L de Tareas),
    devuelve (subtipo, desc, precio, flag). La col L es SIEMPRE 'otro' -> MO/OTR,
    salvo MO HA AING (cuadrilla propia = nómina interna -> ALB)."""
    hoja, col, fila = ref
    if hoja == 'P.MO 09.24':  # jornal usado como referencia de precio en la col 'otro' -> OTR
        cat = PMO.get(fila, {}).get('cat', f'P.MO!{col}{fila}')
        return ('OTR', cat, PMO.get(fila, {}).get('jornal'), '')
    if hoja == 'SUBCONTRATOS':
        s = SUBC.get(fila, {})
        return ('OTR', s.get('desc', f'SUB!{fila}'), s.get('precio'), '')
    if hoja == 'MO HA AING':
        s = MOHA_AING.get(fila, {})
        return ('ALB', s.get('desc', f'AING!{fila}'), s.get('precio'), '')
    if hoja == 'MO HA SILVA':
        s = MOHA_SILVA.get(fila, {})
        return ('OTR', s.get('desc', f'SILVA!{fila}'), s.get('precio'), '')
    return (None, None, None, '')

def expand_tareas(hr, e_item, item_desc, inc_mat=True, inc_alb=True, inc_otr=True):
    """Expande sub-componentes de la cabecera Tareas hr a filas de composicion por-unidad.
    inc_mat/inc_alb/inc_otr: incluir cada tipo segun que columna del '01' referencia Tareas."""
    rows = []
    flags = []
    if not e_item:
        e_item = 1
    for r in tareas_subrange(hr):
        desc = norm(wsT_f.cell(r, 3).value)
        e_sub = wsT_v.cell(r, 5).value
        if not isinstance(e_sub, (int, float)):
            e_sub = 0
        # --- MAT otro (F) / corr (G) ---
        for c, sub in ((6, 'OTR'), (7, 'COR')):
            if not inc_mat:
                break
            pr = wsT_v.cell(r, c).value
            ff = wsT_f.cell(r, c).value
            if isinstance(pr, (int, float)) and pr != 0 and e_sub:
                # desc terminal: si F/G referencia Cotizaciones
                dterm = desc
                refs = parse_refs(ff)
                for (h2, cc, rr) in refs:
                    if h2 == 'Cotizaciones' and rr in COTIZ:
                        dterm = COTIZ[rr]['desc']; break
                # "Consumibles" en Tareas = costo lump en col corralon/otro con precio
                # real (NO el MAT-CONS=1 del APU). Se trata como material normal por columna.
                rows.append({'tipo': 'MAT', 'subtipo': sub,
                             'desc': dterm, 'cant': round(e_sub / e_item, 8), 'desp': 0,
                             'precio': round(pr, 4), 'fuente': 'Tareas'})
        # --- MO/ALB por jornales M/N/O/P (cols 13..16) ---
        for c in (13, 14, 15, 16):
            if not inc_alb:
                break
            cnt = wsT_v.cell(r, c).value
            if isinstance(cnt, (int, float)) and cnt != 0:
                rows.append({'tipo': 'MO', 'subtipo': 'ALB', 'desc': CATMAP[c],
                             'cant': round(cnt * e_sub / e_item, 8), 'desp': 0,
                             'precio': round(SAL[c], 4) if SAL[c] else None, 'fuente': 'Tareas'})
        # --- MO/OTR col L (12) ---
        lpr = wsT_v.cell(r, 12).value
        lf = wsT_f.cell(r, 12).value
        if inc_otr and isinstance(lpr, (int, float)) and lpr != 0 and e_sub:
            refs = parse_refs(lf)
            sub_t, dterm, pterm, flag = ('OTR', desc, lpr, '')
            for ref in refs:
                st, dd, pp, fl = terminal_mo_desc(ref)
                if st:
                    sub_t, dterm, pterm, flag = st, dd, (pp if pp else lpr), fl
                    break
            if flag:
                flags.append((item_desc, dterm, flag))
            rows.append({'tipo': 'MO', 'subtipo': sub_t, 'desc': dterm,
                         'cant': round(e_sub / e_item, 8), 'desp': 0,
                         'precio': round(pterm, 4) if pterm else round(lpr, 4), 'fuente': 'Tareas'})
    return rows, flags

def apu_comp_rows(pnum):
    """Convierte la composicion P-XX (Bloque 1, por-unidad-de-partida-APU) a filas."""
    p = apu_by_num.get(pnum)
    if not p:
        return []
    rows = []
    for m in p['materiales']:
        rows.append({'tipo': 'MAT', 'subtipo': 'COR', 'desc': m['desc'], 'cant': m['cant'],
                     'desp': m['desp'], 'precio': m['precio'], 'fuente': f'APU P-{pnum}'})
    if p['consumibles_total']:
        rows.append({'tipo': 'MAT', 'subtipo': 'CONS', 'desc': 'Consumibles', 'cant': p['consumibles_total'],
                     'desp': 0, 'precio': 1, 'fuente': f'APU P-{pnum}'})
    rend = p['rend'] or 1
    for x in p['mo']:
        rows.append({'tipo': 'MO', 'subtipo': 'ALB', 'desc': x['desc'],
                     'cant': round((x['cant'] or 0) / rend, 8), 'desp': 0,
                     'precio': x['precio'], 'fuente': f'APU P-{pnum}'})
    for x in p['equipos']:
        rows.append({'tipo': 'EQ', 'subtipo': '', 'desc': x['desc'],
                     'cant': round((x['cant'] or 0) / rend, 8), 'desp': 0,
                     'precio': x['precio'], 'fuente': f'APU P-{pnum}'})
    return rows

# ---------- Resolvedor recursivo de celdas '01' (referencias encadenadas, Fase C) ----------
COL_TIPO = {'F': ('MAT', 'OTR'), 'G': ('MAT', 'COR'), 'I': ('MO', 'OTR'),
            'J': ('MO', 'ALB'), 'L': ('EQ', '')}

def _lump_row(tipo, sub, desc, val):
    return {'tipo': tipo, 'subtipo': sub, 'desc': desc, 'cant': 1, 'desp': 0,
            'precio': round(val, 4), 'fuente': 'lump'}

def _mo_clasif_columna(x, sub_def):
    """Clasifica una fila MO por la columna (sub_def: I→OTR, J→ALB), EXCEPTO fuentes con
    clasificación definitiva: MO HA AING = nómina propia hormigón → ALB; SUBCONTRATOS / MO HA
    SILVA = subcontrato por definición → OTR (no se sube a ALB aunque esté en col J).
    ⚠ Pendiente con Arquinering: cuándo cotizan obreros en MO/OTR y subcontratos en MO/ALB."""
    fu = str(x.get('fuente', ''))
    if fu.startswith('MO HA AING') or fu in ('SUBCONTRATOS', 'MO HA SILVA'):
        return  # fuente definitiva, no sobreescribir
    x['subtipo'] = sub_def

def _scale(base, factor):
    out = []
    for x in base:
        y = dict(x); y['cant'] = round((y['cant'] or 0) * factor, 8); out.append(y)
    return out

def resolve_01_cell(row, col_letter, depth=0, target=None):
    """Resuelve la composicion de una celda F/G/I/J/L de '01', siguiendo referencias
    encadenadas a otras celdas '01' (mismo tipo), APU indice, bloques MO HA AING y Tareas.
    Las cadenas cross-tipo o a pools/celdas no resolubles -> lump. Escala a `target` si se da."""
    if col_letter not in COL_TIPO:
        return []
    tipo, sub_def = COL_TIPO[col_letter]
    col = col_idx(col_letter)
    f = wf['01'].cell(row, col).value
    val = wv['01'].cell(row, col).value or 0
    desc = norm(wv['01'].cell(row, 3).value)
    tgt = target if target is not None else val
    if not tgt:
        return []
    if depth > 8 or not isinstance(f, str) or not f.startswith('='):
        return [_lump_row(tipo, sub_def, desc, tgt)]
    refs = _refs_in(f)

    def is_apu(sh):
        return bool(sh) and sh.replace(' ', '').upper().startswith('APU')
    special = any((sh is None) or is_apu(sh) or sh == 'Tareas' or
                  (sh == 'MO HA AING' and rr in MOHA_BLOCKS)
                  for (st, en, sh, c, rr) in refs)
    if not special:  # solo hojas-insumo terminales -> delegar al resolvedor de hojas
        out = resolve_cost(f, val, tipo, sub_def, desc)
        if tipo == 'MO':
            for x in out:
                _mo_clasif_columna(x, sub_def)
        return _scale(out, tgt / val) if val else out

    rows = []
    for i, (st, en, sh, c, rr) in enumerate(refs):
        contrib = _eval_with(f, refs, {j: 0 for j in range(len(refs)) if j != i})
        if contrib is None or abs(contrib) < 1e-9:
            continue
        cU = c.upper()
        if sh is None:  # cadena a otra celda '01' — solo mismo tipo
            tgt_tipo = COL_TIPO.get(cU, (None, None))[0]
            sval = (wv['01'].cell(rr, col_idx(cU)).value or 0) if cU in COL_TIPO else 0
            if tgt_tipo == tipo and sval:
                sub_rows = resolve_01_cell(rr, cU, depth + 1)
                rows += _scale(sub_rows, contrib / sval)
            else:  # cross-tipo / pool / no resoluble -> lump
                rows.append(_lump_row(tipo, sub_def, desc, contrib))
        elif is_apu(sh):
            pnum = INDICE.get(rr, {}).get('pnum')
            base = [dict(x) for x in apu_comp_rows(pnum) if x['tipo'] == tipo] if pnum else []
            bs = sum((x['cant'] or 0) * (x['precio'] or 0) * (1 + (x.get('desp') or 0)) for x in base)
            rows += _scale(base, contrib / bs) if bs else [_lump_row(tipo, sub_def, desc, contrib)]
        elif sh == 'MO HA AING' and rr in MOHA_BLOCKS:
            blk = MOHA_BLOCKS[rr]; sc = contrib / blk['precio'] if blk['precio'] else 0
            for cat, info in blk['cats'].items():
                rows.append({'tipo': 'MO', 'subtipo': 'ALB', 'desc': cat,
                             'cant': round((info['jor'] / blk['rend']) * blk['gg'] * sc, 6),
                             'desp': 0, 'precio': round(info['sal'], 4),
                             'fuente': 'MO HA AING (jornales)'})
        elif sh == 'Tareas':
            tr = expand_tareas(rr, wv['01'].cell(row, 5).value or 1, desc)[0]
            base = [dict(x) for x in tr if x['tipo'] == tipo]
            bs = sum((x['cant'] or 0) * (x['precio'] or 0) * (1 + (x.get('desp') or 0)) for x in base)
            rows += _scale(base, contrib / bs) if bs else [_lump_row(tipo, sub_def, desc, contrib)]
        else:  # hoja-insumo terminal dentro de formula especial -> lump del aporte
            rows.append(_lump_row(tipo, sub_def, desc, contrib))
    if not rows:
        return [_lump_row(tipo, sub_def, desc, tgt)]
    # Clasificación MO por la columna que se resuelve (I→OTR, J→ALB), excepto fuentes con
    # clasificación DEFINITIVA (ver _mo_clasif_columna).
    if tipo == 'MO':
        for x in rows:
            _mo_clasif_columna(x, sub_def)
    # escalar al target (respeta "sin pintura" y la verdad de columnas de total)
    cur = sum((x['cant'] or 0) * (x['precio'] or 0) * (1 + (x.get('desp') or 0)) for x in rows)
    return _scale(rows, tgt / cur) if (cur and abs(cur - tgt) > 1e-6) else rows

# ---------- Iterar items de '01' ----------
ws, wsv01 = wf['01'], wv['01']
items = []
flags_all = []
for r in range(8, ws.max_row + 1):
    a = wsv01.cell(r, 1).value  # nro de item del wb de VALORES: algunos son formula (=+A96+0.01)
    if not isinstance(a, (int, float)) or float(a) == int(a):
        continue
    desc = norm(wsv01.cell(r, 3).value)
    ud = norm(wsv01.cell(r, 4).value)
    cant = wsv01.cell(r, 5).value or 0
    F, G, I, J, L = (ws.cell(r, c).value for c in (6, 7, 9, 10, 12))
    # Targets por-unidad desde las columnas de TOTAL N/O/P/Q/R (14-18) ÷ cant — NO desde
    # F/G/I/J/L. Razon: el presupuesto "sin pintura" anula los totales N-R (costo 0) pero deja
    # los unitarios F/L intactos como referencia. Las columnas de total son la verdad del costo.
    tot_NOPQR = [wsv01.cell(r, c).value or 0 for c in (14, 15, 16, 17, 18)]
    tF, tG, tI, tJ, tL = [(v / cant if cant else 0) for v in tot_NOPQR]
    refs_all = []
    for v in (F, G, I, J):
        refs_all += parse_refs(v)
    sheets_ref = {h for (h, c, rr) in refs_all}

    rows = []
    route = None
    # --- Ruta APU indice ---
    apu_rows_refs = [(h, c, rr) for (h, c, rr) in refs_all if h.startswith('APU')]
    if any(h == 'Tareas' for (h, c, rr) in refs_all):
        route = 'Tareas'
        hr = next(rr for (h, c, rr) in refs_all if h == 'Tareas')
        inc_mat = any(parse_refs(v) and any(h == 'Tareas' for h, c, rr in parse_refs(v)) for v in (F, G))
        inc_alb = bool(parse_refs(J)) and any(h == 'Tareas' for h, c, rr in parse_refs(J))
        inc_otr = bool(parse_refs(I)) and any(h == 'Tareas' for h, c, rr in parse_refs(I))
        rows, flags = expand_tareas(hr, cant, desc, inc_mat, inc_alb, inc_otr)
        flags_all += flags
    else:
        route = 'directo'  # RESOLUCIÓN POR COLUMNA: cada tipo de insumo (MAT/MO/EQ) se resuelve
        # desde SU propia columna/fuente. La MO de una partida APU NO viene obligatoriamente junto
        # al MAT — sólo si esa columna referencia el APU. Y la MO se clasifica por la COLUMNA
        # (I→OTR, J→ALB), no por la fuente. resolve_01_cell maneja APU índice/SUBCONTRATOS/encadenados.
        if tF:
            rows += resolve_01_cell(r, 'F', target=tF)
        if tG:
            rows += resolve_01_cell(r, 'G', target=tG)
        if tI:
            rows += resolve_01_cell(r, 'I', target=tI)
        if tJ:
            rows += resolve_01_cell(r, 'J', target=tJ)
    # --- EQ (L) por columna: si L referencia el índice APU -> equipos de la partida; sino % lump ---
    if tL:
        if any(h.startswith('APU') for (h, c, rr) in parse_refs(L)):
            rows += resolve_01_cell(r, 'L', target=tL)
        else:
            rows.append({'tipo': 'EQ', 'subtipo': '', 'desc': f'Herramientas/consumibles - {desc[:40]}',
                         'cant': 1, 'desp': 0, 'precio': round(tL, 4), 'fuente': 'pct'})

    # A4: flete/traslado -> SIEMPRE MO/OTR (logística/subcontrato), aunque el presupuesto los
    # cargue en la columna Albañil (comentario 4b).
    for x in rows:
        if x['tipo'] == 'MO' and any(k in (x['desc'] or '').lower() for k in ('flete', 'traslado')):
            x['subtipo'] = 'OTR'

    # A1: consolidar consumibles -> UN único MAT-CONS por partida (precio 1, cant = costo).
    # Aplica a APU y Tareas por igual (cualquier fila MAT cuya desc diga "consumible").
    cons_cost = 0.0
    keep = []
    for x in rows:
        if x['tipo'] == 'MAT' and 'consumible' in (x['desc'] or '').lower():
            cons_cost += (x['cant'] or 0) * (x['precio'] or 0) * (1 + (x['desp'] or 0))
        else:
            keep.append(x)
    if cons_cost > 1e-6:
        keep.append({'tipo': 'MAT', 'subtipo': 'CONS', 'desc': 'Consumibles',
                     'cant': round(cons_cost, 4), 'desp': 0, 'precio': 1, 'fuente': 'consumibles'})
    rows = keep

    # Merge de insumos repetidos dentro de la tarea (mismo tipo/subtipo/desc/precio/desp):
    # suma cantidades. Limpia duplicados de las cadenas (I y J que resuelven al mismo insumo).
    merged = {}
    order = []
    for x in rows:
        # descartar equipos de costo 0 (p.ej. APU con deprec=0) — no aportan al costo
        if x['tipo'] == 'EQ' and (x['cant'] or 0) * (x['precio'] or 0) == 0:
            continue
        k = (x['tipo'], x['subtipo'], (x['desc'] or '').strip().lower(),
             round(x['precio'] or 0, 2), round(x.get('desp') or 0, 6))
        if k in merged:
            merged[k]['cant'] = round((merged[k]['cant'] or 0) + (x['cant'] or 0), 8)
        else:
            merged[k] = dict(x); order.append(k)
    rows = [merged[k] for k in order]

    def _sum(t):
        if t == 'MAT':
            return sum((x['cant'] or 0) * (x['precio'] or 0) * (1 + (x['desp'] or 0))
                       for x in rows if x['tipo'] == 'MAT')
        return sum((x['cant'] or 0) * (x['precio'] or 0) for x in rows if x['tipo'] == t)

    # Tareas/directo: escalar al total del '01' por tipo (la composicion-detalle debe
    # sumar lo presupuestado). Preserva items y proporciones; flag si el ajuste es >1%
    # (quirks de la fuente, p.ej. 23.01 duplica, 1.09).
    if route in ('Tareas', 'directo'):
        for t, tgt in (('MAT', tF + tG), ('MO', tI + tJ)):
            rec = _sum(t)
            if not rec:
                continue
            if tgt == 0:           # item excluido del costo (p.ej. "sin pintura") -> anular
                k = 0.0
            elif abs(rec / tgt - 1) > 1e-4:
                k = tgt / rec
            else:
                continue
            for x in rows:
                if x['tipo'] == t:
                    x['cant'] = round((x['cant'] or 0) * k, 8)
            if tgt and abs(rec / tgt - 1) > 0.01:
                flags_all.append((desc, f'{t} escalado x{k:.3f}', 'scaled'))

    # reconciliacion por-unidad
    mat_rec = _sum('MAT')
    mo_rec = _sum('MO')
    eq_rec = _sum('EQ')
    items.append({'item': a, 'desc': desc, 'ud': ud, 'cant': cant, 'route': route,
                  'rows': rows,
                  'tgt_mat': round((tF + tG), 4), 'tgt_mo': round((tI + tJ), 4), 'tgt_eq': round(tL, 4),
                  'rec_mat': round(mat_rec, 4), 'rec_mo': round(mo_rec, 4), 'rec_eq': round(eq_rec, 4)})

json.dump({'items': items, 'flags': flags_all}, open(OUT_JSON, 'w', encoding='utf-8'),
          ensure_ascii=False, indent=1)

# ---------- Reporte de reconciliacion ----------
def dpct(rec, tgt):
    if not tgt:
        return 0 if abs(rec) < 1 else 999
    return (rec / tgt - 1) * 100

lines = []
lines.append(f'Items procesados: {len(items)}')
from collections import Counter
rc = Counter(i['route'] for i in items)
lines.append('Rutas: ' + ', '.join(f'{k}={v}' for k, v in rc.items()))
bad_mat = [i for i in items if abs(dpct(i['rec_mat'], i['tgt_mat'])) > 1 and (i['tgt_mat'] or i['rec_mat'])]
bad_mo = [i for i in items if abs(dpct(i['rec_mo'], i['tgt_mo'])) > 1 and (i['tgt_mo'] or i['rec_mo'])]
lines.append(f'MAT |delta|>1%: {len(bad_mat)} / {len(items)}')
lines.append(f'MO  |delta|>1%: {len(bad_mo)} / {len(items)}')
lines.append('--- peores MAT ---')
for i in sorted(bad_mat, key=lambda x: -abs(dpct(x['rec_mat'], x['tgt_mat'])))[:25]:
    lines.append(f"  {i['item']:>6} {i['route']:8} tgt={i['tgt_mat']:>14.2f} rec={i['rec_mat']:>14.2f} "
                 f"d={dpct(i['rec_mat'], i['tgt_mat']):>8.1f}%  {i['desc'][:34]}")
lines.append('--- peores MO ---')
for i in sorted(bad_mo, key=lambda x: -abs(dpct(x['rec_mo'], x['tgt_mo'])))[:25]:
    lines.append(f"  {i['item']:>6} {i['route']:8} tgt={i['tgt_mo']:>14.2f} rec={i['rec_mo']:>14.2f} "
                 f"d={dpct(i['rec_mo'], i['tgt_mo']):>8.1f}%  {i['desc'][:34]}")
lines.append(f'--- flags ALB/OTR dudosos: {len(flags_all)} ---')
seen = set()
for d, dd, fl in flags_all:
    k = (dd, fl)
    if k in seen:
        continue
    seen.add(k)
    lines.append(f'  [{fl}] {dd[:40]}  (ej item: {d[:30]})')
open('scripts/_pres_parse_report.txt', 'w', encoding='utf-8').write('\n'.join(lines))
print('OK -> scripts/_pres_parse_report.txt')
