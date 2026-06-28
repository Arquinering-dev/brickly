# -*- coding: utf-8 -*-
"""
estandarizar_presupuesto_v8.py — Unifica la hoja 1_Presupuesto al estándar v8 común
en las 3 obras (CH, GDR, SVD). Confirmado con Pedro (2026-06-24).

Cambios:
  * Layout idéntico en las 3 obras (24 columnas A:X), sin diferencias de fórmula.
  * VENTA = input por tarea (P.Unit) en TODAS las obras. CH/GDR: se congela el valor actual
    (=costo×coef) como input → value-preserving exacto. SVD ya era input. El desglose de venta
    (MT/MO/EQ) se deriva proporcional al costo (misma fórmula en todas).
  * Se ELIMINA todo el bloque de avance redundante (manual legacy, % Cert Actual, Imp.*,
    desglose Acum.*, Control, Acum_tot $). Queda SOLO '% Acum Tot' (= SUMIFS de Cert_App_Output),
    necesaria para reproducir EXACTO el % avance físico (incl. SVD, cuyo circuito usa PV-doc).
  * PV subtotal = una sola columna (= P.Unit × Cant).
  * Se recablean a las columnas nuevas: 0_CONFIG B(costo/ppto), Cert_Control_OC!E (% avance físico).

Uso:
    python scripts/estandarizar_presupuesto_v8.py <in.xlsx> <out.xlsx>
Después: excel_recalc.py <out> ; recalc.py <out> ; verificar.
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import unicodedata

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except Exception:
        pass


def norm(v):
    if v is None:
        return ""
    s = unicodedata.normalize("NFKD", str(v))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


# ---- layout canónico: (header, key, tipo) ; tipo: in=input, f=formula, '' chrome ----
CANON = [
    ("Rubro MT", "rub_mt", "in"), ("Rubro MT/Prov", "rub_prov", "in"),
    ("Rubro MO/OTR", "rub_otr", "in"), ("Rubro MO/ALB", "rub_alb", "in"),
    ("Cod. Ítem", "cod", "in"), ("Estado", "estado", "in"), ("Descripción", "desc", "in"),
    ("U", "u", "in"), ("Cant", "cant", "in"),
    ("MT", "cmt", "in"), ("MO/OTR", "cotr", "in"), ("MO/ALB", "calb", "in"), ("EQ", "ceq", "in"),
    ("Costo Unit.", "costo_unit", "f"),
    ("MT", "vmt", "f"), ("MO/OTR", "votr", "f"), ("MO/ALB", "valb", "f"), ("EQ", "veq", "f"),
    ("P.Unit", "pv_unit", "in"), ("Margen", "margen", "f"),
    ("Costo total", "costo_total", "f"), ("PV subtotal", "pv_subtotal", "f"),
    ("Etapa", "etapa", "in"), ("Presupuesto", "presupuesto", "in"), ("% Acum Tot", "avance", "f"),
]
COL = {k: i + 1 for i, (_, k, _) in enumerate(CANON)}        # key -> col index (1-based)
LET = {k: get_column_letter(i + 1) for i, (_, k, _) in enumerate(CANON)}  # key -> letter
HDR_ROW = 3
DATA_ROW = 4

# columnas de input numérico (font azul) y de fórmula (font verde)
NUM_IN = {"cant", "cmt", "cotr", "calb", "ceq", "pv_unit"}
FORMULAS = {"costo_unit", "vmt", "votr", "valb", "veq", "costo_total", "pv_subtotal", "avance"}
MONEY = {"cmt", "cotr", "calb", "ceq", "costo_unit", "vmt", "votr", "valb", "veq", "pv_unit",
         "costo_total", "pv_subtotal"}

BLUE = Font(name="Aptos Narrow", color="FF0000FF")
GREEN = Font(name="Aptos Narrow", color="FF008000")
BLACK = Font(name="Aptos Narrow", color="FF000000")
HFONT = Font(name="Aptos Narrow", color="FFFFFFFF", bold=True)
TFONT = Font(name="Aptos Narrow", color="FF1F3864", bold=True)
HFILL = PatternFill("solid", fgColor="1F4E78")
TFILL = PatternFill("solid", fgColor="D6E4F0")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)


def find_col(ws, hr, *needs, exclude=()):
    needs = [norm(n) for n in needs]
    exc = [norm(e) for e in exclude]
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(row=hr, column=c).value)
        if not h:
            continue
        if any(e in h for e in exc):
            continue
        if all(n in h for n in needs):
            return c
    return None


def read_source(path):
    """Lee 1_Presupuesto de la obra original. Devuelve (rows, title, src).
    rows: lista de dicts con keys de CANON + '_kind' (task/banner/total).
    src: dict campo->col(1-based) de la hoja ORIGINAL (para remapear consumidores)."""
    wbf = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    ws = wbf["1_Presupuesto"]
    wv = wbv["1_Presupuesto"]
    hr = HDR_ROW
    # columnas fijas (idénticas en las 3 obras) por posición
    src = dict(rub_mt=1, rub_prov=2, rub_otr=3, rub_alb=4, cod=6, estado=7, desc=8, u=9,
               cant=10, cmt=11, cotr=12, calb=13, ceq=14, costo_unit=15, pv_unit=16)
    # columnas que difieren -> por header
    src["etapa"] = find_col(ws, hr, "etapa")
    src["presupuesto"] = find_col(ws, hr, "presupuesto")
    src["pv_subtotal"] = find_col(ws, hr, "pv subtotal") or find_col(ws, hr, "subtotal", exclude=("acum",))
    # avance acumulado (la del circuito Cert_*): header exacto "% acum tot"
    src["avance"] = None
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(row=hr, column=c).value) == "% acum tot":
            src["avance"] = c
            break
    src["pv_sub_old"] = src["pv_subtotal"]
    title = ws.cell(row=1, column=1).value

    # último renglón con contenido
    last = hr
    for r in range(hr + 1, ws.max_row + 1):
        if any(wv.cell(row=r, column=src[k]).value not in (None, "")
               for k in ("cod", "desc", "cant", "cmt", "etapa")):
            last = r
    rows = []
    for r in range(DATA_ROW, last + 1):
        desc = wv.cell(row=r, column=src["desc"]).value
        cant = wv.cell(row=r, column=src["cant"]).value
        d = {}
        for k in ("rub_mt", "rub_prov", "rub_otr", "rub_alb", "cod", "estado", "desc",
                  "u", "etapa", "presupuesto"):
            d[k] = wv.cell(row=r, column=src[k]).value
        for k in ("cant", "cmt", "cotr", "calb", "ceq", "pv_unit"):
            d[k] = wv.cell(row=r, column=src[k]).value
        if desc is not None and "total" in norm(desc):
            d["_kind"] = "total"
        elif not isinstance(cant, (int, float)):
            d["_kind"] = "banner"
        else:
            d["_kind"] = "task"
        d["_pv_sub_old"] = wv.cell(row=r, column=src["pv_sub_old"]).value if src["pv_sub_old"] else None
        rows.append(d)
    wbf.close()
    wbv.close()
    return rows, title, src


def build(in_path, out_path):
    rows, title, src = read_source(in_path)
    wb = openpyxl.load_workbook(in_path)
    old = wb["1_Presupuesto"]
    idx = wb.sheetnames.index("1_Presupuesto")
    tabcolor = old.sheet_properties.tabColor
    wb.remove(old)
    ws = wb.create_sheet("1_Presupuesto", idx)
    if tabcolor:
        ws.sheet_properties.tabColor = tabcolor

    # título (fila 1) + headers (fila 3)
    ws.cell(row=1, column=1, value=title or "PRESUPUESTO — COSTO Y PRECIO DE VENTA")
    ws.cell(row=1, column=1).font = TFONT
    for c in range(1, len(CANON) + 1):
        ws.cell(row=1, column=c).fill = TFILL
    for i, (h, k, _t) in enumerate(CANON):
        cell = ws.cell(row=HDR_ROW, column=i + 1, value=h)
        cell.font = HFONT
        cell.fill = HFILL
        cell.alignment = CTR

    # datos
    out_r = DATA_ROW
    task_rows = []
    total_row = None
    for d in rows:
        kind = d["_kind"]
        r = out_r
        if kind == "banner":
            for k in ("rub_mt", "rub_prov", "rub_otr", "rub_alb", "cod", "estado", "desc",
                      "etapa", "presupuesto"):
                if d.get(k) not in (None, ""):
                    ws.cell(row=r, column=COL[k], value=d[k]).font = BLACK
        elif kind == "total":
            total_row = r
            ws.cell(row=r, column=COL["desc"], value=d["desc"]).font = TFONT
        else:  # task
            task_rows.append(r)
            # inputs
            for k in ("rub_mt", "rub_prov", "rub_otr", "rub_alb", "cod", "estado", "desc",
                      "u", "etapa", "presupuesto"):
                v = d.get(k)
                if v not in (None, ""):
                    ws.cell(row=r, column=COL[k], value=v).font = BLACK
            for k in ("cant", "cmt", "cotr", "calb", "ceq"):
                ws.cell(row=r, column=COL[k], value=(d.get(k) if d.get(k) is not None else 0)).font = BLUE
            # venta total = input congelado (valor actual)
            ws.cell(row=r, column=COL["pv_unit"], value=(d.get("pv_unit") or 0)).font = BLUE
            # fórmulas
            S = LET["pv_unit"]; N = LET["costo_unit"]; I = LET["cant"]
            J = LET["cmt"]; K = LET["cotr"]; L = LET["calb"]; M = LET["ceq"]
            f = {
                "costo_unit": f"=SUM({J}{r}:{M}{r})",
                "vmt":  f"=IF(${N}{r}=0,${S}{r},{J}{r}/${N}{r}*${S}{r})",
                "votr": f"=IF(${N}{r}=0,0,{K}{r}/${N}{r}*${S}{r})",
                "valb": f"=IF(${N}{r}=0,0,{L}{r}/${N}{r}*${S}{r})",
                "veq":  f"=IF(${N}{r}=0,0,{M}{r}/${N}{r}*${S}{r})",
                "margen": f'=IFERROR({S}{r}/{N}{r},"")',   # margen verificable por tarea = PV÷Costo
                "costo_total": f"={N}{r}*{I}{r}",
                "pv_subtotal": f"={S}{r}*{I}{r}",
                "avance": (f"=SUMIFS(Cert_App_Output!$G:$G,Cert_App_Output!$E:$E,${LET['cod']}{r},"
                           f"Cert_App_Output!$K:$K,${LET['presupuesto']}{r})"),
            }
            for k, fm in f.items():
                ws.cell(row=r, column=COL[k], value=fm).font = GREEN
        out_r += 1

    # fila total: SUM de costo_total y pv_subtotal sobre las tareas
    if total_row is not None and task_rows:
        a, b = min(task_rows), max(task_rows)
        ws.cell(row=total_row, column=COL["costo_total"],
                value=f"=SUM({LET['costo_total']}{a}:{LET['costo_total']}{b})").font = TFONT
        ws.cell(row=total_row, column=COL["pv_subtotal"],
                value=f"=SUM({LET['pv_subtotal']}{a}:{LET['pv_subtotal']}{b})").font = TFONT

    # formato de número
    last_data = out_r - 1
    for k in MONEY:
        for r in range(DATA_ROW, last_data + 1):
            ws.cell(row=r, column=COL[k]).number_format = '#,##0'
    for r in range(DATA_ROW, last_data + 1):
        ws.cell(row=r, column=COL["avance"]).number_format = '0.0%'
        ws.cell(row=r, column=COL["cant"]).number_format = '#,##0.00'
        ws.cell(row=r, column=COL["margen"]).number_format = '0.0000'
    ws.column_dimensions[LET["margen"]].width = 9

    # anchos + freeze + gridlines
    widths = {"rub_mt": 16, "rub_prov": 16, "rub_otr": 16, "rub_alb": 16, "cod": 9, "estado": 8,
              "desc": 40, "u": 6, "cant": 8}
    for k, w in widths.items():
        ws.column_dimensions[LET[k]].width = w
    for k in MONEY:
        ws.column_dimensions[LET[k]].width = 13
    ws.column_dimensions[LET["etapa"]].width = 20
    ws.column_dimensions[LET["presupuesto"]].width = 12
    ws.column_dimensions[LET["avance"]].width = 11
    ws.freeze_panes = f"{LET['cmt']}{DATA_ROW}"   # congela rubros..cant + filas 1-3
    ws.sheet_view.showGridLines = False

    # recablear consumidores cross-sheet por remapeo de columnas (filas preservadas)
    n_rewired = rewire(wb, src)

    wb.save(out_path)
    return last_data, total_row, len(task_rows), n_rewired


import re
from openpyxl.worksheet.formula import ArrayFormula

_REF = re.compile(r"'1_Presupuesto'!([\$A-Z0-9:]+)")
_COL = re.compile(r"[A-Z]{1,3}")


def rewire(wb, src):
    """Remapea las referencias de columna a 1_Presupuesto en TODAS las hojas consumidoras
    al layout canónico. Como build() preserva el número de fila de cada tarea, sólo cambian
    las letras de columna (un solo pase, sin colisiones). Devuelve # de fórmulas tocadas."""
    # remap: letra vieja -> letra nueva, por campo (src tiene las viejas; LET las nuevas)
    remap = {}
    for key, oldcol in src.items():
        if key == "pv_sub_old" or oldcol is None or key not in LET:
            continue
        remap[get_column_letter(oldcol)] = LET[key]
    # identidad para columnas no mapeadas (se dejan igual)

    def repl_ref(m):
        tail = m.group(1)
        tail2 = _COL.sub(lambda cm: remap.get(cm.group(0), cm.group(0)), tail)
        return "'1_Presupuesto'!" + tail2

    n = 0
    for sh in wb.sheetnames:
        if sh == "1_Presupuesto":
            continue
        ws = wb[sh]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "1_Presupuesto" in v:
                    nv = _REF.sub(repl_ref, v)
                    if nv != v:
                        cell.value = nv
                        n += 1
                elif isinstance(v, ArrayFormula) and v.text and "1_Presupuesto" in v.text:
                    nt = _REF.sub(repl_ref, v.text)
                    if nt != v.text:
                        cell.value = ArrayFormula(v.ref, nt)
                        n += 1
    return n


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("uso: estandarizar_presupuesto_v8.py <in.xlsx> <out.xlsx>")
        sys.exit(1)
    ld, tr, nt, nrw = build(sys.argv[1], sys.argv[2])
    print(f"OK -> {sys.argv[2]}  (last_data={ld}, total_row={tr}, tareas={nt}, refs_recableadas={nrw})")
