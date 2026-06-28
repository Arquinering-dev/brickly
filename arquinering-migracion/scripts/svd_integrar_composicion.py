"""
Integración: poblar 1_Composicion del Resumen v8 de SVD con la COMPOSICIÓN del APU Unificado.

Lee los valores resueltos (data_only) de APU_Unificado_SVD4140_v1.xlsx!COMPOSICIÓN (A-P) y los
escribe en SVD_4140_Resumen_de_Obra_v8_2.xlsx!1_Composicion con la estructura de 19 cols (igual a
CH): A-P + Q (Cant MO/ALB Total = O*P) + R (Cod_Item_Ppto = código de ítem float, clave de join a
1_Presupuesto col E) + S (Rubro = XLOOKUP a 1_Presupuesto, igual fórmula que CH).

Items 26-28 y los partidos (2.01/24.10) quedan huérfanos sobre 1_Presupuesto (S="") pero el dato
queda guardado. NO toca 1_Presupuesto ni otras hojas.
"""
import openpyxl, warnings, sys, re
warnings.filterwarnings('ignore'); sys.stdout.reconfigure(encoding='utf-8')
from openpyxl.styles import Font

APU = 'archivos/output/APU_Unificado_SVD4140_v1.xlsx'
RES = 'archivos/output/SVD_4140_Resumen_de_Obra_v8_2.xlsx'

BLACK = Font(color='FF000000', name='Aptos Narrow')
BLUE = Font(color='FF0000FF', name='Aptos Narrow')   # input
GREEN = Font(color='FF008000', name='Aptos Narrow')  # fórmula
MONEY = '#,##0.00'

# leer COMPOSICIÓN del APU Unificado (valores resueltos)
av = openpyxl.load_workbook(APU, data_only=True)['COMPOSICIÓN']
rows = []
for r in range(2, av.max_row + 1):
    a = av.cell(r, 1).value
    if not a:
        continue
    vals = [av.cell(r, c).value for c in range(1, 17)]  # A..P
    rows.append(vals)
print(f'Filas de composición a integrar: {len(rows)}')

def pto_to_item(cod):
    """PTO-0102 -> 1.02 (float)."""
    m = re.match(r'PTO-(\d{2})(\d{2})', str(cod))
    return int(m.group(1)) + int(m.group(2)) / 100 if m else None

# abrir Resumen (carga default para preservar fórmulas del resto del archivo)
wb = openpyxl.load_workbook(RES)
ws = wb['1_Composicion']
# header S debe ser "Rubro" (el reader del dashboard busca 'rubro' para armar jornales;
# el placeholder de SVD traía "MO/ALB" -> 0 horas presupuestadas). CH usa "Rubro".
ws.cell(1, 19, 'Rubro')
# limpiar filas de datos viejas (2..max)
if ws.max_row >= 2:
    ws.delete_rows(2, ws.max_row - 1)

for i, vals in enumerate(rows, start=2):
    A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P = vals
    R = pto_to_item(A)
    Q = (O or 0) * (P if isinstance(P, (int, float)) else 0)
    out = [A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, round(Q, 8), R]
    for ci, v in enumerate(out, start=1):
        cell = ws.cell(i, ci, v)
        cell.font = BLACK
        if ci in (9, 10, 16, 17):  # Precio, Costo, Cant MO/ALB Total, Cant Ej -> money/num
            cell.number_format = MONEY if ci in (9, 10, 17) else '0.######'
        if ci == 7:  # Cant (input)
            cell.font = BLUE
    # S (col 19): Rubro vía XLOOKUP a 1_Presupuesto (misma fórmula que CH)
    s = ws.cell(i, 19, f"=IFERROR(_xlfn.XLOOKUP(R{i},'1_Presupuesto'!$E:$E,'1_Presupuesto'!$D:$D),\"\")")
    s.font = GREEN

wb.save(RES)
print('OK -> 1_Composicion poblada en', RES)
print('huérfanos esperados: rubros 26-28 + partidos 2.01/24.10 (S vacío)')
