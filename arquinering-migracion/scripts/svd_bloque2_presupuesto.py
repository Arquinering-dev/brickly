"""svd_bloque2_presupuesto.py — Migracion SVD 4140 a v8.
Bloque 1b: clona el esqueleto v8 de GDR v8_12 -> SVD_4140_Resumen_de_Obra_v8_1.xlsx.
Bloque 2 : reescribe _Listas (plan Tezamat) y 1_Presupuesto desde el legacy SVD
           (Pto. Costos + Pto. Vta.), con la convencion de 4 rubros y el reenganche
           de certificacion identico a GDR.

Value-preserving: costos por-unidad del legacy intactos; venta = costo * coef (coef
calculado del legacy y verificado contra los subtotales Pto.Vta).
"""
import sys
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

SRC = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
LEGACY = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"
DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"

# --- _Listas: rama 53 OBRA del plan + especiales (identico a GDR/CH) ----------
CANONICO = [
    ("Preliminares", "53001"), ("Demolición", "53002"), ("Movimiento de Suelos", "53003"),
    ("Hormigón MT", "53004"), ("Homigón MO", "53005"),
    ("Metálica MT", "53006"), ("Metálica MO", "53007"),
    ("Albañilería MT", "53008"), ("Albañilería MO", "53009"),
    ("Durlock MT", "53010"), ("Durlock MO", "53011"), ("Aberturas", "53012"),
    ("Revestimiento MT", "53013"), ("Revestimiento MO", "53014"),
    ("Pintura MT", "53015"), ("Pintura MO", "53016"),
    ("Sanitaria MT", "53017"), ("Sanitaria MO", "53018"),
    ("Eléctrico MT", "53019"), ("Eléctrico MO", "53020"),
    ("Provisiones", "53021"), ("Gastos Generales", "53022"), ("Varios Ferreteria", "53023"),
    ("Termomecánica MT", "53024"), ("Termomecánica MO", "53025"),
    ("Seguridad e Higiene", "53026"), ("Herrería MT", "53027"), ("Herrería MO", "53028"),
    ("Alquiler de Equipos", "53980"), ("Gastos a Reintegrar", "53990"),
    ("Gastos GCBA Construccion", "53999"),
    ("Mov. Variables", "52302"),
    ("Supervisión de Obra MO", "52209"),
]

# rubros con split MT/MO (cualquier costo MAT -> "X MT", costo MO -> "X MO")
SPLIT = {
    "Hormigón": ("Hormigón MT", "Homigón MO"),
    "Albañilería": ("Albañilería MT", "Albañilería MO"),
    "Durlock": ("Durlock MT", "Durlock MO"),
    "Eléctrico": ("Eléctrico MT", "Eléctrico MO"),
    "Pintura": ("Pintura MT", "Pintura MO"),
    "Revestimiento": ("Revestimiento MT", "Revestimiento MO"),
    "Sanitaria": ("Sanitaria MT", "Sanitaria MO"),
    "Termomecánica": ("Termomecánica MT", "Termomecánica MO"),
}
# rubros sin split (mismo nombre canonico)
SINGLE = {
    "Preliminares": "Preliminares",
    "Agrimensura": "Preliminares",                  # Q8 best-criterion (precedente GDR) -> marcar
    "Excavación y Mov. de Suelos": "Movimiento de Suelos",
    "Depresion de napas": "Movimiento de Suelos",   # Q8 best-criterion -> marcar
    "Gastos Generales Obra": "Gastos Generales",
    "Movilidad": "Mov. Variables",
    "Supervisión de Obra": "Supervisión de Obra MO",
    "Seguridad e Higiene": "Seguridad e Higiene",
    "Equipos": "Alquiler de Equipos",
    "Aberturas": "Aberturas",
    "Consumibles y Ferretería": "Varios Ferreteria",
    "Gastos en Personal": "Gastos Generales",       # rama indirecta, sin etapa propia -> marcar
}
# rubros ejecutados con cuadrilla propia (hoja M.O. en legacy) -> MO interna (ALB / col D)
MO_PROPIA = {"Hormigón", "Albañilería", "Eléctrico", "Sanitaria"}


def rubro_canon(legacy_rub):
    """Devuelve (nombre_MT, nombre_MO) canonicos para un rubro legacy."""
    if legacy_rub in SPLIT:
        return SPLIT[legacy_rub]
    if legacy_rub in SINGLE:
        c = SINGLE[legacy_rub]
        return c, c
    return None, None  # desconocido


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _norm(s):
    return str(s or "").strip().lower()


def read_venta_subtotal(wb):
    """Mapa (presupuesto, desc_normalizada) -> venta SUBTOTAL legacy (Pto.Vta col L).
    Match por descripcion (robusto: 191/191 vs 187 por codigo). Suma si repite."""
    vt = wb["Pto. Vta."]
    SEC = [("PTO 01", 9, 55), ("PTO 02", 57, 238)]
    m = {}
    for presup, r0, r1 in SEC:
        for r in range(r0, r1 + 1):
            desc = vt.cell(row=r, column=3).value
            if not desc:
                continue
            L = _num(vt.cell(row=r, column=12).value)  # L = venta subtotal (K*cant)
            k = (presup, _norm(desc))
            m[k] = m.get(k, 0.0) + L
    return m


def read_legacy():
    """Lee Pto.Costos (costos) + Pto.Vta (venta) y devuelve la lista de filas v8.
    La venta legacy (subtotal por descripcion) se reparte entre las filas de costo
    que comparten descripcion, proporcional al costo (preserva el total exacto)."""
    wb = openpyxl.load_workbook(LEGACY, data_only=True)
    co = wb["Pto. Costos"]
    venta_sub = read_venta_subtotal(wb)
    SECCIONES = [("PTO 01", 9, 57), ("PTO 02", 64, 240)]
    rows = []
    sum_costo = sum_venta = 0.0
    sin_venta = []
    for presup, r0, r1 in SECCIONES:
        for r in range(r0, r1 + 1):
            cod = co.cell(row=r, column=2).value
            rub = co.cell(row=r, column=1).value
            desc = co.cell(row=r, column=3).value
            if cod is None and desc is None:
                continue
            # etapa = codigo entero sin rubro
            is_etapa = isinstance(cod, (int, float)) and float(cod) == int(cod) and not rub
            if is_etapa:
                rows.append({"kind": "etapa", "cod": cod, "desc": str(desc or "").strip(),
                             "presup": presup})
                continue
            if not rub:           # fila sin rubro y no-etapa: subtotal/ruido -> skip
                continue
            U = co.cell(row=r, column=4).value
            cant = co.cell(row=r, column=5).value or 0
            mat = co.cell(row=r, column=6).value or 0
            mo = co.cell(row=r, column=8).value or 0
            eq = co.cell(row=r, column=10).value or 0
            try: mat = float(mat)
            except (TypeError, ValueError): mat = 0.0
            try: mo = float(mo)
            except (TypeError, ValueError): mo = 0.0
            try: eq = float(eq)
            except (TypeError, ValueError): eq = 0.0
            try: cantf = float(cant)
            except (TypeError, ValueError): cantf = 0.0
            rows.append({"kind": "task", "cod": cod, "rub_legacy": str(rub).strip(),
                         "desc": str(desc or "").strip(), "ndesc": _norm(desc), "U": U,
                         "cant": cant, "cantf": cantf, "mat": mat, "mo": mo, "eq": eq,
                         "presup": presup, "costo_sub": (mat + mo + eq) * cantf})
            sum_costo += (mat + mo + eq) * cantf

    # --- repartir la venta legacy (por descripcion) entre filas que comparten desc
    from collections import defaultdict
    groups = defaultdict(list)
    for it in rows:
        if it["kind"] == "task":
            groups[(it["presup"], it["ndesc"])].append(it)
    sin_venta = []
    for (presup, nd), grp in groups.items():
        vsub = venta_sub.get((presup, nd))
        if vsub is None:
            for it in grp:
                it["venta_u"] = 0.0
            sin_venta.append((presup, grp[0]["desc"][:30]))
            continue
        tot_cost = sum(it["costo_sub"] for it in grp)
        n = len(grp)
        for it in grp:
            frac = (it["costo_sub"] / tot_cost) if tot_cost > 0 else (1.0 / n)
            alloc = vsub * frac                      # venta subtotal asignada a la fila
            it["venta_u"] = (alloc / it["cantf"]) if it["cantf"] else 0.0
            sum_venta += alloc
    vt = wb["Pto. Vta."]
    venta_p1 = _num(vt["M56"].value)
    venta_p2 = _num(vt["M239"].value)
    return rows, sum_costo, sum_venta, venta_p1, venta_p2, sin_venta


def build():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    rows, sum_costo, sum_venta, venta_p1, venta_p2, sin_venta = read_legacy()

    venta_total = (venta_p1 or 0) + (venta_p2 or 0)
    coef = venta_total / sum_costo if sum_costo else 1.0  # solo informativo (1_GGBB!F67)

    # --- _Listas -------------------------------------------------------------
    L = wb["_Listas"]
    for r in range(1, 61):
        for c in (1, 2, 3):
            L.cell(row=r, column=c).value = None
    L["A1"] = "RUBROS CANÓNICOS (Plan de Cuentas Tezamat — rama 53 OBRA + especiales)"
    L["B1"] = "Código"; L["C1"] = "TIPOS"; L["C2"] = "MT"; L["C3"] = "MO"
    for i, (rub, cod) in enumerate(CANONICO):
        L.cell(row=2 + i, column=1).value = rub
        L.cell(row=2 + i, column=2).value = cod
    last = 1 + len(CANONICO)
    rng = f"$A$2:$A${last}"
    nm = "RUBROS_PLAN"
    try: del wb.defined_names[nm]
    except (KeyError, TypeError): pass
    wb.defined_names[nm] = DefinedName(nm, attr_text=f"_Listas!{rng}")

    # --- 1_Presupuesto -------------------------------------------------------
    p = wb["1_Presupuesto"]
    # limpiar TODAS las filas de datos (desde 4) en cols A..AV
    for r in range(4, p.max_row + 1):
        for c in range(1, 49):
            p.cell(row=r, column=c).value = None
    # P2 / 1_GGBB!F67 = markup global (solo informativo; la venta es input por tarea)
    p["P2"] = "='1_GGBB'!F67"
    wb["1_GGBB"]["F67"] = coef

    sin_match = set()
    cur_etapa = ""
    out_r = 4
    n_task = 0
    for it in rows:
        if it["kind"] == "etapa":
            cur_etapa = it["desc"]
            p.cell(row=out_r, column=6).value = it["cod"]       # F codigo
            p.cell(row=out_r, column=8).value = it["desc"]      # H nombre etapa
            out_r += 1
            continue
        # task
        mt_name, mo_name = rubro_canon(it["rub_legacy"])
        if mt_name is None:
            sin_match.add(it["rub_legacy"])
            mt_name = mo_name = "-"
        # asignar columnas de rubro segun donde hay costo
        A = mt_name if it["mat"] > 0 else "-"          # rubro MT
        B = "-"                                         # Prov: legacy no separa
        mo_propia = it["rub_legacy"] in MO_PROPIA
        C = mo_name if (it["mo"] > 0 and not mo_propia) else "-"   # MO/OTR
        D = mo_name if (it["mo"] > 0 and mo_propia) else "-"       # MO/ALB
        # si no hay costo en ningun lado, igual etiquetar MT con el rubro (para control)
        if A == "-" and C == "-" and D == "-" and mt_name != "-":
            A = mt_name
        p.cell(row=out_r, column=1).value = A
        p.cell(row=out_r, column=2).value = B
        p.cell(row=out_r, column=3).value = C
        p.cell(row=out_r, column=4).value = D
        p.cell(row=out_r, column=6).value = it["cod"]            # F
        p.cell(row=out_r, column=7).value = "OK"                # G estado
        p.cell(row=out_r, column=8).value = it["desc"]          # H
        p.cell(row=out_r, column=9).value = it["U"]             # I
        p.cell(row=out_r, column=10).value = it["cant"]         # J
        p.cell(row=out_r, column=11).value = it["mat"]          # K MT
        # MO -> a OTR (L) o ALB (M) segun cuadrilla
        if mo_propia:
            p.cell(row=out_r, column=12).value = 0
            p.cell(row=out_r, column=13).value = it["mo"]       # M MO/ALB
        else:
            p.cell(row=out_r, column=12).value = it["mo"]       # L MO/OTR
            p.cell(row=out_r, column=13).value = 0
        p.cell(row=out_r, column=14).value = it["eq"]           # N EQ
        rr = out_r
        # O = costo unit; P = venta unit (INPUT legacy por tarea, markup no uniforme)
        p.cell(row=rr, column=15).value = f"=+SUM(K{rr}:N{rr})"          # O Costo Unit
        p.cell(row=rr, column=16).value = it["venta_u"]                  # P P Unit (input)
        # venta por tipo: proporcional al costo (suma exacta = P); si O=0 -> todo a MT
        p.cell(row=rr, column=18).value = f"=IF($O{rr}=0,$P{rr},K{rr}/$O{rr}*$P{rr})"   # R MT
        p.cell(row=rr, column=19).value = f"=IF($O{rr}=0,0,L{rr}/$O{rr}*$P{rr})"        # S MO/OTR
        p.cell(row=rr, column=20).value = f"=IF($O{rr}=0,0,M{rr}/$O{rr}*$P{rr})"        # T MO/ALB
        p.cell(row=rr, column=21).value = f"=IF($O{rr}=0,0,N{rr}/$O{rr}*$P{rr})"        # U EQ
        p.cell(row=rr, column=22).value = f"=+SUM(R{rr}:U{rr})"          # V (= P)
        p.cell(row=rr, column=24).value = f"=+R{rr}*J{rr}"              # X
        p.cell(row=rr, column=25).value = f"=+Z{rr}-X{rr}"             # Y
        p.cell(row=rr, column=26).value = f"=+V{rr}*J{rr}"              # Z
        p.cell(row=rr, column=28).value = cur_etapa                     # AB ETAPA
        p.cell(row=rr, column=29).value = 0                             # AC % Acum Ant
        p.cell(row=rr, column=30).value = 0                             # AD % Cert Actual
        p.cell(row=rr, column=31).value = f"=AC{rr}+AD{rr}"             # AE
        p.cell(row=rr, column=32).value = f'=IF($AE{rr}=1,"🔴","⚪")'  # AF
        p.cell(row=rr, column=33).value = f"=AC{rr}*Z{rr}"             # AG
        p.cell(row=rr, column=34).value = f"=AD{rr}*Z{rr}"             # AH
        p.cell(row=rr, column=35).value = f"=AE{rr}*Z{rr}"             # AI
        p.cell(row=rr, column=37).value = f"=AE{rr}*X{rr}"             # AK
        p.cell(row=rr, column=38).value = f"=AE{rr}*S{rr}*J{rr}"        # AL
        p.cell(row=rr, column=39).value = f"=AE{rr}*T{rr}*J{rr}"        # AM
        p.cell(row=rr, column=40).value = f"=AE{rr}*U{rr}*J{rr}"        # AN
        p.cell(row=rr, column=41).value = f"=AK{rr}+AL{rr}+AM{rr}+AN{rr}"  # AO
        p.cell(row=rr, column=43).value = f"=+O{rr}"                    # AQ Costo_ud
        p.cell(row=rr, column=44).value = f"=+O{rr}*J{rr}"             # AR Costo_total
        p.cell(row=rr, column=45).value = it["presup"]                 # AS Presupuesto
        p.cell(row=rr, column=46).value = (
            f"=SUMIFS(Cert_App_Output!$G:$G,Cert_App_Output!$E:$E,$F{rr},"
            f"Cert_App_Output!$K:$K,$AS{rr})")                          # AT
        p.cell(row=rr, column=47).value = f"=AT{rr}*Z{rr}"             # AU
        p.cell(row=rr, column=48).value = (
            f'=IF(AT{rr}=0,"",IF(ABS(AT{rr}-_xlfn.MAXIFS(Cert_App_Output!$H:$H,'
            f'Cert_App_Output!$E:$E,$F{rr},Cert_App_Output!$K:$K,$AS{rr}))<0.0001,"✓","⚠"))')  # AV
        out_r += 1
        n_task += 1

    pmax = out_r - 1

    # sanity check rubro ∈ plan (1_Presupuesto A/B/C/D)
    def sp(sheet, colrange):
        return "+".join(
            f"SUMPRODUCT(('{sheet}'!{cl}4:{cl}{mx}<>\"\")*('{sheet}'!{cl}4:{cl}{mx}<>\"-\")"
            f"*ISNA(MATCH('{sheet}'!{cl}4:{cl}{mx},{rng},0)))"
            for cl, mx in colrange)
    L["E1"] = "VALIDACIÓN — rubros de tarea ∈ plan de cuentas"
    L["E2"] = "Rubros fuera de plan — 1_Presupuesto (A/B/C/D)"
    L["F2"] = "=" + sp("1_Presupuesto", [("A", pmax), ("B", pmax), ("C", pmax), ("D", pmax)])
    L["G2"] = '=IF(F2=0,"✓","⚠")'

    # dropdowns
    dv = DataValidation(type="list", formula1=nm, allow_blank=True, showErrorMessage=False)
    p.add_data_validation(dv)
    dv.add(f"A4:D{pmax}")

    wb.save(DST)
    print(f"✓ Bloque 1b+2 → {DST}")
    print(f"  · _Listas: {len(CANONICO)} rubros (named range RUBROS_PLAN {rng})")
    print(f"  · 1_Presupuesto: {n_task} tareas, filas 4..{pmax}")
    print(f"  · venta = INPUT por tarea (markup no uniforme). markup global={coef:.4f} (1_GGBB!F67 informativo)")
    print(f"  · venta legacy subtotal: PTO01={venta_p1:,.0f}  PTO02={venta_p2:,.0f}  total={venta_total:,.0f}")
    print(f"  · venta reconstruida (Σ venta_u×cant) = {sum_venta:,.0f}  Δ={venta_total-sum_venta:,.2f}")
    print(f"  · costo legacy total = {sum_costo:,.0f}")
    if sin_match:
        print(f"  ⚠ rubros legacy SIN mapeo: {sorted(sin_match)}")
    else:
        print("  · todos los rubros legacy mapeados")
    if sin_venta:
        print(f"  ⚠ tareas SIN venta legacy ({len(sin_venta)}): {sin_venta[:10]}")
    else:
        print("  · todas las tareas con venta legacy matcheada")


if __name__ == "__main__":
    build()
