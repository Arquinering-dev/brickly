"""svd_bloque7_cert.py — Migracion SVD 4140 a v8. Bloque 7 (circuito Cert_*).
Reescribe las 6 hojas Cert_* con los datos reales de SVD:
  - Cert_OC_Cliente: 3 OC (OC01 Pto1 ant20%, OC02 Pto2 ant40%, OC03 Adic). IVA 10,5%
    (obra civil, como Facturacion). Blanco/Negro derivado por OC de la hoja Facturacion.
  - Cert_App_Output: avance fisico por (cert, tarea) de los 2 docs (Pto01 12 certs, Pto02 8).
    PV ALMACENADO del doc (Subtotal) -> montos certificados exactos, sin depender del match de codigo.
  - Cert_Cabecera / Cert_Calculo: madres (ANT + avance), B/N por madre.
  - Cert_Control_OC: 3 OC + fila TOTAL. Cert_Facturacion: header (cobros se cargan luego).
Adapta el loader de GDR al header de 2 filas de SVD (Descripcion / Anterior-Actual-Acumulado).
"""
import sys
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as gl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
LEGACY = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"
FU = "archivos/fuente/"
PMAX = 221
HF = Font(bold=True, color="FFFFFFFF")
HFILL = PatternFill("solid", fgColor="FF1F4E78")
SEP24 = datetime.datetime(2024, 9, 1)

# OC -> (file, anticipo, monto venta, presupuesto-dim, solo_negro)
OCS = {
    "SVD-OC01": dict(file="SVD 4140_Pto. 01_Cert.12.xlsx", ant=0.2, monto=421306052.0,
                     pto="PTO 01", solo_negro=False, fac_tag="Pto 1"),
    "SVD-OC02": dict(file="SVD 4140_Pto. 02_Cert. 08.xlsx", ant=0.4, monto=736857133.02,
                     pto="PTO 02", solo_negro=False, fac_tag="Pto 2"),
    "SVD-OC03": dict(file=None, ant=0.0, monto=15315354.68, pto="PTO 03",
                     solo_negro=False, fac_tag="Adic"),
}
IVA = 0.105


def norm(v):
    return str(v or "").strip().lower()


def hdr(ws, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h); c.font = HF; c.fill = HFILL
        c.alignment = Alignment(horizontal="center")


# ---- split fiscal Blanco/Negro por OC desde Facturacion -------------------
def fiscal_por_oc():
    f = openpyxl.load_workbook(LEGACY, data_only=True)["Facturacion"]
    agg = {}  # tag -> [blanco, negro]
    for r in range(11, 292):
        B = f.cell(row=r, column=2).value          # nº factura ('0003-..' o 'SF')
        E = f.cell(row=r, column=5).value or 0
        J = f.cell(row=r, column=10).value          # 'Pto 1'/'Pto 2'/'Adic'
        try: E = float(E)
        except (TypeError, ValueError): E = 0.0
        if not B or E == 0:
            continue
        tag = str(J or "").strip()
        agg.setdefault(tag, [0.0, 0.0])
        if str(B).strip().upper() == "SF":
            agg[tag][1] += E
        else:
            agg[tag][0] += E
    out = {}
    for tag, (b, n) in agg.items():
        t = b + n
        out[tag] = (b / t, n / t) if t > 0 else (0.31, 0.69)
    return out


# ---- loader de avance (adaptado al header de 2 filas de SVD) ---------------
def localizar(ws):
    hr = None
    for r in range(1, 15):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("descrip" in v for v in vals) and any(v == "subtotal" for v in vals):
            hr = r
            break
    if hr is None:
        return None
    sub = next((c for c in range(1, ws.max_column + 1)
                if norm(ws.cell(hr, c).value) == "subtotal"), None)
    acum = [c for c in range(1, ws.max_column + 1) if norm(ws.cell(hr + 1, c).value) == "acumulado"]
    actual = [c for c in range(1, ws.max_column + 1) if norm(ws.cell(hr + 1, c).value) == "actual"]
    if sub is None or not acum or not actual:
        return None
    best_c, best_n = 1, -1
    for c in range(1, 6):
        n = sum(1 for r in range(hr + 2, ws.max_row + 1)
                if re.match(r"^\d+(\.\d+)+$", str(ws.cell(r, c).value or "").strip()))
        if n > best_n:
            best_c, best_n = c, n
    return hr, best_c, acum[0], actual[0], sub   # cod, acum%, actual%, pv


def parse_cert(ws):
    loc = localizar(ws)
    if not loc:
        return None, None
    hr, cc, c_acum, c_act, c_pv = loc
    fecha = None
    for rr in range(2, hr + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(rr, c).value
            if isinstance(v, datetime.datetime) and v.year >= 2024:
                fecha = v
    out = {}
    for r in range(hr + 2, ws.max_row + 1):
        raw = str(ws.cell(r, cc).value or "").strip()
        if not re.match(r"^\d+(\.\d+)+$", raw):
            continue
        acum = ws.cell(r, c_acum).value or 0
        act = ws.cell(r, c_act).value or 0
        pv = ws.cell(r, c_pv).value or 0
        try: acum = float(acum)
        except (TypeError, ValueError): acum = 0.0
        try: act = float(act)
        except (TypeError, ValueError): act = 0.0
        try: pv = float(pv)
        except (TypeError, ValueError): pv = 0.0
        out[raw] = (acum, act, pv)
    return out, fecha


def descubrir_certs(wb):
    best = {}
    for s in wb.sheetnames:
        m = re.search(r"Cert\.?\s*0*(\d+)", s)
        if m:
            num = int(m.group(1))
            if num not in best or "(2)" in s:
                best[num] = s
    return [(best[n], n) for n in sorted(best)]


def main():
    fiscal = fiscal_por_oc()
    print("Split fiscal por OC (Blanco/Negro):")
    for tag, (b, n) in fiscal.items():
        print(f"   {tag}: B={b*100:.1f}% N={n*100:.1f}%")

    # ---- parsear avance de los docs ----
    app_rows = []      # (oc, idc, fecha, cod, anterior, actual, pv)
    madres = []        # (idc, oc)
    for oc, info in OCS.items():
        if not info["file"]:
            print(f"[{oc}] sin documento de avance (contrato-only)")
            continue
        wb = openpyxl.load_workbook(FU + info["file"], data_only=True)
        certs = descubrir_certs(wb)
        print(f"[{oc}] {info['file']} — {len(certs)} certs")
        for sheet, cnum in certs:
            data, fecha = parse_cert(wb[sheet])
            if data is None:
                print(f"   ⚠ {sheet}: no se pudo anclar — SALTEADO")
                continue
            idc = f"{oc}-C{cnum:02d}"
            madres.append((idc, oc))
            for cod, (acum, act, pv) in sorted(data.items(), key=lambda kv: kv[0]):
                if abs(act) > 1e-9:
                    app_rows.append((oc, idc, fecha, cod, acum - act, act, pv))

    print(f"\n>>> Cert_App_Output filas: {len(app_rows)}  ·  madres de avance: {len(madres)}")

    wb = openpyxl.load_workbook(DST)
    fb, fn = fiscal.get("Pto 1", (0.31, 0.69))

    # ---------- Cert_OC_Cliente ----------
    for name in ["Cert_OC_Cliente", "Cert_App_Output", "Cert_Cabecera", "Cert_Calculo",
                 "Cert_Facturacion", "Cert_Control_OC"]:
        if name in wb.sheetnames:
            del wb[name]
    oc = wb.create_sheet("Cert_OC_Cliente")
    hdr(oc, ["Obra", "ID OC", "Descripcion", "Presupuesto aprobado", "% anticipo", "Mes base CAC",
             "Indice base CAC", "% Blanco sugerido", "% Negro sugerido", "% desacopio sugerido",
             "% IVA sugerido", "Presupuesto", "Fecha anticipo"])
    DESCR = {"SVD-OC01": "Pto 1 - Preliminares + Mov.Suelo + Hormigón",
             "SVD-OC02": "Pto 2 - Albañilería + Instalaciones + Terminaciones",
             "SVD-OC03": "Adicionales / Trabajos Complementarios"}
    for i, (ocid, info) in enumerate(OCS.items()):
        r = 2 + i
        b, n = fiscal.get(info["fac_tag"], (0.31, 0.69))
        oc.cell(r, 1, "SVD 4140"); oc.cell(r, 2, ocid); oc.cell(r, 3, DESCR[ocid])
        oc.cell(r, 4, info["monto"]).number_format = "\\$#,##0"
        oc.cell(r, 5, info["ant"]).number_format = "0%"
        oc.cell(r, 6, SEP24).number_format = "mmm-yy"
        oc.cell(r, 7, f"=_xlfn.XLOOKUP(F{r},'0_Indice_CAC'!$A:$A,'0_Indice_CAC'!$B:$B)")
        oc.cell(r, 8, round(b, 4)).number_format = "0%"
        oc.cell(r, 9, round(n, 4)).number_format = "0%"
        oc.cell(r, 10, info["ant"]).number_format = "0%"       # desacopio = anticipo
        oc.cell(r, 11, IVA).number_format = "0.0%"
        oc.cell(r, 12, info["pto"])
    oc.freeze_panes = "A2"

    # ---------- Cert_App_Output ----------
    ao = wb.create_sheet("Cert_App_Output")
    hdr(ao, ["Obra", "ID OC", "ID Certif", "Fecha", "Cod. tarea", "% anterior", "% actual",
             "% total", "PV total tarea", "$ base tarea (cert act.)", "Presupuesto", "Control vs Ppto"])
    for i, (ocid, idc, fecha, cod, ant, act, pv) in enumerate(app_rows):
        r = 2 + i
        ao.cell(r, 1, "SVD 4140"); ao.cell(r, 2, ocid); ao.cell(r, 3, idc)
        if fecha:
            ao.cell(r, 4, fecha).number_format = "dd/mm/yyyy"
        ao.cell(r, 5, cod)
        ao.cell(r, 6, ant); ao.cell(r, 7, act)
        ao.cell(r, 8, f"=F{r}+G{r}")
        ao.cell(r, 9, pv).number_format = "\\$#,##0"           # PV almacenado del doc
        ao.cell(r, 10, f"=G{r}*I{r}").number_format = "\\$#,##0"
        ao.cell(r, 11, f"=_xlfn.XLOOKUP(B{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$L:$L)")
        ao.cell(r, 12, f'=IF($E{r}="","",IF(AND(ISNUMBER($I{r}),$H{r}<=1.0001),"✓","⚠"))')
    ao.freeze_panes = "A2"

    # ---------- Cert_Cabecera ----------
    cab = wb.create_sheet("Cert_Cabecera")
    hdr(cab, ["ID Certif", "ID OC", "Fecha", "$ base bruta certif.", "% desacopio (madre)",
              "$ desacopio", "Subtotal neto", "Chequeo"])
    cab_rows = [(f"{ocid}-ANT", 0.0) for ocid in OCS]
    cab_rows += [(idc, OCS[ocid]["ant"]) for idc, ocid in madres]
    for i, (idc, desac) in enumerate(cab_rows):
        r = 2 + i
        cab.cell(r, 1, idc)
        cab.cell(r, 2, f'=LEFT($A{r},FIND("|",SUBSTITUTE($A{r},"-","|",2))-1)')
        cab.cell(r, 3, f'=IFERROR(_xlfn.XLOOKUP($A{r},Cert_App_Output!$C:$C,Cert_App_Output!$D:$D),"")')
        cab.cell(r, 4, f"=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$C:$C,$A{r})")
        cab.cell(r, 5, desac).number_format = "0%"
        cab.cell(r, 6, f"=$D{r}*$E{r}")
        cab.cell(r, 7, f"=$D{r}-$F{r}")
        cab.cell(r, 8, f'=IF(COUNTIF(Cert_Calculo!$A:$A,$A{r})>0,"OK","FALTA en Calculo")')
    cab.freeze_panes = "A2"

    # ---------- Cert_Calculo ----------
    cal = wb.create_sheet("Cert_Calculo")
    hdr(cal, ["ID Certif", "ID Cert+Fact", "Tipo", "Obra", "ID OC", "Fecha", "$ base certif (total)",
              "% facturacion", "$ base de esta parte", "% desacopio", "$ desacopio", "$ base neta",
              "Indice base CAC", "Indice CAC a la fecha", "Ratio CAC", "$ CAC", "$ base + CAC",
              "% IVA", "$ IVA", "$ Total certificacion", "USD MEP a la fecha", "U$ Total", "",
              "Control B/N suma 100%", "Indice CAC override"])
    cal_specs = [(f"{ocid}-ANT", ocid, True) for ocid in OCS]
    cal_specs += [(idc, ocid, False) for idc, ocid in madres]
    r = 2
    for idc, ocid, es_ant in cal_specs:
        info = OCS[ocid]
        b, n = fiscal.get(info["fac_tag"], (0.31, 0.69))
        tipos = [("Negro", n)] if info["solo_negro"] else [("Blanco", b), ("Negro", n)]
        ant_monto = info["monto"] * info["ant"]
        for tipo, pfac in tipos:
            suf = "N" if tipo == "Negro" else "B"
            cal.cell(r, 1, idc); cal.cell(r, 2, f"{idc}-{suf}"); cal.cell(r, 3, tipo)
            cal.cell(r, 4, "SVD 4140" if es_ant else
                     f"=_xlfn.XLOOKUP(A{r},Cert_App_Output!$C:$C,Cert_App_Output!$A:$A)")
            cal.cell(r, 5, ocid)
            if not es_ant:
                cal.cell(r, 6, f"=_xlfn.XLOOKUP(A{r},Cert_App_Output!$C:$C,Cert_App_Output!$D:$D)")
            if es_ant:
                cal.cell(r, 7, ant_monto).number_format = "\\$#,##0"
            else:
                cal.cell(r, 7, f"=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$C:$C,$A{r})")
            cal.cell(r, 8, round(pfac, 4)).number_format = "0%"
            cal.cell(r, 9, f"=$G{r}*H{r}")
            cal.cell(r, 10, f"=_xlfn.XLOOKUP($A{r},Cert_Cabecera!$A:$A,Cert_Cabecera!$E:$E)")
            cal.cell(r, 11, f"=-I{r}*J{r}")
            cal.cell(r, 12, f"=I{r}+K{r}")
            cal.cell(r, 13, f"=_xlfn.XLOOKUP($E{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$G:$G)")
            cal.cell(r, 14, (f"=IF($Y{r}=\"\",_xlfn.XLOOKUP(DATE(YEAR(F{r}),MONTH(F{r}),1),"
                             f"'0_Indice_CAC'!$A:$A,'0_Indice_CAC'!$B:$B),$Y{r})") if not es_ant
                     else f"=IF($Y{r}=\"\",'0_Indice_CAC'!$B$4,$Y{r})")
            cal.cell(r, 15, f"=N{r}/M{r}")
            cal.cell(r, 16, f"=L{r}*(O{r}-1)")
            cal.cell(r, 17, f"=L{r}+P{r}")
            cal.cell(r, 18, IVA if tipo == "Blanco" else 0).number_format = "0.0%"
            cal.cell(r, 19, f"=Q{r}*R{r}")
            cal.cell(r, 20, f"=Q{r}+S{r}")
            cal.cell(r, 21, 1170)   # USD MEP provisional (DÓLAR MEP de los docs)
            cal.cell(r, 22, f'=IF(C{r}="Negro",T{r}/U{r},"")')
            cal.cell(r, 24, f'=IF(ABS(SUMIFS($H:$H,$A:$A,$A{r})-1)<0.0001,"✓","⚠ ≠100%")')
            r += 1
    cal.freeze_panes = "A2"

    # ---------- Cert_Facturacion (header) ----------
    fac = wb.create_sheet("Cert_Facturacion")
    hdr(fac, ["id_cert_fact", "Comprobante", "Tipo", "monto", "moneda", "Fecha cobro", "monto_ars_equiv",
              "Estado conciliado", "$ certificado (control)", "Saldo a facturar", "id_factura", "TC cobro",
              "id_cert_madre", "id_OC", "Clase", "Retención", "Haber conciliado"])
    fac.freeze_panes = "A2"

    # ---------- Cert_Control_OC ----------
    co = wb.create_sheet("Cert_Control_OC")
    hdr(co, ["ID OC", "Presupuesto", "Descripcion", "$ Contrato", "% Avance fisico", "$ Certificado",
             "$ Facturado", "$ Cobrado", "$ Anticipo certif.", "$ Avance s/certificar",
             "Saldo a facturar", "Saldo a cobrar", "Estado"])
    for i in range(3):
        r = 2 + i
        co.cell(r, 1, f"=Cert_OC_Cliente!B{2+i}")
        co.cell(r, 2, f"=_xlfn.XLOOKUP($A{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$L:$L)")
        co.cell(r, 3, f"=_xlfn.XLOOKUP($A{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$C:$C)")
        co.cell(r, 4, f"=_xlfn.XLOOKUP($A{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$D:$D)")
        co.cell(r, 5, (f"=IFERROR(SUMPRODUCT(('1_Presupuesto'!$AS$5:$AS${PMAX}=$B{r})*'1_Presupuesto'!$AT$5:$AT${PMAX}"
                       f"*'1_Presupuesto'!$Z$5:$Z${PMAX})/SUMPRODUCT(('1_Presupuesto'!$AS$5:$AS${PMAX}=$B{r})"
                       f"*'1_Presupuesto'!$Z$5:$Z${PMAX}),0)"))
        co.cell(r, 6, f"=SUMIFS(Cert_Calculo!$T:$T,Cert_Calculo!$E:$E,$A{r})")
        co.cell(r, 7, f'=SUMIFS(Cert_Facturacion!$G:$G,Cert_Facturacion!$A:$A,$A{r}&"-*")')
        co.cell(r, 8, f'=SUMIFS(Cert_Facturacion!$G:$G,Cert_Facturacion!$A:$A,$A{r}&"-*",Cert_Facturacion!$F:$F,">0")')
        co.cell(r, 9, f'=SUMIFS(Cert_Calculo!$T:$T,Cert_Calculo!$E:$E,$A{r},Cert_Calculo!$B:$B,"*ANT*")')
        co.cell(r, 10, f'=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$B:$B,$A{r})-SUMIFS(Cert_Calculo!$I:$I,Cert_Calculo!$E:$E,$A{r},Cert_Calculo!$B:$B,"<>*ANT*")')
        co.cell(r, 11, f"=F{r}-G{r}")
        co.cell(r, 12, f"=G{r}-H{r}")
        co.cell(r, 13, f'=IF(J{r}>1000,"⚠ Certificar",IF(F{r}=0,"🔴 Sin iniciar",IF(K{r}>1000,"🟠 Por facturar",IF(L{r}>1000,"🔵 Por cobrar",IF(E{r}>=0.999,"🟢 Cerrada","🟡 En ejecucion")))))')
    tr = 5
    co.cell(tr, 1, "TOTAL").font = Font(bold=True)
    co.cell(tr, 4, "=SUM(D2:D4)").number_format = "\\$#,##0"
    co.cell(tr, 5, "=SUMPRODUCT($D$2:$D$4,$E$2:$E$4)/SUM($D$2:$D$4)").number_format = "0.0%"
    for c in (6, 7, 8, 9, 10, 11, 12):
        co.cell(tr, c, f"=SUM({gl(c)}2:{gl(c)}4)").number_format = "\\$#,##0"
    co.sheet_properties.tabColor = "FF2E7D32"
    co.freeze_panes = "A2"

    wb.save(DST)
    print(f"\n✓ Bloque 7 → {DST}")
    print(f"  · Cert_OC_Cliente: 3 OC (OC01 ${OCS['SVD-OC01']['monto']:,.0f} ant20% · "
          f"OC02 ${OCS['SVD-OC02']['monto']:,.0f} ant40% · OC03 Adic ${OCS['SVD-OC03']['monto']:,.0f})")
    print(f"  · Cert_App_Output: {len(app_rows)} filas (avance real Pto01/Pto02)")
    print(f"  · Cert_Cabecera: {len(cab_rows)} madres · Cert_Calculo: {r-2} filas B/N")
    print(f"  · Cert_Control_OC: 3 OC + TOTAL · Cert_Facturacion: header (cobros pendientes)")
    print(f"  · OC03 Adicionales: contrato sin avance (no hay doc de certificación) — pendiente")


if __name__ == "__main__":
    main()
