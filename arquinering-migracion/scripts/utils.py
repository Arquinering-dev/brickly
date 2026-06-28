"""
utils.py — Funciones auxiliares compartidas
Arquinering S.R.L. — Proyecto Migración Resumen de Obra

Importar desde scripts de migración:
    from utils import aplicar_color, COLOR, normalizar_rubro, RUBROS_CANONICOS
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------------------------------------------------------------------
# COLOR CODING ESTÁNDAR (ver CLAUDE.md sección 4)
# ---------------------------------------------------------------------------

COLOR = {
    "input":     PatternFill("solid", fgColor="DDEEFF"),  # A - Azul  - input manual
    "ref_ext":   PatternFill("solid", fgColor="DDFFDD"),  # B - Verde - referencia externa
    "formula":   PatternFill("solid", fgColor="FFFFFF"),  # C - Blanco - fórmula calculada
    "estatico":  PatternFill("solid", fgColor="F2F2F2"),  # D - Gris  - label / estático
    "pendiente": PatternFill("solid", fgColor="FFFF99"),  # E - Amarillo - pendiente / alerta
}

FUENTE_TITULO = Font(name="Calibri", bold=True, size=11, color="000000")
FUENTE_NORMAL = Font(name="Calibri", size=10, color="000000")
FUENTE_ALERTA = Font(name="Calibri", size=10, color="CC0000", bold=True)

ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center")
ALINEACION_IZQ    = Alignment(horizontal="left",   vertical="center")
ALINEACION_DER    = Alignment(horizontal="right",  vertical="center")

def borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def aplicar_color(ws, celda_ref, tipo: str):
    """
    Aplica el color de fondo según el tipo de celda.
    tipo: "input" | "ref_ext" | "formula" | "estatico" | "pendiente"
    """
    if tipo not in COLOR:
        raise ValueError(f"Tipo de color desconocido: '{tipo}'. "
                         f"Opciones: {list(COLOR.keys())}")
    ws[celda_ref].fill = COLOR[tipo]

def aplicar_rango_color(ws, desde_fila, hasta_fila, desde_col, hasta_col, tipo: str):
    """
    Aplica color a un rango rectangular de celdas.
    Las columnas se pasan como enteros (1=A, 2=B, etc.)
    """
    relleno = COLOR[tipo]
    for fila in range(desde_fila, hasta_fila + 1):
        for col in range(desde_col, hasta_col + 1):
            ws.cell(row=fila, column=col).fill = relleno


# ---------------------------------------------------------------------------
# RUBROS CANÓNICOS (ver CLAUDE.md sección 5)
# ---------------------------------------------------------------------------

RUBROS_CANONICOS = [
    "Hormigón",
    "Albañilería",
    "Revoque",
    "Cielorraso",
    "Revestimiento",
    "Pisos",
    "Pintura",
    "Carpintería",
    "Herrería",
    "Vidriería",
    "Eléctrico",
    "Sanitaria",
    "Gas",
    "Pluvial",
    "Aislaciones",
    "Impermeabilización",
    "Granito",
    "Varios",
    "Gastos Generales",
]

# Mapa de normalización: variantes conocidas → nombre canónico
# Completar con variantes encontradas en el legacy
_NORMALIZACION_RUBROS = {
    # Legacy CH
    "albanileria":         "Albañilería",
    "albañileria":         "Albañilería",
    "albanilería":         "Albañilería",
    "hormigon":            "Hormigón",
    "carpinteria":         "Carpintería",
    "carpintería":         "Carpintería",
    "impermeabilizacion":  "Impermeabilización",
    "impermeabilizacion":  "Impermeabilización",
    "electricidad":        "Eléctrico",
    "electrico":           "Eléctrico",
    "eléctrico":           "Eléctrico",
    "plomeria":            "Sanitaria",
    "plomería":            "Sanitaria",
    "sanitaria":           "Sanitaria",
    "aislacion":           "Aislaciones",
    "aislaciones":         "Aislaciones",
    "cielorraso":          "Cielorraso",
    "cielo raso":          "Cielorraso",
    "revestimientos":      "Revestimiento",
    "pisos y revestimientos": "Pisos",
    "gastos generales":    "Gastos Generales",
    "gg":                  "Gastos Generales",
    "varios":              "Varios",
    "otros":               "Varios",
}

def normalizar_rubro(rubro_raw: str) -> str:
    """
    Convierte un nombre de rubro del archivo legacy al nombre canónico.
    Devuelve el nombre canónico si lo encuentra, o None si no hay mapeo.
    Loguear los None en logs/pendientes.md para resolución manual.
    """
    if rubro_raw is None:
        return None
    clave = str(rubro_raw).strip().lower()
    # Primero verificar si ya está en forma canónica
    for r in RUBROS_CANONICOS:
        if r.lower() == clave:
            return r
    # Luego buscar en el mapa de variantes
    return _NORMALIZACION_RUBROS.get(clave, None)


# ---------------------------------------------------------------------------
# CATEGORÍAS MO UOCRA (ver CLAUDE.md sección 5)
# ---------------------------------------------------------------------------

CATEGORIAS_MO_UOCRA = [
    "CAPATAZ",
    "OFICIAL",
    "MEDIO OFICIAL",
    "AYUDANTE",
    "JEFE DE OBRA",
]

CODIGOS_MO = {
    "CAPATAZ":       "MO-CAP",
    "OFICIAL":       "MO-OFA",
    "MEDIO OFICIAL": "MO-MOF",
    "AYUDANTE":      "MO-AYU01",
    "JEFE DE OBRA":  "MO-JDO",
}


# ---------------------------------------------------------------------------
# HELPERS GENERALES
# ---------------------------------------------------------------------------

def safe_float(valor, default=0.0):
    """Convierte a float ignorando None, strings vacíos y errores."""
    if valor is None:
        return default
    try:
        return float(valor)
    except (ValueError, TypeError):
        return default

def es_formula(valor):
    """True si el valor de celda es una fórmula Excel."""
    return isinstance(valor, str) and valor.startswith("=")

def copiar_estilo(origen, destino):
    """Copia fill, font, alignment y border de una celda a otra."""
    from copy import copy
    destino.fill      = copy(origen.fill)
    destino.font      = copy(origen.font)
    destino.alignment = copy(origen.alignment)
    destino.border    = copy(origen.border)

def ancho_columna_auto(ws, col_letra, min_ancho=8, max_ancho=40):
    """Ajusta el ancho de una columna al contenido más largo."""
    max_len = 0
    for celda in ws[col_letra]:
        if celda.value:
            largo = len(str(celda.value))
            if largo > max_len:
                max_len = largo
    ws.column_dimensions[col_letra].width = max(min_ancho, min(max_len + 2, max_ancho))

def validar_rubros_presentes(rubros_a_verificar: list) -> dict:
    """
    Recibe una lista de nombres de rubro y devuelve un dict:
      { rubro: "OK" | "NO_CANONICO" | "SIN_MAPEO" }
    """
    resultado = {}
    for r in rubros_a_verificar:
        if r in RUBROS_CANONICOS:
            resultado[r] = "OK"
        else:
            mapeado = normalizar_rubro(r)
            resultado[r] = f"→ {mapeado}" if mapeado else "SIN_MAPEO ⚠️"
    return resultado


if __name__ == "__main__":
    # Test básico
    print("RUBROS CANÓNICOS:")
    for r in RUBROS_CANONICOS:
        print(f"  {r}")

    print("\nPRUEBA normalizar_rubro:")
    tests = ["albanileria", "Hormigón", "plomeria", "rubro_desconocido"]
    for t in tests:
        print(f"  '{t}' → '{normalizar_rubro(t)}'")
