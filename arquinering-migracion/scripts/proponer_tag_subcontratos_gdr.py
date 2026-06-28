# -*- coding: utf-8 -*-
"""
Genera una planilla APARTE con la propuesta de tagueo de pagos a subcontratos
en 2_Movimientos del GDR v8_12. NO modifica el archivo de obra.
Salida: archivos/output/GDR_Propuesta_Tag_Subcontratos_2_Movimientos.xlsx
"""
import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = 'archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx'
OUT = 'archivos/output/GDR_Propuesta_Tag_Subcontratos_2_Movimientos.xlsx'

# Maestro de subcontratos (de 2_Subcontratos)
MAESTRO = {
    'GDR-SC-001': ('ELVIO CESPEDES',    'Durlock/Yeso',          'Yeso',                     14940000),
    'GDR-SC-002': ('ELVIO CESPEDES',    'Revoque Exterior',      'Revoque Exterior',          8360000),
    'GDR-SC-003': ('ANCLAJES CIMA',     'Fundaciones',           'Fundaciones',               3037256.14),
    'GDR-SC-004': ('CELSI VIAL S.A.',   'Exc. y Mov. de Suelos', 'Excavacion y Mov. Suelos',        0),
    'GDR-SC-005': ('CUADRILLA HORMIGON','Hormigon MO',           'Hormigon M.O.',            89779071.33),
}

# Clasificacion propuesta: fila -> (SC_id, TIPO, confianza, nota)
PROP = {
    # --- SC-004 Celsi Vial (cuenta 53003 Mov. de Suelos) ---
    410: ('GDR-SC-004', 'BASE', 'Alta', 'Nombre explicito "Celsi Vial". Excavacion inicial.'),
    412: ('GDR-SC-004', 'BASE', 'Alta', 'Celsi Vial - Certif. #2'),
    413: ('GDR-SC-004', 'BASE', 'Alta', 'Celsi Vial - Certif. #4'),
    414: ('GDR-SC-004', 'BASE', 'Alta', 'Celsi Vial - Certif. #5'),
    415: ('GDR-SC-004', 'BASE', 'Alta', 'Celsi Vial - Certif. #5 (2do pago)'),
    416: ('GDR-SC-004', 'BASE', 'Alta', 'Celsi Vial - Certif. #7'),
    # --- SC-003 Anclajes Cima (cuenta 53005 Hormigon MO) ---
    475: ('GDR-SC-003', 'ANT',  'Alta',  'Anclajes Cima - Anticipo 30%'),
    477: ('GDR-SC-003', 'BASE', 'Alta',  'Anclajes CIMA - Certif. #1'),
    478: ('GDR-SC-003', 'CAC',  'Media', 'Obs truncada "Certif. #1 - C" -> probable CAC del certif #1'),
    479: ('GDR-SC-003', 'CAC',  'Alta',  'Anclajes CIMA - Anticipo CAC'),
    # --- SC-005 Cuadrilla Hormigon (cuenta 53005) ---
    483: ('GDR-SC-005', 'BASE', 'Alta', '"Cuadrilla Hormigon" textual'),
    # Certificados de hormigon estructural en 53005 SIN nombre de proveedor:
    481: ('GDR-SC-005', 'BASE', 'Baja', '"Certif. #1 - Hormigon Fundaciones". ATRIBUCION DUDOSA: podria ser SC-003 (Anclajes=Fundaciones) o SC-005 (cuadrilla). Confirmar.'),
    482: ('GDR-SC-005', 'BASE', 'Baja', '"Certif. #1 - Hormigon Fundaciones" (split). Misma duda SC-003 vs SC-005.'),
    486: ('GDR-SC-005', 'BASE', 'Baja', '"Rafael Meza - Certif. #3". Meza = posible capataz cuadrilla hormigon. Confirmar si es SC-005.'),
    488: ('GDR-SC-005', 'BASE', 'Baja', '"Certif. #2 - Losa 3 Hormigon". Tentativo SC-005.'),
    489: ('GDR-SC-005', 'BASE', 'Baja', '"Certif. #3 - Losa 4 Hormigon". Tentativo SC-005.'),
    490: ('GDR-SC-005', 'BASE', 'Baja', '"Certif. Final - Hormigon". Tentativo SC-005.'),
    # --- SC-002 Revoque (cuenta 53009 Albanileria MO + 53014 Revestimiento) ---
    560: ('GDR-SC-002', 'BASE', 'Media', 'Cespedes Certif. #2 (porcion revoque en Albanileria; mismo dia que yeso r583).'),
    567: ('GDR-SC-002', 'BASE', 'Media', 'Elvio Cespedes Certif. #3 (revoque).'),
    568: ('GDR-SC-002', 'CAC',  'Media', 'Elvio Cespedes - CAC * Certif. (revoque).'),
    569: ('GDR-SC-002', 'CAC',  'Media', 'Elvio Cespedes - CAC * Certif. (revoque, 2da linea).'),
    571: ('GDR-SC-002', 'BASE', 'Media', 'Elvio Cespedes Certif. #4 (revoque).'),
    589: ('GDR-SC-002', 'BASE', 'Alta',  '"Elvio Cespedes - Revoque" en cuenta Revestimiento MO. Revoque claro -> SC-002.'),
    # --- SC-001 Yeso (cuenta 53011 Durlock MO) ---
    580: ('GDR-SC-001', 'BASE', 'Alta', 'Elvio Cespedes - Yesero'),
    581: ('GDR-SC-001', 'BASE', 'Alta', 'Elvio Cespedes - Yesero'),
    582: ('GDR-SC-001', 'BASE', 'Alta', 'Cespedes Elvio - Yesero - Certif.'),
    583: ('GDR-SC-001', 'BASE', 'Alta', 'Cespedes Elvio - Certif. #2 (yeso).'),
    584: ('GDR-SC-001', 'BASE', 'Alta', 'Elvio Cespedes - Certif. #3 (yeso).'),
    585: ('GDR-SC-001', 'CAC',  'Alta', 'Elvio Cespedes - CAC * Certif. (yeso).'),
    586: ('GDR-SC-001', 'BASE', 'Alta', 'Elvio Cespedes - Certif. #4 (yeso).'),
    587: ('GDR-SC-001', 'BASE', 'Alta', 'Elvio Cespedes - Yeso - Certif.'),
}

EXCLUIDOS = [476, 480, 484, 485, 487]  # bonos/adelantos puntuales en cuentas MO

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['2_Movimientos']


def gv(r, c):
    return ws.cell(r, c).value


def fecha(r):
    c = ws.cell(r, 3).value
    return c.strftime('%Y-%m-%d') if hasattr(c, 'strftime') else c


rows_prop = []
for r in sorted(PROP):
    sc, tipo, conf, nota = PROP[r]
    obs_orig = str(gv(r, 5) or '').strip()
    cuenta = gv(r, 1)
    desc_cta = gv(r, 2)
    debe = gv(r, 9) or 0
    prov = MAESTRO[sc][0]
    obs_new = f'{sc} | {tipo} | {obs_orig}'
    rows_prop.append([r, fecha(r), cuenta, desc_cta, debe, obs_orig, sc, prov, tipo, obs_new, conf, nota])

excl_rows = []
for r in range(2, ws.max_row + 1):
    A = str(gv(r, 1) or '')
    if A not in ('53005', '53009'):
        continue
    if r in PROP:
        continue
    obs = str(gv(r, 5) or '').strip()
    if re.search(r'\b[12]Q\b', obs) or 'SEREN' in obs.upper() or 'Sobres' in obs:
        excl_rows.append([r, fecha(r), gv(r, 1), gv(r, 2), gv(r, 9) or 0, obs, 'Nomina/Quincena (circuito 2_Quincenas), NO subcontrato'])
    elif r in EXCLUIDOS:
        excl_rows.append([r, fecha(r), gv(r, 1), gv(r, 2), gv(r, 9) or 0, obs, 'Bono/adelanto nomina, NO subcontrato'])

# ---- Escribir workbook salida ----
out = openpyxl.Workbook()
HDR = PatternFill('solid', fgColor='1F4E78')
HDRF = Font(color='FFFFFF', bold=True, name='Aptos Narrow')
TITLEF = Font(color='1F3864', bold=True, size=13, name='Aptos Narrow')
CONF_FILL = {'Alta': 'C6EFCE', 'Media': 'FFEB9C', 'Baja': 'FFC7CE'}
thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
AN = Font(name='Aptos Narrow')

ws1 = out.active
ws1.title = 'Propuesta_Subcontratos'
ws1['A1'] = 'GDR 3760 - Propuesta de tagueo de pagos a SUBCONTRATOS en 2_Movimientos (v8_12)'
ws1['A1'].font = TITLEF
hdr1 = ['Fila 2_Mov', 'Fecha', 'Cuenta', 'Desc Cuenta', 'Debe (ARS)', 'Observacion original (Tezamat)',
        'SC propuesto', 'Proveedor', 'TIPO', 'OBSERVACIONES A ESCRIBIR (col E)', 'Confianza', 'Nota / motivo']
for j, h in enumerate(hdr1, 1):
    c = ws1.cell(3, j, h)
    c.fill = HDR
    c.font = HDRF
    c.alignment = Alignment(wrap_text=True, vertical='center')
    c.border = BORDER
ri = 4
for row in rows_prop:
    for j, val in enumerate(row, 1):
        c = ws1.cell(ri, j, val)
        c.border = BORDER
        c.font = AN
        if j == 5:
            c.number_format = '#,##0'
        if j == 10:
            c.font = Font(name='Consolas', bold=True)
        if j == 11:
            c.fill = PatternFill('solid', fgColor=CONF_FILL.get(val, 'FFFFFF'))
            c.alignment = Alignment(horizontal='center')
        if j in (6, 12):
            c.alignment = Alignment(wrap_text=True, vertical='top')
    ri += 1
for j, w in enumerate([9, 11, 8, 18, 13, 32, 12, 16, 7, 40, 9, 46], 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
ws1.freeze_panes = 'A4'

ws2 = out.create_sheet('Excluidos_nomina')
ws2['A1'] = 'Filas en cuentas MO (53005/53009) consideradas y DESCARTADAS como subcontrato'
ws2['A1'].font = TITLEF
hdr2 = ['Fila 2_Mov', 'Fecha', 'Cuenta', 'Desc Cuenta', 'Debe (ARS)', 'Observacion original', 'Motivo de exclusion']
for j, h in enumerate(hdr2, 1):
    c = ws2.cell(3, j, h)
    c.fill = HDR
    c.font = HDRF
    c.alignment = Alignment(wrap_text=True, vertical='center')
    c.border = BORDER
ri = 4
for row in excl_rows:
    for j, val in enumerate(row, 1):
        c = ws2.cell(ri, j, val)
        c.border = BORDER
        c.font = AN
        if j == 5:
            c.number_format = '#,##0'
    ri += 1
for j, w in enumerate([9, 11, 8, 18, 13, 34, 44], 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
ws2.freeze_panes = 'A4'

ws3 = out.create_sheet('Resumen_por_SC')
ws3['A1'] = 'Resumen: total propuesto a imputar por subcontrato vs maestro 2_Subcontratos'
ws3['A1'].font = TITLEF
hdr3 = ['SC', 'Proveedor', 'Rubro', 'Monto Presup. (E)', '# movs propuestos',
        'Total propuesto', 'Total BASE+ANT', 'Total CAC', 'Confianza dominante']
for j, h in enumerate(hdr3, 1):
    c = ws3.cell(3, j, h)
    c.fill = HDR
    c.font = HDRF
    c.alignment = Alignment(wrap_text=True, vertical='center')
    c.border = BORDER
agg = defaultdict(lambda: [0, 0.0, 0.0, 0.0, defaultdict(int)])
for row in rows_prop:
    sc = row[6]
    debe = row[4]
    tipo = row[8]
    conf = row[10]
    a = agg[sc]
    a[0] += 1
    a[1] += debe
    if tipo in ('BASE', 'ANT'):
        a[2] += debe
    if tipo == 'CAC':
        a[3] += debe
    a[4][conf] += 1
ri = 4
for sc in sorted(MAESTRO):
    prov, rubro, _, pres = MAESTRO[sc]
    a = agg.get(sc, [0, 0, 0, 0, {}])
    confdom = max(a[4].items(), key=lambda x: x[1])[0] if a[4] else '-'
    vals = [sc, prov, rubro, pres, a[0], a[1], a[2], a[3], confdom]
    for j, val in enumerate(vals, 1):
        c = ws3.cell(ri, j, val)
        c.border = BORDER
        c.font = AN
        if j in (4, 6, 7, 8):
            c.number_format = '#,##0'
    ri += 1
for j, w in enumerate([12, 18, 20, 17, 16, 18, 16, 14, 16], 1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
ws3.freeze_panes = 'A4'

ws4 = out.create_sheet('Instrucciones')
notas = [
    'COMO USAR ESTA PLANILLA',
    '',
    '1) Esta planilla es una PROPUESTA. No modifica GDR_..._v8_12.xlsx.',
    '2) El dashboard lee los pagos a subcontrato parseando la columna E (Observaciones) de 2_Movimientos',
    '   con el formato:  {SC-id} | {TIPO} | {descripcion}',
    '   - 2_Movimientos!Q (mov_id)   = texto antes del primer " | "   -> debe quedar el SC-id (GDR-SC-00X)',
    '   - 2_Movimientos!R (mov_tipo)  = texto entre el 1er y 2do " | " -> BASE / ANT / CAC / CS',
    '   - 2_Subcontratos!H = SUMIFS(BASE)+SUMIFS(ANT);  I = SUMIFS(CAC);  J = SUMIFS(CS)',
    '',
    'TIPOS:',
    '   BASE = certificado base (descuenta saldo del contrato)',
    '   ANT  = anticipo (descuenta saldo)',
    '   CAC  = ajuste por indice CAC (NO descuenta saldo)',
    '   CS   = costo social / extra (NO descuenta saldo) - no se detectaron en GDR',
    '',
    'PARA APLICARLO (cuando lo apruebes): copiar el texto de la columna "OBSERVACIONES A ESCRIBIR"',
    'sobre 2_Movimientos!E de la fila indicada (el texto original queda preservado en 2_Movimientos!P).',
    'Luego recalcular con scripts/recalc.py.',
    '',
    'PUNTOS A CONFIRMAR (CLAUDE.md: no asumir mapeo de subcontratos):',
    '  a) SC-005 vs SC-003 en cuenta 53005 (Hormigon MO): los certificados "Hormigon Fundaciones /',
    '     Losa / Rafael Meza" (filas 481,482,486,488,489,490) - son de la cuadrilla de hormigon (SC-005)',
    '     o de Anclajes Cima/Fundaciones (SC-003)? Asignados TENTATIVAMENTE a SC-005 (confianza Baja).',
    '  b) Cespedes: se separo su certificado en Yeso (SC-001, cuenta Durlock MO) y Revoque (SC-002,',
    '     cuentas Albanileria + Revestimiento MO), porque emite ambos el mismo dia. Confirmar el corte.',
    '  c) SC-004 (Celsi) tiene Monto Presup. = 0 en el maestro pero ~8,7M de pagos: falta cargar el',
    '     monto del contrato en 2_Subcontratos!E7, si no el % consumido / saldo dara incoherente.',
]
for i, t in enumerate(notas, 1):
    c = ws4.cell(i, 1, t)
    c.font = Font(name='Aptos Narrow', bold=(i == 1 or t.endswith(':') or t.startswith('PUNTOS')))
ws4.column_dimensions['A'].width = 112

out.save(OUT)
print('OK ->', OUT)
print('Movs propuestos:', len(rows_prop), '| Excluidos listados:', len(excl_rows))
tot = sum(r[4] for r in rows_prop)
print('Total ARS propuesto a subcontratos:', f'{tot:,.0f}'.replace(',', '.'))
