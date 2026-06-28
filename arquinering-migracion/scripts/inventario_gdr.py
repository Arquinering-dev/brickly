"""Inventario GDR vs CH v8_11 — bloque 1 de la migracion de GDR a v8.
Lista hojas, dimensiones, tab color y primeras filas (titulo/header) de cada libro.
Solo lectura."""
import openpyxl
from openpyxl.utils import get_column_letter

GDR = "archivos/referencia/GDR_3760_Resumen_de_Obra_v8.xlsx"
CH  = "archivos/output/CH_2171_Resumen_de_Obra_v8_11.xlsx"


def first_text_cells(ws, max_rows=3, max_cols=12):
    """Devuelve el texto de las primeras filas (para ver titulo/header)."""
    out = []
    for r in range(1, max_rows + 1):
        row = []
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row.append(f"{get_column_letter(c)}{r}={repr(v)[:40]}")
        if row:
            out.append("  " + " | ".join(row))
    return out


def inventory(path, label):
    wb = openpyxl.load_workbook(path, read_only=False)
    print(f"\n{'='*78}\n{label}: {path}\n{'='*78}")
    print(f"Total hojas: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        tab = ws.sheet_properties.tabColor
        tabc = tab.rgb if tab else "-"
        print(f"\n[{name}]  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}  tab={tabc}")
        for line in first_text_cells(ws):
            print(line)
    return set(wb.sheetnames)


gdr_sheets = inventory(GDR, "GDR (referencia, estructura clasica)")
ch_sheets  = inventory(CH,  "CH v8_11 (MODELO a replicar)")

print(f"\n\n{'#'*78}\nCOMPARACION DE HOJAS\n{'#'*78}")
print("\n-- Hojas en CH (modelo) que NO estan en GDR (=> hay que crear/adaptar):")
for s in ch_sheets - gdr_sheets:
    print(f"   + {s}")
print("\n-- Hojas en GDR que NO estan en CH (=> candidatas a eliminar/maestro/analisis):")
for s in gdr_sheets - ch_sheets:
    print(f"   - {s}")
print("\n-- Hojas en ambos:")
for s in sorted(gdr_sheets & ch_sheets):
    print(f"   = {s}")
