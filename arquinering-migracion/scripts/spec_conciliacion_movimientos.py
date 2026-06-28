"""
SPEC Conciliacion contra 2_Movimientos (ingresos cert + egresos subcontrato).
Entrada: CH_2171_..._v8_8.xlsx  ->  Salida: ..._v8_9.xlsx (v8_8 backup).
Solo forma/estructura computada; convencion de Observaciones {ID} | {TIPO} | {desc}.
 1) Normaliza las 12 filas SC de E al convenio ' | ' (AVANCE->BASE, CAC->CAC); deja quincenas/ingresos.
 2) 2_Movimientos: cols derivadas mov_id (Q) / mov_tipo (R) parseando E.
 3) 2_Subcontratos: pagado BASE+ANT / CAC / CS / total / saldo disp / % / estado (regla CAC-CS no descuenta).
 4) Cert_Facturacion: + Retencion (P, input) + Haber conciliado (Q, ref) ; H repurposado a Estado conciliado.
 5) 0_CONFIG!B49 = tolerancia de match ($, input).
"""
import re
import openpyxl
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci
from openpyxl.styles import Font, PatternFill, Alignment

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_8.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_9.xlsx"
TOL_CELL = "'0_CONFIG'!$B$49"

wb = openpyxl.load_workbook(SRC, data_only=False)
base = wb['Cert_Calculo'].cell(2, 1)
FN, FS = base.font.name or 'Aptos Narrow', base.font.size or 11
HDR_FILL = PatternFill('solid', fgColor='FF1F4E78')
HDR_FONT = Font(name=FN, size=FS, bold=True, color='FFFFFFFF')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
AZUL = lambda: Font(name=FN, size=FS, color='FF0000FF')
VERDE = lambda: Font(name=FN, size=FS, color='FF00B050')
NEGRO = lambda: Font(name=FN, size=FS, color='FF000000')
MONEY, PCT = '\\$#,##0', '0.0%'


def hdr(ws, row, col, text):
    c = ws.cell(row, col, text); c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN
    return c


# ===== 1) Normalizar filas SC de E al convenio ' | ' =====
mv = wb['2_Movimientos']
TIPO_MAP = {'AVANCE': 'BASE', 'CAC': 'CAC', 'ANTICIPO': 'ANT', 'ANT': 'ANT', 'CS': 'CS'}
norm = 0
pat = re.compile(r'^(CH-SC-\d+)\s+(AVANCE|CAC|ANTICIPO|ANT|CS)$', re.I)
for r in range(2, mv.max_row + 1):
    E = mv.cell(r, ci('E')).value
    if not isinstance(E, str):
        continue
    m = pat.match(E.strip())
    if not m:
        continue
    sid, tipo = m.group(1).upper(), TIPO_MAP[m.group(2).upper()]
    desc = str(mv.cell(r, ci('P')).value or '').strip().strip('()')
    mv.cell(r, ci('E')).value = "%s | %s | %s" % (sid, tipo, desc)
    norm += 1

# ===== 2) 2_Movimientos: mov_id (Q) / mov_tipo (R) =====
Q, R = ci('Q'), ci('R')   # 17, 18
hdr(mv, 1, Q, 'mov_id'); hdr(mv, 1, R, 'mov_tipo')
for r in range(2, mv.max_row + 1):
    c = mv.cell(r, Q, '=IFERROR(TRIM(LEFT($E%d,FIND(" | ",$E%d)-1)),"")' % (r, r)); c.font = NEGRO()
    c = mv.cell(r, R, '=IFERROR(TRIM(UPPER(MID($E{r},FIND(" | ",$E{r})+3,'
                      'IFERROR(FIND(" | ",$E{r},FIND(" | ",$E{r})+3),LEN($E{r})+1)'
                      '-FIND(" | ",$E{r})-3))),"")'.format(r=r)); c.font = NEGRO()
mv.column_dimensions['Q'].width = 16; mv.column_dimensions['R'].width = 11

# ===== 3) 2_Subcontratos: control de egresos por contrato =====
sc = wb['2_Subcontratos']
cols = [('H', 'Pagado BASE+ANT'), ('I', 'CAC pagado'), ('J', 'CS pagado'),
        ('K', 'Total pagado'), ('L', 'Saldo disponible'), ('M', '% consumido'), ('N', 'Estado')]
for L, t in cols:
    hdr(sc, 3, ci(L), t)
MOV = "'2_Movimientos'"
for r in range(4, sc.max_row + 1):
    a = sc.cell(r, ci('A')).value
    if not (isinstance(a, str) and a.startswith('CH-SC')):
        continue
    def sif(tipo):
        return 'SUMIFS(%s!$I:$I,%s!$Q:$Q,$A%d,%s!$R:$R,"%s")' % (MOV, MOV, r, MOV, tipo)
    sc.cell(r, ci('H'), '=%s+%s' % (sif('BASE'), sif('ANT'))).font = VERDE()
    sc.cell(r, ci('I'), '=%s' % sif('CAC')).font = VERDE()
    sc.cell(r, ci('J'), '=%s' % sif('CS')).font = VERDE()
    sc.cell(r, ci('K'), '=H%d+I%d+J%d' % (r, r, r)).font = NEGRO()
    sc.cell(r, ci('L'), '=$E%d-H%d' % (r, r)).font = NEGRO()
    sc.cell(r, ci('M'), '=IFERROR(H%d/$E%d,0)' % (r, r)).font = NEGRO()
    sc.cell(r, ci('N'),
            '=IF(L{r}<=0,"🔴 Sin saldo",IF(M{r}>=0.9,"🟠 <10% saldo","🟢 OK"))'.format(r=r)).font = NEGRO()
    for L in 'HIJKL':
        sc.cell(r, ci(L)).number_format = MONEY
    sc.cell(r, ci('M')).number_format = PCT
for L, w in {'H': 15, 'I': 13, 'J': 12, 'K': 14, 'L': 15, 'M': 12, 'N': 15}.items():
    sc.column_dimensions[L].width = w

# ===== 4) Cert_Facturacion: Retencion + Haber conciliado + Estado (repurpose H) =====
fa = wb['Cert_Facturacion']
fa.cell(1, ci('H'), 'Estado conciliado'); fa.cell(1, ci('H')).font = HDR_FONT
fa.cell(1, ci('H')).fill = HDR_FILL; fa.cell(1, ci('H')).alignment = HDR_ALIGN
hdr(fa, 1, ci('P'), 'Retención'); hdr(fa, 1, ci('Q'), 'Haber conciliado')
for r in range(2, fa.max_row + 1):
    if fa.cell(r, ci('A')).value is None:
        continue
    # Retencion: input (default 0)
    rc = fa.cell(r, ci('P'), 0); rc.font = AZUL(); rc.number_format = MONEY
    # Haber conciliado (BASE) desde 2_Movimientos por id_cert_fact
    hc = fa.cell(r, ci('Q'),
                 '=SUMIFS(%s!$J:$J,%s!$Q:$Q,$A%d,%s!$R:$R,"BASE")' % (MOV, MOV, r, MOV))
    hc.font = VERDE(); hc.number_format = MONEY
    # Estado conciliado computado (reemplaza flag manual): Haber + Retencion = certificado +- tol
    est = fa.cell(r, ci('H'),
                  '=IF($I{r}=0,"—",IF(ABS($Q{r}+$P{r}-$I{r})<={tol},"✅ Conciliado",'
                  'IF($Q{r}>0,"🟡 Parcial","🔴 Pendiente")))'.format(r=r, tol=TOL_CELL))
    est.font = NEGRO()
fa.column_dimensions['P'].width = 12; fa.column_dimensions['Q'].width = 15

# ===== 5) 0_CONFIG: parametro tolerancia =====
cf = wb['0_CONFIG']
cf.cell(49, 1, 'Tolerancia conciliación ($)').font = NEGRO()
tc = cf.cell(49, 2, 100); tc.font = AZUL(); tc.number_format = MONEY

wb.save(DST)
print("Guardado:", DST, "| filas E normalizadas:", norm)

# ===== verificacion: saldos esperados por SC (BASE+ANT descuenta) =====
from collections import defaultdict
pagB = defaultdict(float); pagCAC = defaultdict(float)
for r in range(2, mv.max_row + 1):
    E = mv.cell(r, ci('E')).value
    if not isinstance(E, str) or ' | ' not in E:
        continue
    sid = E.split(' | ')[0].strip(); tipo = E.split(' | ')[1].strip().upper()
    debe = mv.cell(r, ci('I')).value or 0
    if tipo in ('BASE', 'ANT'):
        pagB[sid] += debe
    elif tipo == 'CAC':
        pagCAC[sid] += debe
print("\nVerificacion egresos por contrato:")
for r in range(4, sc.max_row + 1):
    a = sc.cell(r, ci('A')).value
    if isinstance(a, str) and a.startswith('CH-SC'):
        pres = sc.cell(r, ci('E')).value or 0
        print("  %-10s pres=%-12s BASE+ANT=%-12s CAC=%-10s saldo=%s" %
              (a, pres, pagB[a], pagCAC.get(a, 0), pres - pagB[a]))
