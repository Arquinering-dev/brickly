"""Bloque Q6 — Reestructura 2_Subcontratos de GDR del maestro viejo (A-M) a la matriz de
conciliación de CH (A-N). Decisión Pedro 2026-06-22: mantener local + matriz CH.
- Preserva los 5 contratos (Proveedor/Rubro/Descripción/Monto/Ajusta CAC/% Anticipo).
- Renombra Contrato# SC-NNN -> GDR-SC-NNN (convención CH, único para futuro maestro cross-obra).
- Rubros col C: se dejan DESCRIPTIVOS (CH no los alinea al plan; el cruce es por Contrato#, no rubro).
- H-N = fórmulas de conciliación CH (cruzan 2_Movimientos por mov_id/mov_tipo). Inertes hasta
  que Tezamat cargue pagos tagueados `GDR-SC-NNN | BASE/ANT/CAC/CS | ...`.
Value-preserving en los datos de contrato. In-place sobre v8_12.

Uso: python scripts/reestructurar_subcontratos_gdr.py
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
HR = 3  # fila header (como CH)

HEADERS = ["Contrato#", "Proveedor", "Rubro", "Descripción", "Monto Presup.", "Ajusta CAC",
           "% Anticipo", "Pagado BASE+ANT", "CAC pagado", "CS pagado", "Total pagado",
           "Saldo disponible", "% consumido", "Estado"]
# fórmulas H-N de CH (parametrizadas por fila r)
FH = ('=SUMIFS(\'2_Movimientos\'!$I:$I,\'2_Movimientos\'!$Q:$Q,$A{r},\'2_Movimientos\'!$R:$R,"BASE")'
      '+SUMIFS(\'2_Movimientos\'!$I:$I,\'2_Movimientos\'!$Q:$Q,$A{r},\'2_Movimientos\'!$R:$R,"ANT")')
FI = '=SUMIFS(\'2_Movimientos\'!$I:$I,\'2_Movimientos\'!$Q:$Q,$A{r},\'2_Movimientos\'!$R:$R,"CAC")'
FJ = '=SUMIFS(\'2_Movimientos\'!$I:$I,\'2_Movimientos\'!$Q:$Q,$A{r},\'2_Movimientos\'!$R:$R,"CS")'
FK = "=H{r}+I{r}+J{r}"
FL = "=$E{r}-H{r}"
FM = "=IFERROR(H{r}/$E{r},0)"
FN = '=IF(L{r}<=0,"🔴 Sin saldo",IF(M{r}>=0.9,"🟠 <10% saldo","🟢 OK"))'


def main():
    wb = openpyxl.load_workbook(PATH)
    sc = wb["2_Subcontratos"]

    # 1) leer los contratos viejos (rows 4..) preservando A-G
    contratos = []
    for r in range(HR + 1, sc.max_row + 1):
        a = sc.cell(r, 1).value
        if a is None or str(a).startswith(("TOTAL", "INSTRU", "►")):
            continue
        sid = str(a).strip()
        if not sid.upper().startswith("SC-"):
            continue
        nuevo = "GDR-" + sid  # SC-001 -> GDR-SC-001
        contratos.append({
            "id": nuevo,
            "prov": sc.cell(r, 2).value,
            "rubro": sc.cell(r, 3).value,
            "desc": sc.cell(r, 4).value,
            "monto": sc.cell(r, 5).value,
            "cac": sc.cell(r, 6).value,
            "ant": sc.cell(r, 7).value,
        })

    # 2) limpiar contenido viejo desde fila HR (header) hacia abajo
    for mc in list(sc.merged_cells.ranges):
        sc.unmerge_cells(str(mc))
    for r in range(HR, sc.max_row + 1):
        for c in range(1, sc.max_column + 1):
            sc.cell(r, c).value = None
    sc.data_validations.dataValidation = []

    # 3) header (fila HR)
    hf = Font(bold=True, color="FFFFFFFF")
    hfill = PatternFill("solid", fgColor="FF1F4E78")
    for j, h in enumerate(HEADERS, start=1):
        c = sc.cell(HR, j, h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    # 4) contratos + fórmulas de conciliación
    for i, ct in enumerate(contratos):
        r = HR + 1 + i
        sc.cell(r, 1, ct["id"])
        sc.cell(r, 2, ct["prov"])
        sc.cell(r, 3, ct["rubro"])
        sc.cell(r, 4, ct["desc"])
        sc.cell(r, 5, ct["monto"]).number_format = "\\$#,##0"
        sc.cell(r, 6, ct["cac"])
        sc.cell(r, 7, ct["ant"]).number_format = "0%"
        for col, f in [(8, FH), (9, FI), (10, FJ), (11, FK), (12, FL), (13, FM), (14, FN)]:
            sc.cell(r, col, f.format(r=r))
        sc.cell(r, 11).number_format = "\\$#,##0"
        sc.cell(r, 12).number_format = "\\$#,##0"
        sc.cell(r, 13).number_format = "0%"

    sc.sheet_properties.tabColor = "FF2E75B6"
    sc.freeze_panes = "A%d" % (HR + 1)
    for col, w in {"A": 13, "B": 20, "C": 18, "D": 22, "E": 15, "F": 11, "G": 10,
                   "H": 16, "I": 13, "J": 13, "K": 14, "L": 15, "M": 12, "N": 14}.items():
        sc.column_dimensions[col].width = w

    wb.save(PATH)
    print(f"✓ 2_Subcontratos reestructurado → matriz CH ({len(contratos)} contratos)")
    for ct in contratos:
        print(f"   {ct['id']}: {ct['prov']} / {ct['rubro']} / ${ct['monto']:,.0f} (CAC {ct['cac']}, ant {ct['ant']})")
    print("   H-N conciliación cableada (inerte hasta pagos tagueados en 2_Movimientos)")


if __name__ == "__main__":
    main()
