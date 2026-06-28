"""Bloque 5 — Reestructura 2_Gastos_DirInd de GDR de matriz (concepto×mes) a tabla plana
(Fecha | Tipo | Concepto | Monto), como CH v8_11. Adaptación de
reestructurar_gastos_dirind.py: in-place sobre v8_12 + banner de sección por MATCH EXACTO
(el título A1 de GDR "GASTOS DIRECTOS E INDIRECTOS…" rompía el startswith del script CH).

Value-preserving: cada celda numérica no-cero de los meses → una fila. Total Real/Descontado,
Ratio CAC, Estado del mes y filas TOTAL se descartan (las deriva el dashboard).
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter as gl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"

wb = openpyxl.load_workbook(PATH)
g = wb["2_Gastos_DirInd"]

# ---- 1) leer la matriz vieja ----
mesescol = {}  # col -> fecha (header fila 3 que es FECHA)
for c in range(2, g.max_column + 1):
    v = g.cell(3, c).value
    if hasattr(v, "year") and hasattr(v, "month"):
        mesescol[c] = v

conceptos_orden = []
seccion = None
recs = []  # (fecha, tipo, concepto, monto)
for r in range(1, g.max_row + 1):
    a = g.cell(r, 1).value
    if a is None:
        continue
    s = str(a).strip()
    up = s.upper()
    if up == "GASTOS DIRECTOS":       # banner exacto (no el título A1)
        seccion = "Directo"; continue
    if up == "GASTOS INDIRECTOS":
        seccion = "Indirecto"; continue
    if up.startswith(("TOTAL", "RATIO", "ESTADO", "CONCEPTO", "CARGAR", "GASTOS ")):
        continue
    if seccion is None:               # filas de título/instrucción antes del 1er banner
        continue
    concepto = s
    if (concepto, seccion) not in conceptos_orden:
        conceptos_orden.append((concepto, seccion))
    for c, fecha in mesescol.items():
        v = g.cell(r, c).value
        if isinstance(v, (int, float)) and v != 0:
            recs.append((fecha, seccion, concepto, float(v)))

recs.sort(key=lambda x: (x[0], x[1], x[2]))
print(f"Conceptos: {len(conceptos_orden)} | registros a migrar: {len(recs)}")
for con, sec in conceptos_orden:
    n = sum(1 for x in recs if x[2] == con)
    print(f"   [{sec[:3]}] {con}: {n} meses")

# ---- 2) lista de conceptos en _Listas col I + named range ----
L = wb["_Listas"]
COL = 9  # col I
L.cell(1, COL, "CONCEPTOS GASTOS DIR/IND").font = Font(bold=True)
L.cell(1, COL + 1, "Tipo").font = Font(bold=True)
for i, (con, sec) in enumerate(conceptos_orden):
    L.cell(2 + i, COL, con)
    L.cell(2 + i, COL + 1, sec)
last = 1 + len(conceptos_orden)
ref = "_Listas!$%s$2:$%s$%d" % (gl(COL), gl(COL), last)
if "GASTOS_CONCEPTO" in wb.defined_names:
    del wb.defined_names["GASTOS_CONCEPTO"]
wb.defined_names["GASTOS_CONCEPTO"] = DefinedName("GASTOS_CONCEPTO", attr_text=ref)

# ---- 3) recrear la hoja DESDE CERO (evita columnas/estilos fantasma A1:AA que rompen el COM Open) ----
idx = wb.sheetnames.index("2_Gastos_DirInd")
del wb["2_Gastos_DirInd"]
g = wb.create_sheet("2_Gastos_DirInd", idx)

# ---- 4) tabla plana ----
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill("solid", fgColor="FF1F4E78")
for j, h in enumerate(["Fecha", "Tipo", "Concepto", "Monto"], start=1):
    c = g.cell(1, j, h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
for i, (fecha, tipo, concepto, monto) in enumerate(recs):
    r = 2 + i
    g.cell(r, 1, fecha).number_format = "dd/mm/yyyy"
    g.cell(r, 2, tipo)
    g.cell(r, 3, concepto)
    g.cell(r, 4, monto).number_format = "\\$#,##0"
for col, w in {"A": 12, "B": 11, "C": 30, "D": 15}.items():
    g.column_dimensions[col].width = w
g.freeze_panes = "A2"

# ---- 5) dropdowns ----
end = max(len(recs) + 1, 500)
dv_tipo = DataValidation(type="list", formula1='"Directo,Indirecto"', allow_blank=True)
g.add_data_validation(dv_tipo); dv_tipo.add("B2:B%d" % end)
dv_con = DataValidation(type="list", formula1="GASTOS_CONCEPTO", allow_blank=True)
g.add_data_validation(dv_con); dv_con.add("C2:C%d" % end)

# verificación value-preserving
total = sum(x[3] for x in recs)
print(f"Σ Monto migrado = {total:,.2f}")
wb.save(PATH)
print("✓ Guardado:", PATH)
