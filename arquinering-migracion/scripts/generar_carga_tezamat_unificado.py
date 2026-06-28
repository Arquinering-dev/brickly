"""generar_carga_tezamat_unificado.py — Deliverable único para Arquinering.
Unifica el catálogo de IDs (subcontratos + ingresos, todas las obras) con las
propuestas POR MOVIMIENTO del mayor de Tezamat (desc original → observación propuesta).
Reemplaza a `Carga_Tezamat_IDs_GDR_CH.xlsx` y `GDR_Propuesta_Tag_Subcontratos_*.xlsx`.

Salida: docs/Carga_Tezamat_IDs.xlsx
Hojas: Instrucciones · Subcontratos (catálogo) · Ingresos (catálogo) ·
       Mov GDR (propuesta) · Mov SVD (propuesta) · Pendientes
Terminología: "Cliente" (no siempre es Fideicomiso).
"""
import sys
import copy
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

VIEJO = "docs/Carga_Tezamat_IDs_GDR_CH.xlsx"
GDRPROP = "archivos/output/GDR_Propuesta_Tag_Subcontratos_2_Movimientos.xlsx"
SVD = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
OUT = "docs/Carga_Tezamat_IDs.xlsx"

HF = Font(bold=True, color="FFFFFFFF")
HFILL = PatternFill("solid", fgColor="FF1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")

# subcontratistas SVD: ID -> substring para matchear en proveedor/obs
SVD_SC = {
    "SVD-SC-001": "CELSI VIAL", "SVD-SC-002": "SERENO", "SVD-SC-003": "SILVIO AGUIRRE",
    "SVD-SC-004": "PASQUARIELLO", "SVD-SC-005": "FERNANDO AMARILLA", "SVD-SC-006": "CESPEDES",
    "SVD-SC-007": "MARCELO FABIAN SILVA", "SVD-SC-008": "EZEQUIEL MENDOZA",
    "SVD-SC-009": "SCHULZE", "SVD-SC-010": "MANSILLA DERQUI",
}


def hdr_row(ws, headers, row=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row, j, h); c.font = HF; c.fill = HFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def autoancho(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = openpyxl.load_workbook(VIEJO)

    # ---- 1) Instrucciones: Fideicomiso -> Cliente, título 3 obras ----------
    ins = wb["Instrucciones"]
    for r in range(1, ins.max_row + 1):
        v = ins.cell(r, 1).value
        if isinstance(v, str) and "Fideicomiso" in v:
            ins.cell(r, 1).value = v.replace("Fideicomiso", "Cliente")
    ins["A1"] = "Carga de IDs en Tezamat — Arquinering (obras GDR, CH y SVD)"
    # nota sobre las hojas de propuesta por movimiento
    last = ins.max_row + 2
    ins.cell(last, 1, "HOJAS 'Mov <obra> (propuesta)': el extracto del mayor de Tezamat, fila por fila, "
                      "con la observación ACTUAL y la PROPUESTA. Copiá la columna 'OBSERVACIÓN A ESCRIBIR' "
                      "al campo Observaciones del movimiento correspondiente.")

    # ---- 2) renombrar hoja subcontratos a 'catálogo' ----------------------
    if "Subcontratos (egresos)" in wb.sheetnames:
        wb["Subcontratos (egresos)"].title = "Subcontratos (catálogo)"

    # ---- 3) Ingresos: Fideicomiso->Cliente + append SVD -------------------
    ig = wb["Ingresos (cobros)"]
    ig.title = "Ingresos (catálogo)"
    for r in (1, 2):
        v = ig.cell(r, 1).value
        if isinstance(v, str):
            ig.cell(r, 1).value = v.replace("Fideicomiso", "Cliente")
    # última fila de datos
    last_ig = ig.max_row
    while last_ig > 4 and ig.cell(last_ig, 2).value is None:
        last_ig -= 1
    style_ig = last_ig
    # SVD: leer Cert_Calculo (B=ID Cert+Fact, C=Tipo)
    cal = openpyxl.load_workbook(SVD, data_only=True)["Cert_Calculo"]
    r = last_ig + 1
    n_svd_ing = 0
    for cr in range(2, cal.max_row + 1):
        idf = cal.cell(cr, 2).value      # SVD-OC01-C01-B
        tipo = cal.cell(cr, 3).value     # Blanco / Negro
        if not idf:
            continue
        clase = "Anticipo" if "ANT" in str(idf) else "Avance"
        parte = "Blanco (c/factura)" if tipo == "Blanco" else "Negro (s/factura)"
        ej = f"{idf} | BASE | cobro {clase.lower()} (separar CAC en fila aparte)"
        vals = ["SVD", idf, parte, clase, ej]
        for c, v in enumerate(vals, start=1):
            cell = ig.cell(r, c, v)
            if ig.cell(style_ig, c).has_style:
                cell.font = copy.copy(ig.cell(style_ig, c).font)
                cell.alignment = copy.copy(ig.cell(style_ig, c).alignment)
        r += 1
        n_svd_ing += 1

    # ---- 4) Mov GDR (propuesta): portar de GDR_Propuesta ------------------
    gp = openpyxl.load_workbook(GDRPROP)["Propuesta_Subcontratos"]
    mg = wb.create_sheet("Mov GDR (propuesta)")
    HEAD = ["Fila 2_Mov", "Fecha", "Cuenta", "Desc Cuenta", "Debe (ARS)",
            "Observación ACTUAL (Tezamat)", "ID propuesto", "Proveedor", "TIPO",
            "OBSERVACIÓN A ESCRIBIR", "Confianza", "Nota / motivo"]
    hdr_row(mg, HEAD)
    out_r = 2
    for gr in range(4, gp.max_row + 1):
        if gp.cell(gr, 1).value is None:
            continue
        for c in range(1, 13):
            mg.cell(out_r, c, gp.cell(gr, c).value)
        out_r += 1
    autoancho(mg, [9, 11, 9, 18, 13, 30, 13, 20, 7, 34, 9, 40])
    mg.freeze_panes = "A2"

    # ---- 5) Mov SVD (propuesta): generar de 2_Movimientos -----------------
    mov = openpyxl.load_workbook(SVD, data_only=True)["2_Movimientos"]
    ms = wb.create_sheet("Mov SVD (propuesta)")
    hdr_row(ms, HEAD)
    out_r = 2
    n_sc = n_ing = 0
    for mr in range(2, mov.max_row + 1):
        cta = mov.cell(mr, 1).value
        if cta is None:
            continue
        desc_cta = mov.cell(mr, 2).value
        obs = mov.cell(mr, 5).value
        prov = mov.cell(mr, 6).value
        debe = mov.cell(mr, 9).value or 0
        fecha = mov.cell(mr, 3).value
        txt = (str(prov or "") + " " + str(obs or "")).upper()
        es_ingreso = str(cta).startswith("4")
        # subcontrato?
        sc_match = next((sid for sid, key in SVD_SC.items() if key in txt), None)
        if sc_match:
            if "CAC" in txt:
                tipo = "CAC"
            elif "CARGA" in txt or " CS" in txt:
                tipo = "CS"
            elif "ANTIC" in txt:
                tipo = "ANT"
            else:
                tipo = "BASE"
            row = [mr, fecha, cta, desc_cta, debe, obs, sc_match, prov, tipo,
                   f"{sc_match} | {tipo} | {str(obs or '')[:40]}", "Alta",
                   "Match por nombre de proveedor/obs."]
            for c, v in enumerate(row, start=1):
                ms.cell(out_r, c, v)
            out_r += 1; n_sc += 1
        elif es_ingreso and debe == 0:   # ingreso (cobro del cliente, va en Haber)
            haber = mov.cell(mr, 10).value or 0
            row = [mr, fecha, cta, desc_cta, haber, obs, "(asignar OC/cert)", prov, "BASE",
                   "SVD-OCxx-Cyy-B/N | BASE | (identificar qué certificación paga este cobro)",
                   "Manual", "Cobro del Cliente — asignar a la certificación (ver hoja Ingresos catálogo)."]
            for c, v in enumerate(row, start=1):
                ms.cell(out_r, c, v)
            out_r += 1; n_ing += 1
    autoancho(ms, [9, 11, 9, 18, 13, 30, 16, 22, 7, 40, 9, 42])
    ms.freeze_panes = "A2"

    # ---- 6) reordenar hojas + guardar -------------------------------------
    orden = ["Instrucciones", "Subcontratos (catálogo)", "Ingresos (catálogo)",
             "Mov GDR (propuesta)", "Mov SVD (propuesta)", "Pendientes"]
    wb._sheets.sort(key=lambda s: orden.index(s.title) if s.title in orden else 99)

    wb.save(OUT)
    print(f"✓ Deliverable unificado → {OUT}")
    print(f"  · Instrucciones: terminología 'Cliente'; nota de hojas de propuesta")
    print(f"  · Subcontratos (catálogo): GDR+CH+SVD")
    print(f"  · Ingresos (catálogo): +{n_svd_ing} filas SVD (cliente)")
    print(f"  · Mov GDR (propuesta): {out_r and ''}portada de GDR_Propuesta")
    print(f"  · Mov SVD (propuesta): {n_sc} pagos a subcontratos + {n_ing} cobros del cliente")
    print(f"  · hojas: {wb.sheetnames}")

    # limpiar archivos superados
    for f in (VIEJO, GDRPROP):
        try:
            os.remove(f)
            print(f"  · eliminado (superado): {f}")
        except OSError:
            pass


if __name__ == "__main__":
    main()
