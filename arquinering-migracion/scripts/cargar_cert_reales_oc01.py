"""
Carga de datos REALES de certificación OC01 (documento Cert #1 y #2) y vaciado de lo inventado.
Entrada: v8_9 -> Salida: v8_10 (v8_9 backup).
 - Cert_App_Output: vaciar todo; cargar avance real C01 (14 tareas) + C02 (18 tareas), key PTO 01.
 - Cert_Calculo: conservar C01/C02/ANT(OC01) (MEP->1276 del doc); vaciar C03/C04 + todo OC02.
 - Cert_Cabecera: conservar madres reales (ANT/C01/C02 OC01); vaciar las 5 inventadas.
 - Cert_Facturacion: vaciar cobros (no hay datos reales). Estructura intacta.
 - Fiscal Blanco/Negro/IVA quedan en default OC (provisional); cobros y split efectivo -> pendientes.
"""
import datetime
from copy import copy
import openpyxl
from openpyxl.utils import column_index_from_string as ci

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_9.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_10.xlsx"
DOC = "archivos/fuente/CH 2171_Presupuesto 01_Cert. #2.xlsx"

wb = openpyxl.load_workbook(SRC)
doc = openpyxl.load_workbook(DOC, data_only=True)

# ---- leer avance real del documento ----
def leer(sheet):
    ws = doc[sheet]; out = {}
    for r in range(8, 59):
        a = ws.cell(r, 1).value
        if a is None or '.' not in str(a):
            continue
        try:
            cod = float(a)
        except (TypeError, ValueError):
            continue
        out[cod] = {'ant': ws.cell(r, ci('J')).value or 0, 'act': ws.cell(r, ci('K')).value or 0}
    return out
c1, c2 = leer('Cert. 1'), leer('Cert. 2')
# filas a cargar: (ID Certif, fecha, codigo, %ant, %act) solo donde %act>0
FECHA = {'CH-OC01-C01': datetime.datetime(2026, 4, 30), 'CH-OC01-C02': datetime.datetime(2026, 6, 2)}
filas = []
for cod in sorted(c1):
    if c1[cod]['act'] > 0:
        filas.append(('CH-OC01-C01', FECHA['CH-OC01-C01'], cod, 0, c1[cod]['act']))
for cod in sorted(c2):
    if c2[cod]['act'] > 0:
        filas.append(('CH-OC01-C02', FECHA['CH-OC01-C02'], cod, c2[cod]['ant'], c2[cod]['act']))
print("Filas reales a cargar:", len(filas), "(C01 %d / C02 %d)" %
      (sum(1 for f in filas if f[0].endswith('C01')), sum(1 for f in filas if f[0].endswith('C02'))))

# ============ Cert_App_Output: vaciar y recargar ============
ao = wb['Cert_App_Output']
NC = ao.max_column
# templates de estilo desde la fila 2 (preserva azul/verde/formatos)
tpl = {c: (copy(ao.cell(2, c).font), ao.cell(2, c).number_format) for c in range(1, NC + 1)}
for r in range(2, ao.max_row + 1):
    for c in range(1, NC + 1):
        ao.cell(r, c).value = None
F = {  # formulas por columna (con {r})
    'H': '=F{r}+G{r}',
    'I': "=SUMIFS('1_Presupuesto'!$X:$X,'1_Presupuesto'!$AQ:$AQ,K{r},'1_Presupuesto'!$F:$F,E{r})",
    'J': '=G{r}*I{r}',
    'K': '=_xlfn.XLOOKUP(B{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$L:$L)',
    'L': '=IF($E{r}="","",IF(AND(ISNUMBER($I{r}),$H{r}<=1.0001),"✓","⚠"))',
    'N': '=IF($C{r}="","",IF(AND(COUNTIF($C$2:$C{r},$C{r})=1,COUNTIF(Cert_Calculo!$A:$A,$C{r})=0),MAX($N$1:N{rm})+1,""))',
}
for i, (idc, fecha, cod, ant, act) in enumerate(filas):
    r = 2 + i
    vals = {'A': 'CH 2171', 'B': 'CH-OC01', 'C': idc, 'D': fecha, 'E': cod, 'F': ant, 'G': act}
    for col, v in vals.items():
        ao.cell(r, ci(col), v)
    for col, f in F.items():
        ao.cell(r, ci(col), f.format(r=r, rm=r - 1))
    for c in range(1, NC + 1):   # reaplicar estilo
        ao.cell(r, c).font = copy(tpl[c][0])
        ao.cell(r, c).number_format = tpl[c][1]

# ============ Cert_Calculo: conservar reales, vaciar inventadas ============
cc = wb['Cert_Calculo']
keep = {'CH-OC01-C01', 'CH-OC01-C02', 'CH-OC01-ANT'}
for r in range(2, cc.max_row + 1):
    a = cc.cell(r, 1).value
    if a in keep:
        cc.cell(r, ci('U'), 1276)          # MEP real del documento (DÓLAR MEP)
    elif a is not None:
        for c in range(1, cc.max_column + 1):
            cc.cell(r, c).value = None      # vaciar fila inventada (C03/C04/OC02)

# ============ Cert_Cabecera: conservar 3 madres reales ============
cab = wb['Cert_Cabecera']
for r in range(2, cab.max_row + 1):
    a = cab.cell(r, 1).value
    if a is not None and a not in keep:
        for c in range(1, cab.max_column + 1):
            cab.cell(r, c).value = None

# ============ Cert_Facturacion: vaciar cobros (sin datos reales) ============
fa = wb['Cert_Facturacion']
for r in range(2, fa.max_row + 1):
    for c in range(1, fa.max_column + 1):
        fa.cell(r, c).value = None

wb.save(DST)
print("Guardado:", DST)
