"""Genera la planilla para Arquinering: cómo cargar los IDs en las Observaciones de Tezamat
(subcontratos = egresos, y certificaciones = ingresos/cobros), para GDR y CH, + pendientes.
Salida: docs/Carga_Tezamat_IDs_GDR_CH.xlsx
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

FILES = {"GDR": "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx",
         "CH":  "archivos/output/CH_2171_Resumen_de_Obra_v8_11.xlsx"}
OUT = "docs/Carga_Tezamat_IDs_GDR_CH.xlsx"

HF = Font(bold=True, color="FFFFFFFF", size=11)
HFILL = PatternFill("solid", fgColor="FF1F4E78")
TITLE = Font(bold=True, size=14, color="FF1F3864")
SECT = Font(bold=True, color="FFFFFFFF")
SECTFILL = PatternFill("solid", fgColor="FF2E75B6")
WRAP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(horizontal="center", vertical="center")
THIN = Border(*[Side(style="thin", color="FFCCCCCC")] * 4)
YELL = PatternFill("solid", fgColor="FFFFF2CC")


def hdr(ws, row, headers, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.font = HF
        c.fill = HFILL
        c.alignment = CEN if not widths else Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w


def read_sc(f):
    ws = openpyxl.load_workbook(f, data_only=True)["2_Subcontratos"]
    out = []
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a or str(a).startswith("TOTAL"):
            continue
        out.append(dict(id=a, prov=ws.cell(r, 2).value, rubro=ws.cell(r, 3).value,
                        monto=ws.cell(r, 5).value or 0, ant=ws.cell(r, 7).value or 0,
                        pagado=ws.cell(r, 11).value or 0))  # col K = Total pagado
    return out


def read_certfact(f):
    ws = openpyxl.load_workbook(f, data_only=True)["Cert_Calculo"]
    out = []
    for r in range(2, ws.max_row + 1):
        b = ws.cell(r, 2).value  # ID Cert+Fact
        tipo = ws.cell(r, 3).value  # Blanco/Negro
        if not b:
            continue
        clase = "Anticipo" if "ANT" in str(b) else "Avance"
        out.append(dict(id=b, fiscal=tipo, clase=clase))
    return out


wb = openpyxl.Workbook()

# ============ Hoja 1: Instrucciones ============
ws = wb.active
ws.title = "Instrucciones"
ws.sheet_view.showGridLines = False
ws["A1"] = "Carga de IDs en Tezamat — Arquinering (obras GDR y CH)"
ws["A1"].font = TITLE
lines = [
    "",
    "OBJETIVO: que cada movimiento de Tezamat lleve, en el campo OBSERVACIONES, un código que",
    "permita identificarlo en la web (dashboard) — a qué subcontrato pertenece, o a qué certificación.",
    "",
    "FORMATO de la observación (separador EXACTO  «  | »  = espacio + barra + espacio):",
    "",
    "        {ID}  |  {TIPO}  |  {descripción libre}",
    "",
    "TIPOS posibles:",
    "    BASE  → pago base / certificado de avance        → DESCUENTA saldo del contrato",
    "    ANT   → anticipo                                 → DESCUENTA saldo",
    "    CAC   → ajuste por CAC                            → NO descuenta (suma aparte)",
    "    CS    → cargas sociales                          → NO descuenta",
    "",
    "REGLA: el monto BASE y el ajuste CAC van en REGISTROS SEPARADOS (dos filas en Tezamat),",
    "cada uno con su TIPO. Lo mismo las cargas sociales (CS).",
    "",
    "DÓNDE: hoja 'Subcontratos (egresos)' = pagos a subcontratistas.",
    "       hoja 'Ingresos (cobros)'       = cobros del Fideicomiso, se taguean con el ID de la certificación.",
    "       hoja 'Pendientes'              = puntos a confirmar.",
    "",
    "EJEMPLOS:",
    "   Pago a Cuadrilla Hormigón (GDR):     GDR-SC-005 | BASE | certificado hormigón",
    "   Anticipo a Anclajes Cima (GDR):      GDR-SC-003 | ANT | anticipo micropilotes",
    "   Ajuste CAC de ese pago:              GDR-SC-003 | CAC | cac micropilotes   (fila aparte)",
    "   Cobro Fideicomiso cert#1 Blanco GDR: GDR-OC01-C01-B | BASE | cobro cert 1",
]
for i, t in enumerate(lines, start=2):
    ws.cell(i, 1, t)
ws.column_dimensions["A"].width = 110

# ============ Hoja 2: Subcontratos (egresos) ============
ws = wb.create_sheet("Subcontratos (egresos)")
ws.sheet_view.showGridLines = False
ws["A1"] = "SUBCONTRATOS — código a poner en Observaciones de cada PAGO a subcontratista"
ws["A1"].font = TITLE
H = ["Obra", "ID (código Tezamat)", "Proveedor", "Rubro", "Monto presup. ($)",
     "% Anticipo", "TIPOS a usar", "Ejemplo de Observación", "Estado en la web hoy"]
hdr(ws, 3, H, [8, 18, 22, 22, 16, 10, 16, 40, 24])
r = 4
for obra in ("GDR", "CH"):
    for sc in read_sc(FILES[obra]):
        tipos = "BASE" + (" / ANT" if sc["ant"] else "") + " / CAC / CS"
        ej = f"{sc['id']} | BASE | (descripción)"
        if sc["monto"] == 0:
            estado = "⚠ Contrato con monto $0 — confirmar"
        elif sc["pagado"] > 0:
            estado = f"Ya tagueado — pagado ${sc['pagado']:,.0f}"
        else:
            estado = "Sin taguear (pagado $0) → falta cargar ID"
        vals = [obra, sc["id"], sc["prov"], sc["rubro"], sc["monto"],
                sc["ant"], tipos, ej, estado]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.border = THIN
            c.alignment = WRAP
            if j == 5:
                c.number_format = "$#,##0"
            if j == 6:
                c.number_format = "0%"
            if sc["monto"] == 0 and j == 9:
                c.fill = YELL
        r += 1
ws.freeze_panes = "A4"

# ============ Hoja 3: Ingresos (cobros) ============
ws = wb.create_sheet("Ingresos (cobros)")
ws.sheet_view.showGridLines = False
ws["A1"] = "INGRESOS / COBROS — código a poner en Observaciones de cada COBRO del Fideicomiso"
ws["A1"].font = TITLE
ws["A2"] = ("Cada depósito del Fideicomiso se identifica con la certificación que paga. Usar el ID de abajo + TIPO "
            "(BASE el monto, CAC el ajuste, en filas separadas). -B = parte Blanca (con factura) / -N = parte Negra.")
ws["A2"].alignment = WRAP
H = ["Obra", "ID Cert+Fact (código Tezamat)", "Parte fiscal", "Clase", "Ejemplo de Observación"]
hdr(ws, 4, H, [8, 30, 14, 12, 42])
r = 5
for obra in ("GDR", "CH"):
    for cf in read_certfact(FILES[obra]):
        fiscal = ("Blanco (c/factura)" if str(cf["fiscal"]) == "Blanco"
                  else "Negro (s/factura)" if str(cf["fiscal"]) == "Negro" else cf["fiscal"])
        ej = f"{cf['id']} | BASE | cobro {cf['clase'].lower()}"
        vals = [obra, cf["id"], fiscal, cf["clase"], ej]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.border = THIN
            c.alignment = WRAP
        r += 1
ws.freeze_panes = "A5"

# ============ Hoja 4: Pendientes ============
ws = wb.create_sheet("Pendientes")
ws.sheet_view.showGridLines = False
ws["A1"] = "PENDIENTES / A CONFIRMAR con Arquinering"
ws["A1"].font = TITLE
H = ["Obra", "Tema", "Detalle / Acción"]
hdr(ws, 3, H, [8, 30, 90])
pend = [
    ("GDR", "Subcontrato con monto $0", "GDR-SC-004 (Celsi Vial S.A. — Exc. y Mov. de Suelos) figura con "
     "monto presupuestado $0. Confirmar si es un contrato real (cargarle el monto) o si quedó sin usar."),
    ("GDR", "Ingresos sin taguear", "Los movimientos de ingreso (depósitos del Fideicomiso) están cargados "
     "pero como monto único (lump). Hay que IDENTIFICAR qué depósito paga qué certificación y reemitir cada "
     "uno con el ID de la hoja 'Ingresos (cobros)' + BASE/CAC. Hasta entonces la web no muestra cobrado por cert."),
    ("GDR", "Subcontratos sin taguear", "Ningún pago a subcontratista tiene todavía el ID GDR-SC-NNN en "
     "Observaciones → la web muestra 'pagado $0' en todos. Al taguear, concilian solos."),
    ("GDR", "CAC y Cargas Sociales", "Donde corresponda, separar el ajuste CAC (TIPO CAC) y las cargas sociales "
     "(TIPO CS) en registros aparte del monto base."),
    ("CH", "Ingresos / cobros", "Ídem GDR: los cobros del Fideicomiso (cta 410222) están como depósitos lump. "
     "Taguear cada uno con el ID de cert (CH-OC01-...) para activar la conciliación de ingresos."),
    ("CH", "Subcontrato sobrepagado", "CH-SC-002 (Alejandro Ardiles / Agrimensura) aparece sobrepagado "
     "(−$50.000): pagado $730k vs presup $680k. Confirmar si es error de carga o pago extra pactado."),
    ("CH", "OC02 sin certificaciones", "El presupuesto 02 de CH existe pero sin avance cargado (falta el "
     "documento de certificaciones de OC02). Sus cobros se taguean cuando se cargue."),
]
r = 4
for obra, tema, det in pend:
    for j, v in enumerate([obra, tema, det], 1):
        c = ws.cell(r, j, v)
        c.border = THIN
        c.alignment = WRAP
        if j == 3:
            c.fill = YELL
    r += 1
ws.freeze_panes = "A4"

wb.save(OUT)
print(f"✓ Planilla generada → {OUT}")
print(f"  Hojas: {wb.sheetnames}")
