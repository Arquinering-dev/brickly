"""svd_bloque8_config_cac.py — Migracion SVD 4140 a v8. Bloque 8.
Reescribe 0_Indice_CAC (serie SVD, base sep-2024) y 0_CONFIG (parametros SVD).
Trabaja sobre el archivo de trabajo SVD (clon de GDR ya con 1_Presupuesto SVD).
"""
import sys
import datetime as dt
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
PMAX = 221                       # ultima fila de datos de 1_Presupuesto SVD
VENTA_LEGACY = 1158163185.0      # venta total legacy (Pto.Vta M56+M239) — sanity

# serie CAC SVD (mes, valor INDEC). Base = sep-2024. Previsiones desde mar-2026.
CAC = [
    (dt.datetime(2024, 9, 1), 14366.9, False),
    (dt.datetime(2024, 10, 1), 14614.9, False),
    (dt.datetime(2024, 11, 1), 15043.0, False),
    (dt.datetime(2024, 12, 1), 15356.4, False),
    (dt.datetime(2025, 1, 1), 15592.5, False),
    (dt.datetime(2025, 2, 1), 15852.4, False),
    (dt.datetime(2025, 3, 1), 15995.4, False),
    (dt.datetime(2025, 4, 1), 16253.5, False),
    (dt.datetime(2025, 5, 1), 16521.3, False),
    (dt.datetime(2025, 6, 1), 16643.4, False),
    (dt.datetime(2025, 7, 1), 16939.5, False),
    (dt.datetime(2025, 8, 1), 17187.6, False),
    (dt.datetime(2025, 9, 1), 17762.3, False),
    (dt.datetime(2025, 10, 1), 18172.8, False),
    (dt.datetime(2025, 11, 1), 18534.9, False),
    (dt.datetime(2025, 12, 1), 18776.5, False),
    (dt.datetime(2026, 1, 1), 19209.4, False),
    (dt.datetime(2026, 2, 1), 19453.0, False),
    (dt.datetime(2026, 3, 1), 19806.240593911236, True),
    (dt.datetime(2026, 4, 1), 20138.793405782737, True),
    (dt.datetime(2026, 5, 1), 20471.34621765424, True),
    (dt.datetime(2026, 6, 1), 20471.34621765424, True),
]


def build():
    wb = openpyxl.load_workbook(DST)

    # --- 0_Indice_CAC --------------------------------------------------------
    ic = wb["0_Indice_CAC"]
    for r in range(4, ic.max_row + 1):           # limpiar datos viejos (GDR)
        for c in (1, 2, 3, 4):
            ic.cell(row=r, column=c).value = None
    ic["A1"] = "INDICE CAC"
    ic["A3"] = "Escala de tiempo"; ic["B3"] = "Valores"; ic["C3"] = "Previsión"; ic["D3"] = "Indice"
    r0 = 4
    for i, (mes, val, prev) in enumerate(CAC):
        r = r0 + i
        ic.cell(row=r, column=1).value = mes
        ic.cell(row=r, column=2).value = val
        if prev:
            ic.cell(row=r, column=3).value = val
        ic.cell(row=r, column=4).value = f"=$B${r0}/B{r}"   # ratio base/corriente
    rlast = r0 + len(CAC) - 1

    # --- 0_CONFIG ------------------------------------------------------------
    cf = wb["0_CONFIG"]
    cf["A1"] = "▌ CONFIGURACIÓN DE OBRA — SVD 4140"
    cf["B3"] = "Edificio El Salvador 4140"
    cf["B4"] = "En ejecución"
    cf["B5"] = dt.datetime(2025, 1, 1); cf["C5"] = "presupuesto base sep-2024"
    cf["B8"] = dt.datetime(2024, 9, 1); cf["C8"] = "fecha base del presupuesto (sep-2024)"
    cf["B9"] = 14366.9
    cf["B10"] = "=INDEX('0_Indice_CAC'!$A$4:$A$33,MATCH(1E+300,'0_Indice_CAC'!$B$4:$B$33))"
    cf["B11"] = "=MIN('0_Indice_CAC'!$D$4:$D$33)"
    cf["B13"] = "='1_GGBB'!F67"
    cf["B15"] = "B31 / N69 / GGN"
    cf["C15"] = "Blanco 31% c/IVA · Negro 69% — DERIVADO de Facturación, revisar"
    cf["B16"] = (f"=SUMPRODUCT(('1_Presupuesto'!$K$5:$K${PMAX}+'1_Presupuesto'!$L$5:$L${PMAX}"
                 f"+'1_Presupuesto'!$M$5:$M${PMAX})*'1_Presupuesto'!$J$5:$J${PMAX})")
    cf["B17"] = f"=SUMPRODUCT('1_Presupuesto'!$O$5:$O${PMAX}*'1_Presupuesto'!$J$5:$J${PMAX})"
    cf["B18"] = f"=SUMPRODUCT('1_Presupuesto'!$P$5:$P${PMAX}*'1_Presupuesto'!$J$5:$J${PMAX})"
    # sanity: la venta es input no-uniforme -> el check es venta total == legacy
    cf["A19"] = "Sanity check venta (debe ser 0)"
    cf["B19"] = f"=+B18-{VENTA_LEGACY:.0f}"
    cf["C19"] = "venta total = subtotales legacy Pto.Vta (markup no uniforme; K solo informativo)"

    wb.save(DST)
    print(f"✓ Bloque 8 → {DST}")
    print(f"  · 0_Indice_CAC: {len(CAC)} meses (sep-2024..jun-2026), base sep-2024=14366.9, filas 4..{rlast}")
    print(f"  · 0_CONFIG: SVD 4140, mes base sep-2024, K='1_GGBB'!F67, fiscal B31/N69 (revisar)")
    print(f"  · B16/B17/B18 SUMPRODUCT sobre 1_Presupuesto $5:${PMAX}; sanity B19=venta-{VENTA_LEGACY:,.0f}")


if __name__ == "__main__":
    build()
