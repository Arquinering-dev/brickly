"""
Reestructura 2_Gastos_DirInd de matriz (concepto × mes) a tabla plana tipo 2_Movimientos.
Entrada v8_10 -> salida v8_11 (v8_10 backup). Columnas: Fecha | Tipo | Concepto | Monto.
- Migra value-preserving cada celda de la matriz (sólo columnas de fecha; excluye Total Real/Descontado).
- Dropdowns: Tipo (Directo/Indirecto, inline) y Concepto (named range GASTOS_CONCEPTO en _Listas).
- La deflactación CAC, el agrupado por mes/tipo y los totales los deriva el dashboard (no se almacenan).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter as gl

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_10.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_11.xlsx"

wb = openpyxl.load_workbook(SRC)
g = wb["2_Gastos_DirInd"]

# ---- 1) leer la matriz vieja ----
mesescol = {}  # col -> fecha (sólo columnas cuyo header de fila 3 es una FECHA)
for c in range(2, g.max_column + 1):
    v = g.cell(3, c).value
    if hasattr(v, "year") and hasattr(v, "month"):
        mesescol[c] = v

conceptos_orden = []      # para la lista del dropdown (en orden de aparición)
seccion = None
recs = []                 # (fecha, tipo, concepto, monto)
for r in range(1, g.max_row + 1):
    a = g.cell(r, 1).value
    if a is None:
        continue
    s = str(a).strip()
    up = s.upper()
    if "▌" in s or up.startswith("GASTOS "):     # banner de sección
        seccion = "Indirecto" if "INDIRECTO" in up else ("Directo" if "DIRECTO" in up else seccion)
        continue
    if up.startswith("TOTAL") or up.startswith("RATIO") or up.startswith("ESTADO") or up.startswith("CONCEPTO"):
        continue
    if seccion is None:
        continue
    concepto = s
    if concepto not in conceptos_orden:
        conceptos_orden.append((concepto, seccion))
    for c, fecha in mesescol.items():
        v = g.cell(r, c).value
        if isinstance(v, (int, float)) and v != 0:
            recs.append((fecha, seccion, concepto, float(v)))

recs.sort(key=lambda x: (x[0], x[1], x[2]))
print("Conceptos:", len(conceptos_orden), "| registros a migrar:", len(recs))

# ---- 2) lista de conceptos en _Listas + named range ----
L = wb["_Listas"]
COL = 9  # columna I (libre)
L.cell(1, COL, "CONCEPTOS GASTOS DIR/IND").font = Font(bold=True)
L.cell(1, COL + 1, "Tipo").font = Font(bold=True)
for i, (con, sec) in enumerate(conceptos_orden):
    L.cell(2 + i, COL, con)
    L.cell(2 + i, COL + 1, sec)
last = 1 + len(conceptos_orden)
ref = "_Listas!$%s$2:$%s$%d" % (gl(COL), gl(COL), last)
if "GASTOS_CONCEPTO" in wb.defined_names:
    del wb.defined_names["GASTOS_CONCEPTO"]
wb.defined_names["GASTOS_CONCEPTO"] = DefinedName("GASTOS_CONCEPTO", attr_text=ref)

# ---- 3) vaciar la hoja vieja (valores, merges, DV) ----
for mc in list(g.merged_cells.ranges):
    g.unmerge_cells(str(mc))
for r in range(1, g.max_row + 1):
    for c in range(1, g.max_column + 1):
        g.cell(r, c).value = None
g.data_validations.dataValidation = []

# ---- 4) escribir la tabla plana ----
HDR_FONT = Font(bold=True, color="FFFFFFFF")
HDR_FILL = PatternFill("solid", fgColor="FF1F4E78")
headers = ["Fecha", "Tipo", "Concepto", "Monto"]
for j, h in enumerate(headers, start=1):
    c = g.cell(1, j, h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
for i, (fecha, tipo, concepto, monto) in enumerate(recs):
    r = 2 + i
    g.cell(r, 1, fecha).number_format = "dd/mm/yyyy"
    g.cell(r, 2, tipo)
    g.cell(r, 3, concepto)
    g.cell(r, 4, monto).number_format = "\\$#,##0"
for col, w in {"A": 12, "B": 11, "C": 30, "D": 15}.items():
    g.column_dimensions[col].width = w
g.freeze_panes = "A2"

# ---- 5) dropdowns ----
nrow = len(recs)
end = max(nrow + 1, 500)
dv_tipo = DataValidation(type="list", formula1='"Directo,Indirecto"', allow_blank=True)
g.add_data_validation(dv_tipo); dv_tipo.add("B2:B%d" % end)
dv_con = DataValidation(type="list", formula1="GASTOS_CONCEPTO", allow_blank=True)
g.add_data_validation(dv_con); dv_con.add("C2:C%d" % end)

wb.save(DST)
print("Guardado:", DST)
