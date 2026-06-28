"""svd_bloque6b_subc.py — Migracion SVD 4140 a v8. Bloque 6b.
2_Subcontratos: matriz desde la hoja Contratistas legacy (10 subcontratistas).
  H-N = formulas SUMIFS (replican GDR; concilian solas cuando se tagueen los pagos
  en 2_Movimientos cols Q/R — pendiente, como GDR).
2_Quincenas: placeholder (limpia data GDR; se carga de las hojas M.O. mas adelante —
  area de maestro cross-obra de horas, pendiente).
"""
import sys
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
LEGACY = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"

# rubro legacy de Contratistas -> rubro plan (best-effort, marcar en pendientes)
RUBRO_SC = {
    "Movimiento de Suelos": "Movimiento de Suelos",
    "ALBAMO": "Albañilería MO",
    "ELECMO": "Eléctrico MO",
    "GESTORIA": "H. Gestoria",
    "Proyectado/Revoque/Yeso/Carp.Monolitico": "Durlock MO",
    "INSTALACION AIRE ACONDICIONADO": "Termomecánica MO",
    "HORMIGON ALISADO - Revestimiento": "Revestimiento MO",
    "CALEFACCION PISO RADIANTE": "Termomecánica MO",
}


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_contratistas():
    ws = openpyxl.load_workbook(LEGACY, data_only=True)["Contratistas"]
    heads = []
    for r in range(1, ws.max_row + 1):
        A = ws.cell(row=r, column=1).value
        B = ws.cell(row=r, column=2).value
        D = ws.cell(row=r, column=4).value
        if A and B is None and D:
            heads.append(r)
    heads.append(ws.max_row + 1)
    out = []
    for i in range(len(heads) - 1):
        r0, r1 = heads[i], heads[i + 1]
        name = str(ws.cell(row=r0, column=1).value).strip()
        rubro_leg = str(ws.cell(row=r0, column=4).value).strip()
        presup = 0.0
        cac = 0.0
        for r in range(r0 + 1, r1):
            B = ws.cell(row=r, column=2).value
            E = _num(ws.cell(row=r, column=5).value)
            D = _num(ws.cell(row=r, column=4).value)
            if B and "presup" in str(B).lower() and E > 0:
                presup = E
            elif E > presup and not (B and "sum" in str(B).lower()):
                presup = max(presup, E)
            cac += D
        out.append({"name": name, "rubro_leg": rubro_leg, "presup": presup, "cac": cac})
    return out


def build():
    wb = openpyxl.load_workbook(DST)
    contratos = parse_contratistas()

    sc = wb["2_Subcontratos"]
    # limpiar filas 4+ (data GDR)
    for r in range(4, sc.max_row + 1):
        for c in range(1, 15):
            sc.cell(row=r, column=c).value = None

    for i, ct in enumerate(contratos):
        r = 4 + i
        cid = f"SVD-SC-{i+1:03d}"
        rubro = RUBRO_SC.get(ct["rubro_leg"], ct["rubro_leg"])
        sc.cell(row=r, column=1).value = cid
        sc.cell(row=r, column=2).value = ct["name"]
        sc.cell(row=r, column=3).value = rubro
        sc.cell(row=r, column=4).value = ct["rubro_leg"]          # Descripción = rubro legacy
        sc.cell(row=r, column=5).value = ct["presup"]
        sc.cell(row=r, column=6).value = "SI" if ct["cac"] > 0 else "NO"
        sc.cell(row=r, column=7).value = 0                        # % anticipo (default, revisar)
        sc.cell(row=r, column=8).value = (
            f"=SUMIFS('2_Movimientos'!$I:$I,'2_Movimientos'!$Q:$Q,$A{r},'2_Movimientos'!$R:$R,\"BASE\")"
            f"+SUMIFS('2_Movimientos'!$I:$I,'2_Movimientos'!$Q:$Q,$A{r},'2_Movimientos'!$R:$R,\"ANT\")")
        sc.cell(row=r, column=9).value = (
            f"=SUMIFS('2_Movimientos'!$I:$I,'2_Movimientos'!$Q:$Q,$A{r},'2_Movimientos'!$R:$R,\"CAC\")")
        sc.cell(row=r, column=10).value = (
            f"=SUMIFS('2_Movimientos'!$I:$I,'2_Movimientos'!$Q:$Q,$A{r},'2_Movimientos'!$R:$R,\"CS\")")
        sc.cell(row=r, column=11).value = f"=H{r}+I{r}+J{r}"
        sc.cell(row=r, column=12).value = f"=$E{r}-H{r}"
        sc.cell(row=r, column=13).value = f"=IFERROR(H{r}/$E{r},0)"
        sc.cell(row=r, column=14).value = (
            f'=IF(L{r}<=0,"🔴 Sin saldo",IF(M{r}>=0.9,"🟠 <10% saldo","🟢 OK"))')

    # 2_Quincenas -> placeholder
    q = wb["2_Quincenas"]
    n_q = 0
    for r in range(4, q.max_row + 1):
        for c in range(1, q.max_column + 1):
            if q.cell(row=r, column=c).value is not None:
                q.cell(row=r, column=c).value = None
                n_q += 1

    wb.save(DST)
    print(f"✓ Bloque 6b → {DST}")
    print(f"  · 2_Subcontratos: {len(contratos)} contratos SVD-SC-NNN (filas 4..{3+len(contratos)})")
    for i, ct in enumerate(contratos):
        rb = RUBRO_SC.get(ct["rubro_leg"], ct["rubro_leg"])
        print(f"     SVD-SC-{i+1:03d} {ct['name'][:22]:23} {rb[:18]:19} presup={ct['presup']:>14,.0f} CAC={'SI' if ct['cac']>0 else 'NO'}")
    print(f"  · 2_Quincenas: placeholder ({n_q} celdas GDR limpiadas) — cargar de M.O. (pendiente)")


if __name__ == "__main__":
    build()
