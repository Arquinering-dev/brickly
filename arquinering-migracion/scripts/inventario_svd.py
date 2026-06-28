"""Inventario SVD 4140 — bloque 1 de la migracion de SVD a v8.
Escanea los 4 insumos: resumen legacy, mayores Tezamat, y los 2 docs de certificacion.
Compara el resumen legacy contra el modelo (CH v8_11 / GDR v8_12). Solo lectura."""
import openpyxl
from openpyxl.utils import get_column_letter

F = "archivos/fuente/"
RESUMEN = F + "SVD 4140 - Resumen de Obra.xlsx"
MAYORES = F + "SVD 4140 Mayores 2026-06-01.xlsx"
CERT01  = F + "SVD 4140_Pto. 01_Cert.12.xlsx"
CERT02  = F + "SVD 4140_Pto. 02_Cert. 08.xlsx"

CH  = "archivos/output/CH_2171_Resumen_de_Obra_v8_11.xlsx"
GDR = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"


def first_text_cells(ws, max_rows=4, max_cols=14):
    out = []
    for r in range(1, max_rows + 1):
        row = []
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row.append(f"{get_column_letter(c)}{r}={repr(v)[:38]}")
        if row:
            out.append("  " + " | ".join(row))
    return out


def inventory(path, label, show_rows=4):
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    except Exception as e:
        print(f"\n!! No pude abrir {label}: {e}")
        return set()
    print(f"\n{'='*80}\n{label}: {path}\n{'='*80}")
    print(f"Total hojas: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        tab = ws.sheet_properties.tabColor
        tabc = tab.rgb if tab else "-"
        print(f"\n[{name}]  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}  tab={tabc}")
        for line in first_text_cells(ws, max_rows=show_rows):
            print(line)
    return set(wb.sheetnames)


print("\n" + "#"*80 + "\n# 1. RESUMEN LEGACY SVD\n" + "#"*80)
svd_sheets = inventory(RESUMEN, "SVD 4140 Resumen (LEGACY a migrar)")

print("\n" + "#"*80 + "\n# 2. MAYORES TEZAMAT (para 2_Movimientos)\n" + "#"*80)
inventory(MAYORES, "SVD 4140 Mayores Tezamat", show_rows=8)

print("\n" + "#"*80 + "\n# 3/4. DOCS DE CERTIFICACION\n" + "#"*80)
inventory(CERT01, "SVD Pto 01 Cert 12", show_rows=6)
inventory(CERT02, "SVD Pto 02 Cert 08", show_rows=6)

print("\n" + "#"*80 + "\n# COMPARACION DE HOJAS: SVD legacy vs modelos\n" + "#"*80)
ch_sheets  = set(openpyxl.load_workbook(CH, read_only=True).sheetnames)
gdr_sheets = set(openpyxl.load_workbook(GDR, read_only=True).sheetnames)

print("\n-- Hojas del MODELO CH v8_11 que NO estan en SVD legacy (hay que crear/adaptar):")
for s in sorted(ch_sheets - svd_sheets):
    print(f"   + {s}")
print("\n-- Hojas en SVD legacy que NO estan en CH (analisis/legacy a evaluar):")
for s in sorted(svd_sheets - ch_sheets):
    print(f"   - {s}")
print("\n-- Hojas en ambos (SVD legacy ∩ CH):")
for s in sorted(svd_sheets & ch_sheets):
    print(f"   = {s}")
