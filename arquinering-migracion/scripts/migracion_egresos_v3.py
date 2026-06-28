"""
migracion_egresos_v3.py — Versión egresos (prototipo estado futuro)
Arquinering S.R.L. — CH 2171

Genera CH_2171_Resumen_de_Obra_v8_3_egresos.xlsx a partir de _v8_2.
Aplica las nuevas definiciones de proceso del CIRCUITO DE EGRESOS:
  - Rubros alineados al plan de cuentas de Tezamat (token único "Rubro MT"/"Rubro MO").
  - 3_Control_Ppto cruza por E (Desc Cuenta) de 2_Gastos, no por reclasificación manual.
  - Subcontratos identificados por ID en Observaciones (CH-SC-001 TIPO).
  - CAC como fila separada (no embebido).
  - Maestros 0_Indice_CAC / 0_Jornales_MO documentados para import futuro.
  - Hojas obsoletas eliminadas: 2_Pagos_Subc, 2_Pagos_Quincena_SC, 2_Certificaciones.
El circuito de INGRESOS no se modifica (se neutralizan solo las celdas que
dependían de 2_Certificaciones, que se elimina).

Un solo wb.save(). Luego: excel_recalc.py + recalc.py.
"""
import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_2.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_3_egresos.xlsx"

# ----------------------------------------------------------------------------
# Mapas de conversión base -> token Tezamat
# ----------------------------------------------------------------------------
MT = {
    "Albañilería": "Albañilería MT", "Hormigón": "Hormigón MT",
    "Durlock/Yeso": "Durlock MT", "Electricidad": "Eléctrico MT",
    "Pintura": "Pintura MT", "Revestimiento": "Revestimiento MT",
    "Sanitaria": "Sanitaria MT", "Agrimensura": "Agrimensura MT",
    "Gastos Generales Obra": "Gastos Generales Obra MT",
    "Termomecánica": "Termomecánica MT",
    "Preliminar": "Preliminares",
    "Excavación y Mov. de Suelos": "Movimiento de Suelos",
    "Movilidad": "Mov. Variables",
}
MO = {
    "Albañilería": "Albañilería MO", "Hormigón": "Homigón MO",  # typo igual a Tezamat 53005
    "Durlock/Yeso": "Durlock MO", "Electricidad": "Eléctrico MO",
    "Pintura": "Pintura MO", "Revestimiento": "Revestimiento MO",
    "Sanitaria": "Sanitaria MO", "Agrimensura": "Agrimensura MO",
    "Gastos Generales Obra": "Gastos Generales Obra MO",
    "Termomecánica": "Termomecánica MO",
    "Seguridad e Higiene": "Seguridad e Higiene MO",
    "Supervisión de Obra": "Supervisión de Obra MO",
    "Mantenimiento": "Mantenimiento MO",
    "Preliminar": "Preliminares",
    "Excavación y Mov. de Suelos": "Movimiento de Suelos",
    "Movilidad": "Mov. Variables",
}
QUINCENA = {  # códigos 2_Quincenas col E -> token MO
    "ALBAMO": "Albañilería MO", "HORMIMO": "Homigón MO",
    "ELECTMO": "Eléctrico MO", "HERRERO": "Herrería MO",
}
CONTROL_JORN = {  # 3_Control_Jornales col A (base -> token MO)
    "Albañilería": "Albañilería MO", "Electricidad": "Eléctrico MO",
    "Sanitaria": "Sanitaria MO", "Pintura": "Pintura MO",
    "Durlock/Yeso": "Durlock MO", "Hormigón": "Homigón MO",
    "Revestimiento": "Revestimiento MO", "Mantenimiento": "Mantenimiento MO",
}

# Tokens de 3_Control_Ppto: 17 canónicos + extras (no canónicos) debajo
CANONICOS = [
    "Preliminares", "Movimiento de Suelos",
    "Albañilería MO", "Albañilería MT", "Durlock MO", "Durlock MT",
    "Eléctrico MO", "Eléctrico MT", "Homigón MO", "Hormigón MT",
    "Pintura MO", "Pintura MT", "Revestimiento MO", "Revestimiento MT",
    "Sanitaria MO", "Sanitaria MT", "Mov. Variables",
]
EXTRAS = [
    "Agrimensura MT", "Agrimensura MO",
    "Gastos Generales Obra MT", "Gastos Generales Obra MO",
    "Termomecánica MT", "Termomecánica MO",
    "Seguridad e Higiene MO", "Supervisión de Obra MO",
    "Mantenimiento MO", "Herrería MO",
]
TOKENS = CANONICOS + EXTRAS  # 27 tokens -> filas 4..30


def conv(val, mapa):
    if val in (None, ""):
        return val
    s = str(val).strip()
    if s not in mapa:
        raise ValueError(f"Rubro base no mapeado: {s!r}")
    return mapa[s]


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    log = []

    # ====================================================================
    # 1) 1_Presupuesto: convertir cols A,B (MT) y C,D (MO) a tokens
    # ====================================================================
    ws = wb["1_Presupuesto"]
    for r in range(4, 213):
        for col, mapa in ((1, MT), (2, MT), (3, MO), (4, MO)):
            c = ws.cell(r, col)
            if c.value not in (None, ""):
                c.value = conv(c.value, mapa)
    log.append("1_Presupuesto A/B->MT, C/D->MO tokens")

    # ====================================================================
    # 2) 2_Quincenas: col E códigos -> token MO
    # ====================================================================
    ws = wb["2_Quincenas"]
    for r in range(4, ws.max_row + 1):
        c = ws.cell(r, 5)
        if c.value not in (None, ""):
            c.value = conv(c.value, QUINCENA)
    log.append("2_Quincenas col E -> tokens MO (fix bug de matching)")

    # ====================================================================
    # 3) 2_Gastos: cruce por E. A,B,U,V,W obsoletas (vaciadas, no borradas
    #    físicamente para preservar integridad de letras de columna).
    #    X repunteada a 'rubro reconocido' por E.
    # ====================================================================
    ws = wb["2_Gastos"]
    ws.cell(1, 1).value = "Rubro (OBSOLETA — cruce por E)"
    ws.cell(1, 2).value = "MT/MO (OBSOLETA — cruce por E)"
    ws.cell(1, 21).value = "Contrato SC (OBSOLETA — ID en Observaciones)"
    ws.cell(1, 22).value = "Monto CAC (OBSOLETA — CAC es fila aparte)"
    ws.cell(1, 23).value = "Monto Base (OBSOLETA — base = T en filas no-CAC)"
    ws.cell(1, 24).value = "Rubro reconocido"
    for r in range(2, ws.max_row + 1):
        for col in (1, 2, 21, 22, 23):
            ws.cell(r, col).value = None
        # X solo en filas con dato (E=Desc Cuenta o L=Debe presentes)
        tiene_dato = (ws.cell(r, 5).value not in (None, "")
                      or ws.cell(r, 12).value not in (None, ""))
        ws.cell(r, 24).value = (
            f"=ISNUMBER(MATCH(E{r},'3_Control_Ppto'!$A$4:$A$35,0))"
            if tiene_dato else None
        )
    log.append("2_Gastos A,B,U,V,W vaciadas+obsoletas; X=rubro reconocido por E")

    # ====================================================================
    # 4) 3_Control_Ppto: fórmulas unificadas, filas 4-35 repobladas
    # ====================================================================
    ws = wb["3_Control_Ppto"]
    # limpiar región de rubros (A..Q, filas 4-35)
    for r in range(4, 36):
        for col in range(1, 18):
            ws.cell(r, col).value = None

    def presup(r):
        return (
            f"=SUMPRODUCT(('1_Presupuesto'!$A$4:$A$212=$A{r})*'1_Presupuesto'!$K$4:$K$212*'1_Presupuesto'!$J$4:$J$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$B$4:$B$212=$A{r})*'1_Presupuesto'!$K$4:$K$212*'1_Presupuesto'!$J$4:$J$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$C$4:$C$212=$A{r})*'1_Presupuesto'!$L$4:$L$212*'1_Presupuesto'!$J$4:$J$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$D$4:$D$212=$A{r})*'1_Presupuesto'!$M$4:$M$212*'1_Presupuesto'!$J$4:$J$212)"
        )

    def avance(r):
        num = (
            f"SUMPRODUCT(('1_Presupuesto'!$A$4:$A$212=$A{r})*'1_Presupuesto'!$AI$4:$AI$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$C$4:$C$212=$A{r})*'1_Presupuesto'!$AJ$4:$AJ$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$D$4:$D$212=$A{r})*'1_Presupuesto'!$AK$4:$AK$212)"
        )
        den = (
            f"SUMPRODUCT(('1_Presupuesto'!$A$4:$A$212=$A{r})*'1_Presupuesto'!$R$4:$R$212*'1_Presupuesto'!$J$4:$J$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$C$4:$C$212=$A{r})*'1_Presupuesto'!$S$4:$S$212*'1_Presupuesto'!$J$4:$J$212)"
            f"+SUMPRODUCT(('1_Presupuesto'!$D$4:$D$212=$A{r})*'1_Presupuesto'!$T$4:$T$212*'1_Presupuesto'!$J$4:$J$212)"
        )
        return f"=IFERROR(({num})/({den}),0)"

    for i, tok in enumerate(TOKENS):
        r = 4 + i
        ws.cell(r, 1).value = tok
        ws.cell(r, 2).value = (
            f'=IF($A{r}="","",IF(RIGHT($A{r},3)=" MT","MT",IF(RIGHT($A{r},3)=" MO","MO","—")))'
        )
        ws.cell(r, 3).value = presup(r)
        ws.cell(r, 5).value = (
            f"=SUMIFS('2_Gastos'!$S:$S,'2_Gastos'!$E:$E,$A{r})"
            f"+SUMIFS('2_Quincenas'!$O:$O,'2_Quincenas'!$E:$E,$A{r})"
        )
        ws.cell(r, 6).value = (
            f"=SUMIFS('2_Gastos'!$T:$T,'2_Gastos'!$E:$E,$A{r})"
            f"+SUMIFS('2_Quincenas'!$N:$N,'2_Quincenas'!$E:$E,$A{r})"
        )
        ws.cell(r, 8).value = f"=MAX(C{r}-E{r},0)"
        ws.cell(r, 9).value = f"=IFERROR(MIN(MAX(1-E{r}/C{r},0),1),0)"
        ws.cell(r, 11).value = f"=MAX(E{r}-C{r},0)"
        ws.cell(r, 12).value = f"=IFERROR(MAX(E{r}/C{r}-1,0),0)"
        ws.cell(r, 13).value = f'=IF(L{r}>0.1,"🔴",IF(L{r}>0.05,"🟡","🟢"))'
        ws.cell(r, 15).value = avance(r)
        ws.cell(r, 17).value = f"=IFERROR(IF(AND(O{r}=1,L{r}<=0),H{r}/C{r},0),0)"
    log.append(f"3_Control_Ppto: {len(TOKENS)} tokens (fórmulas unificadas por E)")

    # nuevo sanity check fila 102 (fila vacía entre 101 y 103)
    ws.cell(102, 1).value = "Costo 2_Gastos sin rubro reconocido (cuenta fuera de plan)"
    ws.cell(102, 5).value = "=SUMIFS('2_Gastos'!$T:$T,'2_Gastos'!$X:$X,FALSE)"
    ws.cell(102, 13).value = '=IF(E102=0,"🟢","⚠ revisar cuentas Tezamat")'
    log.append("3_Control_Ppto fila 102: sanity 'gasto sin rubro reconocido'")

    # ====================================================================
    # 5) 3_Control_Jornales: col A base -> token MO
    # ====================================================================
    ws = wb["3_Control_Jornales"]
    for r in range(4, ws.max_row + 1):
        c = ws.cell(r, 1)
        v = c.value
        if v in (None, "", "Total"):
            continue
        c.value = conv(v, CONTROL_JORN)
    log.append("3_Control_Jornales col A -> tokens MO")

    # ====================================================================
    # 6) 2_Subcontratos: tracking por ID en Observaciones (CH-SC-001 TIPO)
    # ====================================================================
    ws = wb["2_Subcontratos"]
    ws.cell(3, 15).value = "Cargas Sociales"  # col O nueva (trailing)
    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a in (None, ""):
            continue
        s = str(a).strip()
        if s.upper().startswith("SC-"):
            s = "CH-" + s
            ws.cell(r, 1).value = s
        A = f"$A{r}"
        ws.cell(r, 9).value = (   # I Pagado sin CAC = AVANCE + QUINCENA (base)
            f"=SUMIFS('2_Gastos'!$T:$T,'2_Gastos'!$H:$H,\"*\"&{A}&\" AVANCE*\")"
            f"+SUMIFS('2_Gastos'!$T:$T,'2_Gastos'!$H:$H,\"*\"&{A}&\" QUINCENA*\")"
        )
        ws.cell(r, 10).value = f"=E{r}-I{r}"          # J Saldo Disponible
        ws.cell(r, 11).value = (  # K CAC Pagado
            f"=SUMIFS('2_Gastos'!$T:$T,'2_Gastos'!$H:$H,\"*\"&{A}&\" CAC*\")"
        )
        ws.cell(r, 15).value = (  # O Cargas Sociales
            f"=SUMIFS('2_Gastos'!$T:$T,'2_Gastos'!$H:$H,\"*\"&{A}&\" CARGAS SOCIALES*\")"
        )
        ws.cell(r, 12).value = f"=I{r}+K{r}+O{r}"     # L Total Pagado
    log.append("2_Subcontratos: IDs CH-SC-*, tracking por Observaciones, +Cargas Sociales")

    # ====================================================================
    # 7) Neutralizar celdas de INGRESOS que dependían de 2_Certificaciones
    # ====================================================================
    dash = wb["3_Dashboard"]
    nota_ing = Comment("Ingreso: se reconecta en la versión de ingresos. "
                       "2_Certificaciones eliminada en esta versión de egresos.", "migracion")
    for coord in ("C4", "C5", "C6", "C20"):
        dash[coord].value = 0
        dash[coord].comment = nota_ing
    cf = wb["3_Cash_Flow"]
    for row in (4, 5):  # ingresos cobros / ingresos CAC
        for col in range(2, 20):  # B..S
            cf.cell(row, col).value = 0
    cf.cell(4, 1).comment = Comment("Ingresos neutralizados (2_Certificaciones eliminada). "
                                    "Se reconectan en la versión de ingresos.", "migracion")
    log.append("Ingresos (Dashboard C4/C5/C6/C20, Cash_Flow filas 4-5) neutralizados a 0")

    # ====================================================================
    # 8) Eliminar hojas obsoletas (ya sin referencias)
    # ====================================================================
    for sh in ("2_Pagos_Subc", "2_Pagos_Quincena_SC", "2_Certificaciones"):
        del wb[sh]
    log.append("Eliminadas: 2_Pagos_Subc, 2_Pagos_Quincena_SC, 2_Certificaciones")

    # ====================================================================
    # 9) _Listas: actualizar lista de rubros a tokens nuevos
    # ====================================================================
    ws = wb["_Listas"]
    ws.cell(1, 1).value = "RUBROS CANÓNICOS (Tezamat)"
    for r in range(2, 60):
        ws.cell(r, 1).value = None
    for i, tok in enumerate(TOKENS):
        ws.cell(2 + i, 1).value = tok
    log.append("_Listas: lista de rubros -> tokens Tezamat")

    # ====================================================================
    # 10) Maestros: documentar mecanismo de importación (placeholder)
    # ====================================================================
    doc_cac = ("IMPORTACIÓN MAESTRA (pendiente): este índice se actualiza por "
               "proceso Python desde archivo maestro externo (p.ej. "
               "archivos/fuente/CAC_maestro.xlsx). NO usar vínculos externos vivos. "
               "Una sola actualización en el maestro -> import a cada Resumen.")
    doc_jor = ("IMPORTACIÓN MAESTRA (pendiente): tarifas UOCRA se actualizan por "
               "proceso Python desde archivo maestro externo (p.ej. "
               "archivos/fuente/Jornales_maestro.xlsx). NO usar vínculos externos vivos.")
    wb["0_Indice_CAC"].cell(1, 6).value = doc_cac
    wb["0_Jornales_MO"].cell(1, 12).value = doc_jor
    log.append("0_Indice_CAC / 0_Jornales_MO: nota de importación maestra documentada")

    # ====================================================================
    # 11) CHANGELOG
    # ====================================================================
    if "CHANGELOG" in wb.sheetnames:
        del wb["CHANGELOG"]
    ch = wb.create_sheet("CHANGELOG")
    rows = [
        ["CHANGELOG — CH 2171 v8_3 (versión EGRESOS, prototipo estado futuro)"],
        [""],
        ["Principio: rubros alineados al plan de cuentas de Tezamat; ID de SC + tipo de"],
        ["movimiento + CAC cargados de forma estandarizada en Tezamat (aguas arriba)."],
        ["El circuito de INGRESOS no se modifica en esta versión."],
        [""],
        ["CAMBIO", "DETALLE", "MOTIVO"],
        ["Rubros -> token único", "1_Presupuesto A/B='X MT', C/D='X MO'; Hormigón MO='Homigón MO' (typo Tezamat)",
         "Plan de cuentas Tezamat trae rubro+tipo en un solo campo (Desc Cuenta)"],
        ["Cruce Control_Ppto por E", "SUMIFS sobre 2_Gastos!E (Desc Cuenta), no por col A reclasificada",
         "Elimina reclasificación manual; rubro viene alineado de origen"],
        ["2_Gastos A,B,U,V,W obsoletas", "Vaciadas + header (OBSOLETA). NO borradas físicamente",
         "delete_cols corrompe refs por letra (openpyxl no reajusta). Vaciar es reference-safe"],
        ["2_Gastos X repunteada", "X = ISNUMBER(MATCH(E, Control_Ppto rubros)) = 'rubro reconocido'",
         "Alimenta sanity check de gasto sin rubro"],
        ["2_Quincenas col E", "Códigos ALBAMO/HORMIMO/ELECTMO/HERRERO -> tokens MO",
         "Alinea con Control (arregla bug: antes no matcheaba por usar códigos)"],
        ["Control_Ppto fórmulas unificadas", "1 fórmula por columna (sin IF B=MT/MO); 27 tokens (17 canónicos + 10 extras debajo)",
         "Mantenibilidad; reconciliación exacta del presupuesto"],
        ["Sanity 'gasto sin rubro'", "Control_Ppto E102 = SUMIFS(2_Gastos!T, X=FALSE)",
         "Detecta cuentas Tezamat fuera del plan de rubros"],
        ["Subcontratos por Observaciones", "ID CH-SC-001; SUMIFS comodín '*ID TIPO*' (AVANCE/QUINCENA/CAC/CARGAS SOCIALES)",
         "ID vive en Observaciones (col H), no en col U"],
        ["Saldo SC", "Saldo = Presup - (AVANCE + QUINCENA); CAC y Cargas suman a Total pero no descuentan saldo",
         "Definición de proceso nueva"],
        ["CAC fila separada", "El CAC viene como fila aparte marcada en Observaciones; col S deflacta por fila (sin cambios)",
         "Ya no se deriva restando en la misma fila (V/W obsoletas)"],
        ["Hojas eliminadas", "2_Pagos_Subc, 2_Pagos_Quincena_SC, 2_Certificaciones",
         "Deep-dive SC va al dashboard; certificaciones -> versión de ingresos"],
        ["Ingresos neutralizados", "Dashboard C4/C5/C6/C20 y Cash_Flow filas 4-5 = 0 (con nota)",
         "2_Certificaciones eliminada; ingresos se reconectan en otra versión"],
        ["Maestros documentados", "0_Indice_CAC / 0_Jornales_MO: nota de import futuro (sin vínculo vivo)",
         "Una sola actualización en maestro externo"],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ch.cell(r, c).value = val

    wb.save(DST)
    print("✅ Guardado:", DST)
    print("\nResumen de cambios:")
    for x in log:
        print("  -", x)


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    main()
