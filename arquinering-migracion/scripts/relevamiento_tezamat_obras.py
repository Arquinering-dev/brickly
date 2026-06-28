"""relevamiento_tezamat_obras.py — Propuestas de tagueo por movimiento + auditoría de
cobertura, en el deliverable docs/Carga_Tezamat_IDs.xlsx.

Reglas clave:
- La 'Observación ACTUAL (Tezamat)' = col P 'Observaciones original' si está poblada
  (caso CH, que en el Excel ya tiene nuestros tags pero en Tezamat NO), si no col E.
  NUNCA se marca 'ya tagueado': Tezamat todavía no tiene los IDs (los carga Arquinering).
- Cada fila incluye Nº Asiento y Nº Comprobante para ubicar el movimiento en Tezamat.
- Match de subcontratos por apellido/razón social contra (proveedor + obs actual).
- Hoja 'Cobertura (auditoría)': por subcontrato (contrato vs movs identificados, %) y por OC.

Hojas regeneradas: Mov SVD / Mov CH (relevadas) · Mov GDR (se le agregan asiento/comprob a la
proposal hand-curada existente) · Cobertura (auditoría).
"""
import sys
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DOC = "docs/Carga_Tezamat_IDs.xlsx"
OBRAS = {
    "SVD": "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx",
    "CH": "archivos/output/CH_2171_Resumen_de_Obra_v8_11.xlsx",
    "GDR": "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx",
}
SC_KEYS = {
    "SVD": {"SVD-SC-001": ["CELSI"], "SVD-SC-002": ["SEREN"], "SVD-SC-003": ["AGUIRRE"],
            "SVD-SC-004": ["PASQUAR"], "SVD-SC-005": ["AMARILLA"], "SVD-SC-006": ["CESPEDES"],
            "SVD-SC-007": ["SILVA"], "SVD-SC-008": ["MENDOZA"], "SVD-SC-009": ["SCHULZE"],
            "SVD-SC-010": ["MANSILLA", "DERQUI"]},
    "CH": {"CH-SC-001": ["BALACLAV"], "CH-SC-002": ["ARDILES"],
           "CH-SC-003": ["MICROPILOTE", "FEIS"], "CH-SC-004": ["CELSI"]},
}
HF = Font(bold=True, color="FFFFFFFF")
HFILL = PatternFill("solid", fgColor="FF1F4E78")
MOVHEAD = ["Fila 2_Mov", "Nº Asiento", "Nº Comprobante", "Fecha", "Cuenta", "Desc Cuenta",
           "Debe (ARS)", "Observación ACTUAL (Tezamat)", "ID propuesto", "Proveedor", "TIPO",
           "OBSERVACIÓN A ESCRIBIR", "Confianza", "Nota / motivo"]


def U(s):
    return str(s or "").upper()


def hdr(ws, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(1, j, h); c.font = HF; c.fill = HFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def anchos(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def tipo_de(txt):
    if "CAC" in txt:
        return "CAC"
    if "CARGA" in txt or " CS" in txt:
        return "CS"
    if "ANTIC" in txt:
        return "ANT"
    return "BASE"


def match_sc(txt, keys):
    return next((s for s, ks in keys.items() if any(k in txt for k in ks)), None)


def relevar(code):
    """Lee 2_Movimientos de la obra. Devuelve (movs, por_sc, ingresos).
    obs_actual = col P (original) si poblada, si no col E."""
    wb = openpyxl.load_workbook(OBRAS[code], data_only=True)
    mov = wb["2_Movimientos"]
    keys = SC_KEYS.get(code, {})
    # mapa fila -> (asiento, comprob, obs_actual, prov) para resolver CAC desagregado
    info = {}
    for r in range(2, mov.max_row + 1):
        E = mov.cell(r, 5).value
        P = mov.cell(r, 16).value if mov.max_column >= 16 else None
        obs_act = P if (P not in (None, "")) else E
        info[r] = (mov.cell(r, 4).value, mov.cell(r, 8).value, obs_act, mov.cell(r, 6).value)

    movs = []
    por_sc = {sid: {"n": 0, "monto": 0.0} for sid in keys}
    ingresos = {"n": 0, "monto": 0.0}
    for r in range(2, mov.max_row + 1):
        cta = mov.cell(r, 1).value
        if cta is None:
            continue
        asiento, comprob, obs_act, prov = info[r]
        desc_cta = mov.cell(r, 2).value
        debe = mov.cell(r, 9).value or 0
        haber = mov.cell(r, 10).value or 0
        fecha = mov.cell(r, 3).value
        obs_str = str(obs_act or "")
        txt = U(prov) + " | " + U(obs_str)

        if str(cta).startswith("4"):                          # ingreso (cobro del cliente)
            ingresos["n"] += 1; ingresos["monto"] += float(haber or 0)
            movs.append([r, asiento, comprob, fecha, cta, desc_cta, haber, obs_str,
                         "(asignar OC/cert)", prov, "BASE",
                         f"{code}-OCxx-Cyy-B/N | BASE | (identificar qué certificación paga este cobro)",
                         "Manual", "Cobro del Cliente — asignar a la certificación (ver catálogo Ingresos)."])
            continue

        # filas de CAC desagregado que agregamos en el Excel (no existen aparte en Tezamat)
        m = re.search(r"CAC desagregado de fila (\d+)", obs_str)
        if m:
            base_r = int(m.group(1))
            ba, bc, bobs, bprov = info.get(base_r, (None, None, "", ""))
            sid = match_sc(U(bprov) + " | " + U(bobs), keys)
            if sid:
                por_sc[sid]["n"] += 1; por_sc[sid]["monto"] += float(debe or 0)
                movs.append([r, ba, bc, fecha, cta, desc_cta, debe, str(bobs or ""),
                             sid, bprov, "CAC",
                             f"{sid} | CAC | ajuste CAC del certif. (mov base asiento {ba})", "Media",
                             f"CAC que en el Excel separamos de la fila {base_r}. En Tezamat: cargar el "
                             f"ajuste CAC como registro CAC aparte del movimiento base (asiento {ba})."])
            continue

        sid = match_sc(txt, keys)
        if sid:
            tipo = tipo_de(txt)
            por_sc[sid]["n"] += 1; por_sc[sid]["monto"] += float(debe or 0)
            movs.append([r, asiento, comprob, fecha, cta, desc_cta, debe, obs_str, sid, prov, tipo,
                         f"{sid} | {tipo} | {obs_str[:40]}", "Alta",
                         "Match por nombre de proveedor/observación."])
    return movs, por_sc, ingresos


def escribir_mov(wb, code, movs):
    name = f"Mov {code} (propuesta)"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    hdr(ws, MOVHEAD)
    for i, row in enumerate(movs):
        for c, v in enumerate(row, start=1):
            ws.cell(2 + i, c, v)
    anchos(ws, [9, 11, 14, 11, 9, 17, 13, 30, 15, 20, 7, 40, 9, 44])
    ws.freeze_panes = "A2"


def augment_gdr(wb):
    """Agrega Nº Asiento / Nº Comprobante a la hoja Mov GDR ya portada (hand-curada)."""
    name = "Mov GDR (propuesta)"
    if name not in wb.sheetnames:
        return {}
    old = wb[name]
    # idempotencia: si ya está ampliada (col 2 = 'Nº Asiento'), solo recomputar cobertura
    if str(old.cell(1, 2).value or "").startswith("Nº Asiento"):
        agg = {}
        for r in range(2, old.max_row + 1):
            sid = old.cell(r, 9).value
            if sid and str(sid).startswith("GDR-SC"):
                a = agg.setdefault(sid, [0, 0.0]); a[0] += 1; a[1] += float(old.cell(r, 7).value or 0)
        return agg
    # leer filas viejas (layout 12 cols: Fila,Fecha,Cuenta,Desc,Debe,ObsAct,ID,Prov,TIPO,ObsEsc,Conf,Nota)
    rows = []
    for r in range(2, old.max_row + 1):
        if old.cell(r, 1).value is None:
            continue
        rows.append([old.cell(r, c).value for c in range(1, 13)])
    # lookup asiento/comprob desde GDR 2_Movimientos por nº de fila
    mov = openpyxl.load_workbook(OBRAS["GDR"], data_only=True)["2_Movimientos"]
    look = {r: (mov.cell(r, 4).value, mov.cell(r, 8).value) for r in range(2, mov.max_row + 1)}
    del wb[name]
    ws = wb.create_sheet(name)
    hdr(ws, MOVHEAD)
    agg = {}
    for i, old_row in enumerate(rows):
        fila = old_row[0]
        asiento, comprob = look.get(int(fila), (None, None)) if isinstance(fila, (int, float)) else (None, None)
        # reordenar: Fila, Asiento, Comprob, Fecha, Cuenta, Desc, Debe, ObsAct, ID, Prov, TIPO, ObsEsc, Conf, Nota
        new = [old_row[0], asiento, comprob] + old_row[1:]
        for c, v in enumerate(new, start=1):
            ws.cell(2 + i, c, v)
        sid = old_row[6]
        if sid and str(sid).startswith("GDR-SC"):
            a = agg.setdefault(sid, [0, 0.0]); a[0] += 1; a[1] += float(old_row[4] or 0)
    anchos(ws, [9, 11, 14, 11, 9, 17, 13, 30, 15, 20, 7, 40, 9, 44])
    ws.freeze_panes = "A2"
    return agg


def contratos(code):
    sc = openpyxl.load_workbook(OBRAS[code], data_only=True)["2_Subcontratos"]
    out = {}
    for r in range(4, sc.max_row + 1):
        cid = sc.cell(r, 1).value
        if cid:
            out[cid] = (sc.cell(r, 2).value, sc.cell(r, 3).value, sc.cell(r, 5).value or 0)
    return out


def certificado_oc(code):
    try:
        co = openpyxl.load_workbook(OBRAS[code], data_only=True)["Cert_Control_OC"]
    except KeyError:
        return {}
    return {co.cell(r, 1).value: (co.cell(r, 6).value or 0) for r in range(2, co.max_row + 1)
            if co.cell(r, 1).value and co.cell(r, 1).value != "TOTAL"}


def main():
    wb = openpyxl.load_workbook(DOC)
    audit_sc, audit_ing = [], []

    for code in ("SVD", "CH"):
        movs, por_sc, ing = relevar(code)
        escribir_mov(wb, code, movs)
        cons = contratos(code)
        for sid, d in por_sc.items():
            prov, rubro, contrato = cons.get(sid, ("", "", 0))
            if d["n"] == 0:
                estado = "⚠ Sin movimientos en el extracto (¿pagado fuera del período / otro nombre?)"
            elif contrato and d["monto"] >= contrato * 0.98:
                estado = "✓ Cubierto (movs ≈ contrato)"
            else:
                estado = f"Parcial: {d['n']} movs identificados"
            audit_sc.append((code, sid, prov, rubro, contrato, d["n"], d["monto"], estado))
        cert = certificado_oc(code)
        estado_i = ("Depósitos lump del Cliente sin asignar a certificación — asignar manual"
                    if ing["monto"] > 0 else "Sin cobros en el extracto")
        audit_ing.append((code, ing["n"], ing["monto"], sum(cert.values()), estado_i))

    # GDR: augmentar y leer cobertura
    agg = augment_gdr(wb)
    cons = contratos("GDR")
    for sid, (prov, rubro, contrato) in cons.items():
        n, monto = agg.get(sid, [0, 0.0])
        estado = ("⚠ Sin movimientos propuestos" if n == 0
                  else ("✓ Cubierto (movs ≈ contrato)" if contrato and monto >= contrato * 0.98
                        else f"Parcial: {n} movs propuestos"))
        audit_sc.append(("GDR", sid, prov, rubro, contrato, n, monto, estado))
    cert = certificado_oc("GDR")
    audit_ing.append(("GDR", "ver 2_Mov", 0, sum(cert.values()),
                      "Depósitos del Cliente lump — asignar manual"))

    # ---- Cobertura (auditoría) ----
    if "Cobertura (auditoría)" in wb.sheetnames:
        del wb["Cobertura (auditoría)"]
    au = wb.create_sheet("Cobertura (auditoría)")
    au.cell(1, 1, "AUDITORÍA DE COBERTURA — qué movimientos del extracto quedan identificados con la propuesta. "
                  "NOTA: ninguna obra está tagueada en Tezamat todavía (los IDs los carga Arquinering).").font = Font(bold=True)
    au.cell(3, 1, "A) SUBCONTRATOS (egresos)").font = Font(bold=True, color="FF1F4E78")
    cols = ["Obra", "ID subcontrato", "Proveedor", "Rubro", "Contrato ($)", "# movs identif.",
            "$ identificado", "Cobertura % vs contrato", "Estado"]
    for j, h in enumerate(cols, start=1):
        c = au.cell(4, j, h); c.font = HF; c.fill = HFILL; c.alignment = Alignment(horizontal="center", wrap_text=True)
    r = 5
    for (obra, sid, prov, rubro, contrato, n, monto, estado) in audit_sc:
        cov = (monto / contrato) if contrato else 0
        for c, v in enumerate([obra, sid, prov, rubro, contrato, n, monto, cov, estado], start=1):
            cell = au.cell(r, c, v)
            if c in (5, 7):
                cell.number_format = "\\$#,##0"
            if c == 8:
                cell.number_format = "0%"
        r += 1
    r += 1
    au.cell(r, 1, "B) INGRESOS / COBROS DEL CLIENTE").font = Font(bold=True, color="FF1F4E78"); r += 1
    cols2 = ["Obra", "# depósitos en extracto", "$ depósitos", "$ certificado (esperado)", "Estado / acción"]
    for j, h in enumerate(cols2, start=1):
        c = au.cell(r, j, h); c.font = HF; c.fill = HFILL; c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    for (obra, n, monto, cert, estado) in audit_ing:
        for c, v in enumerate([obra, n, monto, cert, estado], start=1):
            cell = au.cell(r, c, v)
            if c in (3, 4) and isinstance(v, (int, float)):
                cell.number_format = "\\$#,##0"
        r += 1
    anchos(au, [7, 16, 24, 22, 16, 14, 16, 18, 50])
    au.sheet_properties.tabColor = "FFB8860B"

    orden = ["Instrucciones", "Subcontratos (catálogo)", "Ingresos (catálogo)",
             "Cobertura (auditoría)", "Mov GDR (propuesta)", "Mov CH (propuesta)",
             "Mov SVD (propuesta)", "Pendientes"]
    wb._sheets.sort(key=lambda s: orden.index(s.title) if s.title in orden else 99)
    wb.save(DOC)

    print(f"✓ Relevamiento + auditoría (CH NO pre-tagueado; +asiento/comprobante) → {DOC}")
    print("\n  COBERTURA SUBCONTRATOS:")
    for (obra, sid, prov, rubro, contrato, n, monto, estado) in audit_sc:
        print(f"   {obra} {sid} {str(prov)[:20]:21} contrato=${contrato:>13,.0f}  {n:2} movs ${monto:>13,.0f}  {estado}")
    print("\n  COBERTURA INGRESOS:")
    for (obra, n, monto, cert, estado) in audit_ing:
        md = f"${monto:,.0f}" if isinstance(monto, (int, float)) else str(monto)
        print(f"   {obra}: {n} depósitos {md} · certificado ${cert:,.0f} · {estado}")
    print(f"\n  hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
