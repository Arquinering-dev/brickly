"""svd_bloque356_mov.py — Migracion SVD 4140 a v8. Bloques 3, 5 y 6.
B3: 1_Composicion -> placeholder (SVD no tiene APU Unificado; se limpia la data de GDR).
B5: 2_Gastos_DirInd -> tabla plana vacia (placeholder, carga manual).
B6: 2_Movimientos  -> carga el extracto Tezamat real de SVD (centro de costo SALVA4140).
"""
import sys
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
LEGACY = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"
TEZAMAT = "archivos/fuente/SVD 4140 Mayores 2026-06-01.xlsx"


def build():
    wb = openpyxl.load_workbook(DST)

    # --- B3: 1_Composicion -> placeholder (limpiar data heredada de GDR) -----
    comp = wb["1_Composicion"]
    n_comp = 0
    for r in range(2, comp.max_row + 1):
        for c in range(1, comp.max_column + 1):
            if comp.cell(row=r, column=c).value is not None:
                comp.cell(row=r, column=c).value = None
                n_comp += 1

    # --- B5: 2_Gastos_DirInd -> tabla plana vacia ---------------------------
    gdi = wb["2_Gastos_DirInd"]
    n_gdi = 0
    for r in range(2, gdi.max_row + 1):
        for c in range(1, gdi.max_column + 1):
            if gdi.cell(row=r, column=c).value is not None:
                gdi.cell(row=r, column=c).value = None
                n_gdi += 1

    # --- B6: 2_Movimientos <- extracto Tezamat SVD --------------------------
    mov = wb["2_Movimientos"]
    # limpiar data heredada de GDR (filas 2+), conservar header (fila 1)
    for r in range(2, mov.max_row + 1):
        for c in range(1, mov.max_column + 1):
            mov.cell(row=r, column=c).value = None
    # leer Mayores (1 sola hoja) y copiar A-N 1:1
    tz = openpyxl.load_workbook(TEZAMAT, data_only=True)
    tws = tz[tz.sheetnames[0]]
    n_mov = 0
    out = 2
    for r in range(2, tws.max_row + 1):
        cuenta = tws.cell(row=r, column=1).value
        if cuenta is None:
            continue
        for c in range(1, 15):  # A..N
            mov.cell(row=out, column=c).value = tws.cell(row=r, column=c).value
        out += 1
        n_mov += 1

    wb.save(DST)
    print(f"✓ Bloques 3/5/6 → {DST}")
    print(f"  · B3 1_Composicion: placeholder ({n_comp} celdas GDR limpiadas; header intacto)")
    print(f"  · B5 2_Gastos_DirInd: tabla plana vacia ({n_gdi} celdas limpiadas)")
    print(f"  · B6 2_Movimientos: {n_mov} movimientos Tezamat SVD cargados (filas 2..{out-1})")


if __name__ == "__main__":
    build()
