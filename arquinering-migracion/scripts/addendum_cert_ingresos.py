"""
Addendum de ajustes al circuito de certificacion/ingresos (HANDOFF §5/§10).
Entrada: CH_2171_Resumen_de_Obra_v8_7.xlsx  ->  Salida: ..._v8_8.xlsx (v8_7 queda backup).
Cambios (solo forma/estructura, value-preserving; NO conciliacion contra 2_Movimientos):
 1) Desacopio a nivel madre: hoja nueva Cert_Cabecera (un % por certificacion);
    Cert_Calculo!J pasa de input a referencia (XLOOKUP a la cabecera).
 2) CAC con override: nueva col input "Indice CAC override" (Y) + N = IF(override,...).
 3) id_factura estructurado colgando del Blanco ({id_cert_fact}-F{NN}).
 4) Cert_Facturacion forma final: monto / moneda / monto_ars_equiv (=USD x TC cobro) +
    id_OC, id_cert_madre, id_factura, Clase como columnas propias. Un solo toque.
 5) Consecuencia: Cert_Control_OC!G/H re-apuntan a monto_ars_equiv.
"""
import openpyxl
from openpyxl.utils import get_column_letter as gl, column_index_from_string as ci
from openpyxl.styles import Font, PatternFill, Alignment

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_7.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_8.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=False)

# ---- estilos base (replicar chrome v8) ----
base_cell = wb['Cert_Calculo'].cell(2, 1)
FN = base_cell.font.name or 'Aptos Narrow'
FS = base_cell.font.size or 11
HDR_FILL = PatternFill('solid', fgColor='FF1F4E78')
HDR_FONT = Font(name=FN, size=FS, bold=True, color='FFFFFFFF')
HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
F_INPUT = lambda: Font(name=FN, size=FS, color='FF0000FF')   # azul
F_REF   = lambda: Font(name=FN, size=FS, color='FF00B050')   # verde
F_CALC  = lambda: Font(name=FN, size=FS, color='FF000000')   # negro


def style_header(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row, j, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN


# ================== 1) Cert_Cabecera ==================
# madres + % desacopio leidos de Cert_Calculo (value-preserving: mismo % que hoy por fila)
cc = wb['Cert_Calculo']
madre_desac = {}
for r in range(2, cc.max_row + 1):
    idc = cc.cell(r, ci('A')).value
    des = cc.cell(r, ci('J')).value
    if idc is None:
        continue
    if idc in madre_desac and madre_desac[idc] != des:
        raise SystemExit("Desacopio distinto entre ramas de %s (%s vs %s)" % (idc, madre_desac[idc], des))
    madre_desac[idc] = des
madres = sorted(madre_desac)  # 8

idx = wb.sheetnames.index('Cert_Calculo') + 1
cab = wb.create_sheet('Cert_Cabecera', idx)
style_header(cab, 1, ['ID Certif', 'ID OC', 'Fecha', '$ base bruta certif.',
                      '% desacopio (madre)', '$ desacopio', 'Subtotal neto', 'Chequeo'])
for i, idc in enumerate(madres):
    r = 2 + i
    cab.cell(r, ci('A'), idc).font = F_INPUT()
    c = cab.cell(r, ci('B'), '=LEFT($A%d,FIND("|",SUBSTITUTE($A%d,"-","|",2))-1)' % (r, r)); c.font = F_CALC()
    c = cab.cell(r, ci('C'), '=IFERROR(_xlfn.XLOOKUP($A%d,Cert_App_Output!$C:$C,Cert_App_Output!$D:$D),"")' % r)
    c.font = F_REF(); c.number_format = 'dd/mm/yyyy'
    c = cab.cell(r, ci('D'), '=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$C:$C,$A%d)' % r)
    c.font = F_REF(); c.number_format = '\\$#,##0'
    c = cab.cell(r, ci('E'), madre_desac[idc]); c.font = F_INPUT(); c.number_format = '0.0%'
    c = cab.cell(r, ci('F'), '=$D%d*$E%d' % (r, r)); c.font = F_CALC(); c.number_format = '\\$#,##0'
    c = cab.cell(r, ci('G'), '=$D%d-$F%d' % (r, r)); c.font = F_CALC(); c.number_format = '\\$#,##0'
    c = cab.cell(r, ci('H'), '=IF(COUNTIF(Cert_Calculo!$A:$A,$A%d)>0,"OK","FALTA en Calculo")' % r); c.font = F_CALC()
cab.freeze_panes = 'A2'
cab.sheet_view.showGridLines = False
for col, w in {'A': 16, 'B': 11, 'C': 12, 'D': 20, 'E': 18, 'F': 16, 'G': 16, 'H': 18}.items():
    cab.column_dimensions[col].width = w

# ================== 2) Cert_Calculo: desacopio->ref, CAC override ==================
OVR = ci('Y')  # 25, libre (W/X = 23/24)
style_header(cc, 1, [None] * (OVR - 1) + ['Indice CAC override'])
for r in range(2, cc.max_row + 1):
    if cc.cell(r, ci('A')).value is None:
        continue
    # J: desacopio ahora referencia la cabecera (un solo % por certificacion madre)
    j = cc.cell(r, ci('J'), '=_xlfn.XLOOKUP($A%d,Cert_Cabecera!$A:$A,Cert_Cabecera!$E:$E)' % r)
    j.font = F_REF(); j.number_format = '0.0%'
    # N: indice CAC = override si esta cargado, si no el lookup automatico (patron tipo MEP)
    n = cc.cell(r, ci('N'),
                '=IF($Y%d="",_xlfn.XLOOKUP(DATE(YEAR(F%d),MONTH(F%d),1),' % (r, r, r) +
                "'0_Indice_CAC'!$A:$A,'0_Indice_CAC'!$B:$B),$Y%d)" % r)
    n.font = F_REF(); n.number_format = '#,##0.00'
    # Y: override input (vacio por defecto -> formula-pure)
    y = cc.cell(r, OVR); y.font = F_INPUT(); y.number_format = '#,##0.00'
cc.column_dimensions[gl(OVR)].width = 16

# ================== 4) Cert_Facturacion: forma final (un solo toque) ==================
fa = wb['Cert_Facturacion']
# renombrar/headear cols existentes + nuevas K..O
fa.cell(1, ci('A'), 'id_cert_fact').font = HDR_FONT
fa.cell(1, ci('D'), 'monto')
fa.cell(1, ci('E'), 'moneda')
fa.cell(1, ci('G'), 'monto_ars_equiv')
for L in ['A', 'D', 'E', 'G']:
    h = fa.cell(1, ci(L)); h.font = HDR_FONT; h.fill = HDR_FILL; h.alignment = HDR_ALIGN
style_header(fa, 1, [None] * (ci('K') - 1) + ['id_factura', 'TC cobro', 'id_cert_madre', 'id_OC', 'Clase'])

for r in range(2, fa.max_row + 1):
    a = fa.cell(r, ci('A')).value
    if a is None:
        continue
    tipo = fa.cell(r, ci('C')).value
    oldD = fa.cell(r, ci('D')).value
    oldE = fa.cell(r, ci('E')).value
    if tipo == 'Negro':
        # monto en moneda del cobro = USD; ARS-equiv preservado via TC = oldD/oldE
        fa.cell(r, ci('D'), oldE)
        fa.cell(r, ci('E'), 'USD')
        fa.cell(r, ci('L'), (oldD / oldE) if oldE else None)   # TC cobro
        fa.cell(r, ci('K'), None)                              # sin id_factura (solo Blanco)
    else:  # Blanco
        fa.cell(r, ci('E'), 'ARS')                             # monto (D) ya es ARS
        fa.cell(r, ci('L'), None)
        fa.cell(r, ci('K'), '%s-F01' % a)                      # id_factura colgando del Blanco
    # monto_ars_equiv (repurpose col G, antes 'Retencion')
    fa.cell(r, ci('G'), '=IF($E%d="USD",$D%d*$L%d,$D%d)' % (r, r, r, r))
    # claves propias derivadas
    fa.cell(r, ci('M'), '=LEFT($A%d,LEN($A%d)-2)' % (r, r))
    fa.cell(r, ci('N'), '=LEFT($M%d,FIND("|",SUBSTITUTE($M%d,"-","|",2))-1)' % (r, r))
    fa.cell(r, ci('O'), '=IF(ISNUMBER(SEARCH("-ANT",$M%d)),"Anticipo","Certificacion")' % r)
    # Saldo a facturar (J) re-apunta a monto_ars_equiv
    fa.cell(r, ci('J'), '=$I%d-SUMIFS($G:$G,$A:$A,$A%d)' % (r, r))
    # estilos/formatos
    fa.cell(r, ci('D')).font = F_INPUT(); fa.cell(r, ci('D')).number_format = '#,##0'
    fa.cell(r, ci('E')).font = F_INPUT(); fa.cell(r, ci('E')).number_format = '@'
    fa.cell(r, ci('G')).font = F_CALC();  fa.cell(r, ci('G')).number_format = '\\$#,##0'
    fa.cell(r, ci('K')).font = F_INPUT(); fa.cell(r, ci('K')).number_format = '@'
    fa.cell(r, ci('L')).font = F_INPUT(); fa.cell(r, ci('L')).number_format = '#,##0.00'
    for L in ['M', 'N', 'O']:
        fa.cell(r, ci(L)).font = F_CALC(); fa.cell(r, ci(L)).number_format = '@'
    fa.cell(r, ci('J')).font = F_CALC()
for col, w in {'K': 18, 'L': 12, 'M': 15, 'N': 11, 'O': 14}.items():
    fa.column_dimensions[col].width = w

# ================== 5) Cert_Control_OC: G/H -> monto_ars_equiv ==================
co = wb['Cert_Control_OC']
for r in (2, 3):
    co.cell(r, ci('G'), '=SUMIFS(Cert_Facturacion!$G:$G,Cert_Facturacion!$A:$A,$A%d&"-*")' % r)
    co.cell(r, ci('H'), '=SUMIFS(Cert_Facturacion!$G:$G,Cert_Facturacion!$A:$A,$A%d&"-*",Cert_Facturacion!$F:$F,">0")' % r)

wb.save(DST)
print("Guardado:", DST)
print("Cert_Cabecera madres:", len(madres), "->", dict(madre_desac))
print("Nuevas cols Cert_Facturacion: K id_factura, L TC cobro, M id_cert_madre, N id_OC, O Clase")
print("Override CAC: Cert_Calculo!Y ; N con IF(override,...)")
