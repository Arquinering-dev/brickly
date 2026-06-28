"""svd_bloque6b2_quincenas.py — Migracion SVD 4140 a v8. Bloque 6b (parte 2).
Carga 2_Quincenas (SOLO HORAS) desde las 4 hojas M.O. del legacy. Horas = dias-persona x 8.
Las columnas de costo (K-O) quedan VACIAS: el dato monetario lo carga Arquinering en Tezamat
(evita doble conteo con 2_Movimientos). Decision Pedro 2026-06-22.
"""
import sys
import datetime
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

DST = "archivos/output/SVD_4140_Resumen_de_Obra_v8_1.xlsx"
LEGACY = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"
HS_DIA = 8

# hoja M.O. -> rubro plan (variante MO)
SHEETS = {
    "M.O. HORM": "Homigón MO",
    "M.O. ALBA": "Albañilería MO",
    "M.O. ELE": "Eléctrico MO",
    "M.O. PLO": "Sanitaria MO",
}
# columna de dias -> categoria UOCRA (C=3 CPZ, D=4 OF, E=5 1/2Of, F=6 Ayu)
CATS = [(3, "CAPATAZ"), (4, "OFICIAL"), (5, "MEDIO OFICIAL"), (6, "AYUDANTE")]
OBRA = "EL SALVADOR"


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def leer_quincenas():
    wb = openpyxl.load_workbook(LEGACY, data_only=True)
    rows = []
    for sh, rubro in SHEETS.items():
        ws = wb[sh]
        for r in range(5, ws.max_row + 1):
            A = ws.cell(row=r, column=1).value
            B = ws.cell(row=r, column=2).value
            if not isinstance(A, datetime.datetime) or B not in ("1Q", "2Q"):
                continue
            for col, cat in CATS:
                dias = _num(ws.cell(row=r, column=col).value)
                if dias > 0:
                    rows.append((A, B, cat, rubro, dias * HS_DIA))
    return rows


def build():
    wb = openpyxl.load_workbook(DST)
    q = wb["2_Quincenas"]
    # limpiar filas 4+ (por las dudas) y reescribir
    for r in range(4, q.max_row + 1):
        for c in range(1, q.max_column + 1):
            q.cell(row=r, column=c).value = None

    rows = leer_quincenas()
    for i, (mes, per, cat, rubro, horas) in enumerate(rows):
        r = 4 + i
        q.cell(r, 1, mes).number_format = "mmm-yy"
        q.cell(r, 2, per)
        q.cell(r, 3, OBRA)
        q.cell(r, 4, cat)
        q.cell(r, 5, rubro)
        q.cell(r, 6, f"=G{r}+H{r}+I{r}")     # F Horas totales
        q.cell(r, 7, horas)                    # G Horas normales (días×8)
        q.cell(r, 8, 0)                        # H extra 50%
        q.cell(r, 9, 0)                        # I extra 100%
        # K-O (costo) quedan vacías: el monto va por Tezamat
    qmax = 3 + len(rows)

    # sanity + dropdown rubros de 2_Quincenas ∈ plan
    L = wb["_Listas"]
    rng = "$A$2:$A$34"
    L["E3"] = "Rubros fuera de plan — 2_Quincenas (E)"
    L["F3"] = (f"=SUMPRODUCT(('2_Quincenas'!E4:E{qmax}<>\"\")*('2_Quincenas'!E4:E{qmax}<>\"-\")"
               f"*ISNA(MATCH('2_Quincenas'!E4:E{qmax},{rng},0)))")
    L["G3"] = '=IF(F3=0,"✓","⚠")'
    try:
        dv = DataValidation(type="list", formula1="RUBROS_PLAN", allow_blank=True, showErrorMessage=False)
        q.add_data_validation(dv)
        dv.add(f"E4:E{qmax}")
    except Exception:
        pass

    wb.save(DST)
    print(f"✓ Bloque 6b (quincenas) → {DST}")
    by = {}
    for _, _, _, rub, h in rows:
        by[rub] = by.get(rub, 0) + h
    print(f"  · 2_Quincenas: {len(rows)} registros (horas), filas 4..{qmax}")
    for rub, h in by.items():
        print(f"     {rub:18} {h:>10,.0f} hs")
    print(f"  · columnas de costo (K-O) VACÍAS — el monto lo carga Arquinering en Tezamat")
    print(f"  · sanity _Listas!F3 (rubros quincenas ∈ plan) + dropdown E4:E{qmax}")


if __name__ == "__main__":
    build()
