"""reconciliar_plan_cuentas.py — Cruza el Plan de Cuentas oficial de Tezamat (AING)
contra los rubros usados en CH 2171 (1_Presupuesto + 2_Movimientos + _Listas).

Genera docs/Reconciliacion_Plan_Cuentas_CH2171.md (reproducible).

Fuente plan:  archivos/fuente/AING - Plan de Cuentas.xlsx
Fuente obra:  dashboard_v2/data/CH_2171_Resumen_de_Obra_v8_6.xlsx
"""
import sys
from collections import Counter, defaultdict
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PLAN = "archivos/fuente/AING - Plan de Cuentas.xlsx"
OBRA = "dashboard_v2/data/CH_2171_Resumen_de_Obra_v8_6.xlsx"
OUT = "docs/Reconciliacion_Plan_Cuentas_CH2171.md"


def norm(s):
    return str(s or "").strip().lower()


def load_plan():
    ws = openpyxl.load_workbook(PLAN, data_only=True)["Hoja1"]
    imput = {}          # code -> desc (solo imputables)
    by_branch = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(row=r, column=1).value or "").strip()
        desc = ws.cell(row=r, column=2).value
        if ws.cell(row=r, column=3).value == "S" and code:
            imput[code] = desc
            by_branch[code[:2]].append((code, desc))
    return imput, by_branch


def load_obra():
    wb = openpyxl.load_workbook(OBRA, data_only=True)
    # rubros de 1_Presupuesto (cols A/C/D, header fila 3)
    ws = wb["1_Presupuesto"]
    presup = Counter()
    for r in range(4, ws.max_row + 1):
        for c in (1, 3, 4):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() not in ("-", ""):
                presup[str(v).strip()] += 1
    # cuentas usadas en 2_Movimientos (col A code, col B desc)
    mv = wb["2_Movimientos"]
    mov = defaultdict(lambda: [0.0, 0, None])
    for r in range(2, mv.max_row + 1):
        code = mv.cell(row=r, column=1).value
        if code is None:
            continue
        desc = mv.cell(row=r, column=2).value
        neto = (mv.cell(row=r, column=9).value or 0) - (mv.cell(row=r, column=10).value or 0)
        a = mov[str(code).strip()]
        a[0] += neto; a[1] += 1; a[2] = desc
    # _Listas
    listas = [ws2 for ws2 in (wb["_Listas"].cell(row=r, column=1).value
              for r in range(2, wb["_Listas"].max_row + 1)) if ws2]
    return presup, mov, listas


def main():
    imput, by_branch = load_plan()
    presup, mov, listas = load_obra()
    # índice plan por nombre normalizado → code
    plan_by_name = {norm(d): c for c, d in imput.items()}
    obra_plan = {c: d for c, d in imput.items() if c[:2] == "53"}  # rama OBRA

    L = []
    L.append("# Reconciliación Plan de Cuentas Tezamat ↔ CH 2171")
    L.append("")
    L.append("> Generado por `scripts/reconciliar_plan_cuentas.py`. Fuente del plan: "
             "`AING - Plan de Cuentas.xlsx`; fuente de la obra: `CH_2171_..._v8_6.xlsx`.")
    L.append("")
    L.append("El plan es el plan contable completo de Arquinering. La rama **53 OBRA** "
             "contiene los rubros de obra (con split MT/MO explícito). Las ramas **50/51/52** "
             "(IMPUESTOS/PERSONAL/ADMINISTRACIÓN) son gastos indirectos.")
    L.append("")

    # 1) cuentas de 2_Movimientos: todas deberían estar en el plan (cruce por CÓDIGO)
    L.append("## 1. Cuentas usadas en `2_Movimientos` (cruce por CÓDIGO)")
    L.append("")
    L.append("| Código | Desc Cuenta | Mov. | Neto $ | Rama | ¿En plan? |")
    L.append("|--------|-------------|------|--------|------|-----------|")
    for code in sorted(mov):
        neto, n, desc = mov[code]
        rama = {"4": "INGRESO", "5": "EGRESO"}.get(code[:1], "?")
        sub = {"50": "Impuestos", "51": "Personal", "52": "Admin (indirecto)",
               "53": "OBRA (rubro)", "41": "Ventas", "42": "Otros ing."}.get(code[:2], code[:2])
        ok = "✓" if code in imput else "✗ FALTA"
        L.append(f"| {code} | {desc} | {n} | {neto:,.0f} | {sub} | {ok} |")
    faltan = [c for c in mov if c not in imput]
    L.append("")
    L.append(f"**Resultado:** {len(mov)-len(faltan)}/{len(mov)} cuentas resuelven contra el plan por código."
             + (" ✅ Todas." if not faltan else f" Faltan: {faltan}"))
    L.append("")

    # 2) rubros de 1_Presupuesto vs rama 53 OBRA (cruce por NOMBRE)
    L.append("## 2. Rubros de `1_Presupuesto` ↔ rama 53 OBRA (cruce por NOMBRE)")
    L.append("")
    L.append("| Rubro en 1_Presupuesto | n | Match en plan | Código | Observación |")
    L.append("|------------------------|---|---------------|--------|-------------|")
    sin_match = []
    for rub in sorted(presup):
        n = presup[rub]
        code = plan_by_name.get(norm(rub))
        if code:
            L.append(f"| {rub} | {n} | ✓ | {code} | — |")
        else:
            sin_match.append(rub)
            L.append(f"| {rub} | {n} | ✗ | — | **revisar** |")
    L.append("")
    L.append(f"**Sin match exacto ({len(sin_match)}):** {', '.join(sin_match)}")
    L.append("")

    # 3) cuentas 53 OBRA del plan no usadas como rubro de presupuesto
    L.append("## 3. Cuentas 53 OBRA del plan NO usadas como rubro en `1_Presupuesto`")
    L.append("")
    presup_names = {norm(x) for x in presup}
    no_usadas = [(c, d) for c, d in sorted(obra_plan.items()) if norm(d) not in presup_names]
    L.append("| Código | Desc | ¿Aparece en 2_Movimientos? |")
    L.append("|--------|------|----------------------------|")
    for c, d in no_usadas:
        usada = "sí (gasto)" if c in mov else "no"
        L.append(f"| {c} | {d} | {usada} |")
    L.append("")

    # 4) decisiones de mapeo tomadas con Pedro (2026-06-19)
    L.append("## 4. Decisiones de mapeo (Pedro, 2026-06-19)")
    L.append("")
    L.append("- **Política de nombres:** alinear los nombres de rubro del v8 (_Listas, "
             "1_Presupuesto, 2_Quincenas) EXACTAMENTE al plan, incluido el typo `Homigón MO`. "
             "Cruce 1:1 por nombre y por código.")
    L.append("- **Mov. Variables:** se mantiene como rubro de obra (mapea a la cuenta `52302`).")
    L.append("")
    L.append("| Rubro v8 actual | → Rubro canónico | Código | Acción |")
    L.append("|-----------------|------------------|--------|--------|")
    L.append("| Gastos Generales Obra MT + MO | Gastos Generales | 53022 | colapsar 2→1 (sin split MT/MO) |")
    L.append("| Mantenimiento MO | Gastos Generales | 53022 | fusionar |")
    L.append("| Seguridad e Higiene MO | Seguridad e Higiene | 53026 | renombrar (drop 'MO') |")
    L.append("| Agrimensura MT + MO | Preliminares | 53001 | fusionar en Preliminares |")
    L.append("| Supervisión de Obra MO | H. Ingeniería (indirecto) | 52209 | reclasificar a indirecto |")
    L.append("| Mov. Variables | Mov. Variables | 52302 | queda rubro (cuenta admin) |")
    L.append("")

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    print(f"✓ {OUT} ({len(L)} líneas)")
    print(f"  movimientos: {len(mov)-len(faltan)}/{len(mov)} en plan; "
          f"rubros presup sin match: {len(sin_match)}; "
          f"cuentas 53 OBRA sin uso en presup: {len(no_usadas)}")


if __name__ == "__main__":
    main()
