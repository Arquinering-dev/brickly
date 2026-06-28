"""Bloque 7a+7b — Circuito Cert_* de GDR (estructura).
 1) Agrega a 1_Presupuesto la dimensión 'Presupuesto' (AS, estática: PTO01 si etapa<=3, PTO02 si >=4)
    + reenganche nuevo: AT '% Acum Tot' (SUMIFS desde Cert_App_Output), AU 'Acum_tot $' (=AT*Z),
    AV 'Control'. (Las cols de avance viejas AC-AI quedan; no las lee el reader ni rompen en bloque 8.)
 2) Crea las 6 hojas Cert_* replicando CH: Cert_OC_Cliente (3 OC reales), Cert_App_Output / Cert_Cabecera /
    Cert_Calculo / Cert_Facturacion (headers; filas las llena el loader 7c), Cert_Control_OC (3 OC, fórmulas).
Adaptaciones CH->GDR: PV subtotal X->Z(26); dimensión AQ->AS(45); % Acum Tot AC->AT(46); código col F(6).
Cert_App_Output!I (PV) será STORED por el loader (no SUMIFS) para soportar OC03 adicionales (fuera del ppto base).

Uso: python scripts/crear_cert_gdr.py
"""
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as get_letter

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
HF = Font(bold=True, color="FFFFFFFF")
HFILL = PatternFill("solid", fgColor="FF1F4E78")


def hdr(ws, headers, row=1):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row, j, h)
        c.font = HF
        c.fill = HFILL
        c.alignment = Alignment(horizontal="center")


def etapa_de(cod):
    try:
        return int(float(cod))
    except (TypeError, ValueError):
        return None


def main():
    wb = openpyxl.load_workbook(PATH)

    # ---------- 1) 1_Presupuesto: dimensión + reenganche ----------
    p = wb["1_Presupuesto"]
    HR = 3
    p.cell(HR, 45, "Presupuesto")     # AS
    p.cell(HR, 46, "% Acum Tot")      # AT  (reader 'cert' lo ignora; usa Acum_tot $)
    p.cell(HR, 47, "Acum_tot $")      # AU  (reader 'cert')
    p.cell(HR, 48, "Control")         # AV
    for c in (45, 46, 47, 48):
        cc = p.cell(HR, c); cc.font = HF; cc.fill = HFILL; cc.alignment = Alignment(horizontal="center")
    n_dim = 0
    for r in range(5, 228):
        cod = p.cell(r, 6).value  # col F = código
        et = etapa_de(cod)
        if et is None:
            continue
        # solo tareas (código con decimales N.NN); las cabeceras de etapa son enteros
        es_tarea = isinstance(cod, (int, float)) and float(cod) != float(int(float(cod)))
        pto = "PTO 01" if et <= 3 else "PTO 02"
        p.cell(r, 45, pto)
        if es_tarea:
            p.cell(r, 46, f"=SUMIFS(Cert_App_Output!$G:$G,Cert_App_Output!$E:$E,$F{r},Cert_App_Output!$K:$K,$AS{r})")
            p.cell(r, 47, f"=AT{r}*Z{r}")
            p.cell(r, 48, f'=IF(AT{r}=0,"",IF(ABS(AT{r}-_xlfn.MAXIFS(Cert_App_Output!$H:$H,Cert_App_Output!$E:$E,$F{r},Cert_App_Output!$K:$K,$AS{r}))<0.0001,"✓","⚠"))')
            n_dim += 1

    # ---------- 2) Cert_OC_Cliente ----------
    for name in ["Cert_OC_Cliente", "Cert_App_Output", "Cert_Cabecera", "Cert_Calculo",
                 "Cert_Facturacion", "Cert_Control_OC"]:
        if name in wb.sheetnames:
            del wb[name]
    oc = wb.create_sheet("Cert_OC_Cliente")
    hdr(oc, ["Obra", "ID OC", "Descripcion", "Presupuesto aprobado", "% anticipo", "Mes base CAC",
             "Indice base CAC", "% Blanco sugerido", "% Negro sugerido", "% desacopio sugerido",
             "% IVA sugerido", "Presupuesto", "Fecha anticipo"])
    dic24 = datetime.datetime(2024, 12, 1)
    feb25 = datetime.datetime(2025, 2, 1)
    OCS = [
        # B,C,D monto,E ant,F mesbase,H blanco,I negro,J desac,K iva,L pto,M fant
        ("GDR-OC01", "Pto 1 - Preliminar y Hormigón", 500907720, 0.4, dic24, 0.7, 0.3, 0.4, 0.21, "PTO 01", feb25),
        ("GDR-OC02", "Pto 2 - Albañilería e Instalaciones", 844845248, 0.1, dic24, 0.7, 0.3, 0.1, 0.21, "PTO 02", feb25),
        ("GDR-OC03", "Adicionales 01", 96300019, 0.0, dic24, 0.0, 1.0, 0.0, 0.0, "PTO 03", None),  # PROVISIONAL (Pto3: anticipo/mesbase a confirmar)
    ]
    for i, (b, c, d, e, f, h, ineg, jd, k, l, m) in enumerate(OCS):
        r = 2 + i
        oc.cell(r, 1, "GDR 3760"); oc.cell(r, 2, b); oc.cell(r, 3, c)
        oc.cell(r, 4, d).number_format = "\\$#,##0"
        oc.cell(r, 5, e).number_format = "0%"
        oc.cell(r, 6, f).number_format = "mmm-yy"
        oc.cell(r, 7, f"=_xlfn.XLOOKUP(F{r},'0_Indice_CAC'!$A:$A,'0_Indice_CAC'!$B:$B)")
        oc.cell(r, 8, h).number_format = "0%"; oc.cell(r, 9, ineg).number_format = "0%"
        oc.cell(r, 10, jd).number_format = "0%"; oc.cell(r, 11, k).number_format = "0%"
        oc.cell(r, 12, l)
        if m:
            oc.cell(r, 13, m).number_format = "dd/mm/yyyy"
    oc.freeze_panes = "A2"

    # ---------- 3) Cert_App_Output (headers; loader llena filas) ----------
    ao = wb.create_sheet("Cert_App_Output")
    hdr(ao, ["Obra", "ID OC", "ID Certif", "Fecha", "Cod. tarea", "% anterior", "% actual",
             "% total", "PV total tarea", "$ base tarea (cert act.)", "Presupuesto", "Control vs Ppto"])
    ao.freeze_panes = "A2"

    # ---------- 4) Cert_Cabecera (headers) ----------
    cab = wb.create_sheet("Cert_Cabecera")
    hdr(cab, ["ID Certif", "ID OC", "Fecha", "$ base bruta certif.", "% desacopio (madre)",
              "$ desacopio", "Subtotal neto", "Chequeo"])
    cab.freeze_panes = "A2"

    # ---------- 5) Cert_Calculo (headers) ----------
    cal = wb.create_sheet("Cert_Calculo")
    hdr(cal, ["ID Certif", "ID Cert+Fact", "Tipo", "Obra", "ID OC", "Fecha", "$ base certif (total)",
              "% facturacion", "$ base de esta parte", "% desacopio", "$ desacopio", "$ base neta",
              "Indice base CAC", "Indice CAC a la fecha", "Ratio CAC", "$ CAC", "$ base + CAC",
              "% IVA", "$ IVA", "$ Total certificacion", "USD MEP a la fecha", "U$ Total", "",
              "Control B/N suma 100%", "Indice CAC override"])
    cal.freeze_panes = "A2"

    # ---------- 6) Cert_Facturacion (headers, vacía) ----------
    fac = wb.create_sheet("Cert_Facturacion")
    hdr(fac, ["id_cert_fact", "Comprobante", "Tipo", "monto", "moneda", "Fecha cobro", "monto_ars_equiv",
              "Estado conciliado", "$ certificado (control)", "Saldo a facturar", "id_factura", "TC cobro",
              "id_cert_madre", "id_OC", "Clase", "Retención", "Haber conciliado"])
    fac.freeze_panes = "A2"

    # ---------- 7) Cert_Control_OC (3 OC, fórmulas) ----------
    co = wb.create_sheet("Cert_Control_OC")
    hdr(co, ["ID OC", "Presupuesto", "Descripcion", "$ Contrato", "% Avance fisico", "$ Certificado",
             "$ Facturado", "$ Cobrado", "$ Anticipo certif.", "$ Avance s/certificar",
             "Saldo a facturar", "Saldo a cobrar", "Estado"])
    for i in range(3):
        r = 2 + i
        oref = 2 + i  # Cert_OC_Cliente row
        co.cell(r, 1, f"=Cert_OC_Cliente!B{oref}")
        co.cell(r, 2, f"=_xlfn.XLOOKUP($A{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$L:$L)")
        co.cell(r, 3, f"=_xlfn.XLOOKUP($A{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$C:$C)")
        co.cell(r, 4, f"=_xlfn.XLOOKUP($A{r},Cert_OC_Cliente!$B:$B,Cert_OC_Cliente!$D:$D)")
        co.cell(r, 5, (f"=IFERROR(SUMPRODUCT(('1_Presupuesto'!$AS$5:$AS$227=$B{r})*'1_Presupuesto'!$AT$5:$AT$227*'1_Presupuesto'!$Z$5:$Z$227)"
                       f"/SUMPRODUCT(('1_Presupuesto'!$AS$5:$AS$227=$B{r})*'1_Presupuesto'!$Z$5:$Z$227),0)"))
        co.cell(r, 6, f"=SUMIFS(Cert_Calculo!$T:$T,Cert_Calculo!$E:$E,$A{r})")
        co.cell(r, 7, f'=SUMIFS(Cert_Facturacion!$G:$G,Cert_Facturacion!$A:$A,$A{r}&"-*")')
        co.cell(r, 8, f'=SUMIFS(Cert_Facturacion!$G:$G,Cert_Facturacion!$A:$A,$A{r}&"-*",Cert_Facturacion!$F:$F,">0")')
        co.cell(r, 9, f'=SUMIFS(Cert_Calculo!$T:$T,Cert_Calculo!$E:$E,$A{r},Cert_Calculo!$B:$B,"*ANT*")')
        co.cell(r, 10, f'=SUMIFS(Cert_App_Output!$J:$J,Cert_App_Output!$B:$B,$A{r})-SUMIFS(Cert_Calculo!$I:$I,Cert_Calculo!$E:$E,$A{r},Cert_Calculo!$B:$B,"<>*ANT*")')
        co.cell(r, 11, f"=F{r}-G{r}")
        co.cell(r, 12, f"=G{r}-H{r}")
        co.cell(r, 13, f'=IF(J{r}>1000,"⚠ Certificar",IF(F{r}=0,"🔴 Sin iniciar",IF(K{r}>1000,"🟠 Por facturar",IF(L{r}>1000,"🔵 Por cobrar",IF(E{r}>=0.999,"🟢 Cerrada","🟡 En ejecucion")))))')
    # fila TOTAL (r5) — el reader del dashboard la usa como agregado (Cert_Control_OC)
    tr = 5
    co.cell(tr, 1, "TOTAL").font = Font(bold=True)
    co.cell(tr, 4, "=SUM(D2:D4)").number_format = "\\$#,##0"
    co.cell(tr, 5, "=SUMPRODUCT($D$2:$D$4,$E$2:$E$4)/SUM($D$2:$D$4)").number_format = "0.0%"
    for c in (6, 7, 8, 9, 10, 11, 12):
        L = get_letter(c)
        co.cell(tr, c, f"=SUM({L}2:{L}4)").number_format = "\\$#,##0"
    co.sheet_properties.tabColor = "FF2E7D32"
    co.freeze_panes = "A2"

    wb.save(PATH)
    print(f"✓ Cert_* estructura creada → {PATH}")
    print(f"  · 1_Presupuesto: dimensión Presupuesto + reenganche en {n_dim} tareas (AS/AT/AU/AV)")
    print(f"  · Cert_OC_Cliente: 3 OC (OC01 $500,9M ant40% · OC02 $844,8M ant10% · OC03 Adic $96,3M provisional)")
    print(f"  · Cert_App_Output/Cabecera/Calculo/Facturacion: headers (loader 7c llena filas)")
    print(f"  · Cert_Control_OC: 3 OC con fórmulas")
    print(f"  · orden hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
