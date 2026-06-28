"""
Bloque 3 — Genera APU_Unificado_SVD4140_v1.xlsx (formula-pure) desde budget_parsed.json.

Replica la estructura de los APU_Unificado de referencia (GDR/CH):
  CONFIG · MATERIALES · MANO_DE_OBRA · EQUIPOS · SUBCONTRATOS_PRY ·
  COMPOSICION (1 fila por insumo, VLOOKUP a maestras) · PARTIDAS (1 fila por tarea,
  SUMIFS sobre COMPOSICION) · PPTO_GENERADOR.

PARTIDAS = 184 items de presupuesto (codigo PTO-RRII; Rend=1 porque la composicion ya es
por-unidad). MO/ALB se normaliza a categoria UOCRA por jornal. Valores originales del
presupuesto (T/U/V) para auditoria Δ.
"""
import openpyxl, json, os, warnings
warnings.filterwarnings('ignore')
from openpyxl.styles import Font, PatternFill, Alignment

SCRATCH = os.environ.get('SCRATCH', '.')
bud = json.load(open(os.path.join(SCRATCH, 'budget_parsed.json'), encoding='utf-8'))
PRES = 'archivos/fuente/El Salvador_Pres 04 (sin pintura).xlsx'
OUT = 'archivos/output/APU_Unificado_SVD4140_v1.xlsx'

def rnd(x, n=4):
    return round(x, n) if isinstance(x, (int, float)) else x

# ---------- Nombres de rubro (grupos de '01') ----------
import re as _re
ws01 = openpyxl.load_workbook(PRES, data_only=True)['01']
# Etapas canónicas (Title Case + tildes) por nº de rubro del '01'. Track de ETAPAS DE OBRA
# (debe coincidir cross-obra cuando se refiere a lo mismo), distinto del track rubros↔plan de
# cuentas Tezamat. Confirmado con Pedro (2026-06-26).
RUBRO_NOMBRES = {
    1: 'Tareas Preliminares', 2: 'Movimiento de Suelos', 3: 'Estructura Resistente',
    4: 'Albañilería', 5: 'Aislaciones', 6: 'Cubiertas', 7: 'Revoques',
    8: 'Contrapisos y Carpetas', 9: 'Escaleras', 10: 'Ventilaciones', 11: 'Pisos',
    12: 'Zócalos y Solias', 13: 'Revestimientos', 14: 'Cielorrasos', 15: 'Pintura',
    16: 'Mármoles y Granitos', 17: 'Carpintería', 18: 'Instalación Eléctrica',
    19: 'Instalación Sanitaria', 20: 'Aire Acondicionado', 21: 'Ascensor',
    22: 'Equipamiento', 23: 'Trabajos Complementarios', 24: 'Ayuda de Gremios',
    25: 'Jefatura de Obra', 26: 'Compra Directa Comitente', 27: 'Varios', 28: 'Calefacción',
}

def pto_code(item):
    rubro = int(item)
    sub = int(round((item - rubro) * 100))
    return f'PTO-{rubro:02d}{sub:02d}'

# ---------- Normalizacion MO/ALB a categoria UOCRA por jornal ----------
SAL_CAT = [(75113.05, 'ESPECIALIZADO', 'MO-ESP'), (65360.25, 'OFICIAL', 'MO-OFI'),
           (54784.07, 'MEDIO OFICIAL', 'MO-MED'), (49934.24, 'AYUDANTE', 'MO-AYU'),
           (172653.21, 'JEFE DE OBRA', 'MO-JEF'), (212743.82, 'COORD OBRAS', 'MO-COO'),
           (131457.37, 'CAPATAZ', 'MO-CAP'), (117225.48, 'SOBRESTANTE', 'MO-SOB')]
def norm_mo_alb(precio):
    for sal, cat, cod in SAL_CAT:
        if precio and abs(precio - sal) < max(50, sal * 0.005):
            return cat, cod, round(sal, 4)
    return None

# ---------- Catalogos ----------
materiales, mano_obra, subc, equipos = {}, {}, {}, {}
mat_n = sub_n = eq_n = mo_x = 0

def get_mat_code(row):
    global mat_n
    if row['subtipo'] == 'CONS':
        materiales.setdefault(('consumibles', 1.0),
                              {'cod': 'MAT-CONS', 'desc': 'Consumibles', 'ud': 'gl',
                               'precio': 1, 'cat': 'Consumible'})
        return 'MAT-CONS'
    desc = row['desc']; precio = rnd(row['precio'] or 0)
    key = (desc.lower().strip(), round(precio, 2))
    if key not in materiales:
        mat_n += 1
        materiales[key] = {'cod': f'MAT-{mat_n:04d}', 'desc': desc, 'ud': row.get('ud', '') or '',
                           'precio': precio, 'cat': 'Corralón' if row['subtipo'] == 'COR' else 'Otros'}
    return materiales[key]['cod']

def get_mo_code(row):
    global mo_x
    nm = norm_mo_alb(row['precio'])
    if nm:
        cat, cod, sal = nm
        mano_obra.setdefault(cod, {'desc': cat, 'jornal': sal})
        return cod
    desc = row['desc']; precio = rnd(row['precio'] or 0)
    for cod, v in mano_obra.items():
        if v['desc'].lower() == desc.lower() and abs((v['jornal'] or 0) - precio) < 0.5:
            return cod
    mo_x += 1
    cod = f'MO-X{mo_x:02d}'
    mano_obra[cod] = {'desc': desc, 'jornal': precio}
    return cod

def get_sub_code(row):
    global sub_n
    precio = rnd(row['precio'] or 0)
    key = (row['desc'].lower().strip(), round(precio, 2))
    if key not in subc:
        sub_n += 1
        subc[key] = {'cod': f'SUB-{sub_n:04d}', 'desc': row['desc'], 'ud': row.get('ud', '') or '',
                     'precio': precio}
    return subc[key]['cod']

def get_eq_code(row):
    global eq_n
    precio = rnd(row['precio'] or 0)
    key = (row['desc'].lower().strip(), round(precio, 2))
    if key not in equipos:
        eq_n += 1
        equipos[key] = {'cod': f'EQ-{eq_n:04d}', 'desc': row['desc'], 'precio': precio}
    return equipos[key]['cod']

# ---------- Filas de composicion + partidas ----------
comp_rows, partidas = [], []
for it in bud['items']:
    code = pto_code(round(it['item'], 4))
    n_comp = 0
    for r in it['rows']:
        if (r['cant'] or 0) == 0:   # filas de costo 0 (p.ej. pintura excluida) -> descartar
            continue
        tipo, sub = r['tipo'], r['subtipo']
        if tipo == 'MAT':
            cod = get_mat_code(r)
        elif tipo == 'MO' and sub == 'ALB':
            cod = get_mo_code(r)
        elif tipo == 'MO':
            cod = get_sub_code(r)
        else:
            cod = get_eq_code(r)
        comp_rows.append({'partida': code, 'tipo': tipo, 'subtipo': sub if sub else '', 'cod': cod,
                          'ud': (r.get('ud') or ('j' if tipo == 'MO' else '')),
                          'cant': rnd(r['cant'] or 0, 8), 'desp': rnd(r['desp'] or 0, 6)})
        n_comp += 1
    partidas.append({'cod': code, 'item': it['item'], 'rubro': RUBRO_NOMBRES.get(int(it['item']), ''),
                     'desc': it['desc'], 'ud': it['ud'], 'cant': it['cant'], 'route': it['route'],
                     'n_comp': n_comp, 'tmat': it['tgt_mat'], 'tmo': it['tgt_mo'], 'teq': it['tgt_eq']})

# ====================== Escribir workbook ======================
wb = openpyxl.Workbook()
HDR = Font(bold=True, color='FFFFFF', name='Aptos Narrow')
HFILL = PatternFill('solid', fgColor='1F4E78')
BLUE = Font(color='FF0000FF', name='Aptos Narrow')   # input
GREEN = Font(color='FF008000', name='Aptos Narrow')  # formula
BLACK = Font(color='FF000000', name='Aptos Narrow')  # dato
MONEY = '#,##0.00'

def setrow_hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h); c.font = HDR; c.fill = HFILL
        c.alignment = Alignment(horizontal='center', wrap_text=True)

def freeze(ws, cell):
    ws.freeze_panes = cell
    ws.sheet_view.showGridLines = False

# ---- CONFIG ----
ws = wb.active; ws.title = 'CONFIG'
ws['A1'] = 'CONFIGURACIÓN DEL APU UNIFICADO — SVD 4140'; ws['A1'].font = Font(bold=True, size=13, color='1F3864')
cfg = [('Parámetro', 'Valor', 'Notas'),
       ('Obra', 'El Salvador 4140 (SVD)', ''),
       ('Mes de referencia', 'Septiembre', 'Mes de precios del APU 09-24'),
       ('Año de referencia', 2024, ''),
       ('Versión', 'v1.0 — APU Unificado SVD', 'Generado desde El Salvador_Pres 04'),
       ('GG%', 0, '0 = se aplica en presupuesto, no en APU'),
       ('BB%', 0, ''),
       ('Coef. Cargas MO', 1, 'Jornal P.MO ya es costo total/día'),
       ('IVA', 0.105, 'Obra civil 10,5%')]
for i, (a, b, c) in enumerate(cfg, 3):
    ws.cell(i, 1, a); ws.cell(i, 2, b); ws.cell(i, 3, c)
    if i == 3:
        for j in (1, 2, 3):
            ws.cell(i, j).font = HDR; ws.cell(i, j).fill = HFILL
    else:
        ws.cell(i, 2).font = BLUE
ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 26; ws.column_dimensions['C'].width = 40
# nombre de celda para coef cargas (C22 en ref); aca B10
COEF_CELL = 'CONFIG!$B$10'

# ---- MATERIALES ----
ws = wb.create_sheet('MATERIALES')
setrow_hdr(ws, 1, ['Código', 'Descripción', 'Unidad', 'Precio $', 'Categoría'])
mats_sorted = sorted(materiales.values(), key=lambda m: m['cod'])
for i, m in enumerate(mats_sorted, 2):
    ws.cell(i, 1, m['cod']).font = BLACK
    ws.cell(i, 2, m['desc']).font = BLACK
    ws.cell(i, 3, m['ud']).font = BLACK
    c = ws.cell(i, 4, m['precio']); c.font = BLUE; c.number_format = MONEY
    ws.cell(i, 5, m['cat']).font = BLACK
for col, w in zip('ABCDE', (12, 44, 8, 14, 12)):
    ws.column_dimensions[col].width = w
freeze(ws, 'A2')

# ---- MANO_DE_OBRA ----
ws = wb.create_sheet('MANO_DE_OBRA')
setrow_hdr(ws, 1, ['Código', 'Descripción', 'Salario/día $', 'Coef. Cargas', 'Costo Jornal $'])
mo_sorted = sorted(mano_obra.items(), key=lambda kv: kv[0])
for i, (cod, v) in enumerate(mo_sorted, 2):
    ws.cell(i, 1, cod).font = BLACK
    ws.cell(i, 2, v['desc']).font = BLACK
    c = ws.cell(i, 3, v['jornal']); c.font = BLUE; c.number_format = MONEY
    ws.cell(i, 4, f'={COEF_CELL}').font = GREEN
    e = ws.cell(i, 5, f'=C{i}*D{i}'); e.font = GREEN; e.number_format = MONEY
for col, w in zip('ABCDE', (12, 34, 14, 12, 14)):
    ws.column_dimensions[col].width = w
freeze(ws, 'A2')

# ---- EQUIPOS ----
ws = wb.create_sheet('EQUIPOS')
setrow_hdr(ws, 1, ['Código', 'Descripción', 'Precio $/día', 'Categoría'])
eq_sorted = sorted(equipos.values(), key=lambda e: e['cod'])
for i, e in enumerate(eq_sorted, 2):
    ws.cell(i, 1, e['cod']).font = BLACK
    ws.cell(i, 2, e['desc']).font = BLACK
    c = ws.cell(i, 3, e['precio']); c.font = BLUE; c.number_format = MONEY
    ws.cell(i, 4, 'Equipo').font = BLACK
for col, w in zip('ABCD', (12, 40, 14, 14)):
    ws.column_dimensions[col].width = w
freeze(ws, 'A2')

# ---- SUBCONTRATOS_PRY ----
ws = wb.create_sheet('SUBCONTRATOS_PRY')
setrow_hdr(ws, 1, ['Código', 'Descripción', 'Ud', 'Precio Unitario $', 'Categoría'])
sub_sorted = sorted(subc.values(), key=lambda s: s['cod'])
for i, s in enumerate(sub_sorted, 2):
    ws.cell(i, 1, s['cod']).font = BLACK
    ws.cell(i, 2, s['desc']).font = BLACK
    ws.cell(i, 3, s['ud']).font = BLACK
    c = ws.cell(i, 4, s['precio']); c.font = BLUE; c.number_format = MONEY
    ws.cell(i, 5, 'Subcontrato').font = BLACK
for col, w in zip('ABCDE', (12, 44, 8, 16, 14)):
    ws.column_dimensions[col].width = w
freeze(ws, 'A2')

# ---- PPTO_GENERADOR (layout de referencia GDR/CH) ----
# A=# B=Rubro C=Desc D=Ud E=Cant | F=MAT/ud G=MO/ud H=EQ/ud I=CD/ud (de PARTIDAS) |
# J=MAT total K=MO total L=EQ total M=Costo Total (=unit x Cant) | Q=Código (link XLOOKUP).
ws = wb.create_sheet('PPTO_GENERADOR')
setrow_hdr(ws, 1, ['#', 'Rubro', 'Descripción', 'Ud.', 'Cant.', 'MAT/ud', 'MO/ud', 'EQ/ud',
                   'CD/ud', 'MAT total', 'MO total', 'EQ total', 'Costo Total', '', '', '',
                   'Código'])
for i, p in enumerate(partidas, 2):
    q = f'Q{i}'
    ws.cell(i, 1, p['item']).font = BLACK
    ws.cell(i, 2, p['rubro']).font = BLACK
    ws.cell(i, 3, p['desc']).font = BLACK
    ws.cell(i, 4, p['ud']).font = BLACK
    c = ws.cell(i, 5, p['cant']); c.font = BLUE; c.number_format = '#,##0.00'
    # MAT/ud, MO/ud, EQ/ud, CD/ud desde PARTIDAS (cols M/N/O/P = 13/14/15/16) por código (col Q)
    ws.cell(i, 6, f'=IFERROR(VLOOKUP({q},PARTIDAS!$A:$P,13,FALSE),0)').font = GREEN
    ws.cell(i, 7, f'=IFERROR(VLOOKUP({q},PARTIDAS!$A:$P,14,FALSE),0)').font = GREEN
    ws.cell(i, 8, f'=IFERROR(VLOOKUP({q},PARTIDAS!$A:$P,15,FALSE),0)').font = GREEN
    ws.cell(i, 9, f'=IFERROR(VLOOKUP({q},PARTIDAS!$A:$P,16,FALSE),0)').font = GREEN
    # totales = unit x Cant
    ws.cell(i, 10, f'=F{i}*$E{i}').font = GREEN
    ws.cell(i, 11, f'=G{i}*$E{i}').font = GREEN
    ws.cell(i, 12, f'=H{i}*$E{i}').font = GREEN
    ws.cell(i, 13, f'=J{i}+K{i}+L{i}').font = GREEN
    ws.cell(i, 17, p['cod']).font = BLACK  # col Q = codigo (para XLOOKUP de COMPOSICIÓN)
    for col in 'FGHIJKLM':
        ws[f'{col}{i}'].number_format = MONEY
# fila TOTAL (verificacion: suma de todas las tareas)
tr = len(partidas) + 2
ws.cell(tr, 3, 'TOTAL OBRA').font = Font(bold=True, name='Aptos Narrow')
for col_l, col_i in (('J', 10), ('K', 11), ('L', 12), ('M', 13)):
    cc = ws.cell(tr, col_i, f'=SUM({col_l}2:{col_l}{tr - 1})')
    cc.font = Font(bold=True, color='FF008000', name='Aptos Narrow'); cc.number_format = MONEY
for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'],
                  (7, 26, 40, 6, 10, 13, 13, 12, 13, 15, 15, 14, 15)):
    ws.column_dimensions[col].width = w
ws.column_dimensions['Q'].width = 12
freeze(ws, 'C2')

# ---- COMPOSICION ----
ws = wb.create_sheet('COMPOSICIÓN')
setrow_hdr(ws, 1, ['Partida', 'Tipo', 'Subtipo', 'Cod_Insumo', 'Descripción', 'Unidad',
                   'Cant', '%Desp', 'Precio', 'Costo', 'Clave', 'Count', 'Clave2',
                   'Rend. Part.', 'Cant MO/ALB Unit', 'Cant. Ej. Part.'])
TBL = {'MAT': ('MATERIALES', 4), 'EQ': ('EQUIPOS', 3)}  # (hoja, col_precio); desc siempre col 2
for i, cr in enumerate(comp_rows, 2):
    tipo, sub = cr['tipo'], cr['subtipo']
    if tipo == 'MAT':
        tbl, pcol = 'MATERIALES', 4
    elif tipo == 'EQ':
        tbl, pcol = 'EQUIPOS', 3
    elif sub == 'ALB':
        tbl, pcol = 'MANO_DE_OBRA', 3   # Salario/día
    else:
        tbl, pcol = 'SUBCONTRATOS_PRY', 4
    ws.cell(i, 1, cr['partida']).font = BLACK
    ws.cell(i, 2, tipo).font = BLACK
    ws.cell(i, 3, sub).font = BLACK
    ws.cell(i, 4, cr['cod']).font = BLACK
    ws.cell(i, 5, f'=IFERROR(VLOOKUP(D{i},{tbl}!A:B,2,FALSE),"?")').font = GREEN
    ws.cell(i, 6, cr['ud']).font = BLACK
    g = ws.cell(i, 7, cr['cant']); g.font = BLUE; g.number_format = '0.######'
    h = ws.cell(i, 8, cr['desp']); h.font = BLUE; h.number_format = '0.####'
    ws.cell(i, 9, f'=IFERROR(VLOOKUP(D{i},{tbl}!A:{chr(64+pcol)},{pcol},FALSE),0)').font = GREEN
    j = ws.cell(i, 10, f'=G{i}*I{i}*(1+H{i})'); j.font = GREEN; j.number_format = MONEY
    ws.cell(i, 11, f'=A{i}&"|"&B{i}').font = GREEN
    ws.cell(i, 12, f'=COUNTIFS($K$2:K{i},K{i})').font = GREEN
    ws.cell(i, 13, f'=K{i}&"|"&L{i}').font = GREEN
    ws.cell(i, 14, f'=IFERROR(VLOOKUP(A{i},PARTIDAS!A:F,6,FALSE),0)').font = GREEN
    ws.cell(i, 15, f'=IF(C{i}="ALB",IF(N{i}=0,0,G{i}/N{i}),0)').font = GREEN
    ws.cell(i, 16, f'=IFERROR(_xlfn.XLOOKUP(A{i},PPTO_GENERADOR!$Q:$Q,PPTO_GENERADOR!$E:$E,"-",0),"-")').font = GREEN
for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P'],
                  (11, 6, 7, 12, 40, 8, 11, 8, 13, 14, 16, 7, 18, 11, 14, 13)):
    ws.column_dimensions[col].width = w
freeze(ws, 'A2')
NCOMP = len(comp_rows) + 1  # ultima fila

# ---- PARTIDAS ----
ws = wb.create_sheet('PARTIDAS')
setrow_hdr(ws, 1, ['Código', 'Ítem', 'Rubro', 'Descripción', 'Ud.', 'Rend.', 'Cant Cons.',
                   '% Cons. Desp.', 'Grado Dif.', 'MAT (sum/ud)', 'MO (sum/día)', 'EQ (sum/día)',
                   'MAT/ud', 'MO/ud', 'EQ/ud', 'CD/ud', '# Comp.', 'Cant Obra', 'Fuente',
                   'Orig MAT', 'Orig MO', 'Orig EQ', 'Δ% CD'])
CR = f'COMPOSICIÓN!$A$2:$A${NCOMP}'
CJ = f'COMPOSICIÓN!$J$2:$J${NCOMP}'
CB = f'COMPOSICIÓN!$B$2:$B${NCOMP}'
for i, p in enumerate(partidas, 2):
    ws.cell(i, 1, p['cod']).font = BLACK
    ws.cell(i, 2, p['item']).font = BLACK
    ws.cell(i, 3, p['rubro']).font = BLACK
    ws.cell(i, 4, p['desc']).font = BLACK
    ws.cell(i, 5, p['ud']).font = BLACK
    ws.cell(i, 6, 1).font = BLUE          # Rend=1 (composicion ya por-unidad)
    ws.cell(i, 7, 0).font = BLUE
    ws.cell(i, 8, 0).font = BLUE
    ws.cell(i, 9, 1).font = BLUE          # Grado
    ws.cell(i, 10, f'=SUMIFS({CJ},{CR},A{i},{CB},"MAT")').font = GREEN
    ws.cell(i, 11, f'=SUMIFS({CJ},{CR},A{i},{CB},"MO")*I{i}').font = GREEN
    ws.cell(i, 12, f'=SUMIFS({CJ},{CR},A{i},{CB},"EQ")').font = GREEN
    ws.cell(i, 13, f'=J{i}').font = GREEN
    ws.cell(i, 14, f'=IF(F{i}=0,0,K{i}/F{i})').font = GREEN
    ws.cell(i, 15, f'=IF(F{i}=0,0,L{i}/F{i})').font = GREEN
    ws.cell(i, 16, f'=M{i}+N{i}+O{i}').font = GREEN
    ws.cell(i, 17, f'=COUNTIF(COMPOSICIÓN!$A:$A,A{i})').font = GREEN
    c = ws.cell(i, 18, p['cant']); c.font = BLACK; c.number_format = '#,##0.00'
    ws.cell(i, 19, p['route']).font = BLACK
    for col, val in ((20, p['tmat']), (21, p['tmo']), (22, p['teq'])):
        cc = ws.cell(i, col, rnd(val)); cc.font = BLACK; cc.number_format = MONEY
    ws.cell(i, 23, f'=IFERROR(P{i}/(T{i}+U{i}+V{i})-1,0)').font = GREEN
    ws.cell(i, 23).number_format = '0.0%'
    for col in 'JKLMNOP':
        ws[f'{col}{i}'].number_format = MONEY
for col, w in zip(['A', 'B', 'C', 'D', 'E'], (11, 7, 26, 40, 6)):
    ws.column_dimensions[col].width = w
freeze(ws, 'D2')

wb.save(OUT)

# ---------- Reporte ----------
lines = [f'Generado: {OUT}',
         f'PARTIDAS: {len(partidas)}  |  COMPOSICIÓN filas: {len(comp_rows)}',
         f'MATERIALES: {len(materiales)}  MANO_DE_OBRA: {len(mano_obra)}  '
         f'EQUIPOS: {len(equipos)}  SUBCONTRATOS_PRY: {len(subc)}',
         'Categorías MO: ' + ', '.join(f'{c}={v["desc"]}' for c, v in sorted(mano_obra.items()))]
open('scripts/_apu_unif_gen_report.txt', 'w', encoding='utf-8').write('\n'.join(lines))
print('OK ->', OUT)
print('\n'.join(lines))
