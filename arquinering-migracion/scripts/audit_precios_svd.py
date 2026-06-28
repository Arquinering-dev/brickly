"""
AUDITORÍA (solo lectura) — Punto 7: validación cruzada de precios de
APU_Unificado_SVD4140_v1.xlsx contra la referencia mensual CORRECTA:
  ARQING - APU 09-24.xlsx  (la fecha de cotización de SVD 4140)
MAT (tabla Materiales), MO (P.MO 09.24 del presupuesto) y EQ (Rend. Equipos).
NO modifica nada.
"""
import openpyxl, warnings, sys
warnings.filterwarnings('ignore'); sys.stdout.reconfigure(encoding='utf-8')

FILE = 'archivos/output/APU_Unificado_SVD4140_v1.xlsx'
APU0924 = 'archivos/fuente/ARQING - APU 09-24.xlsx'
PRES = 'archivos/fuente/El Salvador_Pres 04 (sin pintura).xlsx'

def norm(s):
    return ' '.join(str(s).split()).strip().lower() if s is not None else ''

wv = openpyxl.load_workbook(FILE, data_only=True)

# ---- MAT: precios del archivo vs APU 09-24 Materiales ----
fmat = {}
for r in range(2, wv['MATERIALES'].max_row + 1):
    d = norm(wv['MATERIALES'].cell(r, 2).value); p = wv['MATERIALES'].cell(r, 4).value
    if d and isinstance(p, (int, float)):
        fmat[d] = round(p, 2)
amat = {}
wa = openpyxl.load_workbook(APU0924, data_only=True)['Materiales']
for r in range(8, wa.max_row + 1):
    d = norm(wa.cell(r, 2).value); p = wa.cell(r, 4).value
    if d and isinstance(p, (int, float)):
        amat.setdefault(d, round(p, 2))
mat_match = mat_miss = 0; mat_bad = []
for d, p in fmat.items():
    if d in amat:
        if abs(p - amat[d]) < 0.5:
            mat_match += 1
        else:
            mat_bad.append((d, p, amat[d]))
    else:
        mat_miss += 1  # material no proviene de la tabla APU (viene de Cotizaciones/aux del presupuesto)

# ---- MO: MANO_DE_OBRA del archivo vs P.MO 09.24 del presupuesto ----
pmo = openpyxl.load_workbook(PRES, data_only=True)['P.MO 09.24']
PMO_B = {4: 'ESPECIALIZADO', 5: 'OFICIAL', 6: 'MEDIO OFICIAL', 7: 'AYUDANTE'}
pmo_sal = {PMO_B[r]: round(pmo.cell(r, 2).value, 2) for r in PMO_B}
mo_rows = []
for r in range(2, wv['MANO_DE_OBRA'].max_row + 1):
    cat = norm(wv['MANO_DE_OBRA'].cell(r, 2).value).upper(); sal = round(wv['MANO_DE_OBRA'].cell(r, 3).value, 2)
    ref = pmo_sal.get(cat.upper())
    ok = ref is not None and abs(sal - ref) < 0.5
    mo_rows.append((wv['MANO_DE_OBRA'].cell(r, 2).value, sal, ref, ok))

# ---- EQ: EQUIPOS del archivo vs APU 09-24 Rend. Equipos (costo referencial col E) ----
re_map = {}
wre = openpyxl.load_workbook(APU0924, data_only=True)['Rend. Equipos']
for r in range(7, wre.max_row + 1):
    d = norm(wre.cell(r, 2).value); c = wre.cell(r, 5).value
    if d and isinstance(c, (int, float)) and c > 0:
        re_map.setdefault(d, round(c, 2))
eq_match = eq_nomatch = eq_notfound = 0; eq_bad = []
for r in range(2, wv['EQUIPOS'].max_row + 1):
    d = norm(wv['EQUIPOS'].cell(r, 2).value); p = wv['EQUIPOS'].cell(r, 3).value
    if not d or not isinstance(p, (int, float)):
        continue
    if d in re_map:
        # precio archivo = G/D (incluye deprec) -> suele ser <= costo ref. Verificar que no exceda
        if p <= re_map[d] * 1.02 + 1:
            eq_match += 1
        else:
            eq_nomatch += 1; eq_bad.append((d, p, re_map[d]))
    else:
        eq_notfound += 1

L = []
L.append('=== PUNTO 7 — Validación de precios vs APU 09-24 (referencia correcta SVD) ===')
L.append(f'MAT: {len(fmat)} en archivo | {mat_match} coinciden con APU 09-24 | '
         f'{len(mat_bad)} difieren | {mat_miss} de otra fuente (Cotizaciones/aux del presupuesto)')
for d, p, a in mat_bad[:10]:
    L.append(f'    DIFf MAT: {d[:34]:34} archivo={p} APU0924={a}')
L.append(f'MO: salarios vs P.MO 09.24 del presupuesto:')
for cat, sal, ref, ok in mo_rows:
    L.append(f'    {("OK " if ok else "DIF")} {cat:16} archivo={sal} P.MO0924={ref}')
L.append(f'EQ: {eq_match} coinciden/≤ ref Rend.Equipos 09-24 | {eq_nomatch} exceden | '
         f'{eq_notfound} no en tabla (catálogo aux/otros)')
for d, p, a in eq_bad[:8]:
    L.append(f'    EXCEDE EQ: {d[:34]:34} archivo={p} refRendEq={a}')
print('\n'.join(L))
