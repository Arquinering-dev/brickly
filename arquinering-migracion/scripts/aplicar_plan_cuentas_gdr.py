"""aplicar_plan_cuentas_gdr.py — Bloque 2 de la migracion de GDR a v8.
Alinea los rubros de GDR 3760 al Plan de Cuentas Tezamat (mapeo Q8, confirmado por
Pedro 2026-06-22). Espeja aplicar_plan_cuentas_ch.py pero con rename CONSCIENTE DE
COLUMNA (GDR nombra sin split MT/MO: col A/B -> "X MT", col C/D -> "X MO").

Value-preserving: solo renombra etiquetas de rubro (texto); ningun costo numerico
se toca. Crea la copia de trabajo desde la referencia (referencia intacta).

Uso: python scripts/aplicar_plan_cuentas_gdr.py
"""
import sys
import shutil
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

SRC = "archivos/referencia/GDR_3760_Resumen_de_Obra_v8.xlsx"
DST = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"

# --- rama 53 OBRA del plan + especiales (identico a CH) ----------------------
CANONICO = [
    ("Preliminares", "53001"), ("Demolición", "53002"), ("Movimiento de Suelos", "53003"),
    ("Hormigón MT", "53004"), ("Homigón MO", "53005"),
    ("Metálica MT", "53006"), ("Metálica MO", "53007"),
    ("Albañilería MT", "53008"), ("Albañilería MO", "53009"),
    ("Durlock MT", "53010"), ("Durlock MO", "53011"), ("Aberturas", "53012"),
    ("Revestimiento MT", "53013"), ("Revestimiento MO", "53014"),
    ("Pintura MT", "53015"), ("Pintura MO", "53016"),
    ("Sanitaria MT", "53017"), ("Sanitaria MO", "53018"),
    ("Eléctrico MT", "53019"), ("Eléctrico MO", "53020"),
    ("Provisiones", "53021"), ("Gastos Generales", "53022"), ("Varios Ferreteria", "53023"),
    ("Termomecánica MT", "53024"), ("Termomecánica MO", "53025"),
    ("Seguridad e Higiene", "53026"), ("Herrería MT", "53027"), ("Herrería MO", "53028"),
    ("Alquiler de Equipos", "53980"), ("Gastos a Reintegrar", "53990"),
    ("Gastos GCBA Construccion", "53999"),
    ("Mov. Variables", "52302"),
    ("Supervisión de Obra MO", "52209"),  # indirecto por naturaleza; queda como rubro de obra visible (decision Pedro 2026-06-22)
]

# --- mapeo Q8 (confirmado) ----------------------------------------------------
# rubros con split: col A/B (materiales) -> MT ; col C/D (mano de obra) -> MO
SPLIT = {
    "Hormigón": ("Hormigón MT", "Homigón MO"),
    "Albañilería": ("Albañilería MT", "Albañilería MO"),
    "Pintura": ("Pintura MT", "Pintura MO"),
    "Revestimiento": ("Revestimiento MT", "Revestimiento MO"),
    "Sanitaria": ("Sanitaria MT", "Sanitaria MO"),
    "Durlock/Yeso": ("Durlock MT", "Durlock MO"),
    "Electricidad": ("Eléctrico MT", "Eléctrico MO"),
}
# rubros sin split: mismo nombre canonico en cualquier columna
SINGLE = {
    "Preliminar": "Preliminares",
    "Agrimensura": "Preliminares",
    "Excavación y Mov. de Suelos": "Movimiento de Suelos",
    "Gastos Generales Obra": "Gastos Generales",
    "Movilidad": "Mov. Variables",
    "Supervisión de Obra": "Supervisión de Obra MO",
    "Seguridad e Higiene": "Seguridad e Higiene",  # ya = plan, no-op
}
# 2_Quincenas col E (horas = mano de obra propia) -> variante MO
QUINC = {
    "Albañilería": "Albañilería MO",
    "Electricidad": "Eléctrico MO",
    "Sanitaria": "Sanitaria MO",
    "Durlock/Yeso": "Durlock MO",
    "Revestimiento": "Revestimiento MO",
}


def map_presup(val, col):
    """col 1=A(MT) 2=B(Prov/MT) 3=C(MO/OTR) 4=D(MO/ALB)."""
    if not isinstance(val, str):
        return val, False
    v = val.strip()
    if v in SINGLE:
        nv = SINGLE[v]
        return nv, nv != val
    if v in SPLIT:
        mt, mo = SPLIT[v]
        nv = mt if col in (1, 2) else mo
        return nv, nv != val
    return val, False  # '-' u otros: intactos


def apply():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    # 1) 1_Presupuesto A/B/C/D -- rename consciente de columna
    p = wb["1_Presupuesto"]
    n_p = 0
    cambios = {}
    for r in range(4, p.max_row + 1):
        for c in (1, 2, 3, 4):
            cell = p.cell(row=r, column=c)
            nv, changed = map_presup(cell.value, c)
            if changed:
                cambios[(cell.value, c)] = nv
                cell.value = nv
                n_p += 1

    # 2) 2_Quincenas col E -- rubros de horas a variante MO
    q = wb["2_Quincenas"]
    n_q = 0
    for r in range(4, q.max_row + 1):
        cell = q.cell(row=r, column=5)
        if isinstance(cell.value, str) and cell.value.strip() in QUINC:
            cell.value = QUINC[cell.value.strip()]
            n_q += 1

    # 3) _Listas = espejo del plan (limpiar A,B,C hasta fila 60)
    L = wb["_Listas"]
    for r in range(1, 61):
        for c in (1, 2, 3):
            L.cell(row=r, column=c).value = None
    L["A1"] = "RUBROS CANÓNICOS (Plan de Cuentas Tezamat — rama 53 OBRA + especiales)"
    L["B1"] = "Código"
    L["C1"] = "TIPOS"
    L["C2"] = "MT"
    L["C3"] = "MO"
    for i, (rub, cod) in enumerate(CANONICO):
        L.cell(row=2 + i, column=1).value = rub
        L.cell(row=2 + i, column=2).value = cod
    last = 1 + len(CANONICO)
    rng = f"$A$2:$A${last}"
    nm = "RUBROS_PLAN"
    try:
        del wb.defined_names[nm]
    except (KeyError, TypeError):
        pass
    wb.defined_names[nm] = DefinedName(nm, attr_text=f"_Listas!{rng}")

    # 4) sanity check (rubros de tarea ∈ plan)
    pmax, qmax = p.max_row, q.max_row
    L["E1"] = "VALIDACIÓN — rubros de tarea ∈ plan de cuentas"

    def sp(sheet, colrange):
        return "+".join(
            f"SUMPRODUCT(('{sheet}'!{cl}4:{cl}{mx}<>\"\")*('{sheet}'!{cl}4:{cl}{mx}<>\"-\")"
            f"*ISNA(MATCH('{sheet}'!{cl}4:{cl}{mx},{rng},0)))"
            for cl, mx in colrange)
    L["E2"] = "Rubros fuera de plan — 1_Presupuesto (A/B/C/D)"
    L["F2"] = "=" + sp("1_Presupuesto", [("A", pmax), ("B", pmax), ("C", pmax), ("D", pmax)])
    L["G2"] = '=IF(F2=0,"✓","⚠")'
    L["E3"] = "Rubros fuera de plan — 2_Quincenas (E)"
    L["F3"] = "=" + sp("2_Quincenas", [("E", qmax)])
    L["G3"] = '=IF(F3=0,"✓","⚠")'

    # 5) dropdowns
    def add_dv(ws, sqref):
        dv = DataValidation(type="list", formula1=nm, allow_blank=True, showErrorMessage=False)
        dv.error = "Rubro fuera del plan de cuentas"
        dv.prompt = "Elegí un rubro del plan de cuentas"
        ws.add_data_validation(dv)
        dv.add(sqref)
    add_dv(p, f"A4:D{pmax}")
    add_dv(q, f"E4:E{qmax}")

    wb.save(DST)
    print(f"✓ Bloque 2 aplicado → {DST}")
    print(f"  · 1_Presupuesto: {n_p} celdas renombradas (A/B/C/D)")
    print(f"  · 2_Quincenas: {n_q} celdas (col E → MO)")
    print(f"  · _Listas: {len(CANONICO)} rubros canónicos (A2:A{last}); named range RUBROS_PLAN")
    print(f"  · dropdowns 1_Presupuesto!A4:D{pmax}, 2_Quincenas!E4:E{qmax}; sanity F2/F3")
    print("\n  Renames aplicados (valor_origen[col] → canónico):")
    for (orig, col), nv in sorted(cambios.items()):
        print(f"    {orig!r:34} [{'ABCD'[col-1]}] → {nv!r}")


if __name__ == "__main__":
    apply()
