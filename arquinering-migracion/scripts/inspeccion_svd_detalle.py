"""Inspeccion detallada de hojas fuente clave del SVD legacy. Solo lectura."""
import openpyxl
from openpyxl.utils import get_column_letter

RESUMEN = "archivos/fuente/SVD 4140 - Resumen de Obra.xlsx"
wb = openpyxl.load_workbook(RESUMEN, data_only=False)


def dump(sheet, r0, r1, c0, c1, label=""):
    ws = wb[sheet]
    print(f"\n--- [{sheet}] {label}  filas {r0}-{r1}, cols {get_column_letter(c0)}-{get_column_letter(c1)} ---")
    for r in range(r0, r1 + 1):
        cells = []
        for c in range(c0, c1 + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                cells.append(f"{get_column_letter(c)}{r}={repr(v)[:34]}")
        if cells:
            print("  " + " | ".join(cells))


# Listas: rubros y tipos
dump("Listas", 1, 28, 1, 3, "RUBROS / tipos")

# Pto. Vta. header + primeras filas de datos
dump("Pto. Vta.", 4, 12, 1, 16, "header presupuesto venta")
dump("Pto. Vta.", 50, 60, 1, 16, "muestra filas")

# Pto. Costos header
dump("Pto. Costos", 4, 12, 1, 18, "header presupuesto costo")

# Contratistas
dump("Contratistas", 1, 20, 1, 11, "subcontratos")

# GGBB header
dump("GGBB", 2, 12, 1, 16, "gastos generales y beneficio")

# ADICIONALES
dump("ADICIONALES", 1, 13, 1, 25, "adicionales (OC3?)")

# Facturacion (ingresos/OC)
dump("Facturacion", 1, 20, 1, 16, "OCs / facturacion")

# Resumen Pto #1 y #2 (cabecera de cada presupuesto)
dump("Resumen Pto #1", 3, 18, 2, 16, "cabecera Pto 1")
dump("Resumen Pto #2", 3, 18, 2, 16, "cabecera Pto 2")

# Indice CAC
dump("Indice CAC", 5, 30, 1, 4, "CAC")
