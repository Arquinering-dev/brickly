"""reconciliar_plan_cuentas_gdr.py — Cruza el Plan de Cuentas Tezamat (AING) contra
los rubros usados en GDR 3760 (1_Presupuesto cols A/B/C/D + _Listas). Espeja a
reconciliar_plan_cuentas.py (CH) para que el criterio sea idéntico.

Solo lectura. Genera docs/Reconciliacion_Plan_Cuentas_GDR3760.md (reproducible).
"""
import sys
from collections import Counter, defaultdict
import openpyxl

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PLAN = "archivos/fuente/AING - Plan de Cuentas.xlsx"
GDR = "archivos/referencia/GDR_3760_Resumen_de_Obra_v8.xlsx"
CH = "archivos/output/CH_2171_Resumen_de_Obra_v8_11.xlsx"
OUT = "docs/Reconciliacion_Plan_Cuentas_GDR3760.md"


def norm(s):
    return str(s or "").strip().lower()


def load_plan():
    ws = openpyxl.load_workbook(PLAN, data_only=True)["Hoja1"]
    imput = {}          # code -> desc (solo imputables, col C == "S")
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(row=r, column=1).value or "").strip()
        desc = ws.cell(row=r, column=2).value
        if ws.cell(row=r, column=3).value == "S" and code:
            imput[code] = desc
    return imput


def gdr_presup_rubros():
    """Rubros distintos en 1_Presupuesto cols A/B/C/D (header fila 3, datos desde 4)."""
    ws = openpyxl.load_workbook(GDR, data_only=True)["1_Presupuesto"]
    by_col = {1: Counter(), 2: Counter(), 3: Counter(), 4: Counter()}
    allc = Counter()
    for r in range(4, ws.max_row + 1):
        for c in (1, 2, 3, 4):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip() not in ("-", "", " "):
                by_col[c][str(v).strip()] += 1
                allc[str(v).strip()] += 1
    return allc, by_col


def gdr_listas():
    ws = openpyxl.load_workbook(GDR, data_only=True)["_Listas"]
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip():
            out.append(str(v).strip())
    return out


def ch_listas_plan():
    """Los rubros que CH ya alineó al plan (col A nombre, col B código)."""
    ws = openpyxl.load_workbook(CH, data_only=True)["_Listas"]
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        code = ws.cell(row=r, column=2).value
        if name and str(name).strip():
            out.append((str(name).strip(), str(code).strip() if code else ""))
    return out


def main():
    imput = load_plan()
    plan_by_name = {norm(d): c for c, d in imput.items()}
    obra_plan = {c: d for c, d in imput.items() if c[:2] == "53"}  # rama OBRA

    allc, by_col = gdr_presup_rubros()
    listas_gdr = gdr_listas()
    listas_ch = ch_listas_plan()

    L = []
    L.append("# Reconciliación Plan de Cuentas Tezamat ↔ GDR 3760")
    L.append("")
    L.append("> Generado por `scripts/reconciliar_plan_cuentas_gdr.py`. Plan: "
             "`AING - Plan de Cuentas.xlsx`; obra: `GDR_3760_Resumen_de_Obra_v8.xlsx` (referencia).")
    L.append("> Espeja el criterio usado en CH (`reconciliar_plan_cuentas.py`).")
    L.append("")
    L.append("Rama **53 OBRA** = rubros de obra (split MT/MO en el nombre). "
             "Ramas **50/51/52** = indirectos. Cruce por NOMBRE normalizado.")
    L.append("")

    # 1) rubros de GDR 1_Presupuesto vs plan (por nombre)
    L.append("## 1. Rubros de GDR `1_Presupuesto` (cols A/B/C/D) ↔ plan")
    L.append("")
    L.append("| Rubro en GDR | n | Cols | Match plan | Código | Rama |")
    L.append("|---|---|---|---|---|---|")
    sin_match = []
    for rub in sorted(allc):
        n = allc[rub]
        cols = "".join(letter for letter, cc in zip("ABCD", (1, 2, 3, 4)) if by_col[cc].get(rub))
        code = plan_by_name.get(norm(rub))
        if code:
            L.append(f"| {rub} | {n} | {cols} | ✓ | {code} | {code[:2]} |")
        else:
            sin_match.append(rub)
            L.append(f"| {rub} | {n} | {cols} | ✗ | — | **revisar** |")
    L.append("")
    L.append(f"**Sin match exacto ({len(sin_match)}/{len(allc)}):** {', '.join(sin_match) or '—'}")
    L.append("")

    # 2) _Listas de GDR vs plan
    L.append("## 2. `_Listas` de GDR ↔ plan")
    L.append("")
    L.append("| Rubro _Listas GDR | Match plan | Código |")
    L.append("|---|---|---|")
    listas_sin = []
    for rub in listas_gdr:
        code = plan_by_name.get(norm(rub))
        if code:
            L.append(f"| {rub} | ✓ | {code} |")
        else:
            listas_sin.append(rub)
            L.append(f"| {rub} | ✗ | — |")
    L.append("")
    L.append(f"**Sin match ({len(listas_sin)}/{len(listas_gdr)}):** {', '.join(listas_sin) or '—'}")
    L.append("")

    # 3) cómo lo resolvió CH (referencia de mapeo ya aprobado)
    L.append("## 3. `_Listas` de CH ya alineado al plan (referencia / objetivo)")
    L.append("")
    L.append("| Rubro canónico CH | Código |")
    L.append("|---|---|")
    for name, code in listas_ch:
        L.append(f"| {name} | {code} |")
    L.append("")

    # 4) rama 53 OBRA del plan (universo de rubros válidos)
    L.append("## 4. Rama 53 OBRA del plan (universo de rubros de obra)")
    L.append("")
    L.append("| Código | Desc |")
    L.append("|---|---|")
    for c, d in sorted(obra_plan.items()):
        L.append(f"| {c} | {d} |")
    L.append("")

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(L))
    print(f"✓ {OUT}")
    print(f"  GDR 1_Presupuesto: {len(allc)} rubros distintos, {len(sin_match)} sin match exacto")
    print(f"  GDR _Listas: {len(listas_gdr)} rubros, {len(listas_sin)} sin match")
    print()
    print("  === RUBROS GDR 1_Presupuesto SIN MATCH EXACTO ===")
    for rub in sin_match:
        cols = "".join(letter for letter, cc in zip("ABCD", (1, 2, 3, 4)) if by_col[cc].get(rub))
        print(f"    [{cols:4}] {rub}  (n={allc[rub]})")
    print()
    print("  === RUBROS GDR 1_Presupuesto CON MATCH ===")
    for rub in sorted(allc):
        if plan_by_name.get(norm(rub)):
            print(f"    ✓ {rub} → {plan_by_name[norm(rub)]}")


if __name__ == "__main__":
    main()
