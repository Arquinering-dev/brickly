"""
migracion_egresos_v4.py — Refinamientos sobre v8_3 (movimientos ing+egr)
Arquinering S.R.L. — CH 2171

Aplica los 3 pedidos de Pedro sobre la versión de egresos:
  1) Re-taggear los pagos de subcontrato existentes en Observaciones (CH-SC-00X TIPO)
     y guardar la Observación ORIGINAL en una columna nueva (demo antes/después).
     El CAC embebido (V) se desagrega como fila CAC separada (modelo nuevo).
  2) Ingresos y egresos conviven en la hoja -> renombrar 2_Gastos a 2_Movimientos.
     Cash_Flow toma los ingresos (cobros) desde la hoja. Nueva col Tipo Mov (INGRESO/EGRESO).
  3) Columna "En 3_Control_Ppto" (SÍ/NO) para identificar movimientos no contabilizados.

Clona v8_3 -> v8_4. Un solo save. Luego excel_recalc.py + recalc.py.
"""
import os
import shutil
import openpyxl
from openpyxl.comments import Comment

SRC = "archivos/output/CH_2171_Resumen_de_Obra_v8_3_egresos.xlsx"
DST = "archivos/output/CH_2171_Resumen_de_Obra_v8_4.xlsx"
OLD, NEW = "2_Gastos", "2_Movimientos"

# fila -> nº de SC (índices estables: filas no se movieron desde v8_2)
SC_MAP = {11: "001", 18: "002", 24: "003", 42: "002", 60: "001",
          74: "004", 83: "003", 90: "003", 106: "003", 107: "003"}
# filas con CAC embebido (Monto CAC original) -> desagregar como fila CAC aparte
CAC_SPLIT = {83: 492422.92, 107: 652746.66}

# columnas de meses en Cash_Flow: col 2 (B)=2026-01 ... col 19 (S)=2027-06
def mes_de_col(c):
    n = c - 2
    return 2026 + n // 12, n % 12 + 1


def main():
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    log = []

    # ====================================================================
    # 1) Renombrar hoja + actualizar TODAS las refs '2_Gastos'! -> '2_Movimientos'!
    #    (token exacto con !, no rompe '2_Gastos_DirInd'!)
    # ====================================================================
    wb[OLD].title = NEW
    tok_old, tok_new = f"'{OLD}'!", f"'{NEW}'!"
    nref = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and tok_old in v:
                    c.value = v.replace(tok_old, tok_new)
                    nref += 1
    log.append(f"Hoja {OLD} -> {NEW}; {nref} refs de fórmula actualizadas")

    mv = wb[NEW]
    LAST = 0
    for r in range(2, mv.max_row + 1):
        if mv.cell(r, 4).value is not None or mv.cell(r, 12).value is not None:
            LAST = r

    # ====================================================================
    # 2) Columnas nuevas X(24) En 3_Control_Ppto, Y(25) Tipo Mov, Z(26) Obs original
    # ====================================================================
    mv.cell(1, 24).value = "En 3_Control_Ppto"
    mv.cell(1, 25).value = "Tipo Mov"
    mv.cell(1, 26).value = "Observaciones original"

    def set_flags(r):
        mv.cell(r, 24).value = (
            f"=IF(ISNUMBER(MATCH(E{r},'3_Control_Ppto'!$A$4:$A$35,0)),\"SÍ\",\"NO\")"
        )
        mv.cell(r, 25).value = f'=IF(LEFT(D{r}&"",1)="4","INGRESO","EGRESO")'

    # guardar Observación original (Z) ANTES de re-taggear H; setear flags
    for r in range(2, LAST + 1):
        if mv.cell(r, 5).value in (None, "") and mv.cell(r, 12).value in (None, ""):
            continue
        mv.cell(r, 26).value = mv.cell(r, 8).value  # Z = H original
        set_flags(r)
    log.append("2_Movimientos: cols X (En 3_Control_Ppto SÍ/NO), Y (Tipo Mov), Z (Obs original)")

    # ====================================================================
    # 3) Re-tag de subcontratos en Observaciones (H) + desagregado de CAC
    # ====================================================================
    append_rows = []
    for r, scid in SC_MAP.items():
        full = f"CH-SC-{scid}"
        mv.cell(r, 8).value = f"{full} AVANCE"
        if r in CAC_SPLIT:
            cac = CAC_SPLIT[r]
            l_old = mv.cell(r, 12).value
            mv.cell(r, 12).value = round(l_old - cac, 2)   # avance = base (sin CAC)
            append_rows.append((r, full, cac))
    log.append(f"Re-tag {len(SC_MAP)} pagos SC -> 'CH-SC-00X AVANCE'; "
               f"{len(CAC_SPLIT)} filas CAC desagregadas")

    # append filas CAC al final (sin insertar -> no desplaza nada)
    # Columnas: C=3 Mes, D=4 Cuenta, E=5 Desc, F=6 Fecha, H=8 Obs, L=12 Debe,
    #           M=13 Haber, R=18 Saldo Inicial, S=19 Descontado, T=20 Real, Z=26 Obs orig
    nr = LAST
    for parent, full, cac in append_rows:
        nr += 1
        mv.cell(nr, 3).value = f"=EOMONTH(F{nr},-1)+1"           # C Mes
        mv.cell(nr, 4).value = mv.cell(parent, 4).value          # D Cuenta
        mv.cell(nr, 5).value = mv.cell(parent, 5).value          # E Desc Cuenta
        mv.cell(nr, 6).value = mv.cell(parent, 6).value          # F Fecha
        mv.cell(nr, 8).value = f"{full} CAC"                     # H Observaciones
        mv.cell(nr, 12).value = cac                              # L Debe
        mv.cell(nr, 13).value = 0                                # M Haber
        mv.cell(nr, 18).value = 0                                # R Saldo Inicial
        mv.cell(nr, 19).value = (                                # S Monto Descontado
            f"=(L{nr}-M{nr})*SUMIF('0_Indice_CAC'!$A:$A,C{nr},'0_Indice_CAC'!$D:$D)"
        )
        mv.cell(nr, 20).value = f"=L{nr}-M{nr}"                  # T Monto Real
        mv.cell(nr, 26).value = f"(CAC desagregado de fila {parent})"  # Z
        set_flags(nr)

    # ====================================================================
    # 4) Cash_Flow: ingresos (cobros) desde 2_Movimientos; egresos filtra EGRESO
    # ====================================================================
    cf = wb["3_Cash_Flow"]
    for c in range(2, 20):
        y, m = mes_de_col(c)
        d = f"DATE({y},{m},1)"
        # fila 4 Ingresos (cobros) = -Σ T de movimientos INGRESO del mes
        cf.cell(4, c).value = (
            f"=-SUMIFS('{NEW}'!$T:$T,'{NEW}'!$Y:$Y,\"INGRESO\",'{NEW}'!$C:$C,{d})"
        )
        # fila 5 Ingresos por CAC (placeholder: CAC de ingreso como mov. separado futuro)
        cf.cell(5, c).value = 0
        # fila 6 Egresos = Σ T EGRESO del mes + quincenas
        cf.cell(6, c).value = (
            f"=SUMIFS('{NEW}'!$T:$T,'{NEW}'!$Y:$Y,\"EGRESO\",'{NEW}'!$C:$C,{d})"
            f"+SUMIFS('2_Quincenas'!$N:$N,'2_Quincenas'!$A:$A,{d})"
        )
    cf.cell(4, 1).comment = Comment(
        "Ingresos (cobros) tomados de 2_Movimientos (cuentas serie 4 / Tipo Mov=INGRESO). "
        "Supuesto: los movimientos del Fideicomiso (cta 410222) son cobros.", "migracion")
    cf.cell(5, 1).comment = Comment(
        "CAC de ingresos = 0 (placeholder). Cuando Tezamat cargue el CAC de ingreso como "
        "movimiento separado, sumar por Observaciones '*CAC*' sobre INGRESO.", "migracion")
    log.append("Cash_Flow: ingresos (fila 4) desde 2_Movimientos; egresos (fila 6) filtra EGRESO")

    # ====================================================================
    # 5) 0_CONFIG B45 -> total egresos; Dashboard cobrado/resultado desde movimientos
    # ====================================================================
    cfg = wb["0_CONFIG"]
    cfg.cell(45, 1).value = "Total egresos (2_Movimientos)"
    cfg.cell(45, 2).value = f"=SUMIFS('{NEW}'!$T:$T,'{NEW}'!$Y:$Y,\"EGRESO\")"
    dash = wb["3_Dashboard"]
    dash["C6"].value = f"=-SUMIFS('{NEW}'!$T:$T,'{NEW}'!$Y:$Y,\"INGRESO\")"  # Total cobrado
    dash["C20"].value = "=C6-'3_Control_Ppto'!F36"                            # Resultado acum.
    log.append("0_CONFIG B45 = total egresos; Dashboard C6 (cobrado) y C20 (resultado) desde movimientos")

    # ====================================================================
    # 6) Sanity Control_Ppto E102: egreso sin rubro (excluye ingresos)
    # ====================================================================
    cp = wb["3_Control_Ppto"]
    cp.cell(102, 1).value = "Costo EGRESO sin rubro reconocido (cuenta fuera de plan)"
    cp.cell(102, 5).value = (
        f"=SUMIFS('{NEW}'!$T:$T,'{NEW}'!$X:$X,\"NO\",'{NEW}'!$Y:$Y,\"EGRESO\")"
    )
    cp.cell(102, 13).value = '=IF(E102=0,"🟢","⚠ revisar cuentas Tezamat")'
    log.append("Control_Ppto E102: sanity egreso sin rubro (excluye ingresos)")

    # ====================================================================
    # 7) CHANGELOG append
    # ====================================================================
    ch = wb["CHANGELOG"]
    base = ch.max_row + 2
    extra = [
        ["v8_4 (2026-06-18) — ingresos+egresos conviven en 2_Movimientos"],
        ["CAMBIO", "DETALLE", "MOTIVO"],
        ["Hoja 2_Gastos -> 2_Movimientos", "Refs de fórmula actualizadas en todo el libro",
         "Ingresos y egresos conviven; cashflow sale de acá"],
        ["Col Tipo Mov (Y)", "INGRESO si cuenta serie 4, EGRESO si no", "Separa ing/egr sin umbral mágico"],
        ["Col En 3_Control_Ppto (X)", "SÍ/NO si Desc Cuenta matchea rubro del control",
         "Identifica movimientos no contabilizados para revisar"],
        ["Col Observaciones original (Z)", "Guarda la observación previa (cómo se cargaba antes)",
         "Demo antes/después del proceso Tezamat"],
        ["Re-tag SC en Observaciones", "10 pagos -> 'CH-SC-00X AVANCE'; 2 CAC desagregados como fila aparte",
         "Demostrar el tracking de SC por Observaciones"],
        ["Cash_Flow ingresos", "Fila 4 (cobros) = -Σ T INGRESO por mes; fila 6 egresos filtra EGRESO",
         "Ingresos del cashflow salen de 2_Movimientos"],
        ["Sanity egreso sin rubro", "E102 = SUMIFS(T, X=NO, Tipo=EGRESO)", "Excluye ingresos del flag"],
    ]
    for i, row in enumerate(extra):
        for j, val in enumerate(row, 1):
            ch.cell(base + i, j).value = val

    wb.save(DST)
    print("✅ Guardado:", DST)
    for x in log:
        print("  -", x)


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    main()
