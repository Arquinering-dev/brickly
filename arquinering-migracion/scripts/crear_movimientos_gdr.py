"""Bloque 6 — Crea 2_Movimientos en GDR como ESTRUCTURA PLACEHOLDER (vacía), lista para
recibir el extracto corregido de Tezamat que carga Arquinering (decisión del prompt: no se
migra la 2_Gastos legacy). Layout idéntico a CH (A-R) + fórmulas de parseo mov_id/mov_tipo
sobre la convención `{ID} | {TIPO} | {desc}` (buffer de filas). Cablea 0_CONFIG!B45/B47.

Uso: python scripts/crear_movimientos_gdr.py
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

PATH = "archivos/output/GDR_3760_Resumen_de_Obra_v8_12.xlsx"
BUF = 500  # filas de buffer para las fórmulas de parseo Q/R

HEADERS = [
    "Cuenta", "Desc Cuenta", "Fecha Asiento", "Nº Asiento", "Observaciones",
    "Cliente/Proveedor", "Cód Comprobante", "Nº Comprobante", "Debe", "Haber",
    "Unidad", "Desc Unidad", "Centro de Costo", "Desc Centro Costo", "Saldo Inicial",
    "Observaciones original", "mov_id", "mov_tipo",
]
# parseo de la convención "{ID} | {TIPO} | {desc}" (idéntico a CH, compat Excel 2016)
Q_F = '=IFERROR(TRIM(LEFT($E{r},FIND(" | ",$E{r})-1)),"")'
R_F = ('=IFERROR(TRIM(UPPER(MID($E{r},FIND(" | ",$E{r})+3,'
       'IFERROR(FIND(" | ",$E{r},FIND(" | ",$E{r})+3),LEN($E{r})+1)-FIND(" | ",$E{r})-3))),"")')


def main():
    wb = openpyxl.load_workbook(PATH)
    if "2_Movimientos" in wb.sheetnames:
        del wb["2_Movimientos"]
    ws = wb.create_sheet("2_Movimientos")

    # posicionar después de 1_GGBB (orden CH)
    order = wb.sheetnames
    ws_idx = order.index("2_Movimientos")
    target = order.index("1_GGBB") + 1
    wb.move_sheet("2_Movimientos", offset=target - ws_idx)

    # header
    hf = Font(bold=True, color="FFFFFFFF")
    hfill = PatternFill("solid", fgColor="FF1F4E78")
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(1, j, h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
    # fórmulas de parseo Q/R (buffer, inertes hasta que haya datos en E)
    for r in range(2, BUF + 2):
        ws.cell(r, 17, Q_F.format(r=r))
        ws.cell(r, 18, R_F.format(r=r))
    # chrome
    ws.sheet_properties.tabColor = "FFFFC000"
    ws.freeze_panes = "A2"
    widths = {"A": 10, "B": 22, "C": 13, "D": 9, "E": 34, "F": 22, "G": 13, "H": 14,
              "I": 14, "J": 14, "K": 8, "L": 14, "M": 14, "N": 18, "O": 13, "P": 30,
              "Q": 16, "R": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # cablear 0_CONFIG reconciliación (B45 egresos, B47 diferencia)
    cfg = wb["0_CONFIG"]
    cfg["B45"] = ('=SUMPRODUCT(\'2_Movimientos\'!$I$2:$I$1000,'
                  'N(LEFT(TEXT(\'2_Movimientos\'!$A$2:$A$1000,"0"),1)<>"4"))'
                  '-SUMPRODUCT(\'2_Movimientos\'!$J$2:$J$1000,'
                  'N(LEFT(TEXT(\'2_Movimientos\'!$A$2:$A$1000,"0"),1)<>"4"))')
    cfg["B47"] = "=B45-B46"

    wb.save(PATH)
    print(f"✓ 2_Movimientos creada (placeholder vacío) → {PATH}")
    print(f"  · {len(HEADERS)} columnas (A-R), idénticas a CH; tab naranja; freeze A2")
    print(f"  · fórmulas mov_id/mov_tipo (Q/R) en buffer filas 2:{BUF+1} (inertes sin datos)")
    print(f"  · posicionada tras 1_GGBB · 0_CONFIG!B45 (egresos) + B47 (dif) cableados")
    print(f"  · orden de hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
